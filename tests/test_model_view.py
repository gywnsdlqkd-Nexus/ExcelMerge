# -*- coding: utf-8 -*-
"""모델/뷰 전환 골든 패리티 테스트.

구 QTableWidget populate() 루프의 텍스트/배경색 로직을 오라클로 이식해,
DiffTableModel의 DisplayRole/BackgroundRole이 모든 상태 조합에서 동일한지
검증한다. 추가로 필터 델타, 선택 클램프, 자동 편집 오발동 방지를 확인한다.

실행: QT_QPA_PLATFORM=offscreen python tests/test_model_view.py
"""
import os
import sys
import tempfile

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
# 키 위치 전역 저장(prefs.json)이 실제 %APPDATA%를 오염시키지 않도록 임시 폴더로 격리.
# (컨트롤러 _on_key_col_changed/_on_key_row_changed 가 save_key_prefs 를 호출하므로 필수.)
os.environ["APPDATA"] = tempfile.mkdtemp(prefix="em_test_appdata_")
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QApplication

from excelmerge.diff_model import DiffTableModel, EXTRA_ROWS, EXTRA_COLS
from excelmerge.diff_engine import compute_diff, count_dropped_key_rows
from excelmerge.theme import DIFF_COLORS, EXCLUDED_CELL_BG

# MainWindow는 이제 탭 셸이고 비교 UI/상태는 활성 탭의 DiffView에 있다.
# 이 테스트들은 구 MainWindow API(panel_a/_refresh_tables/_diff_matrix ...)를
# 그대로 검증하므로, MainWindow를 띄우고 활성 DiffView를 돌려주는 헬퍼로 대체한다.
_WINDOWS = []   # MainWindow 참조 유지 — 자식 DiffView가 GC로 파괴되지 않도록.


def _diff_view():
    from excelmerge.main_window import MainWindow
    win = MainWindow()
    _WINDOWS.append(win)
    win.show()   # 전체 위젯 트리 레이아웃 — 크기 의존 검증을 위해
    return win.tabs.currentWidget()


def _wait_diff(win, timeout_ms=5000):
    """_on_loaded/_recompute_diff의 compute_diff가 DiffWorker(백그라운드)로 옮겨져
    비동기가 됐으므로, 테스트가 동기 상태(매트릭스·틀 고정)를 검증하려면 워커 종료를
    기다린 뒤 done 시그널을 처리해야 한다."""
    w = getattr(win, "_diff_worker", None)
    if w is not None:
        w.wait(timeout_ms)
    app = QApplication.instance()
    if app is not None:
        app.processEvents()


def oracle_cell(diff_matrix, which, merged_set, staged, excluded_cols, r, c):
    """구 populate() 핫루프(텍스트/색 결정)를 그대로 이식한 오라클."""
    status, a_val, b_val = diff_matrix[r][c]
    direction = staged.get((r, c))
    if direction == "a_to_b":
        text = a_val
    elif direction == "b_to_a":
        text = b_val
    else:
        text = a_val if which == "a" else b_val
    if c in excluded_cols:
        color = EXCLUDED_CELL_BG   # 제외 열은 회색 배경(v166~)
    elif (r, c) in merged_set:
        color = DIFF_COLORS["merged"]
    elif direction is not None:
        color = DIFF_COLORS["staged"]
    elif status == "added":
        # 신규 색은 값이 있는 쪽 패널에만 (빈 쪽은 흰색=same)
        own = a_val if which == "a" else b_val
        color = DIFF_COLORS["added"] if own != "" else DIFF_COLORS["same"]
    else:
        color = DIFF_COLORS[status]
    return text, color


def dump(model):
    out = []
    for r in range(model.rowCount()):
        row = []
        for c in range(model.columnCount()):
            idx = model.index(r, c)
            text = model.data(idx, Qt.DisplayRole) or ""
            bg = model.data(idx, Qt.BackgroundRole)
            row.append((text, bg.getRgb()[:3] if bg is not None else None))
        out.append(row)
    return out


def make_state():
    # 상태 조합 총망라: same/added/modified + staged 양방향 + merged + excluded
    dm = [
        [("same", "H1", "H1"), ("same", "H2", "H2"), ("modified", "H3", "H3x")],
        [("same", "a", "a"), ("modified", "b1", "b2"), ("added", "", "new")],
        [("modified", "c1", "c2"), ("modified", "d1", "d2"), ("same", "e", "e")],
        [("added", "", "f"), ("modified", "g1", ""), ("modified", "h1", "h2")],
    ]
    meta = [(0, 0), (1, 1), (2, None), (None, 3)]
    staged = {(1, 1): "a_to_b", (2, 0): "b_to_a"}
    merged = {(2, 1)}
    excluded = {2}
    return dm, meta, staged, merged, excluded


def test_golden_roles():
    dm, meta, staged, merged, excluded = make_state()
    for side in ("a", "b"):
        m = DiffTableModel(side)
        m.set_diff_data(dm, meta, staged, merged, excluded)
        assert m.rowCount() == len(dm) + EXTRA_ROWS
        assert m.columnCount() == len(dm[0]) + EXTRA_COLS
        for r in range(len(dm)):
            for c in range(len(dm[0])):
                want_text, want_color = oracle_cell(dm, side, merged, staged, excluded, r, c)
                idx = m.index(r, c)
                got_text = m.data(idx, Qt.DisplayRole) or ""
                got_bg = m.data(idx, Qt.BackgroundRole)
                assert got_text == want_text, f"{side} ({r},{c}) text {got_text!r} != {want_text!r}"
                assert got_bg is not None and got_bg.getRgb() == want_color.getRgb(), \
                    f"{side} ({r},{c}) bg mismatch"
                assert m.data(idx, Qt.TextAlignmentRole) == int(Qt.AlignVCenter | Qt.AlignLeft)
        # 여분 행/열: 텍스트 없음 + 기본 배경(None)
        assert m.data(m.index(len(dm), 0), Qt.DisplayRole) is None
        assert m.data(m.index(len(dm), 0), Qt.BackgroundRole) is None
        assert m.data(m.index(0, len(dm[0])), Qt.BackgroundRole) is None
    print("PASS test_golden_roles")


def test_headers():
    # 헤더 DecorationRole이 QIcon을 생성하므로 QGuiApplication이 먼저 존재해야 한다.
    app = QApplication.instance() or QApplication([])
    dm, meta, staged, merged, excluded = make_state()
    m = DiffTableModel("a")
    m.set_diff_data(dm, meta, staged, merged, excluded)
    m.set_key_col(0)
    # 표시 텍스트는 열 문자만(이모지 제거) — 키/제외 표시는 DecorationRole 아이콘으로.
    assert m.headerData(0, Qt.Horizontal, Qt.DisplayRole) == "A"
    assert m.headerData(0, Qt.Horizontal, Qt.BackgroundRole).getRgb()[:3] == (255, 213, 0)
    assert not m.headerData(0, Qt.Horizontal, Qt.DecorationRole).isNull()   # 키 아이콘
    assert m.headerData(2, Qt.Horizontal, Qt.DisplayRole) == "C"
    assert m.headerData(2, Qt.Horizontal, Qt.ForegroundRole).getRgb()[:3] == (140, 140, 140)
    assert not m.headerData(2, Qt.Horizontal, Qt.DecorationRole).isNull()   # 제외 아이콘
    assert m.headerData(1, Qt.Horizontal, Qt.DisplayRole) == "B"
    assert m.headerData(1, Qt.Horizontal, Qt.DecorationRole) is None        # 일반 열은 아이콘 없음
    assert m.headerData(1, Qt.Horizontal, Qt.BackgroundRole).getRgb()[:3] == (232, 234, 240)
    # 세로: meta 원본 행번호, None이면 "-" (A측: meta[3][0] is None)
    assert m.headerData(0, Qt.Vertical, Qt.DisplayRole) == "1"
    assert m.headerData(3, Qt.Vertical, Qt.DisplayRole) == "-"
    b = DiffTableModel("b")
    b.set_diff_data(dm, meta, staged, merged, excluded)
    assert b.headerData(2, Qt.Vertical, Qt.DisplayRole) == "-"   # meta[2][1] is None
    assert b.headerData(3, Qt.Vertical, Qt.DisplayRole) == "4"
    # 여분 행 라벨은 연속 번호
    assert m.headerData(len(dm), Qt.Vertical, Qt.DisplayRole) == str(len(dm) + 1)
    # 미리보기 모드: 색·아이콘 없는 순수 라벨
    m.set_preview_data([["x", "y"], ["1", "2"]])
    assert m.headerData(0, Qt.Horizontal, Qt.DisplayRole) == "A"
    assert m.headerData(0, Qt.Horizontal, Qt.BackgroundRole) is None
    print("PASS test_headers")


def test_notify_equals_fresh_populate():
    """상태 변형 + notify 후의 모델 스냅샷 == 같은 상태로 새로 populate한 스냅샷."""
    dm, meta, staged, merged, excluded = make_state()
    live = DiffTableModel("b")
    live.set_diff_data(dm, meta, staged, merged, excluded)
    # 변형: 셀 스테이징 추가, 기존 staged 하나 merged로 승격, 편집으로 status 변경
    staged[(3, 2)] = "a_to_b"
    merged.add((1, 1)); del staged[(1, 1)]
    dm[2][0] = ("same", "c2", "c2")
    live.notify_cells({(3, 2), (1, 1), (2, 0)})
    fresh = DiffTableModel("b")
    fresh.set_diff_data(dm, meta, staged, merged, excluded)
    assert dump(live) == dump(fresh), "notify 후 상태가 fresh populate와 다름"
    print("PASS test_notify_equals_fresh_populate")


def test_cell_kind():
    dm, meta, staged, merged, excluded = make_state()
    m = DiffTableModel("a")
    m.set_diff_data(dm, meta, staged, merged, excluded)
    assert m.cell_kind(1, 1) == "staged"
    assert m.cell_kind(2, 1) == "merged"
    assert m.cell_kind(1, 2) == "same"        # excluded 열 → same 취급
    assert m.cell_kind(2, 0) == "staged"
    assert m.cell_kind(1, 0) == "same"
    assert m.cell_kind(3, 0) == "changed"     # added
    assert m.cell_kind(len(dm) + 1, 0) == "same"   # 여분 행
    print("PASS test_cell_kind")


def test_view_behaviors():
    app = QApplication.instance() or QApplication([])
    win = _diff_view()
    dm, meta, staged, merged, excluded = make_state()
    win._diff_matrix = dm
    win._diff_row_meta = meta
    win._staged = staged
    win._merged_cells = merged
    win._excluded_cols = set(excluded)
    win._refresh_tables()
    tbl = win.panel_a.table

    # 1. 필터 델타: 변경행만 보기 토글 시 isRowHidden 벡터가 오라클과 일치.
    #    (틀 고정: 고정 행 0..key_row 는 본체에서 항상 숨김 — 상단 고정 띠에 표시)
    kr = win._key_row
    win._diff_only = True
    win._apply_diff_filter()
    excl = win._excluded_cols
    for r, row in enumerate(dm):
        want_hidden = (r <= kr) or not any(
            st != "same" for c, (st, *_) in enumerate(row) if c not in excl)
        assert tbl.isRowHidden(r) == want_hidden, f"row {r} hidden mismatch"
    assert not tbl._user_row_heights, "_user_row_heights 오염됨"
    win._diff_only = False
    win._apply_diff_filter()
    # 필터 해제해도 고정 행(0..key_row)은 계속 숨김, 나머지는 표시
    for r in range(len(dm)):
        assert tbl.isRowHidden(r) == (r <= kr), f"필터 해제 후 행{r} 상태 오류"

    # 2. Ctrl+Shift+End 클램프: 데이터 영역 밖 선택 없음
    tbl._select_range(0, 0, tbl.rowCount() - 1, tbl.columnCount() - 1)
    sel = tbl.get_selected_cells()
    assert sel, "선택 없음"
    assert max(r for r, _ in sel) == len(dm) - 1
    assert max(c for _, c in sel) == len(dm[0]) - 1

    # 3. 셀 값 표시란은 읽기전용 (직접 편집 기능 제거됨)
    assert win.panel_a.cell_edit.isReadOnly()
    assert not hasattr(win.panel_a, "cell_value_edited")

    # 4. staged 오버라이드 텍스트가 표시/TSV에 반영
    assert tbl.model().display_text(1, 1) == "b1"   # a_to_b → a_val
    assert win.panel_b.table.model().display_text(1, 1) == "b1"

    # 5. 부분 갱신이 선택/스크롤을 보존 (기본 특성 확인)
    tbl._set_current_cell(2, 2)
    app.processEvents()
    before = tbl.get_selected_cells()
    win.panel_a.table.model().notify_cells({(2, 2)})
    app.processEvents()
    assert tbl.get_selected_cells() == before, "notify가 선택을 파괴함"

    win.close()
    print("PASS test_view_behaviors")


def test_header_multiselect_extension():
    """헤더 선택 후 Shift+방향키(한 칸)·Ctrl+Shift+방향키(끝까지) 확장 회귀 테스트.
    수정 전에는 setCurrentIndex의 SelectCurrent가 범위 선택을 1셀로 붕괴시켰다."""
    from PyQt5.QtTest import QTest
    app = QApplication.instance() or QApplication([])
    win = _diff_view()

    # 값은 0~4열/0~7행에만 존재. 5~6열과 8행은 엑셀 유령 셀처럼 빈 값으로 채워
    # '값 기준 마지막' 경계 검증에 쓴다.
    def row_vals(vals):
        return [("same", v, v) for v in vals]
    dm = [row_vals([f"h{c}" for c in range(5)] + ["", ""])]
    for r in range(1, 8):
        dm.append(row_vals([f"a{r}{c}" for c in range(5)] + ["", ""]))
    dm.append(row_vals([""] * 7))
    meta = [(r, r) for r in range(len(dm))]
    win._diff_matrix, win._diff_row_meta = dm, meta
    win._refresh_tables()
    tbl = win.panel_a.table
    win.show()
    app.processEvents()
    row_n = tbl.rowCount()
    col_n = tbl.columnCount()

    # 열 헤더: B열 선택 → Shift+→ 두 번 → [1,2,3]
    tbl._select_col(1)
    tbl._on_h_section_pressed(1)
    QTest.keyClick(tbl, Qt.Key_Right, Qt.ShiftModifier)
    assert tbl._full_columns_selected() == [1, 2], tbl._full_columns_selected()
    QTest.keyClick(tbl, Qt.Key_Right, Qt.ShiftModifier)
    assert tbl._full_columns_selected() == [1, 2, 3], tbl._full_columns_selected()
    assert len(tbl.selectionModel().selectedIndexes()) == 3 * row_n
    # Shift+← 로 앵커 방향 축소
    QTest.keyClick(tbl, Qt.Key_Left, Qt.ShiftModifier)
    assert tbl._full_columns_selected() == [1, 2], tbl._full_columns_selected()
    # Ctrl+Shift+→ : '값이 있는' 마지막 열(4)까지 — 매트릭스 폭(7)이나
    # 그리드 폭이 아닌 실제 값 기준 (유령 셀 열 5~6 제외)
    assert tbl.model().data_cols == 7
    assert tbl.model().col_has_values(4) and not tbl.model().col_has_values(5)
    QTest.keyClick(tbl, Qt.Key_Right, Qt.ControlModifier | Qt.ShiftModifier)
    assert tbl._full_columns_selected() == [1, 2, 3, 4], tbl._full_columns_selected()
    # 경계에서 한 번 더 → 엑셀처럼 현재 위치 기준 재판정해 그리드 끝까지
    QTest.keyClick(tbl, Qt.Key_Right, Qt.ControlModifier | Qt.ShiftModifier)
    assert tbl._full_columns_selected() == list(range(1, col_n)), tbl._full_columns_selected()
    # 그리드 끝에서 한 번 더 → 변화 없음
    QTest.keyClick(tbl, Qt.Key_Right, Qt.ControlModifier | Qt.ShiftModifier)
    assert tbl._full_columns_selected() == list(range(1, col_n)), tbl._full_columns_selected()

    # 행 헤더: 3행 선택 → Shift+↓ → [2,3], Ctrl+Shift+↓ 값 있는 마지막 행(7)까지
    assert tbl.model().data_rows == 9
    assert tbl.model().row_has_values(7) and not tbl.model().row_has_values(8)
    tbl._select_row(2)
    tbl._on_v_section_pressed(2)
    QTest.keyClick(tbl, Qt.Key_Down, Qt.ShiftModifier)
    assert tbl._full_rows_selected() == [2, 3], tbl._full_rows_selected()
    QTest.keyClick(tbl, Qt.Key_Down, Qt.ControlModifier | Qt.ShiftModifier)
    assert tbl._full_rows_selected() == list(range(2, 8)), tbl._full_rows_selected()
    # 경계에서 한 번 더 → 그리드 끝까지
    QTest.keyClick(tbl, Qt.Key_Down, Qt.ControlModifier | Qt.ShiftModifier)
    assert tbl._full_rows_selected() == list(range(2, row_n)), tbl._full_rows_selected()

    win.close()
    print("PASS test_header_multiselect_extension")


def test_ctrl_jump_single_selection():
    """Ctrl+방향키 점프 회귀 테스트 — 엑셀처럼 단일 선택으로 이동해야 한다.
    수정 전에는 Ctrl 수정자 때문에 setCurrentIndex가 Toggle로 동작해
    원래 셀과 대상 셀이 함께 선택됐다."""
    from PyQt5.QtTest import QTest
    app = QApplication.instance() or QApplication([])
    win = _diff_view()
    # 0~2열 값 있음, 3열 빈 값, 4열 값 있음
    dm = [[("same", f"h{c}", f"h{c}") for c in range(5)]]
    for r in range(1, 6):
        dm.append([("same", "" if c == 3 else f"v{r}{c}",
                    "" if c == 3 else f"v{r}{c}") for c in range(5)])
    meta = [(r, r) for r in range(len(dm))]
    win._diff_matrix, win._diff_row_meta = dm, meta
    win._refresh_tables()
    tbl = win.panel_a.table
    win.show()
    app.processEvents()

    # 값 경계 점프: (2,0) → 연속 데이터 끝 (2,2) — 단일 선택
    tbl._move_current_cell(2, 0)
    QTest.keyClick(tbl, Qt.Key_Right, Qt.ControlModifier)
    assert tbl.get_selected_cells() == {(2, 2)}, tbl.get_selected_cells()
    # 빈 열 건너 다음 값 셀 (2,4) — 단일 선택
    QTest.keyClick(tbl, Qt.Key_Right, Qt.ControlModifier)
    assert tbl.get_selected_cells() == {(2, 4)}, tbl.get_selected_cells()
    # 값이 더 없으면 마지막 셀로 — 단일 선택
    QTest.keyClick(tbl, Qt.Key_Right, Qt.ControlModifier)
    assert tbl.get_selected_cells() == {(2, tbl.columnCount() - 1)}, tbl.get_selected_cells()

    # Ctrl+Shift+방향키: 범위 선택 유지 (붕괴 없음)
    tbl._move_current_cell(2, 0)
    QTest.keyClick(tbl, Qt.Key_Right, Qt.ControlModifier | Qt.ShiftModifier)
    assert tbl.get_selected_cells() == {(2, 0), (2, 1), (2, 2)}, tbl.get_selected_cells()

    # Ctrl+Shift+End: 데이터 영역으로 클램프된 범위, 붕괴 없음
    tbl._move_current_cell(1, 1)
    QTest.keyClick(tbl, Qt.Key_End, Qt.ControlModifier | Qt.ShiftModifier)
    sel = tbl.get_selected_cells()
    assert max(r for r, _ in sel) == len(dm) - 1, sel
    assert max(c for _, c in sel) == 4, sel
    assert len(sel) == (len(dm) - 1) * 4, len(sel)

    # Ctrl+Home: 단일 이동
    QTest.keyClick(tbl, Qt.Key_Home, Qt.ControlModifier)
    assert tbl.get_selected_cells() == {(0, 0)}, tbl.get_selected_cells()

    win.close()
    print("PASS test_ctrl_jump_single_selection")


def test_selection_mirror_compact_ranges():
    """선택 미러가 range를 직사각형으로 압축하는지 회귀 테스트.
    수정 전에는 열 전체 미러가 행마다 range를 만들어(수천 개) 이후 모든
    선택 질의·페인팅이 느려졌다 (헤더 클릭 딜레이의 주범)."""
    app = QApplication.instance() or QApplication([])
    win = _diff_view()
    dm = [[("same", f"v{r}{c}", f"v{r}{c}") for c in range(6)] for r in range(120)]
    meta = [(r, r) for r in range(len(dm))]
    win._diff_matrix, win._diff_row_meta = dm, meta
    win._refresh_tables()
    src, dst = win.panel_a.table, win.panel_b.table
    win.show()
    app.processEvents()

    # 열 전체 선택 → 동기화된 반대 패널도 range 1개 + 셀 집합 동일
    src._select_col(2)
    app.processEvents()
    assert len(dst.selectionModel().selection()) == 1, \
        len(dst.selectionModel().selection())
    assert dst.get_selected_cells() == src.get_selected_cells()

    # mirror_selection(셀 집합) 경로도 직사각형 병합: 2×3 블록 → range 1개
    block = {(r, c) for r in range(10, 12) for c in range(1, 4)}
    dst.mirror_selection(block)
    assert dst.get_selected_cells() == block
    assert len(dst.selectionModel().selection()) == 1

    # 비직사각형(L자)도 셀 집합은 정확히 보존
    lshape = {(0, 0), (1, 0), (1, 1)}
    dst.mirror_selection(lshape)
    assert dst.get_selected_cells() == lshape

    # 단일 셀 판정 헬퍼
    src._move_current_cell(5, 5)
    assert src._single_selected_cell() == (5, 5)
    src._select_col(1)
    assert src._single_selected_cell() is None

    win.close()
    print("PASS test_selection_mirror_compact_ranges")


def test_staging_color_with_empty_initial_state():
    """비교 직후(스테이징 0개) 상태에서 병합 준비 시 색이 바뀌는지 회귀 테스트.
    populate의 `staged or {}` falsy 체크가 빈 dict일 때 새 객체를 만들어
    모델이 MainWindow 상태와 분리되던 버그."""
    from excelmerge.theme import DIFF_COLORS
    app = QApplication.instance() or QApplication([])
    win = _diff_view()
    dm = [[("same", "h", "h")], [("modified", "x", "y")]]
    win._diff_matrix = dm
    win._diff_row_meta = [(0, 0), (1, 1)]
    win._staged = {}            # 실제 흐름: 비교 직후엔 빈 dict
    win._merged_cells = set()
    win._refresh_tables()
    model = win.panel_a.table.model()
    assert model._staged is win._staged, "모델이 MainWindow staged와 분리됨"
    assert model._merged is win._merged_cells, "모델이 merged_cells와 분리됨"
    win._staged[(1, 0)] = "a_to_b"
    win._notify_cells({(1, 0)})
    bg = model.data(model.index(1, 0), Qt.BackgroundRole)
    assert bg.getRgb() == DIFF_COLORS["staged"].getRgb(), bg.getRgb()
    win.close()
    print("PASS test_staging_color_with_empty_initial_state")


def test_ctrl_jump_skips_hidden_rows():
    """'변경 행만 보기' 상태에서 Ctrl+↓ 점프가 숨겨진 행에 착지하지 않는지."""
    from PyQt5.QtTest import QTest
    app = QApplication.instance() or QApplication([])
    win = _diff_view()
    # 행 1, 4만 변경 — 필터 켜면 2, 3, 5는 숨겨짐 (행 0은 헤더로 항상 표시)
    dm = [[("same", "h", "h")]]
    for r in range(1, 6):
        changed = r in (1, 4)
        dm.append([("modified" if changed else "same", f"a{r}", f"b{r}" if changed else f"a{r}")])
    win._diff_matrix = dm
    win._diff_row_meta = [(r, r) for r in range(len(dm))]
    win._refresh_tables()
    win._diff_only = True
    win._apply_diff_filter()
    tbl = win.panel_a.table
    win.show()
    app.processEvents()
    assert tbl.isRowHidden(2) and tbl.isRowHidden(3), "필터 전제 불성립"

    tbl._move_current_cell(1, 0)
    QTest.keyClick(tbl, Qt.Key_Down, Qt.ControlModifier)
    r, c = tbl._current_cell()
    assert not tbl.isRowHidden(r), f"숨겨진 행 {r}에 착지"
    assert r == 4, f"기대 4행, 실제 {r}행"   # 1 다음 보이는 값 행
    win.close()
    print("PASS test_ctrl_jump_skips_hidden_rows")


def test_find_in_preview_mode():
    """파일 하나만 로드된 미리보기 상태에서도 Ctrl+F 검색이 동작하는지."""
    app = QApplication.instance() or QApplication([])
    win = _diff_view()
    win.panel_a.preview([["id", "name"], ["1", "apple"], ["2", "banana"]])
    matches = list(win._iter_find_matches(win._make_find_matcher("banana")))
    assert matches == [(2, 1)], matches
    # diff 없이도 _goto_find가 이동시키는지
    win.find_edit.setText("apple")
    win._set_find_enabled(True)
    win._goto_find(+1)
    assert win.panel_a.table._current_cell() == (1, 1)
    win.close()
    print("PASS test_find_in_preview_mode")


def test_row_header_multi_stage():
    """복수 행 헤더 선택 후 병합 준비 시 모든 대상 행의 변경 셀이 staged 되는지."""
    from excelmerge.diff_engine import compute_diff
    app = QApplication.instance() or QApplication([])
    win = _diff_view()
    a = [["ID", "V"], ["1", "a1"], ["2", "a2"], ["3", "a3"], ["4", "a4"]]
    b = [["ID", "V"], ["1", "b1"], ["2", "b2"], ["3", "b3"], ["4", "b4"]]
    win._raw_data["a"], win._raw_data["b"] = a, b
    win._diff_matrix, win._diff_row_meta = compute_diff(a, b, 0)
    win.panel_a._row_meta = win.panel_b._row_meta = win._diff_row_meta
    win._refresh_tables()
    win.show()
    app.processEvents()
    tbl = win.panel_a.table

    # 행 1,2,3 헤더를 다중 선택한 상태를 재현
    tbl._select_rows([1, 2, 3])
    app.processEvents()
    assert tbl._selected_header_rows(2) == [1, 2, 3], tbl._selected_header_rows(2)
    # 다중 선택 중 한 행 우클릭 = 전체 대상 → 스테이징
    tbl._select_rows(tbl._selected_header_rows(2))
    win._stage_selected("b_to_a")
    assert set(win._staged.keys()) == {(1, 1), (2, 1), (3, 1)}, sorted(win._staged.keys())
    assert all(v == "b_to_a" for v in win._staged.values())

    # 선택에 없는 행 우클릭 = 그 행만 대상
    assert tbl._selected_header_rows(4) == [4]
    win.close()
    print("PASS test_row_header_multi_stage")


def test_sheet_path_absolute_target_saves():
    """절대경로 rel Target(openpyxl 등) 파일도 저장이 실제 적용되는지 + 시트 미해결 시 raise."""
    import tempfile
    import zipfile
    import openpyxl
    from excelmerge import xlsx_writer

    tmpdir = tempfile.mkdtemp()
    p = os.path.join(tmpdir, "abs.xlsx")
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["ID", "V"]); ws.append(["1", "a1"]); ws.append(["2", "a2"])
    wb.save(p)   # openpyxl은 Target을 "/xl/worksheets/sheet1.xml"(절대)로 쓴다
    with zipfile.ZipFile(p) as z:
        assert b'Target="/xl/' in z.read("xl/_rels/workbook.xml.rels")
        assert xlsx_writer._find_active_sheet_path(z) == "xl/worksheets/sheet1.xml"

    xlsx_writer._write_patches_to_file(p, {"B2": "b1", "B3": "b2"})
    rows = [[c.value for c in r] for r in openpyxl.load_workbook(p).active.iter_rows()]
    assert rows[1][1] == "b1" and rows[2][1] == "b2", rows

    # 시트 미해결 → 명시적 오류(조용한 유실 방지)
    orig = xlsx_writer._find_active_sheet_path
    xlsx_writer._find_active_sheet_path = lambda z: "xl/worksheets/nope.xml"
    try:
        raised = False
        try:
            xlsx_writer._write_patches_to_file(p, {"B2": "x"})
        except ValueError:
            raised = True
        assert raised, "시트 미해결인데 raise 안 함"
    finally:
        xlsx_writer._find_active_sheet_path = orig
    print("PASS test_sheet_path_absolute_target_saves")


def test_diff_filter_after_key_change():
    """키 열 변경(매트릭스 재계산) 후에도 '변경 행만 보기'가 정확히 적용되는지.
    _apply_diff_filter가 캐시가 아닌 실제 isRowHidden을 조회하므로, 모델 리셋이
    숨김 상태를 초기화하는 플랫폼에서도 변경 안 된 행이 새어 나오지 않아야 한다."""
    from excelmerge.diff_engine import compute_diff
    app = QApplication.instance() or QApplication([])
    win = _diff_view()
    a = [["ID", "Name"], ["1", "a"], ["2", "b"], ["3", "c"], ["4", "d"]]
    b = [["ID", "Name"], ["1", "a"], ["2", "b"], ["3", "C"], ["4", "d"]]
    win._raw_data["a"], win._raw_data["b"] = a, b
    win._diff_matrix, win._diff_row_meta = compute_diff(a, b, 0)
    win.panel_a._row_meta = win.panel_b._row_meta = win._diff_row_meta
    win._refresh_tables()
    win._set_buttons_enabled(True)
    win.diff_only_btn.setChecked(True)
    win.show()
    app.processEvents()
    tbl = win.panel_a.table

    def hidden():
        return {r for r in range(len(win._diff_matrix)) if tbl.isRowHidden(r)}

    # key=0: 행 3만 변경 → 1,2,4 숨김. + 틀 고정으로 고정 행 0(헤더)도 본체에서 숨김.
    assert hidden() == {0, 1, 2, 4}, hidden()

    # 모델 리셋이 숨김을 초기화하는 플랫폼을 흉내: 실제로 다 보이게 만든 뒤 필터 재적용
    for r in range(tbl.rowCount()):
        if tbl.isRowHidden(r):
            tbl.setRowHidden(r, False)
    win._apply_diff_filter()
    assert hidden() == {0, 1, 2, 4}, f"필터 재적용 실패: {hidden()}"

    # 키 열 변경 경로 전체
    win._on_key_col_changed(1)
    app.processEvents()
    # key=1(Name): a/b Name이 c vs C로 갈려 3(A전용)·5(B전용) 신규, 1,2,4 동일 → 숨김
    hid = hidden()
    for r in (1, 2, 4):
        assert r in hid, f"key변경 후 동일 행 {r}이 숨겨지지 않음: {hid}"
    win.close()
    print("PASS test_diff_filter_after_key_change")


def test_goto_changed_focus_and_selection_color():
    """이전/다음 변경점 이동 후: 테이블 포커스 이동 + 선택색이 활성/비활성 동일(파랑)."""
    from PyQt5.QtGui import QPalette
    app = QApplication.instance() or QApplication([])
    win = _diff_view()
    dm = [[("same", "h", "h")]]
    for r in range(1, 6):
        ch = r in (1, 3, 5)
        dm.append([("modified" if ch else "same", f"a{r}", f"b{r}" if ch else f"a{r}")])
    win._diff_matrix = dm
    win._diff_row_meta = [(r, r) for r in range(len(dm))]
    win._refresh_tables()
    win._set_buttons_enabled(True)
    win.show()
    app.processEvents()

    for tbl in (win.panel_a.table, win.panel_b.table):
        pal = tbl.palette()
        act = pal.color(QPalette.Active, QPalette.Highlight).getRgb()
        for grp in (QPalette.Inactive, QPalette.Disabled):
            assert pal.color(grp, QPalette.Highlight).getRgb() == act, grp
            assert (pal.color(grp, QPalette.HighlightedText).getRgb()
                    == pal.color(QPalette.Active, QPalette.HighlightedText).getRgb())

    # '전체 보기'에서 이동
    win._diff_only = False
    win._apply_diff_filter()
    win._goto_changed(+1)
    assert win.panel_a.table._current_cell() == (1, 0)
    assert win.panel_a.table.get_selected_cells() == {(1, 0)}
    assert win.panel_a.table.hasFocus(), "이동 후 테이블 포커스 안 됨"
    win._goto_changed(+1)
    assert win.panel_a.table._current_cell() == (3, 0)
    # 반대 패널도 동일 셀 선택(동기화) — 색은 팔레트 통일로 동일 파랑
    assert win.panel_b.table.get_selected_cells() == {(3, 0)}
    win.close()
    print("PASS test_goto_changed_focus_and_selection_color")


def test_load_xlsx_with_empty_fill():
    """styles.xml에 빈 <fill/>이 있는 파일도 로드되는지 회귀 테스트.
    openpyxl은 빈 fill에 'expected Fill' TypeError를 던진다 — 정제 후 재시도해야 함."""
    import io
    import re
    import zipfile
    import tempfile
    import datetime
    import openpyxl
    from excelmerge.loaders import _load_values_pass_openpyxl, _open_workbook

    tmpdir = tempfile.mkdtemp()
    good = os.path.join(tmpdir, "good.xlsx")
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["날짜", "값", "수식"])
    ws.append([datetime.datetime(2026, 7, 3), 10, "=B2*2"])
    wb.save(good)

    # styles.xml의 fills를 빈 <fill/>로 손상
    with zipfile.ZipFile(good) as z:
        styles = z.read("xl/styles.xml").decode("utf-8")
    bad_styles = re.sub(r"<fills.*?</fills>", '<fills count="1"><fill/></fills>',
                        styles, flags=re.S)
    bad = os.path.join(tmpdir, "bad.xlsx")
    with zipfile.ZipFile(good) as zin, zipfile.ZipFile(bad, "w") as zout:
        for it in zin.infolist():
            data = bad_styles.encode() if it.filename == "xl/styles.xml" else zin.read(it.filename)
            zout.writestr(it, data)

    # 손상 파일이 표준 openpyxl 로드에선 실패함을 확인
    failed = False
    try:
        openpyxl.load_workbook(bad, read_only=True, data_only=True)
    except TypeError:
        failed = True
    assert failed, "테스트 전제(빈 fill이 openpyxl을 깨뜨림) 불성립"

    # _open_workbook / openpyxl 값 로더는 정제(styles.xml sanitize) 후 성공해야 함
    _open_workbook(bad, data_only=True).close()
    values = _load_values_pass_openpyxl(bad)
    assert values[0] == ["날짜", "값", "수식"], values[0]
    assert values[1][1] == "10", values[1]
    assert "2026" in values[1][0], f"날짜 보존 실패: {values[1][0]}"
    print("PASS test_load_xlsx_with_empty_fill")


def test_cell_status_one_sided_is_added():
    """한쪽 파일에만 값이 있으면 방향과 무관하게 'added'로 분류."""
    from excelmerge.diff_engine import _cell_status
    assert _cell_status("", "new") == "added"      # B 전용
    assert _cell_status("old", "") == "added"      # A 전용 (기존엔 modified)
    assert _cell_status("a", "b") == "modified"
    assert _cell_status("x", "x") == "same"
    assert _cell_status("", "") == "same"
    print("PASS test_cell_status_one_sided_is_added")


def test_filter_keeps_merged_rows_visible():
    """저장(병합 확정) 후에도 '변경 행만 보기'가 병합됨 행을 숨기지 않는지.
    저장 시 상태가 same이 되면서 행이 즉시 사라져 병합 결과를 볼 수 없던 문제."""
    app = QApplication.instance() or QApplication([])
    win = _diff_view()
    dm = [
        [("same", "h", "h")],
        [("same", "m", "m")],    # 병합 확정된 행 (merged 셀 보유)
        [("same", "u", "u")],    # 그냥 동일한 행
        [("modified", "x", "y")],
    ]
    win._diff_matrix = dm
    win._diff_row_meta = [(r, r) for r in range(len(dm))]
    win._merged_cells = {(1, 0)}
    win._refresh_tables()
    win._diff_only = True
    win._apply_diff_filter()
    tbl = win.panel_a.table
    assert tbl.isRowHidden(0), "고정(헤더) 행은 본체에서 숨김 → 상단 고정 띠에 표시"
    assert not tbl.isRowHidden(1), "병합됨 행이 숨겨짐"
    assert tbl.isRowHidden(2), "동일 행은 숨겨져야 함"
    assert not tbl.isRowHidden(3), "변경 행"
    win.close()
    print("PASS test_filter_keeps_merged_rows_visible")


def test_sheet_tabs_mark_changed():
    """다중 시트 파일 비교 시, 서로 다른(또는 한쪽에만 있는) 시트 탭만 변경 표시(노랑)."""
    import tempfile, os
    import openpyxl
    from PyQt5.QtCore import QEventLoop, QTimer
    from excelmerge.main_window import MainWindow
    app = QApplication.instance() or QApplication([])

    d = tempfile.mkdtemp()
    pa = os.path.join(d, "A.xlsx"); pb = os.path.join(d, "B.xlsx")
    wb = openpyxl.Workbook(); wb.active.title = "S1"; wb.active.append(["id", "v"]); wb.active.append([1, "x"])
    s2 = wb.create_sheet("S2"); s2.append(["id", "v"]); s2.append([1, "AAA"])
    s3 = wb.create_sheet("S3"); s3.append(["id", "v"]); s3.append([1, "only"])
    wb.save(pa)
    wb = openpyxl.Workbook(); wb.active.title = "S1"; wb.active.append(["id", "v"]); wb.active.append([1, "x"])
    s2 = wb.create_sheet("S2"); s2.append(["id", "v"]); s2.append([1, "BBB"])   # 값 다름
    wb.save(pb)

    win = MainWindow()
    view = win.open_file_compare(pa, pb)
    loop = QEventLoop()

    def _chk():
        if view.sheet_tabs.count() >= 3 and view.sheet_tabs._changed_idx:
            loop.quit()
        else:
            QTimer.singleShot(30, _chk)
    QTimer.singleShot(30, _chk); QTimer.singleShot(8000, loop.quit); loop.exec_()

    names = [view.sheet_tabs.tabText(i) for i in range(view.sheet_tabs.count())]
    changed = sorted(names[i] for i in view.sheet_tabs._changed_idx)
    assert changed == ["S2", "S3"], f"변경 시트 표시 오류: {changed}"   # S1 동일→미표시
    win.close()
    print("PASS test_sheet_tabs_mark_changed")


def test_key_col_reset_when_out_of_range():
    """키 열이 현재 데이터 폭을 벗어나면 _on_loaded가 A열(0)로 리셋한다 (A5/M3)."""
    app = QApplication.instance() or QApplication([])
    win = _diff_view()
    win.panel_a.set_path("a.xlsx"); win.panel_b.set_path("b.xlsx")
    win._key_col = 5                      # 넓은 시트에서 F열을 키로 잡았다고 가정
    win.panel_a.table.set_key_col(5)
    # 이제 2열짜리 좁은 데이터가 로드됨 → 5 >= 2 라 리셋되어야 함
    a = [["ID", "V"], ["1", "a1"], ["2", "a2"]]
    b = [["ID", "V"], ["1", "b1"], ["2", "b2"]]
    win._on_loaded(a, b)                  # ctx 없이 직접 호출 → 무조건 적용
    assert win._key_col == 0, f"키 열이 리셋되지 않음: {win._key_col}"
    _wait_diff(win)                        # 백그라운드 DiffWorker 정리 후 닫기
    win.close()
    print("PASS test_key_col_reset_when_out_of_range")


def test_compute_diff_key_row():
    """key_row=1: 프리앰블(0행)+헤더(1행)는 위치 1:1, 그 아래만 키 열 매칭.
    row_meta 는 원본 파일 행 번호를 그대로 담아야 한다(병합 저장 좌표 일치)."""
    a = [["TITLE", "", ""],
         ["ID", "Name", "Score"],
         ["1", "Alice", "10"],
         ["2", "Bob", "20"],
         ["3", "Carol", "30"]]
    b = [["TITLE", "", ""],
         ["ID", "Name", "Score"],
         ["1", "Alice", "10"],
         ["2", "Bobby", "20"],
         ["4", "Dave", "40"]]
    dm, meta = compute_diff(a, b, key_col=0, key_row=1)
    assert len(dm) == 6, f"행수 {len(dm)} != 6"
    assert meta == [(0, 0), (1, 1), (2, 2), (3, 3), (4, None), (None, 4)], f"row_meta 오류: {meta}"
    assert dm[0][0] == ("same", "TITLE", "TITLE"), f"프리앰블 1:1 실패: {dm[0][0]}"
    assert dm[1][0] == ("same", "ID", "ID"), f"헤더 1:1 실패: {dm[1][0]}"
    assert dm[3][1] == ("modified", "Bob", "Bobby"), f"본문 매칭 실패: {dm[3][1]}"
    assert dm[4][1][0] == "added" and dm[4][1][1] == "Carol", f"A 전용 행 실패: {dm[4][1]}"
    assert dm[5][1][0] == "added" and dm[5][1][2] == "Dave", f"B 전용 행 실패: {dm[5][1]}"
    # 드롭 없음
    assert count_dropped_key_rows(a, b, 0, key_row=1) == 0
    # 헤더 아래에 빈 키/중복 키가 있으면 드롭에 잡히되 프리앰블/헤더는 세지 않음
    a2 = a + [["", "Ghost", "0"], ["1", "Dup", "9"]]   # 빈 키 1 + 중복 키("1") 1
    assert count_dropped_key_rows(a2, b, 0, key_row=1) == 2, \
        count_dropped_key_rows(a2, b, 0, key_row=1)
    print("PASS test_compute_diff_key_row")


def test_key_row_header_rendering():
    """세로(행) 헤더가 키 행에 아이콘/노랑 배경/원본 행번호를 반환한다."""
    app = QApplication.instance() or QApplication([])
    a = [["T", ""], ["ID", "V"], ["1", "a"], ["2", "b"]]
    b = [["T", ""], ["ID", "V"], ["1", "a"], ["2", "B"]]
    dm, meta = compute_diff(a, b, key_col=0, key_row=1)
    m = DiffTableModel("a")
    m.set_diff_data(dm, meta, {}, set(), set())

    seen = []
    m.headerDataChanged.connect(lambda o, s, e: seen.append((o, s, e)))
    m.set_key_row(1)
    assert any(o == Qt.Vertical for (o, s, e) in seen), "set_key_row가 Qt.Vertical 통지 안 함"

    dec = m.headerData(1, Qt.Vertical, Qt.DecorationRole)
    assert dec is not None and not dec.isNull(), "키 행 아이콘 없음"
    bg = m.headerData(1, Qt.Vertical, Qt.BackgroundRole)
    assert bg is not None and bg.getRgb()[:3] == (255, 213, 0), f"키 행 배경 오류: {bg}"
    assert m.headerData(1, Qt.Vertical, Qt.DisplayRole) == "2", "키 행 원본번호(2) 표시 오류"
    # 키 행이 아닌 행은 아이콘 없음
    assert m.headerData(0, Qt.Vertical, Qt.DecorationRole) is None, "비키행에 아이콘이 뜸"
    print("PASS test_key_row_header_rendering")


def test_key_row_reset_when_out_of_range():
    """키 행이 현재 데이터 행수를 벗어나면 _on_loaded가 1행(0)으로 리셋한다."""
    app = QApplication.instance() or QApplication([])
    win = _diff_view()
    win.panel_a.set_path("a.xlsx"); win.panel_b.set_path("b.xlsx")
    win._key_row = 10                     # 큰 시트에서 11행을 헤더로 잡았다고 가정
    win.panel_a.table.set_key_row(10)
    a = [["ID", "V"], ["1", "a1"], ["2", "a2"]]   # 3행짜리 → 10 >= 3 라 리셋
    b = [["ID", "V"], ["1", "b1"], ["2", "b2"]]
    win._on_loaded(a, b)
    assert win._key_row == 0, f"키 행이 리셋되지 않음: {win._key_row}"
    _wait_diff(win)                        # 백그라운드 DiffWorker 정리 후 닫기
    win.close()
    print("PASS test_key_row_reset_when_out_of_range")


def test_freeze_main_hides_frozen():
    """틀 고정: 본체(main)는 고정 행(0..key_row)·고정 열(0..key_col)을 '숨겨' 스크롤 영역만 렌더
    (가림 방지). 변경된 비고정 행은 보이고, 변경 없는 비고정 행은 필터로 숨겨진다."""
    app = QApplication.instance() or QApplication([])
    win = _diff_view()
    win.panel_a.set_path("a.xlsx"); win.panel_b.set_path("b.xlsx")
    win._key_row = 1; win._key_col = 1        # 앵커 B2
    for p in (win.panel_a, win.panel_b):
        p.table.set_key_row(1); p.table.set_key_col(1)
    a = [["pre", "x", "y"], ["ID", "V", "W"], ["1", "a", "p"], ["2", "b", "q"]]
    b = [["pre", "x", "y"], ["ID", "V", "W"], ["1", "a", "p"], ["2", "B", "q"]]  # key"2" 변경
    win._on_loaded(a, b)    # 필터 기본 ON
    _wait_diff(win)
    t = win.panel_a.table
    # 고정 열: 본체에서 숨김(고정 띠에 표시)
    assert t.isColumnHidden(0) and t.isColumnHidden(1), "고정 열(A,B)이 본체에서 숨겨지지 않음"
    assert not t.isColumnHidden(2), "비고정 열(C)은 본체에 표시"
    # 고정 행: 본체에서 숨김
    assert t.isRowHidden(0) and t.isRowHidden(1), "고정 행(0,1)이 본체에서 숨겨지지 않음"
    # 비고정 행: 변경 없는 2행 숨김, 변경 3행 표시
    assert t.isRowHidden(2), "변경 없는 본문(2행)은 숨겨져야"
    assert not t.isRowHidden(3), "변경 행(3행)은 표시"
    win.close()
    print("PASS test_freeze_main_hides_frozen")


def test_frozen_span():
    """FreezeController.frozen_span: 앵커→(고정행수,고정열수), 범위 클램프, key_col<0→열0."""
    from excelmerge.widgets import FreezeController as FC
    assert FC.frozen_span(0, 0, 10, 5) == (1, 1)
    assert FC.frozen_span(1, 1, 10, 5) == (2, 2)     # 앵커 B2 → 1~2행, A~B열
    assert FC.frozen_span(1, -1, 10, 5) == (2, 0)    # ROW 순서 → 열 고정 없음
    assert FC.frozen_span(3, 2, 2, 5) == (2, 3)      # key_row가 데이터 밖 → 행수 클램프
    print("PASS test_frozen_span")


def test_freeze_setup():
    """틀 고정 헬퍼 구조: corner는 헤더(고정 눈금) 표시, top/left는 헤더 숨김, 모델 공유,
    앵커대로 스팬, 스크롤 동기."""
    app = QApplication.instance() or QApplication([])
    win = _diff_view()
    win.panel_a.set_path("a.xlsx"); win.panel_b.set_path("b.xlsx")
    win._key_row = 1; win._key_col = 1
    for p in (win.panel_a, win.panel_b):
        p.table.set_key_row(1); p.table.set_key_col(1)
    a = [["pre", "x", "y"], ["ID", "V", "W"]] + [[str(i), "a", "p"] for i in range(20)]
    b = [["pre", "x", "y"], ["ID", "V", "W"]] + [[str(i), "a", "p"] for i in range(20)]
    b[5][1] = "CH"
    win._on_loaded(a, b)
    _wait_diff(win)
    fc = win._freeze["a"]; t = win.panel_a.table
    assert fc.active and len(fc._views) == 3
    assert fc.top.model() is t.model(), "헬퍼가 모델 공유 안 함"
    assert (fc._n_rows, fc._n_cols) == (2, 2)
    # corner는 눈금(양 헤더) 표시, top/left는 숨김
    assert fc.corner.horizontalHeader().isVisible() and fc.corner.verticalHeader().isVisible()
    assert not fc.top.horizontalHeader().isVisible()
    assert not fc.left.verticalHeader().isVisible()
    # corner는 고정 열만, top은 스크롤 열만 표시
    assert not fc.corner.isColumnHidden(1) and fc.corner.isColumnHidden(2), "corner 고정 열 오류"
    assert fc.top.isColumnHidden(1) and not fc.top.isColumnHidden(2), "top 스크롤 열 오류"
    # 스크롤 동기
    win.diff_only_btn.setChecked(False)   # 전체 행 표시 → 세로 스크롤 여지
    t.verticalScrollBar().setValue(3); t.horizontalScrollBar().setValue(1)
    fc._sync_scroll()
    assert fc.left.verticalScrollBar().value() == t.verticalScrollBar().value()
    assert fc.top.horizontalScrollBar().value() == t.horizontalScrollBar().value()
    win.close()
    print("PASS test_freeze_setup")


def test_freeze_bulk_guard():
    """성능 회귀 방지: 필터의 setRowHidden 벌크(_applying_sizes) 중에는 FreezeController의
    sectionResized 핸들러가 발화하면 안 된다(발화 시 숨김 해제 행마다 O(C)+reposition → O(R×C) 폭주)."""
    app = QApplication.instance() or QApplication([])
    win = _diff_view()
    win.panel_a.set_path("a.xlsx"); win.panel_b.set_path("b.xlsx")
    win._key_row = 1; win._key_col = 1
    for p in (win.panel_a, win.panel_b):
        p.table.set_key_row(1); p.table.set_key_col(1)
    a = [["pre", "x"], ["ID", "V"], ["1", "a"], ["2", "b"]]
    b = [["pre", "x"], ["ID", "V"], ["1", "a"], ["2", "B"]]
    win._on_loaded(a, b)
    _wait_diff(win)
    fc = win._freeze["a"]; t = win.panel_a.table
    calls = []
    fc._sync_sizes = lambda: calls.append(1)   # 발화 여부 감지용 몽키패치
    t._applying_sizes = True                    # 벌크 진행 중 상태 모사
    fc._on_row_resized(2, 0, 22)
    fc._on_col_resized(1, 0, 50)
    assert calls == [], "벌크(_applying_sizes) 중 freeze 핸들러가 발화함 — 성능 회귀"
    t._applying_sizes = False                   # 벌크 종료 → 정상 발화
    fc._on_row_resized(2, 0, 22)
    assert calls == [1], "벌크 종료 후엔 freeze 핸들러가 발화해야 함"
    win.close()
    print("PASS test_freeze_bulk_guard")


def test_freeze_corner_menu_nonmodal():
    """크래시 회귀 방지: 고정(corner) 헤더 우클릭 메뉴는 비모달 popup으로 키 열/행 신호를 emit해야
    하고, 모달 exec_는 쓰면 안 된다(자식 오버레이 헤더에서 exec_ 시 access violation 크래시)."""
    from PyQt5.QtWidgets import QMenu
    from PyQt5.QtCore import QPoint
    app = QApplication.instance() or QApplication([])
    win = _diff_view()
    win.panel_a.set_path("a.xlsx"); win.panel_b.set_path("b.xlsx")
    win._key_row = 1; win._key_col = 1
    for p in (win.panel_a, win.panel_b):
        p.table.set_key_row(1); p.table.set_key_col(1)
    a = [["pre", "x", "y"], ["ID", "V", "W"], ["1", "a", "p"], ["2", "b", "q"]]
    b = [["pre", "x", "y"], ["ID", "V", "W"], ["1", "a", "p"], ["2", "B", "q"]]
    win._on_loaded(a, b)
    _wait_diff(win)
    fc = win._freeze["a"]; t = win.panel_a.table
    ch = fc.corner.horizontalHeader(); cv = fc.corner.verticalHeader()
    ch.logicalIndexAt = lambda pos: 0     # 오프스크린 레이아웃 무관하게 섹션 0 지정
    cv.logicalIndexAt = lambda pos: 0
    exec_calls = []
    orig_exec, orig_popup = QMenu.exec_, QMenu.popup
    QMenu.exec_ = lambda self, *a, **k: (exec_calls.append(1), None)[1]
    QMenu.popup = lambda self, *a, **k: (self.actions()[0].trigger() if self.actions() else None)
    col_sig, row_sig = [], []
    t.key_col_changed.connect(col_sig.append)
    t.key_row_changed.connect(row_sig.append)
    try:
        fc._corner_row_menu(QPoint(1, 1))   # 원본 행0(≠key_row1) → 키 행으로 설정 → emit 0
        fc._corner_col_menu(QPoint(1, 1))   # 열0(≠key_col1) → 키 열로 설정 → emit 0
    finally:
        QMenu.exec_, QMenu.popup = orig_exec, orig_popup
    assert row_sig == [0], f"corner 키 행 신호 오류: {row_sig}"
    assert col_sig == [0], f"corner 키 열 신호 오류: {col_sig}"
    assert exec_calls == [], "corner 메뉴가 모달 exec_ 사용 — 크래시 회귀 위험"
    win.close()
    print("PASS test_freeze_corner_menu_nonmodal")


def test_freeze_cells_selectable():
    """회귀 방지: 틀 고정된 셀(키 열/행)도 선택 가능해야 한다.
    고정 오버레이(top/left/corner)는 본체와 '선택 모델을 공유'하고 ExtendedSelection이라,
    거기서 클릭하면 본체 선택이 갱신되고 셀값란·미러 등 기존 배선이 동작한다.
    (과거엔 NoSelection이라 고정 셀을 아예 선택할 수 없었다.)"""
    from PyQt5.QtWidgets import QAbstractItemView
    from PyQt5.QtCore import Qt, QItemSelectionModel
    from PyQt5.QtTest import QTest
    app = QApplication.instance() or QApplication([])
    win = _diff_view(); win.resize(1200, 700)
    win.panel_a.set_path("a.xlsx"); win.panel_b.set_path("b.xlsx")
    win._key_row = 1; win._key_col = 1
    for p in (win.panel_a, win.panel_b):
        p.table.set_key_row(1); p.table.set_key_col(1)
    a = [["pre", "x", "y"], ["ID", "V", "W"]] + [[str(i), "a%d" % i, "p"] for i in range(15)]
    b = [["pre", "x", "y"], ["ID", "V", "W"]] + [[str(i), "a%d" % i, "p"] for i in range(15)]
    b[5][1] = "CH"
    win._on_loaded(a, b); _wait_diff(win)
    fc = win._freeze["a"]; host = win.panel_a.table
    app.processEvents(); fc.refresh(); app.processEvents()

    # 1) 세 오버레이 모두 본체와 선택 모델 공유 + 선택 가능 모드
    for name, v in (("corner", fc.corner), ("top", fc.top), ("left", fc.left)):
        assert v.selectionModel() is host.selectionModel(), f"{name} 선택모델 미공유"
        assert v.selectionMode() != QAbstractItemView.NoSelection, f"{name} 선택 불가"

    # 2) 공유 선택 모델로 고정 셀(키 열, 스크롤 행)을 선택 → 패널 셀값란 갱신 (배선)
    host.selectionModel().select(host.model().index(4, 0),
                                 QItemSelectionModel.ClearAndSelect)
    host.selectionModel().setCurrentIndex(host.model().index(4, 0),
                                          QItemSelectionModel.NoUpdate)
    app.processEvents()
    assert win.panel_a._selected_cell == (4, 0), \
        f"고정 열 셀 선택이 패널에 반영 안 됨: {win.panel_a._selected_cell}"
    assert win.panel_a.cell_edit.isEnabled(), "고정 셀 선택 시 셀값란이 비활성"

    # 3) 키 행 셀을 top 오버레이에서 실제 클릭 → 선택 반영
    idx = host.model().index(1, 2)          # 키 행(1)의 스크롤 열(2)
    rect = fc.top.visualRect(idx)
    if rect.width() > 0:                    # 오프스크린 레이아웃이 잡힐 때만 클릭 검증
        QTest.mouseClick(fc.top.viewport(), Qt.LeftButton, Qt.NoModifier, rect.center())
        app.processEvents()
        assert win.panel_a._selected_cell == (1, 2), \
            f"키 행 셀 클릭 선택 실패: {win.panel_a._selected_cell}"
    win.close()
    print("PASS test_freeze_cells_selectable")


def test_large_selection_queries_range_based():
    """회귀 방지: 대량 선택 조회가 range/staged 기반이라 즉시 반환하고 결과가 정확하다.
    (과거 selectedColumns/selectedRows/전셀순회로 O(R×C) → 수 초 걸리던 경로)"""
    import time
    app = QApplication.instance() or QApplication([])
    win = _diff_view()
    win.panel_a.set_path("a.xlsx"); win.panel_b.set_path("b.xlsx")
    N = 4000
    a = [["ID", "V", "W", "X"]] + [[str(i), "v%d" % i, "p", "q"] for i in range(N)]
    b = [["ID", "V", "W", "X"]] + [[str(i), "v%d" % i, "p", "q"] for i in range(N)]
    b[10][1] = "CHANGED"      # modified 셀 하나
    win._on_loaded(a, b); _wait_diff(win)
    host = win.panel_a.table
    cols = host.columnCount()

    host.selectAll()
    t = time.perf_counter()
    fc = host._full_columns_selected()
    shc = host._selected_header_cols(0)
    shr = host._selected_header_rows(0)
    hs = host._has_staged_selection()
    dt = time.perf_counter() - t
    assert dt < 0.5, f"대량 선택 조회가 느림(회귀 의심): {dt:.2f}s"
    assert fc == list(range(cols)), f"전체 열 판정 오류: {len(fc)}/{cols}"
    assert shc == list(range(cols)), "헤더 대상 열 오류"
    assert len(shr) == host.rowCount(), "헤더 대상 행 오류"
    assert hs is False, "staged 없는데 True"

    # 부분 선택 정확성 (비연속)
    host._select_cols([1, 3])
    assert host._full_columns_selected() == [1, 3]
    assert host._selected_header_cols(3) == [1, 3], "선택에 포함된 앵커 → 전체"
    assert host._selected_header_cols(2) == [2], "선택 밖 앵커 → 단일"
    win.close()
    print("PASS test_large_selection_queries_range_based")


def test_char_range_cache_staging_safe():
    """diff_char_ranges 캐시가 staging 후에도 스테일 하이라이트를 남기지 않는다.
    (게이팅이 캐시 조회 전에 재평가되므로 modified→stage→[] , unstage→원래 range)"""
    from excelmerge.diff_model import DiffTableModel
    m = DiffTableModel("a")
    staged: dict = {}
    matrix = [[("modified", "abcX", "abcY")]]
    m.set_diff_data(matrix, [(0, 0)], staged, set(), set())
    ranges1 = m.diff_char_ranges(0, 0)
    assert ranges1 == [(3, 4)], f"modified 셀 char range 오류: {ranges1}"
    # 참조 공유된 staged를 변형(= notify_cells 경로) — 리셋 없음
    staged[(0, 0)] = "a_to_b"
    assert m.diff_char_ranges(0, 0) == [], "staged 셀은 게이팅으로 [] (캐시 스테일 아님)"
    del staged[(0, 0)]
    assert m.diff_char_ranges(0, 0) == ranges1, "unstage 후 캐시 결과 동일해야"
    print("PASS test_char_range_cache_staging_safe")


if __name__ == "__main__":
    test_golden_roles()
    test_headers()
    test_notify_equals_fresh_populate()
    test_cell_kind()
    test_view_behaviors()
    test_header_multiselect_extension()
    test_ctrl_jump_single_selection()
    test_selection_mirror_compact_ranges()
    test_staging_color_with_empty_initial_state()
    test_ctrl_jump_skips_hidden_rows()
    test_find_in_preview_mode()
    test_cell_status_one_sided_is_added()
    test_filter_keeps_merged_rows_visible()
    test_load_xlsx_with_empty_fill()
    test_goto_changed_focus_and_selection_color()
    test_diff_filter_after_key_change()
    test_row_header_multi_stage()
    test_sheet_path_absolute_target_saves()
    test_key_col_reset_when_out_of_range()
    test_sheet_tabs_mark_changed()
    test_compute_diff_key_row()
    test_key_row_header_rendering()
    test_key_row_reset_when_out_of_range()
    test_freeze_main_hides_frozen()
    test_frozen_span()
    test_freeze_setup()
    test_freeze_bulk_guard()
    test_freeze_corner_menu_nonmodal()
    test_freeze_cells_selectable()
    test_large_selection_queries_range_based()
    test_char_range_cache_staging_safe()
    print("ALL MODEL/VIEW TESTS PASS")
