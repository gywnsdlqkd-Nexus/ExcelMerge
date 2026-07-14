# -*- coding: utf-8 -*-
"""loaders 커버리지 보강 — calamine→openpyxl 폴백, 시트명 목록, 값 캐시.

기존 테스트는 xlsx 정상 경로만 다뤘다. 여기서는 calamine 이 실패했을 때 openpyxl
폴백이 동일 결과를 내는지(핵심 견고성 계약)와 시트명/캐시 유틸을 검증한다.
"""
import openpyxl

from excelmerge import loaders
from excelmerge.loaders import (
    load_values_any, list_sheet_names, clear_values_cache,
)


def _make_xlsx(path, rows, sheet="Sheet1"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    for r in rows:
        ws.append(r)
    wb.save(str(path))
    return str(path)


def _make_multisheet(path, sheets):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, rows in sheets:
        ws = wb.create_sheet(title=name)
        for r in rows:
            ws.append(r)
    wb.save(str(path))
    return str(path)


def test_calamine_to_openpyxl_fallback(tmp_path, monkeypatch):
    """calamine 이 사용 불가일 때 openpyxl 폴백이 같은 값 매트릭스를 낸다."""
    p = _make_xlsx(tmp_path / "v.xlsx", [["ID", "V"], ["1", "hello"], ["2", "world"]], sheet="S")

    clear_values_cache()
    via_calamine = load_values_any(p, None, "S")

    # calamine 비활성화 → _load_values_pass 가 openpyxl 경로로 폴백
    clear_values_cache()
    monkeypatch.setattr(loaders, "_get_calamine", lambda: None)
    via_openpyxl = load_values_any(p, None, "S")

    assert via_openpyxl == via_calamine, (via_calamine, via_openpyxl)
    assert via_openpyxl[1] == ["1", "hello"], via_openpyxl


def test_list_sheet_names_order(tmp_path):
    p = _make_multisheet(tmp_path / "m.xlsx", [
        ("Alpha", [["a"]]),
        ("Beta", [["b"]]),
        ("Gamma", [["c"]]),
    ])
    assert list_sheet_names(p) == ["Alpha", "Beta", "Gamma"]


def test_list_sheet_names_non_excel_returns_empty(tmp_path):
    p = tmp_path / "x.json"
    p.write_text("[]", encoding="utf-8")
    assert list_sheet_names(str(p)) == []


def test_clear_values_cache_reload(tmp_path):
    """캐시를 비운 뒤에도 재로딩이 정상 동작하고, 파일 변경이 반영된다."""
    p = _make_xlsx(tmp_path / "c.xlsx", [["ID"], ["old"]], sheet="S")
    assert load_values_any(p, None, "S")[1] == ["old"]

    # 같은 경로를 새 내용으로 덮어쓰고 캐시를 비우면 새 값이 보여야 한다
    _make_xlsx(tmp_path / "c.xlsx", [["ID"], ["new"]], sheet="S")
    clear_values_cache()
    assert load_values_any(p, None, "S")[1] == ["new"]
