# -*- coding: utf-8 -*-
"""xlsx_writer 서식 병합 + 행/열 삭제 + 신규행 append 테스트.

`_write_patches_to_file` 를 공개 진입점으로 사용해, ExcelMerge.md 가 광고하는
'서식 보존 병합'(표시형식·폰트·채우기)이 실제로 대상 파일에 반영되는지 openpyxl 로
재로딩해 검증한다. 이전에는 값 패치만 테스트됐고 _StyleMerger 경로는 커버리지 0이었다.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill

from excelmerge.xlsx_writer import _write_patches_to_file, _promote_empty_cols_to_delete
from excelmerge.loaders import load_values_any, clear_values_cache


def _make_xlsx(path, rows, sheet="Sheet1"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    for r in rows:
        ws.append(r)
    wb.save(str(path))
    return str(path)


def _reload(path, sheet="Sheet1"):
    return openpyxl.load_workbook(str(path))[sheet]


def test_style_merge_copies_numfmt_font_fill(tmp_path):
    """소스 셀의 커스텀 표시형식(numFmt>=164)·굵은 폰트·노란 채우기가
    대상 셀로 병합되는지 확인 (_StyleMerger.map_index + _map_numfmt 커스텀 경로)."""
    src = str(tmp_path / "src.xlsx")
    dst = str(tmp_path / "dst.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "ID"; ws["B1"] = "V"
    c = ws["B2"]
    c.value = 1.5
    c.number_format = '0.000"x"'                       # 커스텀 → id>=164 강제
    c.font = Font(bold=True)
    c.fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    wb.save(src)

    _make_xlsx(dst, [["ID", "V"], ["1", "plain"]], sheet="S")

    _write_patches_to_file(
        dst, {"B2": "1.5"}, sheet_name="S",
        src_path=src, src_sheet_name="S",
        patch_style_src={"B2": "B2"},
    )

    ws2 = _reload(dst, "S")
    cell = ws2["B2"]
    assert cell.number_format == '0.000"x"', f"표시형식 미병합: {cell.number_format!r}"
    assert cell.font.bold is True, "굵은 폰트 미병합"
    assert (cell.fill.fgColor.rgb or "").endswith("FFFF00"), f"채우기 미병합: {cell.fill.fgColor.rgb!r}"


def test_delete_row(tmp_path):
    """delete_row_nums 로 지정한 <row> 가 실제로 사라지는지."""
    p = _make_xlsx(tmp_path / "d.xlsx",
                   [["ID", "V"], ["1", "a"], ["2", "b"], ["3", "c"]], sheet="S")
    _write_patches_to_file(p, {}, delete_row_nums={3}, sheet_name="S")  # 2번째 데이터행(row 3)
    clear_values_cache()
    vals = load_values_any(p, None, "S")
    joined = ["|".join(r) for r in vals]
    assert not any(row.startswith("2|b") for row in joined), f"삭제된 행이 남음: {vals}"
    assert any(row.startswith("3|c") for row in joined), f"다른 행이 잘못 삭제됨: {vals}"


def test_delete_col(tmp_path):
    """delete_col_letters 로 지정한 열의 모든 셀이 제거되는지."""
    p = _make_xlsx(tmp_path / "c.xlsx",
                   [["A", "B", "C"], ["1", "2", "3"]], sheet="S")
    _write_patches_to_file(p, {}, delete_col_letters={"B"}, sheet_name="S")
    clear_values_cache()
    ws = _reload(p, "S")
    # B열 셀은 비어야(값 없음) 한다
    assert ws["B1"].value in (None, ""), f"B1 미삭제: {ws['B1'].value!r}"
    assert ws["B2"].value in (None, ""), f"B2 미삭제: {ws['B2'].value!r}"
    # A/C는 유지
    assert ws["A1"].value == "A"
    assert ws["C1"].value == "C"


def test_insert_new_row_appended(tmp_path):
    """insert_rows 로 파일 끝에 새 행이 추가되는지."""
    p = _make_xlsx(tmp_path / "i.xlsx", [["ID", "V"], ["1", "a"]], sheet="S")
    # (col_idx, value) — 0-based 열 인덱스
    _write_patches_to_file(p, {}, insert_rows=[[(0, "2"), (1, "b")]], sheet_name="S")
    clear_values_cache()
    vals = load_values_any(p, None, "S")
    joined = ["|".join(r) for r in vals]
    assert any(row.startswith("2|b") for row in joined), f"신규 행 미추가: {vals}"


def test_formula_value_written_as_formula(tmp_path):
    """'=' 로 시작하는 패치 값은 <f>(수식)으로 기록되어야 한다."""
    p = _make_xlsx(tmp_path / "fm.xlsx", [["ID", "V"], ["1", "0"]], sheet="S")
    _write_patches_to_file(p, {"B2": "=A2+1"}, sheet_name="S")
    ws = openpyxl.load_workbook(p, data_only=False)["S"]
    assert ws["B2"].value == "=A2+1", f"수식으로 기록 안 됨: {ws['B2'].value!r}"


def test_promote_empty_row_to_delete(tmp_path):
    """패치 후 전 셀이 빈값이 되는 행은 delete_row_nums 로 승격된다."""
    p = _make_xlsx(tmp_path / "e.xlsx",
                   [["ID", "V"], ["1", "a"], ["2", "b"]], sheet="S")
    # row 2(=["1","a"])의 두 셀을 모두 빈값으로 → 행 삭제 승격 기대
    patches = {"A2": "", "B2": ""}
    new_patches, new_deletes, del_cols = _promote_empty_cols_to_delete(
        patches, set(), p, sheet_name="S")
    assert 2 in new_deletes, f"빈 행이 삭제로 승격 안 됨: deletes={new_deletes}"
