import sys
import os
import re
import json
import struct
import zipfile
from collections import defaultdict
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from lxml import etree
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QFileDialog, QTableWidget, QTableWidgetItem,
    QSplitter, QLineEdit, QPlainTextEdit, QHeaderView, QFrame, QStatusBar, QMessageBox,
    QMenu, QShortcut, QStyle, QAbstractItemView, QScrollBar, QStyleOptionSlider
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QSize, QRect, QPoint, QItemSelection, QItemSelectionRange, QItemSelectionModel
from PyQt5.QtGui import QColor, QFont, QIcon, QPixmap, QKeySequence, QPainter, QPen
import base64 as _b64


_APP_ICON_B64 = "iVBORw0KGgoAAAANSUhEUgAAAJQAAACUCAIAAAD6XpeDAAAQAElEQVR4Aex9CVxUVfv/c+cyA7IjoKKYgOKCu0maoVi4ppWWpRmi9leycjdp0VRcMnnNwi0jc6U3ezOlxD38ZbmGqajhggsGCYoo+zJwZ/7fc+/MOOKwqDDL+w6fZ+6ce+5Znvt859nOuTPI1Jf7WMlCJSAj65/FSsAKnsVCR2QFzwqeBUvAglm3ap4VPAuWgAWzbtU8K3gWLAELZr0ONc+CpWIhrFvBsxCgDLFpBc+QVCykzgqehQBliE0reIakYiF1VvAsBChDbFrBMyQVC6mzgmchQBli0wqeIalYSJ1lgmchwq1rNq3g1bWE63B8K3h1KNy6HtoKXl1LuA7Ht4JXh8Kt66Gt4NW1hOtw/P9S8DgVVU11KFLjDf1fBJ6EloIjBaeyaZBbXD/ttsPVTNeL6Z4SoYwa1OMqE7DYEo01MLMqC3vJLIxfg+wCNiAh54HN9zvqz17WfOwMl9ApjV6Z0OKFMU37jPCSCGXUDJvQGFffmtdi7hdN0TjpvDvgJDnPUMQgGApE2j/9srbOfN4tHzwFB+kDhicHNXuim8+Id70XfWG3aasiPqE08VRBckpRekahSKUoo+aXQ8WbtspiYu3mL3ND4079n3Bt59k2JACQ7/2t2cV0T6aXOhTVZi0fs2au+s+4goPqQMlGvOt18mw9sX1e10DXBYue/Xr9mLj4aaCEgx+KNBNHnIIWLOo7aUqXroFO3l4CUT6RTXIKB8gHhDq+ML758IluKzY2gqXVqKMZK59MvGELPECmCg5S7tTfMz7BlqgwdFTzuPi3zyQvXbI0rFHjhikXUzdt3g9atXqHjn7++Y+E/Sdwt716+6PZroTIhIMz0StyUWDoKB9vL0q5lLV1h83kjxu0Dm4AVcb48JQai4pumBRHsyHLBA9ClPNnkptN/tidiA/wL4c+zVv41tmkCxHvrf72298g3g7duoWFD69ALw4f0K13D9+ADqn/qOO2HYpavH7l8k2As3Fjn7HjXgSWGGfSlEBRI8uhypM/doWnfOuj5tBvjMl0kb2Zy8sywVPLVGr3BavkQA5GcsXaKVlZdye+/UU5OYeOf7VNh5a795yI/fqHNZ+tXxn1rUQoS3Rg96Hjvx6B+Lv17hHx4dh/LXt//Fsv4XRX/H4ADyChlAePfnj0+PTIRd29vXh4yphYZ+j3q9P8NRDio4MOZkCWCR6nSk1Xbt0B+Snfeqevp6f7+LEb7J3sjx9OBDaoDQsfPmHGWFBEZJhE0+eMA6FGUj60AYTzIjfMnL4E+ofT5wf3Xfnl1CEvB+XkqN6b8eXatdv69XkKKMJ3BnYuQ4OtO2wBIcJUjSHVh1C/jKbGIssET84X5NeHiPxbOnXr1g4FUJNGbsBmxJuDOnb1t3fg/bzqieTYyFlx/fylHd/t/iVuX9bfV1D5XFDzN4Y/PXf+qJUrJ69aMxcGE9337j4K3cXR1VUWuWDalGkTtm//49131slUmVu2f/DD9pEDnm9IVBIT69D8mYafb/DTBKXoCTJRUGqZ4BEdTCwlsinOL8rKyrZ3UAx9pQnMIDCDJEF+Xo6oZGRvF7spHsFLh07NewR1xqW5H39+PvkaaPu2Izvjj/xx/DwUd9DgHlNnjIxa+k7/gU9fSM6IeG919OdrUIYuujdsA0VMTr46d+4waCH8K+adPscRQenFa/VZLINBTUQyE837GNMquD0H3RBKECnSM+4eOXTa3t5u9NjX4OSkQaFqUgFHhYK2b/sr6tND48fGDBm85OCBRJjHkODID6auAKJokJR06uuvfoqYHg0gAXYzH6//Fz4UmI0MfRVaOPHtL+BN4RcDAvwiI7feuJG65af5iGiIirbucGod7Hwmudk9/IxuPC0NPDGxGzjKHaFKn6Cyb7558eNZu29n5T3Vrc0zz7TatOpb4KFPOTklER8OIion8iBqGLv5SuRHP0yaEhI+eZint++zz3V5I/RlKGWr1k8Ay3mzv0LfoqISHJv5eIW//TJQzLxxE34RjhBlBKVQXPjFuPhJRDeJHDoOkOOTpMEPxrNK/DBs7ZJFgafgYKlGToTvsUdiF7N+/OihtgN75iPugMShMcAA+GXmKaGLOjEBIZg7ottE/8BppWdk3b1zB5C4OqjR7LvY3X5+Xj2COjk51Qt+LhCnqNT1RQHDQvNQgBa6uspQXr/2Z+QkZ5KX+7cEGzb4JN3TP+CHpsYiywEPy2B5LhNnNcRqCKzWH8fGPOF+nc/8fOWnytTks9HLfpDwAwBTRy2EJ9MXIPCDrOPiZwHFhIML3OqzYAd+Dl3QrE2ALwh5Pcr5+cU46pPUBiNA8xKPX42IiEGCgQY/bU/4Me595BJEtuM+dErLsDP+AreFgMepkNhFfOrxyyEwXLpn50uBXR25m/+mcsGvYd66pQUrond8suh7CBqQwC1t/X4fbGBR4T0VhD9r157ZSdjDKdNfzcnJgp+7npoBndsZzyIX4IHsEEEKBkHZIE2dMfLdd15AOINegPn48XMbt0yGDUg8JVv7vafxU3jIwiCf5lQJR6KWRUbbYzWZqPSbb/r07++lurpKVnYDno+K1B2bZl05LFw6daB39w8RRnp4OkctmzJseD9IefHCzYAQN1PBGM6ICB877kUssmCF5czpKyAgN3BAV0CLxpURcEUDBKXo1SOoM/J6nE6a0hOWAPELUz45X1nfuqg3b/AAG0jOr/jOa/4yZyJhxsxOb45uwZArucSQk0RSTn5uuTu/KHjz9byQ4MVffPZvSDmoZxsYul69/RFiLF64XDKkOgjRAHL/cPYopHQIQEL6dp03d8ygwT2k8ao+QonRC8rt7e33fwdOhrzYj6g8OcXmr0vSynjVvWvzqnmDB/8v5xHOiYmBMPQV96VRz6vSvpfpIydJo5yc3XIjx13fvTn3SMJeqCDSOFwZNqT7tq1zP5o15veDvyPi+CZmO0wldFFCERACCaDYrr0vCjhFl5oQGr/0crcV0YnQVySXXQObEClv3UH8UpPetdbGvMETXd2M+V5QssDONjErh6lvbJflY2XSkHVCOlAuDHgy+7cNuRHvZI8f+8WE8NVb445BVH7NGiyNCo/97oPWAV7I3mBOP4uKOXf2moSiBCSa1ZDQ3tXVTi53QbQCF4hemek54BAFI5N5gyfnY+PsYJGIhAUL+7vZpnLZ26oUE0/lxAkFk17PEM7LXww6vXjONxwXvvCT7Vev33J3JijiksUjgSJS9SOHTsGivjn6U5hZWD9JIwGnBAAQqkBSPRrAAkfO2bxx/X8QrWB1JulECjZ7cbVja5YgomA0MmPwOFVunssPPzuKsihv2bIJd/tnsVzhgA1VqQYFiRBAqBHOTHrlRtzamxujy/bGxTX3mT0qLHpr3DEJReQGCB2RtCH6QOiB/ghekC9CKZGSw0dCNUEwsxKhjEpcQgOgDh/59sx3BVunMxezYpbHEzkHdubatyYqAwMYzEh0DzwjTVjzaeT89XTX+AQbqB06CbcPycqvVKZ2KnlDlV1LRvLGaCx24aGFTesVhfVN+33jP6f35nfxPx/9yZfNfSJatpoFSBCXQo3gvbA6M2hwDwQvWKdet/EDIIq1Mahmz+Ce7Tu2Rv6HI8qoRJwCsEPDBtt7eBYVCocTkkKHzk25VESkXLs4X8ZlE5y0OL1xDmYMHlHCH/BtApbzA/wVf5789QGJ4GMuqOSNVc5BnFMg59COkWMnnKJSxE/sUU5AEenEtGGXEZEm7ZFPGpV2ZM9/QoLffz5kLpJupBM7448ASyyzoQP8WcdOviDEq6ChL/fAEQRlxdrN1YzivQf+2rJu55CQyMhZX82Zbos8D72aeecYWe0wqVmDt3azHVSt37PukVEjNm7hVFxDIgAGthmpeReVczDn3J2zcRXPBVILxMlxygFCpoICq5deIoTOTnc6PHFt0sh/dq2/Dae4c31B73YnhDv7tq1bMWlcdIeAcY09X/NwCR/QJzI8fHnEzI06GvHassae7/TrPnVIyIw9K79yvL734MeX1evOjn4lJ8AfQSZ/9JSMjJvkEZntr/6JDi85RQkWAwKeGBqcvft3Yf+f9uSATxsggcI1JOdnOLk7AwyNKhAnB34V6sRTHlpIhRxSe5n6ZocWqfCLq+Zm7lxTvGND2uWfSnZv5uAjw1680q5Jkpf9MbviX4BTk8z4Yb6n10woOPjxrTOLUnfOuBo97Gwv3zQSbDyIGjdUE/HGzxNwO5AFjuZHcv7YaVcifKgLn+zizeUeSdrjMiC07EyKH9nwsIqwkySzM4ycdDcyOzTT11Sp+t4RulikZlgWqTkh18/pTnP/ogFPZocNSkewGvlu6sJJl1bNuxE9/NT8IefeeioRBMDae2Xb0l02iGBDKrUj3W7cCDuL/PETGI5VG/NlruAR/X7CBp9oGCWPRi5Ukgpzt2VVSccBZWf+9iU7H+L4qpCTRIhmUqH6I5yrqJSAoEgNvWTaCQUt5VjXMp5AQEsiVqV5yQpLm3qzNmcusMdq2Nq05oox3swSPNFmJp0FePnDxzwl3E3iVIVYgx7eP3tPrBz4bT3YqHrZiP6PCKhU37bSFsDS26XSq+KFlr4svbuT45afj9ZilbEOZgkeUWY2H5/AHocdGOIjK74Ks8YEUlTev8u1lKPdN2++9NHSG0VKe05eE/7hI1nvx3rJOMPd03M7tcKHrDivgMvOVRluU2e1Nbn5Opu8soHl/PHTWOR17hro7tvMRa3MFBtCh3i12rG5d/7Pm1o2bmT7VP/k/+zmYD9rBqE4xkMfqgfex+kORk3P4C9dtSE5mMSZkcgo4D38vRxOtCPiO3aq5+EpcEV/aQcQ1DIH4uTwdhND3f/YG5DwS/KLYZd+S+QNQMicYhmx1ILXdq/tdxnHfCERFjkx9C3rwjR8PvZdY2LZR75H9xZCXhon5AJISIeRnRiqoKQqsZfnfvVZm88Xtt2dkN5v+GVAWNGQlt1Gw8cjXuWANLyqMRzpdveuCDjpynUZKZE2VNW4dq+Zn+bJ+ZRUiMCZqPTZXi3h8LQ3DDh5JOCaU+gfJ1eXqVr4KT95r/GX/2rz53nFmCmnJUNKaugccWU3RdTRUdPpEd+gYQZ7qsAnu9CqOSvczmJZKTs31sv8wCPa95tCFHqRbwtXtcbhMQDUvCPL7e4XDfCDFYUjnBZWvCG6k7tDNrsOaNWCqLI4q0uzSSQrLHWvz8C7leOEyYxJ5geeUn3uogNRceioziTcER0epM9ILcfyGAEqAwLimBbaK4pCetVnDTheXZZloFldVtnbitEmJx7rciLd2GYGHqdKy6536i+EmiWhI1qLDq9ADDrAsEByD1IzFcRJZcQUEddUJcjr8V6bpDWSWFi5N6yMK1XW37aDPXPmWK+Q1RtxY8HMwJPzd+82TDyFtNeuZcsmssJkETmoHRPLPYfHzqp5sbz+MTN03Qz6sKGsc4F8eSm5LYpxOJToCT5bt+Q0TjpcAwAAEABJREFUT9/qOtZxQVbH4z/08D/ug9mxfXl4CydHTl18WdcfewgET6Y7r7KgFgqxn15lk6ov6vRbVxDbyzgC8eUEIvrtWtMey1ttudCSCKEK7+/DYk4y4p/Zgbf7/9hyVMcOHsjwtA8aCfhcM4eH1K0mokEzZQa61KStwTZ5+fURHInEmGGAIfaV1YeegYDZllOt+nz25Btr277eu2zSqDQi3r+lbUu/ciNv6ZkTeOKSZuIprBAK3Tq5C3kQCmQL5ESzCYeHs4pUybkATyn2quR6FdXA7MhVm82/uK3c5vzZGtdJa56YsrU9qE+ML/SsS1SXt7cEHL1kP65vzq6JidNHXLtwSU2kaNU0z8cbQXIVA9f+JXMCTy5uaRJEcAfbQKLDww3zog6JGV510QpaM1ILMpbhseIjvGBv2cZQ3zRs9c2YkLNsePa4Xn+DVg+5smHE+ZMRJ/+KSMB+3ojOF7E9VEAel67VxyzPBZf9rz8GcfSUI6Dyb9kM20D3OzzHBzM8iMwgqctzMIjBSzWrFNgmXzk7clyB7a1r7Rvk6kizmYeByvCpomsFigtXoADFIU/BQqDWqISJjTpfpZOJNvNkEnZfhbfH+7BtIKY9TEDwKMzhVdrz/gsah3d/5cOdaSZlncpJWr1kZUOvpAt26RkU4C/7X3+GRdwGwvZKeVDPtnzZFe36CGQmsAwP7zUhmFYBDq8mTatvo8pny2yVtvN2kRbQe3fPdrHLYY+OGTFDB1dmo3mabSAhsLMdtoFUubqdBEKSwDI8oAJ+qyPYTNHhSdpTXeuqr9uwpa9Km8i4PIcm4gI6362rDUk/mmTEDB2MmQ14ROKnWNGqHXN42iQBHJJmGwj2kJ1V+dK0qSX3A5uZnmtgPuTpIL48NR+hih1W8jTPShsXOTBmHuCJ30mIicUik/Bs76ZY0gRnWhKIdySgUjPNI5bhabs+1rvAbKYA7eMqDoM8HVXeLuJjpfXg8Iz/rDTmB5kHeETiNpAH0V1sA6lvHgRn9wgZXg2RI2kbiGrjTzS80LAHx9JWnj7DrnXtWGL8JIFNTGRq8CQPL+fFbSCwpGDbQIXnUBIJBpBnDk88qfagLstGlka1sqRpQ7LCSpa7oHly4Upxg/NXWGw8cgiYrJa1OmlgavDgJ4CfUi2uUwgvD28Hm6ndRMUN81jvIBn8CsrVETOtCA5rSZTlRAYdnpaLlLx64mJQkU/TfCOvimlZMLnmiYxgG+jXY+5YPgwPa4VVMTFJAAaMHiLDg2mtheceRIakA9JwKJlU1j/CbHq7nL/mRqQYHIL1aA4xlf51o5VNrXm4UXEbKDlFRWRb390JGR7qRNPHs4LCi4AKK1X3UpfpqWx1jau+jmwzs6pkUeVgu2ufLZHQpWORrHY/MVUzdv9VMwCPSNwG4oe+4urbzEV157iOQ5bh8dhV11VUUxAdHvS1mmbVX5ZsZmVqJxcKyOOXQ4iNFU93LmAZXvUj1kkLU4MHh0d04HdIi/z8vLCkKWO/8aC5VZbhaYrVvXE80nOogqiy1TWu4XWYxwdbAlFvl6PXkOE5IzY2/jaQPkemBo/oaprToURsm5X27d0MS5oic6LBJIFleDWMVtCt9swXy/Akh/cgfmLNz3tL8CkJCiw1/jYQblRHpgZPLj5oTNgGKsI2EJeXqOOMFeDw2FuNXqLDq1HLahohL9clCdAz/dYicnkOTf6+4Yrq53ramCrDw+wgU4NHJG0DBfi7engKeusjgsbh1Sxaqc0MD1KpLEkAlnIB20Di9yiUr/QzsfRMOr24KiZtA7FvA+WlyUrufeucObwa2sxazfBUd8tIgPZxALEiiZqHbSAitwB/tZvbzUfM8CqO+4jnJgVPXBWLT2BiGih+G0iMOLR3giVNbbGa91p/0E8EyfCk3i7SAnpQt9Km7sWG2xir1qTgyfnUNCfcaddAdw93Z+3D0agQ6aEcnqqQamVVjKrcBiJChhcTy4G/dq0KCdtAWCHCiYnIpOARbduDfFjRurUrljRl+cjwEGciUXtIh1eeIy7K1JIIK3N4UEe5cC6/JRF4VvZ6CktxtTTjow4je9SOj91PdHiHjmOdQoltIDFJAGwYFvgRWxWr4VOacHi1myRU5vDAmreL+NvW9by9yFTbQOBCR6YAD4k5SM6+DST+NFUJtoG0q2JaxrANBFS0Z9W8C1UtZVXTV/8ywhQkCdAw/Uq9cqmyvvg9Cno+RNwG0rtkkqIpwNP6CfHrr+yuYTNVRTdZSfti20A1SxJIVVJrzz2UU1U7CTIuXbA5xb5HoXwmEEm6llfTvZsCPNwt8FOqxZ+/UIaOasu2gfS//sq7UI2TBO2qGAatY+LLUzTbQEK3TsUsScBd1PGcVQ9vIvA4VW6Jq7gNVM4cXt59X39lGV7VXOuuQjs1Dk/yl7oLj1Jgq2ICTCcLJiv2hy29tw2kbuReC9NVnOLhz00EHlHqPzbiNhC1b9tI+3C0ln07H22pBu8CHB5iHFANGlfZhG2dA6RK2iBJiP2BYYZtIBdnQw8mVdKx7qoNgVd3s+lGlvMHjrsQKfoEybENpC44rbtCJD73AJXSq6q0qC4THV6l1x/uQmVJAkaRcdgGOin+i74eXUShmdpmMqbwMjYh1CTav599eDs+3dbDUxC3gXiRDWR4jmKhBod720A1aFyTJkLlNpMvF7eBsICu7N7JBD/wZ5B98UNk8EpdVubmuVy+6UekZNtAedK3gTTzsQyvhtEKemgcHkqPS8zhVW4zydvlyEkVrEJgZxszsZm4YdOAl52rKs4vIsrTfv0VnGgJGZ62WO17nWwDGZoVGV5yCtYThIHPMoNhqodWKrBmGvDyCxTS7zKDGzHWx7uGWIanKVbzVpvbQFVneEQX5C5bd6CR6beB9IViGvDs7NT6TOjK2MNjGV5NohWsv6ixusjCP133Ry4wm1mFw5MLt/NaETkE+MtMvg2kf4+mAU9uU+bt5UCkuJ2dp8eNHhLARu+CgSIALkk1UP9IVdUkCa0abPo+E9z27p7d1NTbQPr3ZxrwfLwVzo5QPruzf2XK7MVfV2FM8eLjX6xEwEZ8r+Ig/t5DFddr7RLn6rZpKwQlGP/3Hqq+B/BUdYM6uKqWybjsrh1L8FmO33W2sNBObw4BnoyqVTsxSRC3gXi9vo9RrCzDQ/wpF86k1yNi0Uq/Xko2h9oUQmMTV3wZmY9703ftiGiTLp65e7OoIdnw9y4Up9RE7Yg5vHudHqfEHF6ZyACgenAgbxfxuQdkeIX+Phxb0nywjYlqTAEekvQyITiQfZaxQhb7U6HKprn2AQgeKybsl6eqVb6SWnJ4NtqtcxlnEAIkCeJzDxQ2jBkMliSAf4NNjV5pCvBEs+PIfmO0kEgxd9axq9md9ZVPln+8GuOpKqlNh6ezmYbwwzbQiSSWJIT0FL80BORE/o2OlIEJTQGeyIa7iyzA314slk+b/ZfKPkirfKgTZHmH1Lovej2ghdofOOLR9HEJuBgcAiYUJBcKbRtKS5pmsg2kz6zJwHPmygJaw/8XLVjU92Ka89j3MlmSR1K2wFCRlVxS5ycyE4rIU4efVIBf1LTUv5dHKTOHZzDDk7RQ8/VXxeCQ8kbmsQ2kf5MmAk8t42wLW7HfGLXJzy/+Me79A4frTfzUqcJ/QoD/k+UfYRAqM7EQw0iZSbm/6i1k69/Lo5RZhld5P+02EK/ZBlKbSFyVcGg6bhSc+Buj/MlTmY2bOO9KiFy92Wl4hAvDz0ZilukfSoAKEMKQMso/wgkFRJpLuPq4BIcH82hwFL1tIPZtIINtTFppOvCI3F0QAvC/7E3LySnx8HQ+k/xpjrJF4Gv2SWme5IAUXhIMcJJIdyoVHv8o/mQAkgTJQj44Hl+++7ScCElC/tOdVWaVJEjMmhK8BvWR6sHJZRUVKouKGH7790z3CWjfqb+w54gH+0KzRgUlVmv9yFdlM1Xq0ga+vx0DeDx2jNk2kJnZTIjDdOCVCf16Yfa7RM4/bU+wt9ess/y4ZfIP26cPHCWHC8xVuZE9pxeFguFK6aEv4JMBm2mwGwypXLho02T1Zg+i4lGv6i/AGuxgmkqIzzQTI9vFItmc6TyR7aaNJ84nX1PAPom8DBvS/ejxeVez2/YY2mTTTm8WhULQ4qXaPFSWJGAOGFJvl1n/wueJHxyifimkwAxtJmMTL5NRmTD6lRxvL0q5lL/4k+36bHR/qsXunR9M++i10VPqDZpQb8+f7lorCjOr3/AxypkFFb8NBIUDYUi+/PNjneITbFGcOLbEPG0meDOd5mFyIr+m+UtnZxAhbEmdOjH66vVbYrXmMG5Mj9u5q1t0eX7ygha9xrjsPemriUUfUxFtiKV31ws10+i/yQXiy5ee6jZ9Tj1Uh4cW9u913cj/6gLz1pBMCh5CgDJh+At35kyHU1HEbr747jvrKuCH2/hw9igkgj1C+g8IlSMWXfFj44tZHuQqY9/3YCg+pC4Cubtlsou37vuRdhL/5EKp4Dknrt3MFW5EimEvFEV9cNs8DabIrsl/h0XEL3Jq2sbobCL7PbtuDugXvTXumMQcjkolIRC1d1BMnTESucTb08fuOhrQZ4TX82+4b9rfFChqPCJQBFW/7CJQeoEs+Q77FU04Nkwgkfg77VtOteoS1WXBz81QFxSYuXZxlku9O/DNODVPMqnmSSIR8Qsb+s+WVelEpSmXSl8dumb+J/++er8JlSAcNLjHyi+nbtwyuceA1xauatw6WDVmbiPoIpyiBkhEpxIBS33CXHByJzJJspZyZh5hIUvJ7ewtl6/+CBz0md/rK3zFp4wowF+9dmmpmSOHGzID8MCFiB/s55XDN4MCs4g8sNXQ3Gf22g1HGGbaLAINcYpjMx+v10MHHjy8KOHgkva9R/96ruucZU1eGN9c1ta15+uNZ69sgRgVDvLM376gKyn21w6VnN1RcjbR5myG+29/e4GgZDCPfaK62Y1v/WZc3313u454P/zl4b5EWG4VvltR2MrXrHUOQgDJ8DILAn5KNeKX+PVFyxdIYYvb+LEbRgxbsjP+CDjUJYIoSwQg2wT4vjH86aio8G++nbfm63Fx8e+HDAnLKOq+ItZnQGi9jgNyQS1esvMLb9xhloaCFzQOXuAzc6vv0dKng0aN+tfKsUuWhsXETMaY275nP1i3bH5xh4Dr5uzqwKpEZgOexI5SDWM1aXTmlcP/zJl+m8jhRGLO+LFbWjSb/k3M9uupiEtJH0XgB0JXOEWoo59fk2dCOo54c9AnK97dd2zOvmOr4xJWgGK3L5YI5biEJfuOfbJu+6yIyDA07v9cW3wCDv1+HrMQOb8z6vaU0bkMOXyYMK5E5no0M/AgJkgNKtgoB1HM6b2ZkCbqiOw/nrU/JPjzebO/khTR1dVOH0W0KSpUZubB6KF4j+wdeFADLxeJUAbduwwD7emM9YGZ02OJ+AD/8g8m5mLpwJyDFH3mzQ88HQYlUUUAAAKhSURBVHdKdcc22asW3bxy+J9l8wsCO7P1jtjNqeOZIk4Iff1ToAhdBGagqxkFDyKnG6myQiNnxR/Hz4cER6ZcKsJawba1mU29Ssw2q3vwLswVPOgfmFWqIUq/RjnTxlz9cc3lLasyggLvoJqoIZIKeMTRI5ZHvLd6XuSGwwlJtzJyiwprlPMBM7RMOpEStXj9kMEroH6Icr+OusGCFMwoTmARB3MFT194ALJMaOpRiHD09x8zcs5lbIy+FR5aEtjZMT2jdM+uG9u+vxY566fQoUuGhMzo1/2jjyat2rJu547vdgNRIKQjnKI+au6mDgHvoeXMiZuhx0S2SMYvHLwzIPguc3X685p92RLAk4QICKEWYkSDpPDLRTnQxaQ9ZcvmK/sEqYhKiGBXmyDoQIyzLubPFdGJkbN+BEI6winqsX1IhAUUtKQu7Qv2xBYgGdfoHKaQ5rKQo4WApy9WlJVqWfkt6CJieljU/f8+L1zPv3Dw1u7NqUgzsNgWHloWNkw57AUaHMIHdrbr0r4MBZyictbUkuULMtAy51zWnzuvY+kS8S2Ms4XgdR+bFgLefTyLJ4AQJOoiRA8soT0Dgu8izUCYCr1cPi8bKhUbnQkFPfBdDgo4Xf9Z7sKJl9EGLRlmZQL6WkpsKd72fQeLBe++uyAGgASkiAewBDYSQUGlgotdjuYnhcU2rEuFQSzt9L8FPJ3coY4o4/ggoV4iXKpQkE5NcHysKf/rwHssaVhYZyt4FgaYPrtW8PSlYWFlK3gWBpg+u1bw9KVhYWUreBYGmD67VvD0pWFhZSt4FgaYPrtW8PSlYWHlqsGzsJv5X2PXCp4FI24FzwqeBUvAglm3ap4VPAuWgAWzbtU8K3gWLAELZt2qeVbwHl4C1h6PL4H/DwAA//8Iz7RoAAAABklEQVQDALA6jvZzMqdeAAAAAElFTkSuQmCC"

def _load_app_icon():
    """앱 아이콘을 단일 소스(app_icon.ico)에서 로드.
    - PyInstaller 번들: sys._MEIPASS 임시 폴더의 ico 사용
    - 개발 모드: 스크립트와 같은 폴더의 ico 사용
    ICO는 멀티 사이즈(16/32/48/256 등)를 포함하므로 작업 표시줄·타이틀바 모두
    동일 이미지에서 적합한 사이즈를 가져온다. 실패 시 base64 PNG로 폴백.
    """
    base_dir = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    ico_path = os.path.join(base_dir, "app_icon.ico")
    if os.path.isfile(ico_path):
        icon = QIcon(ico_path)
        if not icon.isNull():
            return icon
    # 폴백: 임베디드 base64 PNG (ico 누락 시)
    raw = _b64.b64decode(_APP_ICON_B64)
    pix = QPixmap()
    pix.loadFromData(raw, "PNG")
    return QIcon(pix)

EXTRA_ROWS = 20   # 데이터 끝에 추가할 빈 행 수
EXTRA_COLS = 5    # 데이터 끝에 추가할 빈 열 수

# 자동 컬럼 너비 상한 — 150px
# 데이터가 긴 셀 때문에 열이 화면을 가리지 않도록 제한.
# 사용자가 헤더 드래그로 직접 넓힌 열은 _user_col_widths 에 기록되어 이 상한 무시.
# 새로고침 시에는 _run_refresh()가 _user_col_widths를 비우므로 모든 열이 디폴트로 복귀.
MAX_AUTO_COL_WIDTH_PX = 150

DIFF_COLORS = {
    "added":    QColor(198, 239, 206),   # 연두   - B에만 있음
    "modified": QColor(255, 235, 156),   # 노랑   - 값 변경 (삭제됨 포함)
    "staged":   QColor(255, 185,  80),   # 주황   - 저장 대기 중
    "merged":   QColor(173, 216, 230),   # 연파랑 - 병합 완료
    "same":     QColor(255, 255, 255),   # 흰색   - 동일
}


# ── 데이터 로직 ────────────────────────────────────────────────────────────────

def _com_val_to_str(val) -> str:
    """Excel COM에서 반환된 값을 문자열로 변환 (정수형 float 처리 포함)."""
    if val is None:
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val)


def _cell_to_str(v) -> str:
    """openpyxl 캐시값을 문자열로 변환 (정수형 float은 정수로)."""
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v)


def load_sheet_with_formulas(path: str) -> tuple[list[list], list[list]]:
    """
    한 번의 호출로 (계산값 시트, 수식 시트)를 함께 반환한다.

    핵심 개선:
      - read_only=True 로 두 워크북 모두 로드 → 메모리/속도 모두 향상
        (기본 모드 대비 수~수십 배 빠름).
      - Excel COM 의존 완전 제거 — win32com 호출 없음.
      - 미캐시 수식 셀만 _eval_formula_with_row 로 같은 행 컨텍스트 계산
        (전체 시트를 셀 단위로 COM 호출하던 기존 폴백 제거).

    수식 시트는 NetmarbleCompare.md의 '수식 보존 병합' 기능을 위해
    저장 시 원본 수식(=...)을 그대로 기록하는 데 사용되므로 항상 함께 반환.
    """
    # 1차 — 캐시값
    wb_val = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws_val = wb_val.worksheets[0]
    values: list[list[str]] = [
        [_cell_to_str(v) for v in row]
        for row in ws_val.iter_rows(values_only=True)
    ]
    try:
        wb_val.close()
    except Exception:
        pass

    # 2차 — 수식 텍스트 (data_only=False)
    wb_fml = openpyxl.load_workbook(path, read_only=True, data_only=False)
    ws_fml = wb_fml.worksheets[0]
    formulas: list[list[str]] = []
    needs_eval: list[tuple[int, int, str]] = []   # (r, c, formula_text)
    for r_idx, row in enumerate(ws_fml.iter_rows(values_only=True)):
        out_row = []
        for c_idx, v in enumerate(row):
            if v is None:
                out_row.append("")
                continue
            s = str(v)
            out_row.append(s)
            # 같은 좌표의 캐시값이 비었고 수식 텍스트면 자체 계산 대상
            if (
                s.startswith("=")
                and r_idx < len(values)
                and c_idx < len(values[r_idx])
                and values[r_idx][c_idx] == ""
            ):
                needs_eval.append((r_idx, c_idx, s))
        formulas.append(out_row)
    try:
        wb_fml.close()
    except Exception:
        pass

    # 미캐시 수식만 자체 계산 — 캐시가 모두 채워진 일반 파일은 0회
    for r_idx, c_idx, fml in needs_eval:
        if r_idx < len(values) and c_idx < len(values[r_idx]):
            values[r_idx][c_idx] = _eval_formula_with_row(fml, values[r_idx])

    return values, formulas


def load_sheet(path: str) -> list[list]:
    """하위 호환용 — 계산값 시트만 반환."""
    values, _ = load_sheet_with_formulas(path)
    return values


def _eval_formula_with_row(formula: str, row_data: list) -> str:
    """수식을 같은 행 데이터를 컨텍스트로 계산. 실패 시 수식 문자열 그대로 반환."""
    try:
        import formulas as _formulas
        import numpy as _numpy
        # A22, B5 등 행 번호를 전부 1로 정규화 (같은 행 데이터로 매핑하기 위해)
        normalized = re.sub(r'([A-Za-z]+)\d+', lambda m: m.group(1).upper() + '1', formula)
        col_refs = re.findall(r'([A-Z]+)1', normalized)
        kwargs = {}
        for ref in set(col_refs):
            c_idx = column_index_from_string(ref) - 1
            if 0 <= c_idx < len(row_data):
                val = row_data[c_idx]
                try:
                    fv = float(val)
                    val = int(fv) if fv == int(fv) else fv
                except (ValueError, TypeError):
                    pass
                kwargs[f'{ref}1'] = _numpy.array([[val]])
        func = _formulas.Parser().ast(normalized)[1].compile()
        result = func(**kwargs)
        v = list(result.flat)[0] if hasattr(result, 'flat') else result
        if isinstance(v, float) and v == int(v):
            return str(int(v))
        return str(v)
    except Exception:
        return formula


def load_sheet_formulas(path: str) -> list[list]:
    """하위 호환용 — load_sheet_with_formulas()의 두 번째 결과만 반환."""
    _, formulas = load_sheet_with_formulas(path)
    return formulas


def _json_value_to_str(v) -> str:
    """JSON 스칼라/구조 값을 셀 표시용 문자열로 변환.
    - dict/list 등 비스칼라는 compact JSON 으로 직렬화 — 셀 한 칸에 들어가도록.
    """
    if v is None:
        return ""
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)):
        return _cell_to_str(v)
    if isinstance(v, str):
        return v
    return json.dumps(v, ensure_ascii=False)


def _flatten_json(node, prefix: str = "") -> list[tuple[str, str]]:
    """객체/배열을 점-경로 평탄화 — `[ (경로, 값문자열), ... ]`. 폴백 표기용."""
    rows: list[tuple[str, str]] = []
    if isinstance(node, dict):
        if not node:
            rows.append((prefix or "(empty object)", "{}"))
            return rows
        for k, v in node.items():
            key = f"{prefix}.{k}" if prefix else str(k)
            if isinstance(v, (dict, list)):
                rows.extend(_flatten_json(v, key))
            else:
                rows.append((key, _json_value_to_str(v)))
    elif isinstance(node, list):
        if not node:
            rows.append((prefix or "(empty array)", "[]"))
            return rows
        for i, v in enumerate(node):
            key = f"{prefix}[{i}]" if prefix else f"[{i}]"
            if isinstance(v, (dict, list)):
                rows.extend(_flatten_json(v, key))
            else:
                rows.append((key, _json_value_to_str(v)))
    else:
        rows.append((prefix or "(value)", _json_value_to_str(node)))
    return rows


def load_json_as_matrix(path: str) -> list[list[str]]:
    """JSON 파일을 비교용 2D 매트릭스로 변환.
    - 최상위가 객체 배열 [ {...}, {...}, ... ] 이면: 첫 행=키 union(객체 등장 순서 보존),
      이후 각 객체를 한 행으로 — 표 형태로 행 단위 비교 가능.
    - 그 외(객체/스칼라/스칼라 배열 등)는 점-경로 평탄화로 [경로, 값] 2열 폴백.
    """
    with open(path, "r", encoding="utf-8-sig") as f:
        node = json.load(f)

    if isinstance(node, list) and node and all(isinstance(x, dict) for x in node):
        # 키 union — 첫 등장 순서 보존
        keys: list[str] = []
        seen: set[str] = set()
        for obj in node:
            for k in obj.keys():
                if k not in seen:
                    seen.add(k)
                    keys.append(k)
        matrix: list[list[str]] = [list(keys)]
        for obj in node:
            row = [_json_value_to_str(obj.get(k, "")) for k in keys]
            matrix.append(row)
        return matrix

    # 폴백: 평탄화 2열
    matrix = [["path", "value"]]
    for path_str, val in _flatten_json(node):
        matrix.append([path_str, val])
    return matrix


_UASSET_MAGIC = 0x9E2A83C1


def _read_fstring(buf: bytes, offset: int) -> tuple[str, int]:
    """UE FString 디코드 — Int32 length + (ASCII | UTF-16LE) + null terminator.
    음수 length는 UTF-16, 양수는 ASCII. 길이에 null terminator 포함.
    반환: (문자열, 다음 offset). 실패 시 ("", offset+4)로 진행해 풀 손상 시에도
    가능한 한 많은 엔트리를 노출.
    """
    if offset + 4 > len(buf):
        return "", offset + 4
    (n,) = struct.unpack_from("<i", buf, offset)
    offset += 4
    if n == 0:
        return "", offset
    if n > 0:
        # ASCII (length 에 null 포함)
        end = offset + n
        if end > len(buf):
            return "", offset
        raw = buf[offset:end - 1] if n >= 1 else b""
        try:
            s = raw.decode("ascii", errors="replace")
        except Exception:
            s = ""
        return s, end
    # UTF-16
    count = -n
    end = offset + count * 2
    if end > len(buf):
        return "", offset
    raw = buf[offset:end - 2] if count >= 1 else b""
    try:
        s = raw.decode("utf-16-le", errors="replace")
    except Exception:
        s = ""
    return s, end


class _UAssetParseError(Exception):
    """uasset DataTable 본격 파싱 실패 — 호출자가 폴백 경로로 전환."""


_KNOWN_PROP_TYPES = {
    "IntProperty", "Int8Property", "Int16Property", "Int64Property",
    "UInt16Property", "UInt32Property", "UInt64Property", "ByteProperty",
    "BoolProperty", "FloatProperty", "DoubleProperty",
    "NameProperty", "StrProperty", "TextProperty", "ObjectProperty",
    "SoftObjectProperty", "EnumProperty", "StructProperty",
    "ArrayProperty", "MapProperty", "SetProperty",
}


def _try_decode_name_table(
    buf: bytes, off: int, count: int, hash_len: int,
) -> tuple[list[str], int] | None:
    """주어진 (offset, count, hash_len) 조합으로 NameTable 엄격 디코드.
    각 엔트리: FString(4B len + bytes + null) + hash_len bytes.
    잘못된 length가 1개라도 나오면 None — drift 누적으로 흡수되는 가짜 통과 차단.
    """
    decoded: list[str] = []
    cur = off
    n_buf = len(buf)
    for _ in range(count):
        if cur + 4 > n_buf:
            return None
        n = struct.unpack_from("<i", buf, cur)[0]
        cur += 4
        if n > 0:
            byte_len = n
            if byte_len > 8192 or cur + byte_len > n_buf:
                return None
            try:
                s = buf[cur:cur + byte_len - 1].decode("utf-8", errors="replace")
            except Exception:
                return None
            cur += byte_len
        elif n < 0:
            byte_len = (-n) * 2
            if byte_len > 16384 or cur + byte_len > n_buf:
                return None
            try:
                s = buf[cur:cur + byte_len - 2].decode("utf-16-le", errors="replace")
            except Exception:
                return None
            cur += byte_len
        else:
            s = ""
        cur += hash_len
        if cur > n_buf:
            return None
        decoded.append(s)
    return decoded, cur


def _parse_package_summary(buf: bytes) -> dict | None:
    """UE5 PackageFileSummary 정확 파싱 — LegacyFileVersion <= -8 대상.
    반환: {name_count, name_offset, total_header_size, ...} 또는 None (실패).
    """
    try:
        if len(buf) < 0x40:
            return None
        off = 0
        magic = struct.unpack_from("<I", buf, off)[0]; off += 4
        if magic != _UASSET_MAGIC:
            return None
        legacy = struct.unpack_from("<i", buf, off)[0]; off += 4
        if legacy > 0 or legacy < -10:
            return None
        # legacy_ue3_version (항상 존재, -7 이전엔 없지만 보통은 있음)
        off += 4  # legacy_ue3_version
        off += 4  # file_version_ue4
        if legacy <= -8:
            off += 4  # file_version_ue5
        off += 4  # file_version_licensee_ue4
        # CustomVersions Array (count + count * 20bytes)
        custom_count = struct.unpack_from("<i", buf, off)[0]; off += 4
        if custom_count < 0 or custom_count > 200:
            return None
        off += custom_count * 20
        if off + 8 > len(buf):
            return None
        total_header_size = struct.unpack_from("<i", buf, off)[0]; off += 4
        if total_header_size <= 0 or total_header_size > len(buf):
            return None
        # FolderName FString
        fn_len = struct.unpack_from("<i", buf, off)[0]; off += 4
        if fn_len > 0:
            if fn_len > 4096 or off + fn_len > len(buf):
                return None
            off += fn_len
        elif fn_len < 0:
            sz = (-fn_len) * 2
            if sz > 8192 or off + sz > len(buf):
                return None
            off += sz
        if off + 12 > len(buf):
            return None
        off += 4  # package_flags
        name_count = struct.unpack_from("<i", buf, off)[0]; off += 4
        name_offset = struct.unpack_from("<i", buf, off)[0]; off += 4
        if name_count <= 0 or name_count > 1_000_000:
            return None
        if name_offset <= 0 or name_offset >= len(buf):
            return None
        return {
            "name_count": name_count,
            "name_offset": name_offset,
            "total_header_size": total_header_size,
            "legacy": legacy,
        }
    except Exception:
        return None


def _scan_name_table(buf: bytes) -> tuple[list[str], int, int]:
    """NameTable 디코드 — 1차로 정확 파싱, 2차로 휴리스틱(1바이트 step).
    반환: (names, name_offset, name_data_end). 실패 시 _UAssetParseError.
    """
    if len(buf) < 0x40:
        raise _UAssetParseError("buffer too short")

    summary = _parse_package_summary(buf)
    if summary is not None:
        cnt = summary["name_count"]
        off = summary["name_offset"]
        for hash_len in (8, 4, 0, 16):
            res = _try_decode_name_table(buf, off, cnt, hash_len)
            if res is not None:
                decoded, end = res
                return decoded, off, end

    best: tuple[list[str], int, int] | None = None
    scan_end = min(len(buf), 0x400)
    for probe in range(0x20, scan_end - 8):
        cnt, off = struct.unpack_from("<ii", buf, probe)
        if cnt <= 0 or cnt > 200000:
            continue
        if off <= probe or off >= len(buf):
            continue
        for hash_len in (8, 4, 0, 16):
            res = _try_decode_name_table(buf, off, cnt, hash_len)
            if res is None:
                continue
            decoded, end = res
            if best is None or len(decoded) > len(best[0]):
                best = (decoded, off, end)
            break

    if best is None or not best[0]:
        raise _UAssetParseError("name table decode failed")
    return best


def _read_fname(buf: bytes, off: int, names: list[str]) -> tuple[str, int]:
    """FName 8B (NameIndex int32 + Number int32) → 문자열.
    Number > 0 이면 'Name_{Number-1}' 형식.
    """
    if off + 8 > len(buf):
        raise _UAssetParseError("fname truncated")
    idx, num = struct.unpack_from("<ii", buf, off)
    if idx < 0 or idx >= len(names):
        raise _UAssetParseError(f"name index out of range: {idx}")
    name = names[idx]
    if num > 0:
        name = f"{name}_{num - 1}"
    return name, off + 8


def _read_property_tag(
    buf: bytes, off: int, names: list[str],
) -> tuple[dict | None, int, int]:
    """UPropertyTagged 헤더 1개 파싱.
    반환: (tag_dict | None, value_off, value_end). tag_dict가 None이면 'None' 종료자.
    """
    name, off = _read_fname(buf, off, names)
    if name == "None":
        return None, off, off

    type_name, off = _read_fname(buf, off, names)
    if type_name not in _KNOWN_PROP_TYPES:
        raise _UAssetParseError(f"unknown property type: {type_name}")

    if off + 8 > len(buf):
        raise _UAssetParseError("size/index truncated")
    size, array_index = struct.unpack_from("<ii", buf, off)
    off += 8
    if size < 0 or size > 100 * 1024 * 1024:
        raise _UAssetParseError(f"unreasonable size: {size}")

    tag: dict = {"name": name, "type": type_name, "size": size,
                 "array_index": array_index}

    if type_name == "StructProperty":
        struct_name, off = _read_fname(buf, off, names)
        tag["struct_name"] = struct_name
        if off + 16 > len(buf):
            raise _UAssetParseError("struct guid truncated")
        off += 16
    elif type_name == "BoolProperty":
        if off + 1 > len(buf):
            raise _UAssetParseError("bool truncated")
        tag["bool_value"] = buf[off] != 0
        off += 1
    elif type_name in ("ByteProperty", "EnumProperty"):
        enum_name, off = _read_fname(buf, off, names)
        tag["enum_name"] = enum_name
    elif type_name in ("ArrayProperty", "SetProperty"):
        inner_type, off = _read_fname(buf, off, names)
        tag["inner_type"] = inner_type
    elif type_name == "MapProperty":
        key_type, off = _read_fname(buf, off, names)
        val_type, off = _read_fname(buf, off, names)
        tag["key_type"] = key_type
        tag["value_type"] = val_type

    if off + 1 > len(buf):
        raise _UAssetParseError("hasguid truncated")
    has_guid = buf[off]
    off += 1
    if has_guid:
        if off + 16 > len(buf):
            raise _UAssetParseError("propguid truncated")
        off += 16

    value_off = off
    value_end = value_off + size
    if value_end > len(buf):
        raise _UAssetParseError("value truncated")
    return tag, value_off, value_end


def _fmt_num(v) -> str:
    if isinstance(v, float) and v == int(v) and abs(v) < 1e15:
        return str(int(v))
    return str(v)


def _read_struct_value(buf, off, end, names) -> str:
    """중첩 struct = TaggedProperties 시퀀스. {k=v, k=v} 표기."""
    parts = []
    cur = off
    safety = 0
    while cur < end and safety < 256:
        safety += 1
        try:
            tag, vo, ve = _read_property_tag(buf, cur, names)
        except Exception:
            break
        if tag is None:
            break
        val = _read_property_value(buf, vo, ve, tag, names)
        parts.append(f"{tag['name']}={val}")
        cur = ve
    return "{" + ", ".join(parts) + "}"


def _read_array_value(buf, off, end, tag, names) -> str:
    """ArrayProperty/SetProperty 본문. 스칼라 inner는 inline, 그 외는 count 표기."""
    try:
        count = struct.unpack_from("<i", buf, off)[0]
        if count < 0 or count > 1000000:
            return f"array:{end-off}B"
        cur = off + 4
        inner = tag.get("inner_type", "")
        if inner == "StructProperty":
            return f"[{count} structs]"
        items: list[str] = []
        for _ in range(min(count, 5)):
            if cur >= end:
                break
            if inner == "IntProperty":
                if cur + 4 > end: break
                items.append(str(struct.unpack_from("<i", buf, cur)[0])); cur += 4
            elif inner == "Int64Property":
                if cur + 8 > end: break
                items.append(str(struct.unpack_from("<q", buf, cur)[0])); cur += 8
            elif inner == "FloatProperty":
                if cur + 4 > end: break
                items.append(_fmt_num(struct.unpack_from("<f", buf, cur)[0])); cur += 4
            elif inner == "DoubleProperty":
                if cur + 8 > end: break
                items.append(_fmt_num(struct.unpack_from("<d", buf, cur)[0])); cur += 8
            elif inner == "ByteProperty":
                if cur >= end: break
                items.append(str(buf[cur])); cur += 1
            elif inner == "BoolProperty":
                if cur >= end: break
                items.append("true" if buf[cur] else "false"); cur += 1
            elif inner == "NameProperty":
                if cur + 8 > end: break
                n, cur = _read_fname(buf, cur, names)
                items.append(n)
            elif inner == "StrProperty":
                s, cur = _read_fstring(buf, cur)
                items.append(s)
            elif inner == "ObjectProperty":
                if cur + 4 > end: break
                items.append(f"Obj({struct.unpack_from('<i', buf, cur)[0]})"); cur += 4
            else:
                break
        suffix = ", …" if count > len(items) else ""
        return f"[{', '.join(items)}{suffix}]"
    except Exception:
        return f"array:{end-off}B"


def _read_map_value(buf, off, end, tag, names) -> str:
    try:
        if off + 8 <= end:
            _, n_entries = struct.unpack_from("<ii", buf, off)
            return f"map(count={n_entries})"
    except Exception:
        pass
    return f"map:{end-off}B"


def _read_property_value(buf, value_off, value_end, tag, names) -> str:
    """tag 타입별로 값 디코드. 실패 시 hex: 폴백."""
    t = tag["type"]
    try:
        if t == "IntProperty":
            return str(struct.unpack_from("<i", buf, value_off)[0])
        if t == "Int8Property":
            return str(struct.unpack_from("<b", buf, value_off)[0])
        if t == "Int16Property":
            return str(struct.unpack_from("<h", buf, value_off)[0])
        if t == "Int64Property":
            return str(struct.unpack_from("<q", buf, value_off)[0])
        if t == "UInt16Property":
            return str(struct.unpack_from("<H", buf, value_off)[0])
        if t == "UInt32Property":
            return str(struct.unpack_from("<I", buf, value_off)[0])
        if t == "UInt64Property":
            return str(struct.unpack_from("<Q", buf, value_off)[0])
        if t == "FloatProperty":
            return _fmt_num(struct.unpack_from("<f", buf, value_off)[0])
        if t == "DoubleProperty":
            return _fmt_num(struct.unpack_from("<d", buf, value_off)[0])
        if t == "BoolProperty":
            return "true" if tag.get("bool_value") else "false"
        if t == "ByteProperty":
            if tag.get("enum_name") and tag["enum_name"] != "None":
                name, _ = _read_fname(buf, value_off, names)
                return name
            return str(buf[value_off])
        if t == "EnumProperty":
            name, _ = _read_fname(buf, value_off, names)
            return name
        if t == "NameProperty":
            name, _ = _read_fname(buf, value_off, names)
            return name
        if t == "StrProperty":
            s, _ = _read_fstring(buf, value_off)
            return s
        if t == "TextProperty":
            return f"text({value_end - value_off}B)"
        if t == "ObjectProperty":
            idx = struct.unpack_from("<i", buf, value_off)[0]
            return f"Obj({idx})"
        if t == "SoftObjectProperty":
            try:
                name, off2 = _read_fname(buf, value_off, names)
                sub, _ = _read_fstring(buf, off2)
                return f"{name}{':' + sub if sub else ''}"
            except Exception:
                return f"soft({value_end - value_off}B)"
        if t == "StructProperty":
            return _read_struct_value(buf, value_off, value_end, names)
        if t in ("ArrayProperty", "SetProperty"):
            return _read_array_value(buf, value_off, value_end, tag, names)
        if t == "MapProperty":
            return _read_map_value(buf, value_off, value_end, tag, names)
    except Exception:
        pass
    raw = buf[value_off:value_end]
    return f"hex:{raw[:16].hex()}{'…' if len(raw) > 16 else ''}"


def _row_name_is_valid(name: str) -> bool:
    """RowName이 정상 키 형태인가. 패키지 경로/Property 타입명/None은 거부."""
    if not name or name == "None":
        return False
    if name.startswith("/") or name.endswith("Property"):
        return False
    return True


def _try_count_rows(
    buf: bytes, row_map_off: int, num_rows: int, names: list[str], cap: int,
) -> int:
    """후보 위치에서 실제로 몇 row를 안전하게 디코드할 수 있는지 카운트.
    cap 만큼만 시도해 빠르게 비교."""
    cur = row_map_off
    n_buf = len(buf)
    decoded = 0
    target = min(num_rows, cap)
    for _ in range(target):
        if cur + 8 > n_buf:
            break
        try:
            key, cur = _read_fname(buf, cur, names)
        except Exception:
            break
        if not _row_name_is_valid(key):
            break
        safety = 0
        ok = True
        while safety < 1000:
            safety += 1
            try:
                tag, vo, ve = _read_property_tag(buf, cur, names)
            except Exception:
                ok = False
                break
            if tag is None:
                cur = vo
                break
            cur = ve
        if not ok:
            break
        decoded += 1
    return decoded


def _find_row_map_start(
    buf: bytes, scan_from: int, names: list[str],
) -> tuple[int, int] | None:
    """Export 데이터 영역에서 RowMap 시작점 탐색.
    검증: int32 NumRows + FName RowName(키 형태) + 첫 PropertyTag 정상 +
          후보 위치에서 실제 row 디코드 카운트 산정.
    가장 많이 디코드되는 후보 채택.
    반환: (row_map_offset = NumRows 직후 = 첫 RowName 시작, num_rows) | None
    """
    n_buf = len(buf)
    name_count = len(names)
    if scan_from < 0:
        scan_from = 0
    if scan_from >= n_buf - 24:
        return None

    best: tuple[int, int, int] | None = None  # (off, num_rows, score)
    pos = scan_from
    end = n_buf - 24
    while pos < end:
        num_rows, name_idx, name_num = struct.unpack_from("<iii", buf, pos)
        if 1 <= num_rows <= 100000 and 0 <= name_idx < name_count and 0 <= name_num <= 1024:
            row_name = names[name_idx]
            if _row_name_is_valid(row_name):
                try:
                    tag, _, _ = _read_property_tag(buf, pos + 12, names)
                    if tag is not None and tag.get("type") in _KNOWN_PROP_TYPES \
                            and 0 < tag.get("size", 0) <= 1024 * 1024:
                        cap = min(num_rows, 64)
                        score = _try_count_rows(buf, pos + 4, num_rows, names, cap)
                        if score >= max(1, cap // 2):
                            if best is None or score > best[2]:
                                best = (pos + 4, num_rows, score)
                                if score >= cap:
                                    return best[0], best[1]
                except Exception:
                    pass
        pos += 1
    if best is not None:
        return best[0], best[1]
    return None


def _parse_row_map(
    buf: bytes, row_map_off: int, num_rows: int, names: list[str],
) -> tuple[list[str], list[dict]]:
    """RowMap 전체 파싱. row 단위 try/except로 일부 실패는 빈 row로 진행."""
    row_keys: list[str] = []
    rows: list[dict[str, str]] = []
    cur = row_map_off

    for _ in range(num_rows):
        if cur + 8 > len(buf):
            break
        try:
            key, cur = _read_fname(buf, cur, names)
        except Exception:
            break
        row_props: dict[str, str] = {}
        safety = 0
        row_failed = False
        while safety < 1000:
            safety += 1
            try:
                tag, vo, ve = _read_property_tag(buf, cur, names)
            except Exception:
                row_failed = True
                break
            if tag is None:
                cur = vo
                break
            try:
                val = _read_property_value(buf, vo, ve, tag, names)
            except Exception:
                val = ""
            row_props[tag["name"]] = val
            cur = ve
        row_keys.append(key)
        rows.append(row_props)
        if row_failed:
            break

    if not row_keys:
        raise _UAssetParseError("no rows decoded")
    return row_keys, rows


def _build_datatable_matrix(
    row_keys: list[str], rows: list[dict[str, str]],
) -> list[list[str]]:
    """헤더=['RowName'] + 컬럼 합집합(첫 등장 순서 보존)."""
    cols: list[str] = []
    seen: set[str] = set()
    for r in rows:
        for k in r.keys():
            if k not in seen:
                seen.add(k)
                cols.append(k)
    matrix: list[list[str]] = [["RowName"] + cols]
    for key, props in zip(row_keys, rows):
        matrix.append([key] + [props.get(c, "") for c in cols])
    return matrix


def _load_uasset_fallback_matrix(buf: bytes) -> list[list[str]]:
    """본격 파싱 실패 시 [field, value] 헤더 + NameTable 덤프."""
    matrix: list[list[str]] = [["field", "value"]]

    if len(buf) < 8:
        matrix.append(["error", "파일이 너무 짧습니다"])
        return matrix

    (magic,) = struct.unpack_from("<I", buf, 0)
    matrix.append(["magic", f"0x{magic:08X}"])
    if magic != _UASSET_MAGIC:
        matrix.append(["error", "UE 패키지 시그니처가 아닙니다"])
        return matrix

    (legacy,) = struct.unpack_from("<i", buf, 4)
    matrix.append(["legacy_file_version", str(legacy)])

    try:
        names, name_offset, _ = _scan_name_table(buf)
    except _UAssetParseError:
        names, name_offset = [], 0

    matrix.append(["name_count", str(len(names))])
    matrix.append(["name_offset", f"0x{name_offset:08X}" if name_offset else "0"])
    matrix.append(["file_size", str(len(buf))])

    if names:
        matrix.append(["--- name_table ---", ""])
        for i, s in enumerate(names):
            matrix.append([f"name[{i}]", s])
    else:
        matrix.append(["warning", "Name Table 디코드 실패 — 헤더만 표시"])
    return matrix


def load_uasset_as_matrix(path: str) -> list[list[str]]:
    """UE5 DataTable .uasset → RowName + RowStruct 프로퍼티 매트릭스.
    본격 파싱 흐름:
      1) magic 검증
      2) NameTable 디코드 (hash 길이 자동 판별)
      3) Export 영역에서 RowMap 시작점 탐지
      4) RowMap 파싱 (row 단위 안전망)
      5) 컬럼 합집합으로 매트릭스 빌드
    어느 단계든 실패하면 기존 [field/value] 헤더 덤프 폴백 — 앱 크래시 방지.
    """
    with open(path, "rb") as f:
        buf = f.read()

    if len(buf) < 8:
        return _load_uasset_fallback_matrix(buf)
    (magic,) = struct.unpack_from("<I", buf, 0)
    if magic != _UASSET_MAGIC:
        return _load_uasset_fallback_matrix(buf)

    try:
        names, _, name_end = _scan_name_table(buf)
        summary = _parse_package_summary(buf)
        scan_from = summary["total_header_size"] if summary else name_end
        hit = _find_row_map_start(buf, scan_from, names)
        if hit is None:
            raise _UAssetParseError("row map not found")
        row_map_off, num_rows = hit
        row_keys, rows = _parse_row_map(buf, row_map_off, num_rows, names)
        return _build_datatable_matrix(row_keys, rows)
    except Exception:
        return _load_uasset_fallback_matrix(buf)


def load_sheet_with_formulas_any(path: str) -> tuple[list[list], list[list]]:
    """확장자에 따라 (values, formulas) 매트릭스를 반환하는 통합 디스패처.
    - xlsx 계열: 기존 load_sheet_with_formulas() 사용 (수식 별도 추출).
    - json/uasset: 수식 개념이 없으므로 동일 매트릭스를 두 번 반환.
    """
    ext = os.path.splitext(path)[1].lower()
    if ext in _EXCEL_EXTS:
        return load_sheet_with_formulas(path)
    if ext == ".json":
        m = load_json_as_matrix(path)
        return m, [list(row) for row in m]
    if ext == ".uasset":
        m = load_uasset_as_matrix(path)
        return m, [list(row) for row in m]
    # 알려지지 않은 확장자 — 기존 로직(엑셀)로 폴백
    return load_sheet_with_formulas(path)


def _compute_diff_row_order(
    a_data: list[list], b_data: list[list],
) -> tuple[list[list], list[tuple[int | None, int | None]]]:
    """키 없음 — 행 순서 그대로 1:1 매칭."""
    cols = max(
        (max(len(r) for r in a_data) if a_data else 0),
        (max(len(r) for r in b_data) if b_data else 0),
    )
    n = max(len(a_data), len(b_data))
    diff_matrix: list[list] = []
    row_meta: list[tuple] = []
    for i in range(n):
        a_row = a_data[i] if i < len(a_data) else []
        b_row = b_data[i] if i < len(b_data) else []
        row = []
        for c in range(cols):
            av = a_row[c] if c < len(a_row) else ""
            bv = b_row[c] if c < len(b_row) else ""
            row.append((_cell_status(av, bv), av, bv))
        diff_matrix.append(row)
        a_idx = i if i < len(a_data) else None
        b_idx = i if i < len(b_data) else None
        row_meta.append((a_idx, b_idx))
    return diff_matrix, row_meta


def compute_diff(
    a_data: list[list], b_data: list[list], key_col: int = 0
) -> tuple[list[list], list[tuple[int | None, int | None]]]:
    """
    key_col 열 값을 키로 행을 매칭하여 diff를 계산한다.
    key_col == -1 이면 행 순서 기반(ROW order) 비교를 수행한다.
    반환값:
      diff_matrix : list of rows, 각 row = [(status, a_val, b_val), ...]
      row_meta    : [(orig_a_row, orig_b_row), ...]  — None 은 해당 파일에 없는 행
    """
    if not a_data and not b_data:
        return [], []
    if key_col == -1:
        return _compute_diff_row_order(a_data, b_data)

    cols = max(
        (max(len(r) for r in a_data) if a_data else 0),
        (max(len(r) for r in b_data) if b_data else 0),
    )

    def get_key(row):
        return row[key_col] if row and key_col < len(row) else ""

    # 헤더 행(첫 행)은 키 매칭 없이 그대로 1:1
    a_header = a_data[0] if a_data else None
    b_header = b_data[0] if b_data else None

    # 나머지 행을 첫 번째 열 값으로 인덱싱
    a_body = a_data[1:] if len(a_data) > 1 else []
    b_body = b_data[1:] if len(b_data) > 1 else []

    # 키 → (원본 인덱스, 행 데이터) 매핑 (원본 파일 기준 행 번호는 1-based body index + 1)
    a_map: dict[str, tuple[int, list]] = {}
    for i, row in enumerate(a_body):
        key = get_key(row)
        if key == "":
            continue   # 키가 없는 행(빈 행 등)은 키 비교 대상에서 제외
        if key not in a_map:   # 중복 키는 첫 번째 행 사용
            a_map[key] = (i + 1, row)   # +1: 헤더 row가 row 0

    b_map: dict[str, tuple[int, list]] = {}
    for i, row in enumerate(b_body):
        key = get_key(row)
        if key == "":
            continue   # 키가 없는 행(빈 행 등)은 키 비교 대상에서 제외
        if key not in b_map:
            b_map[key] = (i + 1, row)

    # 표시 순서: A의 순서를 기준으로, A에 없고 B에만 있는 키는 뒤에 추가
    all_keys: list[str] = []
    seen: set[str] = set()
    for row in a_body:
        k = get_key(row)
        if k == "" or k in seen:
            continue
        all_keys.append(k)
        seen.add(k)
    for row in b_body:
        k = get_key(row)
        if k == "" or k in seen:
            continue
        all_keys.append(k)
        seen.add(k)

    def make_row(a_row, b_row):
        row = []
        for c in range(cols):
            av = a_row[c] if c < len(a_row) else ""
            bv = b_row[c] if c < len(b_row) else ""
            row.append((_cell_status(av, bv), av, bv))
        return row

    diff_matrix: list[list] = []
    row_meta: list[tuple] = []

    # 헤더 행
    if a_header is not None or b_header is not None:
        ah = a_header or []
        bh = b_header or []
        diff_matrix.append(make_row(ah, bh))
        row_meta.append((0 if a_header is not None else None,
                         0 if b_header is not None else None))

    # 본문 행
    for key in all_keys:
        a_idx, a_row = a_map[key] if key in a_map else (None, [])
        b_idx, b_row = b_map[key] if key in b_map else (None, [])
        diff_matrix.append(make_row(a_row, b_row))
        row_meta.append((a_idx, b_idx))

    return diff_matrix, row_meta


_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_COL_RE = re.compile(r"([A-Z]+)(\d+)")


def _cell_ref(r: int, c: int) -> str:
    return f"{get_column_letter(c + 1)}{r + 1}"


def _cell_status(a_val: str, b_val: str) -> str:
    if a_val == "" and b_val != "":
        return "added"
    if a_val != "" and b_val == "":
        return "modified"
    if a_val != b_val:
        return "modified"
    return "same"


def _promote_empty_cols_to_delete(
    patches: dict[str, str],
    delete_row_nums: set[int],
    path: str | None,
) -> tuple[dict[str, str], set[int], set[str]]:
    """
    1단계: 패치 적용 후 모든 셀이 빈값인 행 → delete_row_nums로 승격 (patches에서 제거)
    2단계: 행 삭제 반영 후 모든 셀이 빈값인 열 → delete_col_letters 반환
    """
    if not path:
        return patches, delete_row_nums, set()

    try:
        with zipfile.ZipFile(path, "r") as zin:
            sheet_name = _find_active_sheet_path(zin)
            xml_data = zin.read(sheet_name)
        tree = etree.fromstring(xml_data)
        ns = _NS
        sheetdata = tree.find(f"{{{ns}}}sheetData")
        file_cells: dict[str, str] = {}
        file_row_refs: dict[int, set[str]] = defaultdict(set)
        if sheetdata is not None:
            for row_el in sheetdata:
                rn = int(row_el.get("r", 0))
                for c_el in row_el:
                    ref = c_el.get("r", "")
                    if not ref:
                        continue
                    v_el  = c_el.find(f"{{{ns}}}v")
                    f_el  = c_el.find(f"{{{ns}}}f")
                    is_el = c_el.find(f"{{{ns}}}is")
                    if f_el is not None and f_el.text:
                        val = "=" + f_el.text
                    elif v_el is not None and v_el.text:
                        val = v_el.text
                    elif is_el is not None:
                        t_el = is_el.find(f"{{{ns}}}t")
                        val  = t_el.text if (t_el is not None and t_el.text) else ""
                    else:
                        val = ""
                    file_cells[ref] = val
                    file_row_refs[rn].add(ref)
    except Exception:
        return patches, delete_row_nums, set()

    merged: dict[str, str] = {**file_cells, **patches}

    new_patches = dict(patches)
    new_deletes = set(delete_row_nums)

    # ── 1단계: 빈 행 → 행 삭제 승격 ──────────────────────────────────────────
    patch_rows: dict[int, set[str]] = defaultdict(set)
    for ref in patches:
        m = _COL_RE.match(ref)
        if m:
            patch_rows[int(m.group(2))].add(ref)

    for row_num, patched_refs in patch_rows.items():
        if row_num in new_deletes:
            continue
        # 이 행의 모든 패치값이 빈값인지 확인
        if any(merged.get(ref, "") != "" for ref in patched_refs):
            continue
        # 파일의 이 행에 패치 외 다른 값이 있는지 확인
        all_refs_in_row = file_row_refs.get(row_num, set())
        non_patched_refs = all_refs_in_row - patched_refs
        if any(merged.get(ref, "") != "" for ref in non_patched_refs):
            continue
        # 행 전체가 빈값 → 행 삭제로 전환
        for ref in patched_refs:
            new_patches.pop(ref, None)
        new_deletes.add(row_num)

    # ── 2단계: 행 삭제 반영 후 빈 열 감지 ────────────────────────────────────
    col_vals: dict[str, list[str]] = defaultdict(list)
    for ref, val in merged.items():
        m = _COL_RE.match(ref)
        if not m:
            continue
        row_num = int(m.group(2))
        if row_num in new_deletes:
            continue
        col_vals[m.group(1)].append(val)

    delete_col_letters: set[str] = set()
    for col_letter, vals in col_vals.items():
        if all(v == "" for v in vals):
            delete_col_letters.add(col_letter)

    return new_patches, new_deletes, delete_col_letters


def _is_file_locked(path: str) -> bool:
    """파일이 다른 프로세스에 의해 열려 있는지 확인한다."""
    try:
        with open(path, "a"):
            return False
    except (IOError, PermissionError):
        return True


def _write_patches_to_file(
    path_base: str,
    patches: dict[str, str],
    insert_rows: list[list[tuple[int, str]]] | None = None,
    delete_row_nums: set[int] | None = None,
    delete_col_letters: set[str] | None = None,
) -> None:
    """
    patches            : {cell_ref: value}          — 기존 셀 덮어쓰기
    insert_rows        : [[(col_idx, value), ...]]   — 파일 끝에 새 행 추가
    delete_row_nums    : {1-based row number}        — 해당 <row> 요소 자체 삭제
    delete_col_letters : {'A', 'B', ...}             — 해당 열의 모든 <c> 삭제
    """
    if _is_file_locked(path_base):
        raise PermissionError(f"파일이 열려 있어 저장할 수 없습니다:\n{path_base}")

    tmp = path_base + ".tmp_merge"
    try:
        with zipfile.ZipFile(path_base, "r") as zin, \
             zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            sheet_name = _find_active_sheet_path(zin)
            for item in zin.infolist():
                # calcChain.xml 항상 제거 — 수식 패치/삽입/삭제 모두 계산 체인을 무효화함
                if item.filename == "xl/calcChain.xml":
                    continue
                data = zin.read(item.filename)
                if item.filename == sheet_name:
                    data = _patch_sheet_xml(
                        data, patches,
                        insert_rows or [], delete_row_nums or set(),
                        delete_col_letters or set(),
                    )
                zout.writestr(item, data)
        os.replace(tmp, path_base)
    except Exception:
        if os.path.exists(tmp):
            os.remove(tmp)
        raise


def _find_active_sheet_path(zin: zipfile.ZipFile) -> str:
    try:
        ns_wb = _NS
        ns_r  = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
        wb_xml  = zin.read("xl/workbook.xml")
        wb_root = etree.fromstring(wb_xml)
        active_tab = 0
        for bv in wb_root.iter(f"{{{ns_wb}}}bookView"):
            active_tab = int(bv.get("activeTab", 0))
            break
        rids = [sh.get(f"{{{ns_r}}}id")
                for sh in wb_root.iter(f"{{{ns_wb}}}sheet")]
        if not rids:
            return "xl/worksheets/sheet1.xml"
        rid = rids[min(active_tab, len(rids) - 1)]
        rels_xml  = zin.read("xl/_rels/workbook.xml.rels")
        rels_root = etree.fromstring(rels_xml)
        for rel in rels_root:
            if rel.get("Id") == rid and rel.get("Type", "").endswith("/worksheet"):
                target = rel.get("Target", "")
                if not target.startswith("xl/"):
                    target = "xl/" + target
                return target
    except Exception:
        pass
    return "xl/worksheets/sheet1.xml"


def _is_numeric(val: str) -> bool:
    try:
        float(val)
        return True
    except (ValueError, TypeError):
        return False


def _set_cell_value(c_el, tag_v, tag_is, tag_t, tag_f, new_val: str):
    for child in list(c_el):
        c_el.remove(child)
    c_el.attrib.pop("t", None)
    if new_val == "":
        return
    if new_val.startswith("="):
        f_el = etree.SubElement(c_el, tag_f)
        f_el.text = new_val[1:]   # '=' 제외한 수식 본문
    elif _is_numeric(new_val):
        v_el = etree.SubElement(c_el, tag_v)
        v_el.text = new_val
    else:
        # t="str": sharedStrings.xml 변경 없이 Excel이 안전하게 수용하는 문자열 타입
        c_el.set("t", "str")
        v_el = etree.SubElement(c_el, tag_v)
        v_el.text = new_val


def _patch_sheet_xml(
    data: bytes,
    patches: dict[str, str],
    insert_rows: list[list[tuple[int, str]]] = [],
    delete_row_nums: set[int] = set(),
    delete_col_letters: set[str] = set(),
) -> bytes:
    """
    patches            : {cell_ref: value}          기존 셀 덮어쓰기
    insert_rows        : [[(col_idx, value), ...]]   파일 끝에 새 행 추가
    delete_row_nums    : {1-based row number}        해당 <row> 요소 자체 삭제
    delete_col_letters : {'A', 'B', ...}             해당 열의 모든 <c> 삭제
    """
    tree = etree.fromstring(data)
    ns    = _NS
    tag_c  = f"{{{ns}}}c"
    tag_f  = f"{{{ns}}}f"
    tag_is = f"{{{ns}}}is"
    tag_t  = f"{{{ns}}}t"
    tag_v  = f"{{{ns}}}v"

    existing: dict[str, etree._Element] = {}
    row_map:  dict[int, etree._Element] = {}

    sheetdata = tree.find(f"{{{ns}}}sheetData")
    if sheetdata is None:
        return data

    for row_el in sheetdata:
        r_idx = int(row_el.get("r", 0))
        row_map[r_idx] = row_el
        for c_el in row_el:
            ref = c_el.get("r", "")
            if ref:
                existing[ref] = c_el

    # 행 삭제 (실제 삭제된 행 번호를 추적)
    actually_deleted: set[int] = set()
    for row_num in delete_row_nums:
        row_el = row_map.get(row_num)
        if row_el is not None:
            sheetdata.remove(row_el)
            actually_deleted.add(row_num)

    # 덮어쓰기 패치
    for ref, new_val in patches.items():
        m = _COL_RE.match(ref)
        if not m:
            continue
        row_num = int(m.group(2))

        if ref in existing:
            c_el = existing[ref]
            if new_val == "":
                # 빈값 패치 시 <c> 요소 자체를 부모 행에서 제거
                parent = c_el.getparent()
                if parent is not None:
                    parent.remove(c_el)
            else:
                _set_cell_value(c_el, tag_v, tag_is, tag_t, tag_f, new_val)
        else:
            if row_num not in row_map:
                row_el = etree.SubElement(sheetdata, f"{{{ns}}}row")
                row_el.set("r", str(row_num))
                row_map[row_num] = row_el
                sheetdata[:] = sorted(sheetdata, key=lambda e: int(e.get("r", 0)))
            else:
                row_el = row_map[row_num]
            if new_val != "":
                c_el = etree.SubElement(row_el, tag_c)
                c_el.set("r", ref)
                _set_cell_value(c_el, tag_v, tag_is, tag_t, tag_f, new_val)
                row_el[:] = sorted(row_el, key=lambda e: column_index_from_string(
                    _COL_RE.match(e.get("r", "A1")).group(1)
                ))

    # 새 행 삽입 (실제 셀이 있는 마지막 행 바로 다음 행 번호부터 — 빈 <row> 요소 무시)
    if insert_rows:
        last_data_row = max(
            (int(row_el.get("r", 0)) for row_el in sheetdata if list(row_el)),
            default=0,
        )
        next_row = last_data_row + 1 if last_data_row > 0 else 1
        for cells in insert_rows:
            row_el = etree.SubElement(sheetdata, f"{{{ns}}}row")
            row_el.set("r", str(next_row))
            for col_idx, val in sorted(cells, key=lambda x: x[0]):
                if val == "":
                    continue
                ref = _cell_ref(next_row - 1, col_idx)
                c_el = etree.SubElement(row_el, tag_c)
                c_el.set("r", ref)
                _set_cell_value(c_el, tag_v, tag_is, tag_t, tag_f, val)
            next_row += 1

    # 빈 열 삭제: 지정된 열 문자에 속하는 <c> 요소를 모든 행에서 제거
    if delete_col_letters:
        for row_el in sheetdata:
            to_remove = [
                c_el for c_el in row_el
                if (m := _COL_RE.match(c_el.get("r", ""))) and m.group(1) in delete_col_letters
            ]
            for c_el in to_remove:
                row_el.remove(c_el)

    # VBA .Delete처럼 삭제된 행 번호를 위로 당겨 빈 행 번호 없애기
    if actually_deleted:
        sorted_deleted = sorted(actually_deleted)
        for row_el in sheetdata:
            rn = int(row_el.get("r", 0))
            offset = sum(1 for d in sorted_deleted if d < rn)
            if offset > 0:
                new_rn = rn - offset
                row_el.set("r", str(new_rn))
                for c_el in row_el:
                    ref = c_el.get("r", "")
                    m = _COL_RE.match(ref)
                    if m:
                        c_el.set("r", f"{m.group(1)}{new_rn}")

    return etree.tostring(tree, xml_declaration=True, encoding="UTF-8", standalone=True)


# ── Worker 스레드 ──────────────────────────────────────────────────────────────

class PreviewWorker(QThread):
    """단일 파일을 로드해 미리보기 데이터를 돌려준다."""
    done = pyqtSignal(str, list, list)   # side('a'|'b'), data(값), formula_data(수식)
    error = pyqtSignal(str)

    def __init__(self, side: str, path: str):
        super().__init__()
        self.side = side
        self.path = path

    def run(self):
        try:
            # 통합 함수 — 같은 파일 1회 파싱으로 값/수식 모두 획득
            # xlsx/json/uasset 모두 디스패처 한 곳에서 처리
            data, formulas = load_sheet_with_formulas_any(self.path)
            self.done.emit(self.side, data, formulas)
        except Exception as e:
            self.error.emit(str(e))


class LoadWorker(QThread):
    done = pyqtSignal(list, list, list, list)   # a_data, b_data, a_formulas, b_formulas
    error = pyqtSignal(str)
    progress = pyqtSignal(str)                  # 단계별 상태바 메시지

    def __init__(self, path_a, path_b):
        super().__init__()
        self.path_a = path_a
        self.path_b = path_b

    def run(self):
        try:
            from concurrent.futures import ThreadPoolExecutor

            def _load(path):
                if not path:
                    return [], []
                return load_sheet_with_formulas_any(path)

            self.progress.emit("A/B 파일 병렬 로딩 중...")
            # A/B 두 파일을 스레드 2개로 동시 파싱 (openpyxl은 IO 바운드라 GIL 영향 적음)
            with ThreadPoolExecutor(max_workers=2) as ex:
                fa = ex.submit(_load, self.path_a)
                fb = ex.submit(_load, self.path_b)
                a, af = fa.result()
                b, bf = fb.result()

            self.progress.emit("로딩 완료 — 비교 계산 중...")
            self.done.emit(a, b, af, bf)
        except Exception as e:
            self.error.emit(str(e))


class StagedMergeWorker(QThread):
    """staged 병합 + 직접 편집값을 한 번에 파일에 기록."""
    done = pyqtSignal(int)
    error = pyqtSignal(str)

    def __init__(self, path_a, path_b, diff_matrix, row_meta, staged: dict, edited: dict,
                 formula_data_a: list = None, formula_data_b: list = None):
        # staged:   {(display_r, c): 'a_to_b'|'b_to_a'}
        # edited:   {'a': {(display_r, c): new_val}, 'b': {(display_r, c): new_val}}
        # row_meta: [(orig_a_row, orig_b_row), ...]
        super().__init__()
        self.path_a = path_a
        self.path_b = path_b
        self.diff_matrix = diff_matrix
        self.row_meta = row_meta
        self.staged = staged
        self.edited = edited
        self.formula_data_a = formula_data_a or []
        self.formula_data_b = formula_data_b or []

    def _formula_val(self, formula_data: list, row_idx: int, col_idx: int, fallback: str) -> str:
        try:
            v = formula_data[row_idx][col_idx]
            return v if v != "" else fallback
        except (IndexError, TypeError):
            return fallback

    def _meta(self, display_r: int) -> tuple:
        if display_r < len(self.row_meta):
            return self.row_meta[display_r]
        return (None, None)

    def run(self):
        try:
            a2b = {k for k, v in self.staged.items() if v == "a_to_b"}
            b2a = {k for k, v in self.staged.items() if v == "b_to_a"}

            # ── B 파일에 쓸 내용 ──────────────────────────────────────────────
            patches_b: dict[str, str] = {}
            insert_rows_b: dict[int, list[tuple[int, str]]] = {}
            delete_rows_b: set[int] = set()   # 1-based 행 번호

            # a2b 병합: display_r 단위로 행 전체가 a2b 대상인지 판별
            a2b_display_rows = {r for (r, c) in a2b}
            for r in a2b_display_rows:
                row = self.diff_matrix[r]
                a_orig, b_orig = self._meta(r)
                # 해당 display 행의 모든 셀이 a2b 대상인지 확인
                row_cells_in_a2b = [(r, c) for c in range(len(row)) if (r, c) in a2b]
                if not row_cells_in_a2b:
                    continue

                if b_orig is not None:
                    # 스테이징된 셀을 patches에 추가
                    # 편집된 값 > 수식 데이터 > diff_matrix a_val 순으로 우선 사용
                    for (_, c) in row_cells_in_a2b:
                        _, a_val, _ = self.diff_matrix[r][c]
                        edited_a = self.edited.get("a", {}).get((r, c))
                        if edited_a is not None:
                            val = edited_a
                        else:
                            val = self._formula_val(self.formula_data_a, a_orig, c, a_val)
                        patches_b[_cell_ref(b_orig, c)] = val
                else:
                    # B에 행 없음(삭제됨 행) → 새 행 삽입
                    for (_, c) in row_cells_in_a2b:
                        _, a_val, _ = self.diff_matrix[r][c]
                        edited_a = self.edited.get("a", {}).get((r, c))
                        if edited_a is not None:
                            val = edited_a
                        else:
                            val = self._formula_val(self.formula_data_a, a_orig, c, a_val)
                        insert_rows_b.setdefault(r, []).append((c, val))

            for (r, c), val in self.edited.get("b", {}).items():
                if (r, c) in a2b:
                    continue
                _, b_orig = self._meta(r)
                if b_orig is not None:
                    patches_b[_cell_ref(b_orig, c)] = val

            # 패치 후 빈 열 감지 및 삭제 준비 (쓸 내용이 있을 때만 파일 접근)
            if (patches_b or insert_rows_b or delete_rows_b) and self.path_b:
                patches_b, delete_rows_b, del_cols_b = _promote_empty_cols_to_delete(
                    patches_b, delete_rows_b, self.path_b
                )
                if (patches_b or insert_rows_b or delete_rows_b or del_cols_b):
                    _write_patches_to_file(
                        self.path_b, patches_b,
                        list(insert_rows_b.values()),
                        delete_rows_b,
                        del_cols_b,
                    )

            # ── A 파일에 쓸 내용 ──────────────────────────────────────────────
            patches_a: dict[str, str] = {}
            insert_rows_a: dict[int, list[tuple[int, str]]] = {}
            delete_rows_a: set[int] = set()

            b2a_display_rows = {r for (r, c) in b2a}
            for r in b2a_display_rows:
                row = self.diff_matrix[r]
                a_orig, b_orig = self._meta(r)
                row_cells_in_b2a = [(r, c) for c in range(len(row)) if (r, c) in b2a]
                if not row_cells_in_b2a:
                    continue

                if a_orig is not None:
                    # 스테이징된 셀을 patches에 추가
                    # 편집된 값 > 수식 데이터 > diff_matrix b_val 순으로 우선 사용
                    for (_, c) in row_cells_in_b2a:
                        _, _, b_val = self.diff_matrix[r][c]
                        edited_b = self.edited.get("b", {}).get((r, c))
                        if edited_b is not None:
                            val = edited_b
                        else:
                            val = self._formula_val(self.formula_data_b, b_orig, c, b_val)
                        patches_a[_cell_ref(a_orig, c)] = val
                else:
                    # A에 행 없음(추가됨 행) → 새 행 삽입
                    for (_, c) in row_cells_in_b2a:
                        _, _, b_val = self.diff_matrix[r][c]
                        edited_b = self.edited.get("b", {}).get((r, c))
                        if edited_b is not None:
                            val = edited_b
                        else:
                            val = self._formula_val(self.formula_data_b, b_orig, c, b_val)
                        insert_rows_a.setdefault(r, []).append((c, val))

            for (r, c), val in self.edited.get("a", {}).items():
                if (r, c) in b2a:
                    continue
                a_orig, _ = self._meta(r)
                if a_orig is not None:
                    patches_a[_cell_ref(a_orig, c)] = val

            if (patches_a or insert_rows_a or delete_rows_a) and self.path_a:
                patches_a, delete_rows_a, del_cols_a = _promote_empty_cols_to_delete(
                    patches_a, delete_rows_a, self.path_a
                )
                if (patches_a or insert_rows_a or delete_rows_a or del_cols_a):
                    _write_patches_to_file(
                        self.path_a, patches_a,
                        list(insert_rows_a.values()),
                        delete_rows_a,
                        del_cols_a,
                    )

            total = (len(patches_b) + len(insert_rows_b) + len(delete_rows_b)
                     + len(patches_a) + len(insert_rows_a) + len(delete_rows_a))
            self.done.emit(total)
        except Exception as e:
            self.error.emit(str(e))


# ── 위젯 ──────────────────────────────────────────────────────────────────────

class MinimapScrollBar(QScrollBar):
    """수직 또는 수평 스크롤바 위에 변경된 셀(행/열)의 위치를 색상 마커로 오버레이.
    paintEvent에서 super 호출 후, orientation에 맞춰 트랙(groove) 영역에
    비율 위치(0.0~1.0)별로 가는 막대를 그린다."""
    _MARKER_COLOR = QColor(255, 140, 0, 220)   # 진한 주황 — 범례의 주황(staged)보다 채도↑

    def __init__(self, orientation, parent=None):
        super().__init__(orientation, parent)
        self._ratios: list = []   # 0.0~1.0 사이 변경 위치 목록

    def set_change_ratios(self, ratios):
        # 변경된 경우에만 repaint (불필요한 페인트 방지)
        if list(ratios) != self._ratios:
            self._ratios = list(ratios)
            self.update()

    def paintEvent(self, e):
        super().paintEvent(e)
        if not self._ratios:
            return
        # QStyle을 통해 정확한 trough(groove) 영역을 얻는다
        opt = QStyleOptionSlider()
        self.initStyleOption(opt)
        groove = self.style().subControlRect(
            QStyle.CC_ScrollBar, opt, QStyle.SC_ScrollBarGroove, self)
        if groove.width() <= 0 or groove.height() <= 0:
            return
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing, False)
        painter.setPen(Qt.NoPen)
        painter.setBrush(self._MARKER_COLOR)
        if self.orientation() == Qt.Vertical:
            track_top = groove.top()
            track_h = groove.height()
            x = groove.left() + 2
            w = max(1, groove.width() - 4)
            denom = max(0, track_h - 2)
            for ratio in self._ratios:
                y = track_top + int(ratio * denom)
                painter.fillRect(x, y, w, 2, self._MARKER_COLOR)
        else:
            track_left = groove.left()
            track_w = groove.width()
            y = groove.top() + 2
            h = max(1, groove.height() - 4)
            denom = max(0, track_w - 2)
            for ratio in self._ratios:
                x = track_left + int(ratio * denom)
                painter.fillRect(x, y, 2, h, self._MARKER_COLOR)
        painter.end()


class ExcelTableWidget(QTableWidget):
    stage_requested   = pyqtSignal(str)   # direction: 'a_to_b' | 'b_to_a'
    unstage_requested = pyqtSignal()
    key_col_changed   = pyqtSignal(int)   # 키 열 변경 요청
    columns_exclude_set = pyqtSignal(list, bool)   # (cols, exclude) — True: 제외 추가, False: 제외 해제
    column_resized    = pyqtSignal(int, int)   # (col, new_width) — 사용자 조작에 의한 변경만
    row_resized       = pyqtSignal(int, int)   # (row, new_height) — 사용자 조작에 의한 변경만
    edit_focus_requested = pyqtSignal()   # F2 — 패널 cell_edit 으로 포커스 이동 요청
    delete_cell_requested = pyqtSignal(int, int)   # (row, col) — Delete 키로 셀 값 비우기 요청

    # RGB 튜플로 비교 — QColor 객체 hash 충돌 크래시 방지
    _STAGED_RGB  = (255, 185,  80)
    _MERGED_RGB  = (173, 216, 230)
    _CHANGED_RGBS = {
        (198, 239, 206),   # added
        (255, 235, 156),   # modified
    }

    def __init__(self, side: str, parent=None):
        super().__init__(parent)
        self.side = side
        self._populating = False
        self._key_col: int = 0
        self._excluded_cols: set[int] = set()   # 변경 검사 제외 열 (display 인덱스)
        # 사용자가 직접 조정한 열/행 크기 — 세션 동안만 유지 (재로드/저장/새로고침 후 복원)
        self._user_col_widths: dict[int, int] = {}
        self._user_row_heights: dict[int, int] = {}
        # 외부(다른 패널)에서 크기를 강제 적용 중일 때 sectionResized 재방출 방지
        self._applying_sizes: bool = False
        # 헤더 다중 선택의 anchor (Shift+방향 확장의 고정점)
        self._header_anchor_col: int | None = None
        self._header_anchor_row: int | None = None
        self.setFont(QFont("맑은 고딕", 9))
        self.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.verticalHeader().setDefaultSectionSize(22)
        self.setAlternatingRowColors(False)
        self.setEditTriggers(QTableWidget.NoEditTriggers)
        self.setSelectionBehavior(QTableWidget.SelectItems)
        self.setSelectionMode(QTableWidget.ExtendedSelection)
        self.setContextMenuPolicy(Qt.CustomContextMenu)
        self.customContextMenuRequested.connect(self._show_context_menu)
        self.horizontalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        self.horizontalHeader().customContextMenuRequested.connect(
            self._show_header_context_menu)
        self.verticalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        self.verticalHeader().customContextMenuRequested.connect(
            self._show_row_header_context_menu)
        # 헤더 크기 변경 추적 — 사용자 조작 시에만 저장/시그널 발행
        self.horizontalHeader().sectionResized.connect(self._on_section_h_resized)
        self.verticalHeader().sectionResized.connect(self._on_section_v_resized)
        # 헤더 클릭 시 anchor 갱신 (Shift 없는 클릭 → 새 anchor / Shift 클릭 → 기존 유지)
        self.horizontalHeader().sectionPressed.connect(self._on_h_section_pressed)
        self.verticalHeader().sectionPressed.connect(self._on_v_section_pressed)

    # ── 사용자 헤더 크기 추적 ────────────────────────────────────────────────
    def _on_section_h_resized(self, logical_index: int, _old: int, new_size: int):
        # populate 중이거나 다른 패널에서 강제 적용 중인 변경은 무시
        if self._populating or self._applying_sizes:
            return
        self._user_col_widths[logical_index] = new_size
        self.column_resized.emit(logical_index, new_size)

    def _on_section_v_resized(self, logical_index: int, _old: int, new_size: int):
        if self._populating or self._applying_sizes:
            return
        self._user_row_heights[logical_index] = new_size
        self.row_resized.emit(logical_index, new_size)

    def _on_h_section_pressed(self, logical_index: int):
        """열 헤더 클릭: Shift/Ctrl 없으면 anchor 갱신, 동반이면 유지."""
        mods = QApplication.keyboardModifiers()
        if not (mods & (Qt.ShiftModifier | Qt.ControlModifier)):
            self._header_anchor_col = logical_index
        elif self._header_anchor_col is None:
            # Shift+클릭인데 anchor가 없으면 현재 클릭 지점을 anchor로
            self._header_anchor_col = logical_index
        # 행 anchor는 무관 — 열 헤더 클릭은 행 헤더 모드를 종료시킴
        self._header_anchor_row = None

    def _on_v_section_pressed(self, logical_index: int):
        mods = QApplication.keyboardModifiers()
        if not (mods & (Qt.ShiftModifier | Qt.ControlModifier)):
            self._header_anchor_row = logical_index
        elif self._header_anchor_row is None:
            self._header_anchor_row = logical_index
        self._header_anchor_col = None

    def apply_column_width(self, col: int, width: int):
        """반대 패널에서의 열 너비 변경을 동기 적용 (시그널 재방출 안 함)."""
        self._user_col_widths[col] = width
        if 0 <= col < self.columnCount() and self.columnWidth(col) != width:
            self._applying_sizes = True
            try:
                self.setColumnWidth(col, width)
            finally:
                self._applying_sizes = False

    def apply_row_height(self, row: int, height: int):
        """반대 패널에서의 행 높이 변경을 동기 적용."""
        self._user_row_heights[row] = height
        if 0 <= row < self.rowCount() and self.rowHeight(row) != height:
            self._applying_sizes = True
            try:
                self.setRowHeight(row, height)
            finally:
                self._applying_sizes = False

    def _apply_user_sizes(self):
        """저장된 사용자 크기를 현재 테이블에 다시 적용 (populate 후 호출)."""
        # 0-크기 값은 hidden 행/열에 대한 sectionResized 시그널이 남긴 오염값일 수
        # 있으므로 무시한다. UI상 0으로 만드는 사용자 조작은 없다.
        self._applying_sizes = True
        try:
            for col, w in self._user_col_widths.items():
                if 0 <= col < self.columnCount() and w > 0:
                    self.setColumnWidth(col, w)
            for row, h in self._user_row_heights.items():
                if 0 <= row < self.rowCount() and h > 0:
                    self.setRowHeight(row, h)
        finally:
            self._applying_sizes = False

    def _clip_auto_column_widths(self):
        """자동 너비 계산 결과를 MAX_AUTO_COL_WIDTH_PX 로 상한 클립.
        sectionResized 시그널이 사용자 변경으로 오해해 _user_col_widths 에
        저장하지 않도록 _applying_sizes 플래그로 차단한다."""
        self._applying_sizes = True
        try:
            for c in range(self.columnCount()):
                if self.columnWidth(c) > MAX_AUTO_COL_WIDTH_PX:
                    self.setColumnWidth(c, MAX_AUTO_COL_WIDTH_PX)
        finally:
            self._applying_sizes = False

    @staticmethod
    def _rgb(item) -> tuple:
        c = item.background().color()
        return (c.red(), c.green(), c.blue())

    def set_key_col(self, col: int):
        self._key_col = col
        self._refresh_key_col_header()

    def set_excluded_cols(self, cols: set):
        """외부(MainWindow)에서 제외 열 집합을 갱신하고 헤더를 다시 칠한다."""
        self._excluded_cols = set(cols)
        self._refresh_key_col_header()

    def _refresh_key_col_header(self):
        for c in range(self.columnCount()):
            item = self.horizontalHeaderItem(c)
            if item is None:
                item = QTableWidgetItem()
                self.setHorizontalHeaderItem(c, item)
            col_letter = get_column_letter(c + 1)
            if c == self._key_col:
                item.setText(f"🔑 {col_letter}")
                item.setBackground(QColor(255, 213, 0))
                item.setForeground(QColor(0, 0, 0))
                item.setFont(QFont("맑은 고딕", 9, QFont.Bold))
            elif c in self._excluded_cols:
                item.setText(f"⊘ {col_letter}")
                item.setBackground(QColor(220, 220, 220))
                item.setForeground(QColor(140, 140, 140))
                item.setFont(QFont("맑은 고딕", 9))
            else:
                item.setText(col_letter)
                item.setBackground(QColor(232, 234, 240))
                item.setForeground(QColor(0, 0, 0))
                item.setFont(QFont("맑은 고딕", 9))

    def _col_stage_items(self, col: int):
        """지정 열의 변경·스테이징된 아이템 목록 반환."""
        changed, staged = [], []
        for r in range(self.rowCount()):
            item = self.item(r, col)
            if item is None:
                continue
            rgb = self._rgb(item)
            if rgb in self._CHANGED_RGBS:
                changed.append(item)
            elif rgb == self._STAGED_RGB:
                staged.append(item)
        return changed, staged

    def _row_stage_items(self, row: int):
        """지정 행의 변경·스테이징된 아이템 목록 반환."""
        changed, staged = [], []
        for c in range(self.columnCount()):
            item = self.item(row, c)
            if item is None:
                continue
            rgb = self._rgb(item)
            if rgb in self._CHANGED_RGBS:
                changed.append(item)
            elif rgb == self._STAGED_RGB:
                staged.append(item)
        return changed, staged

    def _select_col(self, col: int):
        """해당 열 전체 셀을 선택 상태로 설정."""
        sm = self.selectionModel()
        rows = self.rowCount()
        cols = self.columnCount()
        if sm is None or rows == 0 or cols == 0 or not (0 <= col < cols):
            return
        model = self.model()
        sel = QItemSelection(model.index(0, col), model.index(rows - 1, col))
        sm.select(sel, QItemSelectionModel.ClearAndSelect)

    def _select_row(self, row: int):
        """해당 행 전체 셀을 선택 상태로 설정."""
        sm = self.selectionModel()
        rows = self.rowCount()
        cols = self.columnCount()
        if sm is None or rows == 0 or cols == 0 or not (0 <= row < rows):
            return
        model = self.model()
        sel = QItemSelection(model.index(row, 0), model.index(row, cols - 1))
        sm.select(sel, QItemSelectionModel.ClearAndSelect)

    def _selected_header_cols(self, anchor_col: int) -> list[int]:
        """우클릭 시 대상 열 집합 결정.
        - 우클릭한 열이 현재 헤더 다중 선택에 포함되어 있으면 그 선택 전체.
        - 아니면 우클릭한 단일 열만.
        """
        sel_model = self.selectionModel()
        cols: set[int] = set()
        if sel_model is not None:
            # selectedColumns()는 한 칼럼당 한 인덱스만 반환 — 헤더 클릭으로 전체열 선택 시 채워짐.
            for idx in sel_model.selectedColumns():
                cols.add(idx.column())
            # 셀 선택 모드에서 헤더를 Shift/Ctrl-클릭한 경우 selectedIndexes()도 보충.
            if not cols:
                for idx in sel_model.selectedIndexes():
                    cols.add(idx.column())
        if anchor_col in cols and len(cols) > 1:
            return sorted(cols)
        return [anchor_col]

    def _show_header_context_menu(self, pos):
        col = self.horizontalHeader().logicalIndexAt(pos)
        if col < 0:
            return

        target_cols = self._selected_header_cols(col)
        multi = len(target_cols) > 1
        col_letter = get_column_letter(col + 1)
        if multi:
            cols_label = ", ".join(get_column_letter(c + 1) for c in target_cols)
        else:
            cols_label = col_letter

        is_excluded = col in self._excluded_cols
        # 다중 선택 시 제외 토글은 "혼합 상태"를 다룬다: 하나라도 비제외면 일괄 제외, 전부 제외면 일괄 해제.
        if multi:
            any_not_excluded = any(c not in self._excluded_cols for c in target_cols)
            multi_action_exclude = any_not_excluded   # True → 제외 추가, False → 제외 해제
        else:
            multi_action_exclude = not is_excluded

        changed, staged = self._col_stage_items(col)
        # 제외된 열은 stage/unstage 액션을 표시하지 않는다 — 변경이 'same'으로 노출되므로 의미 없음.
        # 다중 선택일 땐 stage/key 액션은 단순화를 위해 노출하지 않는다 (제외 토글만 일괄 처리).
        has_changed = bool(changed) and not is_excluded and not multi
        has_staged  = bool(staged) and not is_excluded and not multi

        menu = QMenu(self)
        menu.setStyleSheet("QMenu { font-family: '맑은 고딕'; font-size: 9pt; }")

        # ── 병합 준비 항목 ──
        act_a2b = act_b2a = act_unstage = None
        if has_changed:
            act_a2b = menu.addAction(f"선택 열: A → B  병합 준비  [{col_letter}열]")
            act_b2a = menu.addAction(f"선택 열: B → A  병합 준비  [{col_letter}열]")
        if has_staged:
            if has_changed:
                menu.addSeparator()
            act_unstage = menu.addAction(f"선택 열: 병합 준비 취소  [{col_letter}열]")

        if has_changed or has_staged:
            menu.addSeparator()

        # ── 키 열 항목 (단일 선택일 때만) ──
        act_key_clear = act_reset = act_key_set = None
        if not multi:
            if col == self._key_col:
                act_key_cur = menu.addAction(f"[키 열]  {col_letter}열 — 현재 키 열")
                act_key_cur.setEnabled(False)
                act_key_clear = menu.addAction("🔓  키 열 해제 (ROW 순서 기반 비교)")
                act_reset = menu.addAction("↩  A열(기본값)으로 초기화") if col != 0 else None
            else:
                act_key_set = menu.addAction(f"키 열로 설정  [{col_letter}열]")

        # ── 변경 검사 제외 토글 ──
        menu.addSeparator()
        if multi_action_exclude:
            label = f"⊘  변경 검사에서 제외  [{cols_label}열]"
        else:
            label = f"↺  검사 제외 해제  [{cols_label}열]"
        act_excl = menu.addAction(label)

        act = menu.exec_(self.horizontalHeader().mapToGlobal(pos))
        if act is None:
            return

        if act_a2b is not None and act == act_a2b:
            self._select_col(col)
            self.stage_requested.emit("a_to_b")
        elif act_b2a is not None and act == act_b2a:
            self._select_col(col)
            self.stage_requested.emit("b_to_a")
        elif act_unstage is not None and act == act_unstage:
            self._select_col(col)
            self.unstage_requested.emit()
        elif act_key_clear is not None and act == act_key_clear:
            self.key_col_changed.emit(-1)
        elif act_reset is not None and act == act_reset:
            self.key_col_changed.emit(0)
        elif act_key_set is not None and act == act_key_set:
            self.key_col_changed.emit(col)
        elif act == act_excl:
            self.columns_exclude_set.emit(target_cols, multi_action_exclude)

    def _show_row_header_context_menu(self, pos):
        row = self.verticalHeader().logicalIndexAt(pos)
        if row < 0:
            return

        changed, staged = self._row_stage_items(row)
        has_changed = bool(changed)
        has_staged  = bool(staged)

        if not has_changed and not has_staged:
            return

        menu = QMenu(self)
        menu.setStyleSheet("QMenu { font-family: '맑은 고딕'; font-size: 9pt; }")

        act_a2b = act_b2a = act_unstage = None
        if has_changed:
            act_a2b = menu.addAction("선택 행: A → B  병합 준비")
            act_b2a = menu.addAction("선택 행: B → A  병합 준비")
        if has_staged:
            if has_changed:
                menu.addSeparator()
            act_unstage = menu.addAction("선택 행: 병합 준비 취소")

        act = menu.exec_(self.verticalHeader().mapToGlobal(pos))
        if act is None:
            return

        if act_a2b is not None and act == act_a2b:
            self._select_row(row)
            self.stage_requested.emit("a_to_b")
        elif act_b2a is not None and act == act_b2a:
            self._select_row(row)
            self.stage_requested.emit("b_to_a")
        elif act_unstage is not None and act == act_unstage:
            self._select_row(row)
            self.unstage_requested.emit()

    def _show_context_menu(self, pos):
        selected = self.selectedItems()
        if not selected:
            return

        has_changed = any(self._rgb(item) in self._CHANGED_RGBS for item in selected)
        has_staged  = any(self._rgb(item) == self._STAGED_RGB   for item in selected)

        if not has_changed and not has_staged:
            return

        menu = QMenu(self)
        menu.setStyleSheet("QMenu { font-family: '맑은 고딕'; font-size: 9pt; }")

        act_a2b     = menu.addAction("선택 셀: A -> B  병합 준비") if has_changed else None
        act_b2a     = menu.addAction("선택 셀: B -> A  병합 준비") if has_changed else None
        act_unstage = None
        if has_staged:
            if has_changed:
                menu.addSeparator()
            act_unstage = menu.addAction("선택 셀: 병합 준비 취소")

        act = menu.exec_(self.viewport().mapToGlobal(pos))
        if act is None:
            return
        if act_a2b is not None and act == act_a2b:
            self.stage_requested.emit("a_to_b")
        elif act_b2a is not None and act == act_b2a:
            self.stage_requested.emit("b_to_a")
        elif act_unstage is not None and act == act_unstage:
            self.unstage_requested.emit()

    # ── 엑셀식 키보드 네비/선택/병합 단축키 ──────────────────────────────────
    def _is_empty_cell(self, r: int, c: int) -> bool:
        if r < 0 or r >= self.rowCount() or c < 0 or c >= self.columnCount():
            return True
        item = self.item(r, c)
        return item is None or item.text() == ""

    def _jump_target(self, r: int, c: int, dr: int, dc: int) -> tuple:
        """엑셀의 Ctrl+방향키 시맨틱으로 점프 대상 (row, col) 반환."""
        max_r = self.rowCount() - 1
        max_c = self.columnCount() - 1
        if max_r < 0 or max_c < 0:
            return (max(0, r), max(0, c))
        nr, nc = r + dr, c + dc
        # 범위 밖이면 그대로
        if nr < 0 or nr > max_r or nc < 0 or nc > max_c:
            return (max(0, min(r, max_r)), max(0, min(c, max_c)))
        cur_empty = self._is_empty_cell(r, c)
        next_empty = self._is_empty_cell(nr, nc)
        if cur_empty:
            # 다음 비어있지 않은 셀까지
            while 0 <= nr <= max_r and 0 <= nc <= max_c and self._is_empty_cell(nr, nc):
                nr += dr; nc += dc
            if nr < 0 or nr > max_r or nc < 0 or nc > max_c:
                # 못 찾으면 끝까지
                return (max(0, min(nr - dr, max_r)), max(0, min(nc - dc, max_c)))
            return (nr, nc)
        if next_empty:
            # 빈 구간 건너 다음 비어있지 않은 셀까지
            while 0 <= nr <= max_r and 0 <= nc <= max_c and self._is_empty_cell(nr, nc):
                nr += dr; nc += dc
            if nr < 0 or nr > max_r or nc < 0 or nc > max_c:
                return (max(0, min(nr - dr, max_r)), max(0, min(nc - dc, max_c)))
            return (nr, nc)
        # 연속 데이터의 마지막 비어있지 않은 셀까지
        while 0 <= nr + dr <= max_r and 0 <= nc + dc <= max_c \
                and not self._is_empty_cell(nr + dr, nc + dc):
            nr += dr; nc += dc
        return (nr, nc)

    def _select_range(self, r1: int, c1: int, r2: int, c2: int):
        """(r1,c1)~(r2,c2) 직사각형의 셀들을 모두 선택 상태로 설정 (기존 선택은 클리어)."""
        rs, re_ = sorted((r1, r2))
        cs, ce_ = sorted((c1, c2))
        self.clearSelection()
        for rr in range(rs, re_ + 1):
            for cc in range(cs, ce_ + 1):
                item = self.item(rr, cc)
                if item is not None:
                    item.setSelected(True)

    def _has_changed_selection(self) -> bool:
        return any(self._rgb(it) in self._CHANGED_RGBS for it in self.selectedItems())

    def _has_staged_selection(self) -> bool:
        return any(self._rgb(it) == self._STAGED_RGB for it in self.selectedItems())

    # ── 헤더 다중 선택 지원 ──────────────────────────────────────────────────
    def _full_columns_selected(self) -> list[int]:
        """selectionModel().selectedColumns()는 한 칼럼이 모든 행에 걸쳐 선택된
        경우만 반환 → 비어있지 않으면 '열 헤더 선택' 상태."""
        sm = self.selectionModel()
        if sm is None:
            return []
        return sorted({idx.column() for idx in sm.selectedColumns()})

    def _full_rows_selected(self) -> list[int]:
        sm = self.selectionModel()
        if sm is None:
            return []
        return sorted({idx.row() for idx in sm.selectedRows()})

    def _select_column_range(self, c1: int, c2: int):
        """[c1..c2] 모든 열의 모든 행 셀을 선택."""
        sm = self.selectionModel()
        rows = self.rowCount()
        cols = self.columnCount()
        if sm is None or rows == 0 or cols == 0:
            return
        cs, ce_ = sorted((c1, c2))
        cs = max(0, cs); ce_ = min(cols - 1, ce_)
        if cs > ce_:
            return
        model = self.model()
        sel = QItemSelection(model.index(0, cs), model.index(rows - 1, ce_))
        sm.select(sel, QItemSelectionModel.ClearAndSelect)

    def _select_row_range(self, r1: int, r2: int):
        sm = self.selectionModel()
        rows = self.rowCount()
        cols = self.columnCount()
        if sm is None or rows == 0 or cols == 0:
            return
        rs, re_ = sorted((r1, r2))
        rs = max(0, rs); re_ = min(rows - 1, re_)
        if rs > re_:
            return
        model = self.model()
        sel = QItemSelection(model.index(rs, 0), model.index(re_, cols - 1))
        sm.select(sel, QItemSelectionModel.ClearAndSelect)

    def keyPressEvent(self, event):
        if self._populating:
            return super().keyPressEvent(event)
        key = event.key()
        mods = event.modifiers()
        ctrl = bool(mods & Qt.ControlModifier)
        shift = bool(mods & Qt.ShiftModifier)
        alt = bool(mods & Qt.AltModifier)

        cur_r = self.currentRow()
        cur_c = self.currentColumn()

        # ── 헤더 다중 선택 확장 (Shift / Ctrl+Shift + 방향키) ──
        # 열 전체가 선택된 상태에서 Shift+←/→ 는 열 단위 확장,
        # 행 전체가 선택된 상태에서 Shift+↑/↓ 는 행 단위 확장.
        if shift and not alt and key in (Qt.Key_Left, Qt.Key_Right, Qt.Key_Up, Qt.Key_Down):
            full_cols = self._full_columns_selected()
            full_rows = self._full_rows_selected()
            is_col_mode = bool(full_cols) and key in (Qt.Key_Left, Qt.Key_Right)
            is_row_mode = bool(full_rows) and key in (Qt.Key_Up, Qt.Key_Down)

            if is_col_mode and self.columnCount() > 0 and self.rowCount() > 0:
                # anchor 초기화: 단일 열만 선택돼있고 anchor 없음 → 그 열을 anchor로
                if self._header_anchor_col is None:
                    if len(full_cols) == 1:
                        self._header_anchor_col = full_cols[0]
                    else:
                        # 다중 열 이미 선택 — currentIndex와 가장 먼 끝을 anchor로
                        cur_col_idx = cur_c if cur_c >= 0 else full_cols[-1]
                        self._header_anchor_col = (
                            full_cols[0] if cur_col_idx == full_cols[-1] else full_cols[-1]
                        )
                # 현재 확장 끝점 = currentIndex 또는 anchor 반대편 끝
                if cur_c >= 0 and cur_c in full_cols:
                    cur_end = cur_c
                else:
                    cur_end = full_cols[-1] if self._header_anchor_col == full_cols[0] else full_cols[0]
                if ctrl:
                    target = 0 if key == Qt.Key_Left else self.columnCount() - 1
                else:
                    delta = -1 if key == Qt.Key_Left else 1
                    target = max(0, min(self.columnCount() - 1, cur_end + delta))
                self._select_column_range(self._header_anchor_col, target)
                self.setCurrentCell(max(0, cur_r if cur_r >= 0 else 0), target)
                event.accept(); return

            if is_row_mode and self.columnCount() > 0 and self.rowCount() > 0:
                if self._header_anchor_row is None:
                    if len(full_rows) == 1:
                        self._header_anchor_row = full_rows[0]
                    else:
                        cur_row_idx = cur_r if cur_r >= 0 else full_rows[-1]
                        self._header_anchor_row = (
                            full_rows[0] if cur_row_idx == full_rows[-1] else full_rows[-1]
                        )
                if cur_r >= 0 and cur_r in full_rows:
                    cur_end = cur_r
                else:
                    cur_end = full_rows[-1] if self._header_anchor_row == full_rows[0] else full_rows[0]
                if ctrl:
                    target = 0 if key == Qt.Key_Up else self.rowCount() - 1
                else:
                    delta = -1 if key == Qt.Key_Up else 1
                    target = max(0, min(self.rowCount() - 1, cur_end + delta))
                self._select_row_range(self._header_anchor_row, target)
                self.setCurrentCell(target, max(0, cur_c if cur_c >= 0 else 0))
                event.accept(); return

        # 헤더 anchor 라이프사이클: 헤더 모드 분기에 들어가지 않은 일반 키는 anchor 무효화
        # (단순 Shift 아닌 키, 혹은 헤더가 아닌 일반 셀 선택 상태일 때)
        if key in (Qt.Key_Left, Qt.Key_Right, Qt.Key_Up, Qt.Key_Down):
            if not shift:
                self._header_anchor_col = None
                self._header_anchor_row = None

        # ── 병합 단축키 (Alt+Left / Alt+Right / Alt+Backspace) ──
        if alt and not ctrl and not shift:
            if key == Qt.Key_Right:
                if self._has_changed_selection():
                    self.stage_requested.emit("a_to_b")
                event.accept(); return
            if key == Qt.Key_Left:
                if self._has_changed_selection():
                    self.stage_requested.emit("b_to_a")
                event.accept(); return
            if key in (Qt.Key_Backspace, Qt.Key_Delete):
                if self._has_staged_selection():
                    self.unstage_requested.emit()
                event.accept(); return

        # ── F2: 셀 편집란 포커스 요청 ──
        if key == Qt.Key_F2 and not ctrl and not shift and not alt:
            self.edit_focus_requested.emit()
            event.accept(); return

        # ── Delete: 단일 셀 값 비우기 ──
        if key == Qt.Key_Delete and not ctrl and not shift and not alt:
            sel = self.selectedItems()
            if len(sel) == 1:
                it = sel[0]
                self.delete_cell_requested.emit(it.row(), it.column())
                event.accept(); return
            # 다중 선택 일괄 삭제는 사고 위험 — 무시 (기본 동작도 막음)
            event.accept(); return

        # ── Enter/Return: 엑셀처럼 아래 칸으로 이동 ──
        if key in (Qt.Key_Return, Qt.Key_Enter) and not ctrl and not alt:
            if cur_r >= 0 and cur_c >= 0 and cur_r + 1 < self.rowCount():
                self.setCurrentCell(cur_r + 1, cur_c)
            event.accept(); return

        # ── Shift+Space: 행 전체, Ctrl+Space: 열 전체 ──
        if key == Qt.Key_Space and shift and not ctrl and not alt and cur_r >= 0:
            self._select_row(cur_r)
            event.accept(); return
        if key == Qt.Key_Space and ctrl and not shift and not alt and cur_c >= 0:
            self._select_col(cur_c)
            event.accept(); return

        # ── Ctrl(+Shift)+방향키: 데이터 경계 점프 (Excel 시맨틱) ──
        if ctrl and not alt and key in (Qt.Key_Left, Qt.Key_Right, Qt.Key_Up, Qt.Key_Down):
            if cur_r < 0 or cur_c < 0:
                return super().keyPressEvent(event)
            dr = -1 if key == Qt.Key_Up else (1 if key == Qt.Key_Down else 0)
            dc = -1 if key == Qt.Key_Left else (1 if key == Qt.Key_Right else 0)
            tr, tc = self._jump_target(cur_r, cur_c, dr, dc)
            if shift:
                # anchor = 현재 selection의 처음 시작점 추정 (currentIndex 기준)
                anchor = self.currentIndex()
                ar = anchor.row() if anchor.isValid() else cur_r
                ac = anchor.column() if anchor.isValid() else cur_c
                self._select_range(ar, ac, tr, tc)
                self.setCurrentCell(tr, tc)
            else:
                self.setCurrentCell(tr, tc)
            event.accept(); return

        # ── Ctrl+Home / Ctrl+End ──
        if ctrl and not alt and key == Qt.Key_Home:
            tr, tc = 0, 0
            if shift and cur_r >= 0 and cur_c >= 0:
                self._select_range(cur_r, cur_c, tr, tc)
            self.setCurrentCell(tr, tc)
            event.accept(); return
        if ctrl and not alt and key == Qt.Key_End:
            tr, tc = max(0, self.rowCount() - 1), max(0, self.columnCount() - 1)
            if shift and cur_r >= 0 and cur_c >= 0:
                self._select_range(cur_r, cur_c, tr, tc)
            self.setCurrentCell(tr, tc)
            event.accept(); return

        super().keyPressEvent(event)

    def populate(self, diff_matrix: list[list], which: str,
                 merged_set: set = None, staged: dict = None,
                 row_meta: list = None, excluded_cols: set = None):
        if not diff_matrix:
            self._safe_clear()
            return

        merged_set = merged_set or set()
        staged     = staged or {}
        excluded_cols = set(excluded_cols) if excluded_cols else set()
        self._excluded_cols = set(excluded_cols)
        rows = len(diff_matrix)
        cols = len(diff_matrix[0])
        side_idx = 0 if which == "a" else 1

        total_cols = cols + EXTRA_COLS
        total_rows = rows + EXTRA_ROWS
        self._populating = True
        self._header_anchor_col = None
        self._header_anchor_row = None
        # 렌더 최적화: 매 setItem 마다 발생하는 리페인트/시그널/정렬 갱신을 차단
        prev_updates = self.updatesEnabled()
        prev_sorting = self.isSortingEnabled()
        self.setUpdatesEnabled(False)
        self.setSortingEnabled(False)
        self.blockSignals(True)
        try:
            self.setRowCount(0)           # 기존 셀·selectionModel 완전 초기화
            self.setColumnCount(total_cols)
            self.setRowCount(total_rows)
            self.setHorizontalHeaderLabels([get_column_letter(c + 1) for c in range(total_cols)])
            if row_meta:
                labels = []
                for r in range(rows):
                    orig = row_meta[r][side_idx] if r < len(row_meta) else None
                    labels.append(str(orig + 1) if orig is not None else "-")
                for r in range(EXTRA_ROWS):
                    labels.append(str(rows + r + 1))
                self.setVerticalHeaderLabels(labels)
            else:
                self.setVerticalHeaderLabels([str(r + 1) for r in range(total_rows)])

            # 핫루프 — 지역변수 바인딩으로 속성 조회 비용 절감
            _set_item = self.setItem
            _is_a = (which == "a")
            _align = Qt.AlignVCenter | Qt.AlignLeft
            _color_merged = DIFF_COLORS["merged"]
            _color_staged = DIFF_COLORS["staged"]
            _diff_colors = DIFF_COLORS
            for r in range(rows):
                row_data = diff_matrix[r]
                for c in range(cols):
                    status, a_val, b_val = row_data[c]
                    direction = staged.get((r, c))
                    if direction == "a_to_b":
                        text = a_val
                    elif direction == "b_to_a":
                        text = b_val
                    else:
                        text = a_val if _is_a else b_val
                    item = QTableWidgetItem(text)
                    if c in excluded_cols:
                        # 제외 열은 status 무관하게 흰색(same) 처리
                        color = _diff_colors["same"]
                    elif (r, c) in merged_set:
                        color = _color_merged
                    elif direction is not None:
                        color = _color_staged
                    else:
                        color = _diff_colors[status]
                    item.setBackground(color)
                    item.setTextAlignment(_align)
                    _set_item(r, c, item)
            # 1) 모든 열에 자동 너비 계산 → 2) MAX_AUTO_COL_WIDTH_PX 상한 클립
            # → 3) 사용자가 직접 조정한 열/행만 그 위에 덮어쓰기 (상한 무시).
            # 이렇게 해야 "사용자가 만진 적 없는 열"은 재비교 후에도 상한 유지된다.
            # 새로고침(_run_refresh)은 _user_col_widths/_user_row_heights를 미리 비우므로
            # 그 경로에서는 3)이 건너뛰어져 모든 열이 디폴트(자동+상한)로 복귀한다.
            # ※ resizeColumnsToContents()가 sectionResized를 발화시키므로 _populating=True 유지 필수.
            self.resizeColumnsToContents()
            self._clip_auto_column_widths()
            if self._user_col_widths or self._user_row_heights:
                self._apply_user_sizes()
            self._refresh_key_col_header()
        finally:
            self.blockSignals(False)
            self.setSortingEnabled(prev_sorting)
            self.setUpdatesEnabled(prev_updates)
            self._populating = False

    def populate_preview(self, data: list[list]):
        if not data:
            self._safe_clear()
            return
        rows = len(data)
        cols = max((len(r) for r in data), default=0)
        total_cols = cols + EXTRA_COLS
        total_rows = rows + EXTRA_ROWS

        self._populating = True
        self._header_anchor_col = None
        self._header_anchor_row = None
        prev_updates = self.updatesEnabled()
        prev_sorting = self.isSortingEnabled()
        self.setUpdatesEnabled(False)
        self.setSortingEnabled(False)
        self.blockSignals(True)
        try:
            self.setRowCount(0)
            self.setColumnCount(total_cols)
            self.setRowCount(total_rows)
            self.setHorizontalHeaderLabels([get_column_letter(c + 1) for c in range(total_cols)])
            self.setVerticalHeaderLabels([str(r + 1) for r in range(total_rows)])
            _set_item = self.setItem
            _align = Qt.AlignVCenter | Qt.AlignLeft
            _bg = DIFF_COLORS["same"]
            for r in range(rows):
                row = data[r]
                row_len = len(row)
                for c in range(cols):
                    val = row[c] if c < row_len else ""
                    item = QTableWidgetItem(val)
                    item.setBackground(_bg)
                    item.setTextAlignment(_align)
                    _set_item(r, c, item)

            # populate()와 동일: 자동 너비 → 상한 클립 → 사용자 수동값 복원.
            # resizeColumnsToContents()의 sectionResized가 사용자 변경으로
            # 오해되지 않도록 _populating=True 상태에서 수행한다.
            self.resizeColumnsToContents()
            self._clip_auto_column_widths()
            if self._user_col_widths or self._user_row_heights:
                self._apply_user_sizes()
        finally:
            self.blockSignals(False)
            self.setSortingEnabled(prev_sorting)
            self.setUpdatesEnabled(prev_updates)
            self._populating = False

    def _safe_clear(self):
        self._populating = True
        try:
            self.setRowCount(0)
            self.setColumnCount(0)
        finally:
            self._populating = False
        self._header_anchor_col = None
        self._header_anchor_row = None

    def get_selected_cells(self) -> set:
        return {(item.row(), item.column()) for item in self.selectedItems()}

    def mirror_selection(self, cells: set):
        # 셀 집합을 row별로 묶고 연속 column 구간을 QItemSelectionRange로 만들어
        # 한 번의 select() 호출로 일괄 적용 — 헤더 클릭처럼 N=수천 셀일 때 결정적.
        self._populating = True
        prev_updates = self.updatesEnabled()
        self.setUpdatesEnabled(False)
        try:
            sm = self.selectionModel()
            if sm is None:
                return
            row_max = self.rowCount() - 1
            col_max = self.columnCount() - 1
            if row_max < 0 or col_max < 0:
                sm.clearSelection()
                return
            by_row: dict[int, list[int]] = {}
            for (r, c) in cells:
                if 0 <= r <= row_max and 0 <= c <= col_max:
                    by_row.setdefault(r, []).append(c)
            sel = QItemSelection()
            model = self.model()
            for r, cs in by_row.items():
                cs.sort()
                start = prev = cs[0]
                for c in cs[1:]:
                    if c == prev + 1:
                        prev = c
                        continue
                    sel.append(QItemSelectionRange(model.index(r, start), model.index(r, prev)))
                    start = prev = c
                sel.append(QItemSelectionRange(model.index(r, start), model.index(r, prev)))
            sm.select(sel, QItemSelectionModel.ClearAndSelect)
        finally:
            self.setUpdatesEnabled(prev_updates)
            self._populating = False


_EXCEL_EXTS = {".xlsx", ".xls", ".xlsm", ".xlsb"}
_SUPPORTED_EXTS = _EXCEL_EXTS | {".json", ".uasset"}


def _extract_supported_path(mime_data) -> str:
    if mime_data.hasUrls():
        for url in mime_data.urls():
            path = url.toLocalFile()
            if os.path.splitext(path)[1].lower() in _SUPPORTED_EXTS:
                return path
    return ""


class DropLineEdit(QLineEdit):
    file_dropped = pyqtSignal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)

    def dragEnterEvent(self, event):
        if _extract_supported_path(event.mimeData()):
            event.acceptProposedAction()
        else:
            event.ignore()

    def dragMoveEvent(self, event):
        if _extract_supported_path(event.mimeData()):
            event.acceptProposedAction()
        else:
            event.ignore()

    def dropEvent(self, event):
        path = _extract_supported_path(event.mimeData())
        if path:
            self.file_dropped.emit(path)
            event.acceptProposedAction()


class CellEditWidget(QPlainTextEdit):
    """셀 값 편집란 — Enter: 적용, Alt+Enter: 줄바꿈 입력.
    항상 2줄 고정 높이. 3줄 이상은 세로 스크롤.
    """
    apply_requested = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        line_h = self.fontMetrics().lineSpacing()
        # 2줄이 잘리지 않게 여유 패딩 포함
        self.setFixedHeight(line_h * 2 + 12)
        self.setLineWrapMode(QPlainTextEdit.WidgetWidth)
        self.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self._inserting_newline = False
        self._block_auto_scroll = False   # ensureCursorVisible 차단 플래그

    def ensureCursorVisible(self):
        # Alt+Enter 삽입 중에는 Qt 자동 스크롤을 차단
        if self._block_auto_scroll:
            return
        super().ensureCursorVisible()

    def keyPressEvent(self, event):
        if event.key() in (Qt.Key_Return, Qt.Key_Enter) and not (event.modifiers() & Qt.AltModifier):
            self.apply_requested.emit()
        elif event.key() in (Qt.Key_Return, Qt.Key_Enter) and (event.modifiers() & Qt.AltModifier):
            vsb = self.verticalScrollBar()
            cur_scroll = vsb.value()
            line_h = self.fontMetrics().lineSpacing()
            # 문서 전체에서 텍스트가 있는 줄이 2줄 이상일 때 스크롤 대상
            doc = self.document()
            non_empty_lines = sum(
                1 for i in range(doc.blockCount())
                if doc.findBlockByNumber(i).text().strip()
            )
            has_two_or_more_lines = non_empty_lines >= 2
            # 이미 최하단에 도달한 경우 스크롤 생략
            at_bottom = cur_scroll >= vsb.maximum()
            should_scroll = has_two_or_more_lines and not at_bottom
            # 자동 스크롤 차단 후 삽입, 직접 스크롤 값 설정
            self._block_auto_scroll = True
            self.textCursor().insertText("\n")
            self._block_auto_scroll = False
            if should_scroll:
                vsb.setValue(cur_scroll + line_h)
        else:
            super().keyPressEvent(event)
            # 일반 타이핑 시 스크롤 위치 유지 (현재 블록이 2번째 이내면 맨 위 고정)
            self._clamp_scroll_if_not_last()



    def _clamp_scroll_if_not_last(self):
        """커서가 마지막 블록이 아니면 스크롤을 상단으로 고정."""
        cursor = self.textCursor()
        doc = self.document()
        if cursor.blockNumber() < doc.blockCount() - 1:
            self.verticalScrollBar().setValue(0)

    def text(self):
        return self.toPlainText()

    def setText(self, val: str):
        self.setPlainText(val if val is not None else "")
        # 텍스트 설정 후 항상 맨 위부터 표시
        self.verticalScrollBar().setValue(0)

    def clear(self):
        super().clear()


class FilePanel(QWidget):
    file_loaded = pyqtSignal(str)
    cell_value_edited = pyqtSignal(int, int, str)   # row, col, new_value

    def __init__(self, label: str, side: str, parent=None):
        super().__init__(parent)
        self.side = side
        self.setAcceptDrops(True)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(4, 4, 4, 4)
        layout.setSpacing(4)

        header = QHBoxLayout()
        title = QLabel(label)
        title.setFont(QFont("맑은 고딕", 10, QFont.Bold))
        header.addWidget(title)
        self._drop_hint = QLabel("  엑셀/JSON/uasset 파일을 여기에 끌어다 놓으세요")
        self._drop_hint.setStyleSheet("color: #888; font-size: 9pt;")
        self._drop_hint.setFont(QFont("맑은 고딕", 9))
        header.addWidget(self._drop_hint)
        header.addStretch()
        layout.addLayout(header)

        file_row = QHBoxLayout()
        self.path_edit = DropLineEdit()
        self.path_edit.setPlaceholderText("엑셀/JSON/uasset 파일을 드래그하거나 찾아보기 버튼을 클릭하세요...")
        self.path_edit.setReadOnly(True)
        self.path_edit.setFocusPolicy(Qt.NoFocus)
        self.path_edit.setContextMenuPolicy(Qt.CustomContextMenu)
        self.path_edit.customContextMenuRequested.connect(self._on_path_context_menu)
        self.path_edit.file_dropped.connect(self._on_file_dropped)
        browse_btn = QPushButton()
        browse_btn.setIcon(self.style().standardIcon(QStyle.SP_DirOpenIcon))
        browse_btn.setFixedSize(32, 32)
        browse_btn.setIconSize(QSize(18, 18))
        browse_btn.setToolTip("찾아보기")
        browse_btn.clicked.connect(self._browse)

        self.save_btn = QPushButton()
        self.save_btn.setIcon(self.style().standardIcon(QStyle.SP_DialogSaveButton))
        self.save_btn.setFixedSize(32, 32)
        self.save_btn.setIconSize(QSize(18, 18))
        self.save_btn.setToolTip("파일 저장")
        self.save_btn.setEnabled(False)
        self.save_btn.setObjectName("save_btn")

        file_row.addWidget(self.path_edit)
        file_row.addWidget(browse_btn)
        file_row.addWidget(self.save_btn)
        layout.addLayout(file_row)

        # 셀 값 편집 행 — 라벨은 표시하지 않고 입력란만 노출
        edit_row = QHBoxLayout()
        self.cell_edit = CellEditWidget()
        self.cell_edit.setPlaceholderText("셀 선택 후 F2로 편집 (Enter 적용 / Alt+Enter 줄바꿈)")
        self.cell_edit.setFont(QFont("맑은 고딕", 9))
        self.cell_edit.setEnabled(False)
        self.cell_edit.apply_requested.connect(self._apply_cell_edit)
        edit_row.addWidget(self.cell_edit)
        layout.addLayout(edit_row)
        self._selected_cell: tuple | None = None   # (row, col) 현재 선택 셀
        self._formula_data: list[list] = []
        self._row_meta: list = []   # [(orig_a_row, orig_b_row), ...]
        self._staged_display: dict[tuple, str] = {}   # (r,c) → 병합 준비 셀의 셀값란 표시 문자열
        self._edited_values: dict[tuple, str] = {}   # (r,c) → 직접 편집된 값 (셀값란 표시 우선)

        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        layout.addWidget(line)

        self.table = ExcelTableWidget(side)
        self.table.itemSelectionChanged.connect(self._on_table_selection_changed)
        self.table.edit_focus_requested.connect(self._focus_cell_edit)
        self.table.delete_cell_requested.connect(self._on_delete_cell_requested)
        layout.addWidget(self.table)

        copy_sc = QShortcut(QKeySequence("Ctrl+C"), self)
        copy_sc.activated.connect(self._on_copy_shortcut)

    def _get_formula(self, display_r: int, c: int) -> str:
        """표시 행(display_r) → 원본 파일 행 인덱스로 변환 후 수식 문자열만 반환.
        수식이 아닌 일반 값 셀은 빈 문자열 반환 (편집값 우선 표시를 위해)."""
        side_idx = 0 if self.side == "a" else 1
        try:
            orig = self._row_meta[display_r][side_idx]
        except (IndexError, TypeError):
            orig = display_r
        if orig is None:
            return ""
        try:
            v = self._formula_data[orig][c]
            return v if v.startswith("=") else ""
        except (IndexError, TypeError):
            return ""

    def _on_table_selection_changed(self):
        if self.table._populating:
            return
        # 다른 셀로 이동 시 편집 중인 값 자동 적용
        # cell_edit 값과 비교할 때 수식 표시 중일 수 있으므로 원래 표시값도 함께 확인
        if self._selected_cell is not None:
            current_text = self.cell_edit.text()
            pr, pc = self._selected_cell
            item = self.table.item(pr, pc)
            original_text = item.text() if item else ""
            original_formula = self._get_formula(pr, pc)
            staged_override = self._staged_display.get((pr, pc))
            if (current_text != original_text
                    and current_text != original_formula
                    and (staged_override is None or current_text != staged_override)):
                self._apply_cell_edit()

        items = self.table.selectedItems()
        if len(items) == 1:
            item = items[0]
            r, c = item.row(), item.column()
            self._selected_cell = (r, c)
            rgb = ExcelTableWidget._rgb(item)
            if rgb in (ExcelTableWidget._STAGED_RGB, ExcelTableWidget._MERGED_RGB):
                override = self._staged_display.get((r, c))
                self.cell_edit.setText(override if override is not None else item.text())
            else:
                # 직접 편집된 값이 있으면 최우선 표시
                edited_val = self._edited_values.get((r, c))
                if edited_val is not None:
                    self.cell_edit.setText(edited_val)
                else:
                    formula = self._get_formula(r, c)
                    self.cell_edit.setText(formula if formula else item.text())
            self.cell_edit.setEnabled(True)
        else:
            self._selected_cell = None
            self.cell_edit.clear()
            self.cell_edit.setEnabled(False)

    def _apply_cell_edit(self):
        if self._selected_cell is None:
            return
        r, c = self._selected_cell
        new_val = self.cell_edit.text()
        self.cell_value_edited.emit(r, c, new_val)
        self.cell_edit.clearFocus()

    def _sync_cell_edit(self):
        """mirror_selection 후 cell_edit 값을 현재 선택 셀에 맞게 갱신 (포커스 이동 없음)."""
        items = self.table.selectedItems()
        if len(items) == 1:
            item = items[0]
            r, c = item.row(), item.column()
            self._selected_cell = (r, c)
            rgb = ExcelTableWidget._rgb(item)
            if rgb in (ExcelTableWidget._STAGED_RGB, ExcelTableWidget._MERGED_RGB):
                override = self._staged_display.get((r, c))
                self.cell_edit.setText(override if override is not None else item.text())
            else:
                edited_val = self._edited_values.get((r, c))
                if edited_val is not None:
                    self.cell_edit.setText(edited_val)
                else:
                    formula = self._get_formula(r, c)
                    self.cell_edit.setText(formula if formula else item.text())
            self.cell_edit.setEnabled(True)
        else:
            self._selected_cell = None
            self.cell_edit.clear()
            self.cell_edit.setEnabled(False)

    def dragEnterEvent(self, event):
        if _extract_supported_path(event.mimeData()):
            event.acceptProposedAction()
            self._set_drop_highlight(True)
        else:
            event.ignore()

    def dragMoveEvent(self, event):
        if _extract_supported_path(event.mimeData()):
            event.acceptProposedAction()
        else:
            event.ignore()

    def dragLeaveEvent(self, _event):
        self._set_drop_highlight(False)

    def dropEvent(self, event):
        path = _extract_supported_path(event.mimeData())
        self._set_drop_highlight(False)
        if path:
            self._on_file_dropped(path)
            event.acceptProposedAction()

    def _set_drop_highlight(self, active: bool):
        if active:
            self.setStyleSheet("FilePanel { border: 2px dashed #0078d4; border-radius: 6px; background: #e8f4ff; }")
        else:
            self.setStyleSheet("")

    def _on_file_dropped(self, path: str):
        self.path_edit.setText(path)
        self.file_loaded.emit(path)

    def _copy_path(self):
        path = self.path_edit.text().strip()
        if path:
            QApplication.clipboard().setText(path)

    def _on_copy_shortcut(self):
        """Ctrl+C — 테이블에 포커스 시 선택 영역 TSV 복사, 그 외엔 경로 복사."""
        focused = QApplication.focusWidget()
        is_table_focus = False
        w = focused
        while w is not None:
            if w is self.table:
                is_table_focus = True
                break
            w = w.parentWidget()
        if is_table_focus:
            self._copy_selection_as_tsv()
        else:
            self._copy_path()

    def _copy_selection_as_tsv(self):
        """선택 셀들을 bounding box 기준 TSV로 클립보드에 복사."""
        items = self.table.selectedItems()
        if not items:
            return
        rows = [it.row() for it in items]
        cols = [it.column() for it in items]
        r1, r2 = min(rows), max(rows)
        c1, c2 = min(cols), max(cols)
        sel_set = {(it.row(), it.column()) for it in items}
        lines = []
        for r in range(r1, r2 + 1):
            cells = []
            for c in range(c1, c2 + 1):
                if (r, c) in sel_set:
                    it = self.table.item(r, c)
                    cells.append(it.text() if it is not None else "")
                else:
                    cells.append("")
            lines.append("\t".join(cells))
        QApplication.clipboard().setText("\r\n".join(lines))

    def _focus_cell_edit(self):
        if self.cell_edit.isEnabled():
            self.cell_edit.setFocus()
            # 전체 선택해서 바로 덮어쓰기 가능하게
            cursor = self.cell_edit.textCursor()
            cursor.select(cursor.Document)
            self.cell_edit.setTextCursor(cursor)

    def _on_delete_cell_requested(self, r: int, c: int):
        """Delete 키로 단일 셀 값 비우기 — 기존 편집 흐름(cell_value_edited) 재사용."""
        self.cell_value_edited.emit(r, c, "")

    def _on_path_context_menu(self, pos):
        menu = QMenu(self)
        menu.setStyleSheet("QMenu { font-family: '맑은 고딕'; font-size: 9pt; }")
        act = menu.addAction("경로 복사  (Ctrl+C)")
        act.setEnabled(bool(self.path_edit.text().strip()))
        if menu.exec_(self.path_edit.mapToGlobal(pos)) == act:
            self._copy_path()

    def _browse(self):
        current = self.path_edit.text().strip()
        init_dir = os.path.dirname(current) if current and os.path.exists(current) else os.path.expanduser("~")
        path, _ = QFileDialog.getOpenFileName(
            self, "비교할 파일 선택", init_dir,
            "Supported (*.xlsx *.xls *.xlsm *.json *.uasset);;"
            "Excel (*.xlsx *.xls *.xlsm);;"
            "JSON (*.json);;"
            "Unreal Asset (*.uasset);;"
            "All Files (*)",
        )
        if path:
            self._on_file_dropped(path)

    def get_path(self) -> str:
        return self.path_edit.text().strip()

    def set_path(self, path: str):
        self.path_edit.setText(path)

    def populate(self, diff_matrix: list[list], merged_set: set = None,
                 staged: dict = None, row_meta: list = None,
                 excluded_cols: set = None):
        self.table.populate(diff_matrix, self.side, merged_set, staged, row_meta,
                            excluded_cols)

    def preview(self, data: list[list]):
        self.table.populate_preview(data)


# ── 메인 윈도우 ───────────────────────────────────────────────────────────────

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("ExcelMerge")
        # 작업 표시줄·Alt+Tab에서도 타이틀바와 동일한 앱 아이콘이 보이도록 명시 지정.
        self.setWindowIcon(_load_app_icon())
        self.resize(1400, 800)
        self._load_worker:          LoadWorker | None         = None
        self._preview_worker_a:    PreviewWorker | None      = None
        self._preview_worker_b:    PreviewWorker | None      = None
        self._staged_merge_worker: StagedMergeWorker | None = None
        self._saving_side: str = "a"
        self._diff_matrix: list[list] = []
        self._diff_row_meta: list = []   # [(orig_a_row, orig_b_row), ...]
        self._merged_cells: set = set()
        self._staged: dict = {}          # {(r, c): 'a_to_b' | 'b_to_a'}
        self._edited: dict = {"a": {}, "b": {}}   # {side: {(r,c): new_val}}
        self._preview_data: dict = {"a": [], "b": []}   # 미리보기 raw data
        self._formula_data: dict = {"a": [], "b": []}   # 수식 원문 데이터
        self._diff_only: bool = False
        self._undo_stack: list = []   # [(side, r, c, old_val)]
        self._raw_data: dict = {"a": [], "b": []}   # 키 열 변경 시 재계산용 캐시
        self._key_col: int = 0
        self._excluded_cols: set[int] = set()   # 변경 검사에서 제외할 (display) 열 인덱스

        self._build_ui()
        self._apply_style()
        undo_sc = QShortcut(QKeySequence("Ctrl+Z"), self)
        undo_sc.activated.connect(self._undo)
        diff_only_sc = QShortcut(QKeySequence("Ctrl+D"), self)
        diff_only_sc.setContext(Qt.ApplicationShortcut)
        diff_only_sc.activated.connect(self._toggle_diff_only_shortcut)
        # F5 — 새로고침 (refresh_btn이 enabled 일 때만 실행)
        refresh_sc = QShortcut(QKeySequence("F5"), self)
        refresh_sc.setContext(Qt.ApplicationShortcut)
        refresh_sc.activated.connect(self._on_refresh_shortcut)
        # Alt+↑ / Alt+↓ — 이전/다음 변경 셀로 이동
        prev_diff_sc = QShortcut(QKeySequence("Alt+Up"), self)
        prev_diff_sc.setContext(Qt.ApplicationShortcut)
        prev_diff_sc.activated.connect(self._on_prev_diff_shortcut)
        next_diff_sc = QShortcut(QKeySequence("Alt+Down"), self)
        next_diff_sc.setContext(Qt.ApplicationShortcut)
        next_diff_sc.activated.connect(self._on_next_diff_shortcut)
        # Ctrl+F — 찾기 입력란으로 포커스 이동
        find_sc = QShortcut(QKeySequence("Ctrl+F"), self)
        find_sc.setContext(Qt.ApplicationShortcut)
        find_sc.activated.connect(self._focus_find)

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(8, 8, 8, 8)
        root.setSpacing(6)

        # ── 툴바 ──
        toolbar = QHBoxLayout()

        self.diff_only_btn = QPushButton("변경 행만 보기")
        self.diff_only_btn.setFixedHeight(36)
        self.diff_only_btn.setFont(QFont("맑은 고딕", 10))
        self.diff_only_btn.setCheckable(True)
        self.diff_only_btn.setObjectName("toggle_btn")
        self.diff_only_btn.setEnabled(False)
        self.diff_only_btn.setToolTip("변경된 행만 표시 / 전체 표시 전환 (Ctrl+D)")
        self.diff_only_btn.toggled.connect(self._on_diff_only_toggled)
        toolbar.addWidget(self.diff_only_btn)

        self.refresh_btn = QPushButton("새로고침")
        self.refresh_btn.setFixedHeight(36)
        self.refresh_btn.setFont(QFont("맑은 고딕", 10))
        self.refresh_btn.setToolTip("지정된 경로의 파일을 다시 불러와 비교합니다 (F5)")
        self.refresh_btn.setEnabled(False)
        self.refresh_btn.clicked.connect(self._run_refresh)
        toolbar.addWidget(self.refresh_btn)

        # 이전/다음 변경 셀 이동 버튼
        self.prev_diff_btn = QPushButton("◀ 이전 변경")
        self.prev_diff_btn.setFixedHeight(36)
        self.prev_diff_btn.setFont(QFont("맑은 고딕", 10))
        self.prev_diff_btn.setEnabled(False)
        self.prev_diff_btn.setToolTip("이전 변경 셀로 이동 (Alt+↑)")
        self.prev_diff_btn.clicked.connect(lambda: self._goto_changed(-1))
        toolbar.addWidget(self.prev_diff_btn)

        self.next_diff_btn = QPushButton("다음 변경 ▶")
        self.next_diff_btn.setFixedHeight(36)
        self.next_diff_btn.setFont(QFont("맑은 고딕", 10))
        self.next_diff_btn.setEnabled(False)
        self.next_diff_btn.setToolTip("다음 변경 셀로 이동 (Alt+↓)")
        self.next_diff_btn.clicked.connect(lambda: self._goto_changed(+1))
        toolbar.addWidget(self.next_diff_btn)

        # 찾기 — 검색란 + 옵션 토글 + 이전/다음 찾기 버튼
        toolbar.addSpacing(16)
        find_box = QHBoxLayout()
        find_box.setSpacing(4)

        self.find_edit = QLineEdit()
        self.find_edit.setObjectName("find_edit")
        self.find_edit.setPlaceholderText("찾을 내용 (Ctrl+F)")
        self.find_edit.setFixedHeight(36)
        self.find_edit.setFixedWidth(200)
        self.find_edit.setFont(QFont("맑은 고딕", 10))
        self.find_edit.setClearButtonEnabled(True)
        self.find_edit.setEnabled(False)
        self.find_edit.setToolTip("셀 값 검색 — Enter: 다음 찾기, Shift+Enter: 이전 찾기")
        self.find_edit.returnPressed.connect(lambda: self._goto_find(+1))
        find_prev_sc = QShortcut(QKeySequence("Shift+Return"), self.find_edit)
        find_prev_sc.setContext(Qt.WidgetShortcut)
        find_prev_sc.activated.connect(lambda: self._goto_find(-1))
        find_box.addWidget(self.find_edit)

        def _find_btn(kind: str, tooltip: str, checkable: bool) -> QPushButton:
            btn = QPushButton()
            btn.setObjectName("find_btn")
            btn.setFixedSize(36, 36)
            btn.setIcon(self._make_find_icon(kind))
            btn.setIconSize(QSize(22, 22))
            btn.setCheckable(checkable)
            btn.setEnabled(False)
            btn.setToolTip(tooltip)
            find_box.addWidget(btn)
            return btn

        self.find_case_btn = _find_btn(
            "case",
            "대소문자 무시 (Ignore case)\n"
            "켜짐: 대소문자를 구분하지 않고 검색\n"
            "꺼짐: 대소문자가 정확히 일치할 때만 검색",
            checkable=True,
        )
        self.find_case_btn.setChecked(True)

        self.find_word_btn = _find_btn(
            "word",
            "전체 단어 일치 (Match whole word only)\n"
            "켜짐: 검색어가 독립된 단어로 존재할 때만 찾음\n"
            "꺼짐: 부분 문자열도 찾음",
            checkable=True,
        )

        self.find_prev_btn = _find_btn("prev", "이전 찾기 (Shift+Enter)", checkable=False)
        self.find_prev_btn.clicked.connect(lambda: self._goto_find(-1))

        self.find_next_btn = _find_btn("next", "다음 찾기 (Enter)", checkable=False)
        self.find_next_btn.clicked.connect(lambda: self._goto_find(+1))

        toolbar.addLayout(find_box)

        toolbar.addStretch()

        # 범례
        for lbl, key in [
            ("추가됨", "added"),
            ("변경됨", "modified"), ("준비 중", "staged"), ("병합됨", "merged"),
        ]:
            dot = QLabel("  ")
            dot.setFixedSize(20, 20)
            dot.setStyleSheet(
                f"background:{DIFF_COLORS[key].name()};"
                "border:1px solid #aaa; border-radius:3px;"
            )
            txt = QLabel(lbl)
            txt.setFont(QFont("맑은 고딕", 9))
            toolbar.addWidget(dot)
            toolbar.addWidget(txt)
            toolbar.addSpacing(8)

        root.addLayout(toolbar)

        # ── 좌우 패널 ──
        splitter = QSplitter(Qt.Horizontal)
        self.panel_a = FilePanel("A 파일 (원본)", "a")
        self.panel_b = FilePanel("B 파일 (비교)", "b")
        splitter.addWidget(self.panel_a)
        splitter.addWidget(self.panel_b)
        splitter.setSizes([700, 700])
        root.addWidget(splitter, 1)

        # ── 변경 셀 위치 미니맵: 양쪽 테이블의 세로/가로 스크롤바를 커스텀으로 교체 ──
        # 스크롤 동기화 시그널 연결 전에 교체해야 verticalScrollBar()/horizontalScrollBar()
        # 핸들이 새 객체를 가리킨다.
        for tbl in (self.panel_a.table, self.panel_b.table):
            tbl.setVerticalScrollBar(MinimapScrollBar(Qt.Vertical, tbl))
            tbl.setHorizontalScrollBar(MinimapScrollBar(Qt.Horizontal, tbl))

        # ── 상태바 ──
        self.status = QStatusBar()
        self.setStatusBar(self.status)
        self.status.showMessage(
            "파일을 선택하면 자동으로 비교합니다.  "
            "| 셀 선택 후 우클릭 → 병합 준비 → 저장"
        )

        # 스크롤 동기화
        for src, dst in [
            (self.panel_a.table.horizontalScrollBar(), self.panel_b.table.horizontalScrollBar()),
            (self.panel_b.table.horizontalScrollBar(), self.panel_a.table.horizontalScrollBar()),
            (self.panel_a.table.verticalScrollBar(),   self.panel_b.table.verticalScrollBar()),
            (self.panel_b.table.verticalScrollBar(),   self.panel_a.table.verticalScrollBar()),
        ]:
            src.valueChanged.connect(dst.setValue)

        # 열/행 크기 양방향 동기화 — 사용자 조작에 의한 변경만 (apply_*는 시그널 미발행)
        self.panel_a.table.column_resized.connect(self.panel_b.table.apply_column_width)
        self.panel_b.table.column_resized.connect(self.panel_a.table.apply_column_width)
        self.panel_a.table.row_resized.connect(self.panel_b.table.apply_row_height)
        self.panel_b.table.row_resized.connect(self.panel_a.table.apply_row_height)

        # 우클릭 → 스테이징 / 언스테이징
        self.panel_a.table.stage_requested.connect(self._stage_selected)
        self.panel_b.table.stage_requested.connect(self._stage_selected)
        self.panel_a.table.unstage_requested.connect(self._unstage_selected)
        self.panel_b.table.unstage_requested.connect(self._unstage_selected)

        # 열 헤더 우클릭 → 키 열 변경
        self.panel_a.table.key_col_changed.connect(self._on_key_col_changed)
        self.panel_b.table.key_col_changed.connect(self._on_key_col_changed)

        # 열 헤더 우클릭 → 변경 검사 제외 일괄 토글
        self.panel_a.table.columns_exclude_set.connect(self._on_columns_exclude_set)
        self.panel_b.table.columns_exclude_set.connect(self._on_columns_exclude_set)

        # 선택 셀 동기화
        self._syncing_selection = False
        self.panel_a.table.itemSelectionChanged.connect(
            lambda: self._sync_selection(self.panel_a.table, self.panel_b.table)
        )
        self.panel_b.table.itemSelectionChanged.connect(
            lambda: self._sync_selection(self.panel_b.table, self.panel_a.table)
        )

        # 패널 내 저장 버튼 연결
        self.panel_a.save_btn.clicked.connect(lambda: self._save_staged("a"))
        self.panel_b.save_btn.clicked.connect(lambda: self._save_staged("b"))

        # 파일 선택 시: 양쪽 모두 있으면 자동 비교, 한쪽만 있으면 미리보기
        self.panel_a.file_loaded.connect(lambda p: self._on_file_loaded("a", p))
        self.panel_b.file_loaded.connect(lambda p: self._on_file_loaded("b", p))

        # 셀 값 직접 편집
        self.panel_a.cell_value_edited.connect(lambda r, c, v: self._on_cell_edited("a", r, c, v))
        self.panel_b.cell_value_edited.connect(lambda r, c, v: self._on_cell_edited("b", r, c, v))

    def _apply_style(self):
        self.setStyleSheet("""
            QMainWindow { background: #f5f5f5; }
            QPushButton {
                background: #0078d4; color: white;
                border: none; border-radius: 4px; padding: 6px 14px;
            }
            QPushButton:hover   { background: #005fa3; }
            QPushButton:pressed { background: #004880; }
            QPushButton:disabled { background: #b0b0b0; color: #e0e0e0; }
            QPushButton#save_btn { background: #c47700; }
            QPushButton#save_btn:hover   { background: #a06000; }
            QPushButton#save_btn:pressed { background: #7d4a00; }
            QPushButton#save_btn:disabled { background: #b0b0b0; color: #e0e0e0; }
            QPushButton#toggle_btn { background: #555; }
            QPushButton#toggle_btn:checked { background: #0078d4; }
            QPushButton#toggle_btn:hover   { background: #333; }
            QPushButton#toggle_btn:disabled { background: #b0b0b0; color: #e0e0e0; }
            QPushButton#find_btn {
                background: #ffffff; border: 1px solid #c8c8c8; padding: 0;
            }
            QPushButton#find_btn:hover   { background: #eef5fc; border-color: #0078d4; }
            QPushButton#find_btn:pressed { background: #dcebf9; }
            QPushButton#find_btn:checked { background: #0078d4; border-color: #005fa3; }
            QPushButton#find_btn:checked:hover { background: #005fa3; }
            QPushButton#find_btn:disabled { background: #f2f2f2; border-color: #dedede; }
            QLineEdit#find_edit:focus { border: 1px solid #0078d4; }
            QLineEdit#find_edit:disabled { background: #f2f2f2; color: #aaa; }
            QLineEdit {
                border: 1px solid #ccc; border-radius: 4px;
                padding: 4px 8px; background: #fff;
            }
            CellEditWidget {
                border: 1px solid #ccc; border-radius: 4px;
                padding: 2px 6px; background: #fff;
            }
            CellEditWidget:disabled { background: #f5f5f5; color: #aaa; }
            QTableWidget { border: 1px solid #ddd; gridline-color: #e0e0e0; }
            QHeaderView::section {
                background: #e8eaf0; border: none;
                border-right: 1px solid #ccc; border-bottom: 1px solid #ccc;
                padding: 3px 6px; font-weight: bold;
            }
        """)

    # ── 파일 미리보기 (비교 전 단독 표시) ────────────────────────────────────────

    def _reset_compare_state(self):
        """비교 결과를 초기화하고 버튼 상태를 되돌린다."""
        self._diff_matrix = []
        self._diff_row_meta = []
        self._merged_cells = set()
        self._staged = {}
        self._edited = {"a": {}, "b": {}}
        self._preview_data = {"a": [], "b": []}
        self._formula_data = {"a": [], "b": []}
        self._excluded_cols.clear()
        self.panel_a.table.set_excluded_cols(self._excluded_cols)
        self.panel_b.table.set_excluded_cols(self._excluded_cols)
        self.panel_a._row_meta = []
        self.panel_b._row_meta = []
        self.panel_a._staged_display = {}
        self.panel_b._staged_display = {}
        self.panel_a._edited_values = {}
        self.panel_b._edited_values = {}
        self.diff_only_btn.setChecked(False)
        self.diff_only_btn.setEnabled(False)
        self.prev_diff_btn.setEnabled(False)
        self.next_diff_btn.setEnabled(False)
        self._set_find_enabled(False)
        self._update_minimap()
        self._set_save_btn_state()

    def _run_preview(self, side: str, path: str):
        """파일 선택 즉시 해당 패널에 원본 데이터를 색상 없이 표시한다."""
        # 비교 결과가 있으면 비교 상태를 리셋하고 반대쪽 패널도 미리보기로 전환
        if self._diff_matrix:
            self._reset_compare_state()
            other_side = "b" if side == "a" else "a"
            other_path = self.panel_b.get_path() if other_side == "b" else self.panel_a.get_path()
            if other_path:
                self._run_preview(other_side, other_path)

        worker = PreviewWorker(side, path)
        worker.done.connect(self._on_preview_done)
        worker.error.connect(lambda msg: self.status.showMessage(f"파일 로드 오류: {msg}"))
        worker.finished.connect(worker.deleteLater)
        if side == "a":
            self._preview_worker_a = worker
        else:
            self._preview_worker_b = worker
        worker.start()
        self.status.showMessage(f"{'A' if side == 'a' else 'B'} 파일 로딩 중...")

    def _on_preview_done(self, side: str, data: list[list], formula_data: list[list]):
        self._preview_data[side] = data
        self._formula_data[side] = formula_data
        panel = self.panel_a if side == "a" else self.panel_b
        panel._formula_data = formula_data
        panel._row_meta = []   # 미리보기 모드: row_meta 없음 (행 인덱스 = 원본 인덱스)
        panel.preview(data)
        rows = len(data)
        cols = max((len(r) for r in data), default=0)
        self.status.showMessage(
            f"{'A' if side == 'a' else 'B'} 파일 로드 완료 — {rows}행 × {cols}열  "
            "| '비교 실행'을 클릭해 두 파일을 비교하세요."
        )
        self._set_save_btn_state()

    # ── 비교 ──────────────────────────────────────────────────────────────────

    def _on_refresh_shortcut(self):
        """F5 단축키 — 새로고침 버튼이 활성 상태일 때만 동작."""
        if self.refresh_btn.isEnabled():
            self._run_refresh()

    def _run_refresh(self):
        # 새로고침은 디폴트(자동너비+MAX_AUTO_COL_WIDTH_PX 상한) 상태로 복귀시킨다.
        # 세션 중 사용자가 헤더를 드래그해 늘려놓은 열/행 크기도 함께 리셋되어야
        # populate() 종료부의 _apply_user_sizes() 분기를 건너뛴다.
        for tbl in (self.panel_a.table, self.panel_b.table):
            tbl._user_col_widths.clear()
            tbl._user_row_heights.clear()

        path_a = self.panel_a.get_path()
        path_b = self.panel_b.get_path()
        if path_a and path_b:
            self._run_compare()
        elif path_a:
            self._run_preview("a", path_a)
        elif path_b:
            self._run_preview("b", path_b)

    def _run_compare(self):
        path_a = self.panel_a.get_path()
        path_b = self.panel_b.get_path()
        if not path_a and not path_b:
            QMessageBox.warning(self, "경고", "A 파일과 B 파일 중 하나 이상 선택하세요.")
            return

        self._set_buttons_enabled(False)
        self._merged_cells = set()
        self._staged = {}
        self._edited = {"a": {}, "b": {}}
        self._diff_matrix = []
        self._diff_row_meta = []   # 미리보기 잠금 해제
        self._excluded_cols.clear()
        self.panel_a.table.set_excluded_cols(self._excluded_cols)
        self.panel_b.table.set_excluded_cols(self._excluded_cols)
        self.status.showMessage("파일 로딩 중...")

        self._load_worker = LoadWorker(path_a, path_b)
        self._load_worker.done.connect(self._on_loaded)
        self._load_worker.error.connect(self._on_error)
        self._load_worker.progress.connect(self.status.showMessage)
        self._load_worker.finished.connect(self._load_worker.deleteLater)
        self._load_worker.start()

    def _on_loaded(self, a_data, b_data, a_formulas, b_formulas):
        self._raw_data["a"] = a_data
        self._raw_data["b"] = b_data
        self._formula_data["a"] = a_formulas
        self._formula_data["b"] = b_formulas
        self.panel_a._formula_data = a_formulas
        self.panel_b._formula_data = b_formulas
        self._diff_matrix, self._diff_row_meta = compute_diff(
            a_data, b_data, self._key_col)
        self.panel_a._row_meta = self._diff_row_meta
        self.panel_b._row_meta = self._diff_row_meta
        self._refresh_tables()

        rows = len(self._diff_matrix)
        cols = len(self._diff_matrix[0]) if self._diff_matrix else 0
        changed = self._count_changed()

        self._set_buttons_enabled(True)
        # 비교 완료 시 디폴트로 "변경 행만 보기" ON.
        # 이미 ON이었다면 setChecked는 시그널이 발생하지 않으므로 _apply_diff_filter를 직접 호출.
        if self.diff_only_btn.isChecked():
            self._apply_diff_filter()
        else:
            self.diff_only_btn.setChecked(True)
        self.status.showMessage(
            f"비교 완료 — {rows}행 × {cols}열 | 변경된 셀: {changed}개  "
            "| 셀 선택 후 우클릭 → 병합 준비 → 선택 병합 저장"
        )

    # ── 키 열 변경 ────────────────────────────────────────────────────────────

    def _on_key_col_changed(self, col: int):
        if col == self._key_col:
            return
        self._key_col = col
        self.panel_a.table.set_key_col(col)
        self.panel_b.table.set_key_col(col)
        if self._raw_data["a"] or self._raw_data["b"]:
            self._recompute_diff()

    def _recompute_diff(self):
        self._merged_cells = set()
        self._staged = {}
        self._edited = {"a": {}, "b": {}}
        # 키 열이 바뀌면 동일 인덱스가 다른 의미가 될 수 있으므로 제외 상태도 리셋.
        self._excluded_cols.clear()
        self.panel_a.table.set_excluded_cols(self._excluded_cols)
        self.panel_b.table.set_excluded_cols(self._excluded_cols)
        # 키 열 변경 시 사용자의 현재 토글 상태(diff_only_btn)는 그대로 유지한다.
        self._diff_matrix, self._diff_row_meta = compute_diff(
            self._raw_data["a"], self._raw_data["b"], self._key_col)
        self.panel_a._row_meta = self._diff_row_meta
        self.panel_b._row_meta = self._diff_row_meta
        self._refresh_tables()
        self._set_buttons_enabled(True)
        self._apply_diff_filter()
        rows = len(self._diff_matrix)
        cols = len(self._diff_matrix[0]) if self._diff_matrix else 0
        changed = self._count_changed()
        key_letter = get_column_letter(self._key_col + 1) if self._key_col >= 0 else "없음(ROW 순서)"
        self.status.showMessage(
            f"키 열: {key_letter}  |  {rows}행 × {cols}열  |  변경된 셀: {changed}개  "
            "| 셀 선택 후 우클릭 → 병합 준비 → 선택 병합 저장"
        )

    # ── 변경 행 필터 ──────────────────────────────────────────────────────────

    def _toggle_diff_only_shortcut(self):
        if self.diff_only_btn.isEnabled():
            self.diff_only_btn.setChecked(not self.diff_only_btn.isChecked())

    def _on_diff_only_toggled(self, checked: bool):
        self._diff_only = checked
        self._apply_diff_filter()

    def _apply_diff_filter(self):
        if not self._diff_matrix:
            self._update_minimap()
            return
        # setRowHidden은 sectionResized(_, _, 0)을 emit해 _user_row_heights를 0으로
        # 오염시킨다. _applying_sizes 플래그로 _on_section_v_resized 측 기록을 차단.
        tables = (self.panel_a.table, self.panel_b.table)
        for tbl in tables:
            tbl._applying_sizes = True
        try:
            excl = self._excluded_cols
            for r, row in enumerate(self._diff_matrix):
                is_header = (r == 0)   # 최상단 행은 항상 표시
                is_changed = any(
                    status != "same"
                    for c, (status, *_) in enumerate(row)
                    if c not in excl
                )
                hidden = self._diff_only and not is_changed and not is_header
                self.panel_a.table.setRowHidden(r, hidden)
                self.panel_b.table.setRowHidden(r, hidden)
        finally:
            for tbl in tables:
                tbl._applying_sizes = False
        self._update_minimap()

    # ── 변경 셀 탐색(이전/다음) ───────────────────────────────────────────────

    def _on_prev_diff_shortcut(self):
        if self.prev_diff_btn.isEnabled():
            self._goto_changed(-1)

    def _on_next_diff_shortcut(self):
        if self.next_diff_btn.isEnabled():
            self._goto_changed(+1)

    def _iter_changed_cells(self):
        """변경된 (r, c) 셀을 행 우선 순서로 yield. 숨겨진 행/제외 열은 제외."""
        if not self._diff_matrix:
            return
        excl = self._excluded_cols
        for r, row in enumerate(self._diff_matrix):
            if self.panel_a.table.isRowHidden(r):
                continue
            for c, cell in enumerate(row):
                if c in excl:
                    continue
                if cell[0] != "same":
                    yield (r, c)

    def _current_anchor(self):
        """현재 선택 셀(우선순위: panel_a → panel_b).
        없으면 (0, -1) 반환 — A1부터 검사하기 위한 sentinel."""
        for tbl in (self.panel_a.table, self.panel_b.table):
            r, c = tbl.currentRow(), tbl.currentColumn()
            if r >= 0 and c >= 0:
                return (r, c)
        return (0, -1)

    def _goto_changed(self, direction: int):
        """direction=+1: 다음 변경 셀, -1: 이전 변경 셀."""
        if not self._diff_matrix:
            return
        cells = list(self._iter_changed_cells())
        if not cells:
            self.status.showMessage("변경된 셀이 없습니다.")
            return
        anchor = self._current_anchor()
        if direction > 0:
            target = next((p for p in cells if p > anchor), None)
            if target is None:
                self.status.showMessage("마지막 변경 셀입니다.")
                return
        else:
            target = next((p for p in reversed(cells) if p < anchor), None)
            if target is None:
                self.status.showMessage("첫 변경 셀입니다.")
                return
        r, c = target
        # 양쪽 패널 동기 선택 + 화면 중앙으로 스크롤
        for tbl in (self.panel_a.table, self.panel_b.table):
            tbl.setCurrentCell(r, c)
            item = tbl.item(r, c)
            if item:
                tbl.scrollToItem(item, QAbstractItemView.PositionAtCenter)

    # ── 찾기 ──
    def _make_find_icon(self, kind: str) -> QIcon:
        """찾기 버튼용 아이콘을 QPainter로 렌더링 (HiDPI 2배 해상도).
        체크 시 배경이 파란색으로 바뀌므로 Off=진회색 / On=흰색 두 벌을 등록."""
        def render(color: QColor) -> QPixmap:
            s = 32
            pm = QPixmap(s * 2, s * 2)
            pm.setDevicePixelRatio(2)
            pm.fill(Qt.transparent)
            p = QPainter(pm)
            p.setRenderHint(QPainter.Antialiasing)
            p.setRenderHint(QPainter.TextAntialiasing)
            if kind == "case":
                p.setPen(color)
                p.setFont(QFont("Segoe UI", 12, QFont.Bold))
                p.drawText(QRect(0, 0, s, s), Qt.AlignCenter, "Aa")
            elif kind == "word":
                p.setPen(color)
                p.setFont(QFont("Segoe UI", 10, QFont.Bold))
                p.drawText(QRect(0, 0, s, s - 8), Qt.AlignCenter, "ab")
                p.setPen(QPen(color, 1.8, Qt.SolidLine, Qt.RoundCap))
                y = s - 7
                p.drawLine(QPoint(7, y), QPoint(s - 7, y))
                p.drawLine(QPoint(7, y), QPoint(7, y - 4))
                p.drawLine(QPoint(s - 7, y), QPoint(s - 7, y - 4))
            else:  # "prev" / "next" — 셰브론 화살표
                p.setPen(QPen(color, 2.6, Qt.SolidLine, Qt.RoundCap, Qt.RoundJoin))
                m = s // 2
                d = -1 if kind == "prev" else 1
                p.drawLine(QPoint(m - 3 * d, m - 7), QPoint(m + 4 * d, m))
                p.drawLine(QPoint(m + 4 * d, m), QPoint(m - 3 * d, m + 7))
            p.end()
            return pm

        ic = QIcon()
        ic.addPixmap(render(QColor("#3b3b3b")), QIcon.Normal, QIcon.Off)
        ic.addPixmap(render(QColor("#ffffff")), QIcon.Normal, QIcon.On)
        ic.addPixmap(render(QColor("#b8b8b8")), QIcon.Disabled, QIcon.Off)
        ic.addPixmap(render(QColor("#b8b8b8")), QIcon.Disabled, QIcon.On)
        return ic

    def _focus_find(self):
        """Ctrl+F — 찾기 입력란 포커스 + 전체 선택."""
        if self.find_edit.isEnabled():
            self.find_edit.setFocus()
            self.find_edit.selectAll()

    def _make_find_matcher(self, term: str):
        """검색 옵션(대소문자 무시/전체 단어)에 맞는 판별 함수를 반환."""
        ignore_case = self.find_case_btn.isChecked()
        whole_word = self.find_word_btn.isChecked()
        if whole_word:
            # \b는 검색어가 특수문자로 시작/끝나면 동작하지 않으므로 lookaround 사용
            flags = re.IGNORECASE if ignore_case else 0
            pat = re.compile(r"(?<!\w)" + re.escape(term) + r"(?!\w)", flags)
            return lambda text: pat.search(text) is not None
        if ignore_case:
            needle = term.casefold()
            return lambda text: needle in text.casefold()
        return lambda text: term in text

    def _iter_find_matches(self, match):
        """검색어와 일치하는 (r, c) 셀을 행 우선 순서로 yield.
        숨겨진 행/제외 열은 제외. A/B 표시 텍스트 중 한쪽이라도 일치하면 매치."""
        if not self._diff_matrix:
            return
        excl = self._excluded_cols
        tbl_a, tbl_b = self.panel_a.table, self.panel_b.table
        for r, row in enumerate(self._diff_matrix):
            if tbl_a.isRowHidden(r):
                continue
            for c in range(len(row)):
                if c in excl:
                    continue
                for tbl in (tbl_a, tbl_b):
                    item = tbl.item(r, c)
                    if item is not None and match(item.text()):
                        yield (r, c)
                        break

    def _goto_find(self, direction: int):
        """direction=+1: 다음 찾기, -1: 이전 찾기. 끝에 도달하면 반대편에서 순환."""
        term = self.find_edit.text()
        if not term or not self._diff_matrix:
            return
        cells = list(self._iter_find_matches(self._make_find_matcher(term)))
        if not cells:
            self.status.showMessage(f'"{term}" — 일치 항목이 없습니다.')
            return
        anchor = self._current_anchor()
        wrapped = ""
        if direction > 0:
            target = next((p for p in cells if p > anchor), None)
            if target is None:
                target = cells[0]
                wrapped = " — 처음부터 다시 검색합니다."
        else:
            target = next((p for p in reversed(cells) if p < anchor), None)
            if target is None:
                target = cells[-1]
                wrapped = " — 끝에서부터 다시 검색합니다."
        r, c = target
        # 양쪽 패널 동기 선택 + 화면 중앙으로 스크롤
        for tbl in (self.panel_a.table, self.panel_b.table):
            tbl.setCurrentCell(r, c)
            item = tbl.item(r, c)
            if item:
                tbl.scrollToItem(item, QAbstractItemView.PositionAtCenter)
        idx = cells.index(target) + 1
        self.status.showMessage(f'찾기: "{term}" {idx}/{len(cells)}개 일치{wrapped}')

    def _update_minimap(self):
        """현재 표시 중인 행을 기준으로 변경 행/열의 비율 위치를 양쪽
        세로/가로 스크롤바에 전달한다.
        - 세로: 가시 행 중 변경 셀이 하나라도 있는 행의 위치.
        - 가로: 가시 행 안에서 변경 셀이 있는 열의 위치.
        """
        row_ratios = []
        col_ratios = []
        if self._diff_matrix:
            excl = self._excluded_cols
            visible_rows = [
                r for r in range(len(self._diff_matrix))
                if not self.panel_a.table.isRowHidden(r)
            ]

            def _row_has_changed(r):
                return any(
                    st != "same"
                    for c, (st, *_) in enumerate(self._diff_matrix[r])
                    if c not in excl
                )

            n = len(visible_rows)
            if n == 1:
                r = visible_rows[0]
                if _row_has_changed(r):
                    row_ratios.append(0.0)
            elif n > 1:
                denom = n - 1
                for vi, r in enumerate(visible_rows):
                    if _row_has_changed(r):
                        row_ratios.append(vi / denom)

            cols_total = len(self._diff_matrix[0]) if self._diff_matrix else 0
            if cols_total > 0 and visible_rows:
                if cols_total == 1:
                    if 0 not in excl and any(self._diff_matrix[r][0][0] != "same" for r in visible_rows):
                        col_ratios.append(0.0)
                else:
                    denom_c = cols_total - 1
                    for c in range(cols_total):
                        if c in excl:
                            continue
                        if any(self._diff_matrix[r][c][0] != "same" for r in visible_rows):
                            col_ratios.append(c / denom_c)
        for tbl in (self.panel_a.table, self.panel_b.table):
            v = tbl.verticalScrollBar()
            if isinstance(v, MinimapScrollBar):
                v.set_change_ratios(row_ratios)
            h = tbl.horizontalScrollBar()
            if isinstance(h, MinimapScrollBar):
                h.set_change_ratios(col_ratios)

    # ── 선택 셀 스테이징 (우클릭) ─────────────────────────────────────────────

    def _stage_selected(self, direction: str):
        if not self._diff_matrix:
            return

        cells = (
            self.panel_a.table.get_selected_cells()
            | self.panel_b.table.get_selected_cells()
        )
        cells = {
            (r, c) for (r, c) in cells
            if r < len(self._diff_matrix)
            and c < len(self._diff_matrix[r])
            and c not in self._excluded_cols
            and self._diff_matrix[r][c][0] != "same"
        }
        if not cells:
            QMessageBox.information(self, "알림", "선택한 셀 중 변경된 셀이 없습니다.")
            return

        # undo 스택에 stage 동작 기록 (셀 목록과 방향을 한 번에 저장)
        self._undo_stack.append(("stage", list(cells), direction))

        for cell in cells:
            self._staged[cell] = direction

        # staged 셀에 대해 양쪽 패널의 셀값란 표시값(수식 우선) 미리 계산
        def _resolve_display(fd, meta, side_idx, r, c, fallback):
            try:
                orig = meta[r][side_idx] if r < len(meta) else r
            except (IndexError, TypeError):
                orig = r
            if orig is not None:
                try:
                    v = fd[orig][c]
                    if v:
                        return v
                except (IndexError, TypeError):
                    pass
            return fallback

        for (r, c) in cells:
            dir_ = self._staged[r, c]
            try:
                _, a_val, b_val = self._diff_matrix[r][c]
            except (IndexError, TypeError):
                a_val, b_val = "", ""

            # 편집된 값이 있으면 _formula_data보다 우선 사용
            a_edited = self.panel_a._edited_values.get((r, c))
            b_edited = self.panel_b._edited_values.get((r, c))

            if dir_ == "a_to_b":
                if a_edited is not None:
                    a_display = a_edited
                else:
                    a_display = _resolve_display(self._formula_data["a"], self.panel_a._row_meta, 0, r, c, a_val)
                b_display = a_display
            else:
                if b_edited is not None:
                    b_display = b_edited
                else:
                    b_display = _resolve_display(self._formula_data["b"], self.panel_b._row_meta, 1, r, c, b_val)
                a_display = b_display
            self.panel_a._staged_display[r, c] = a_display
            self.panel_b._staged_display[r, c] = b_display

        self._refresh_tables()
        self.panel_a.table.clearSelection()
        self.panel_b.table.clearSelection()
        self.panel_a._selected_cell = None
        self.panel_b._selected_cell = None
        self.panel_a.cell_edit.clear()
        self.panel_a.cell_edit.setEnabled(False)
        self.panel_b.cell_edit.clear()
        self.panel_b.cell_edit.setEnabled(False)
        self._set_save_btn_state()
        self.status.showMessage(
            f"병합 준비 완료 — {len(self._staged)}개 셀 대기 중  | '저장'을 클릭하면 파일에 저장됩니다."
        )

    # ── 선택 셀 언스테이징 (병합 준비 취소) ───────────────────────────────────────

    def _unstage_selected(self):
        if not self._staged:
            return
        cells = (
            self.panel_a.table.get_selected_cells()
            | self.panel_b.table.get_selected_cells()
        )
        removed = {c for c in cells if c in self._staged}
        if not removed:
            QMessageBox.information(self, "알림", "선택한 셀 중 병합 준비된 셀이 없습니다.")
            return
        for cell in removed:
            del self._staged[cell]
            self.panel_a._staged_display.pop(cell, None)
            self.panel_b._staged_display.pop(cell, None)
        sel_a = self.panel_a.table.get_selected_cells()
        sel_b = self.panel_b.table.get_selected_cells()
        self._refresh_tables()
        if sel_a:
            self.panel_a.table.mirror_selection(sel_a)
        if sel_b:
            self.panel_b.table.mirror_selection(sel_b)
        self.panel_a._sync_cell_edit()
        self.panel_b._sync_cell_edit()
        self._set_save_btn_state()
        self.status.showMessage(
            f"병합 준비 취소 — {len(removed)}개 셀 제거됨 | 남은 대기 셀: {len(self._staged)}개"
        )

    # ── 저장 ─────────────────────────────────────────────────────────────────

    def _save_staged(self, side: str):
        """side: 'a' 또는 'b' — 해당 파일만 저장."""
        other = "b" if side == "a" else "a"
        path = self.panel_a.get_path() if side == "a" else self.panel_b.get_path()

        # JSON/uasset 등 비-xlsx 는 저장 미지원 — 사용자에게 안내 후 중단
        if path and os.path.splitext(path)[1].lower() not in _EXCEL_EXTS:
            QMessageBox.information(
                self, "저장 미지원",
                f"{'A' if side == 'a' else 'B'} 파일은 비교 전용 형식입니다.\n"
                f"저장(병합)은 Excel(.xlsx/.xls/.xlsm) 파일에서만 지원됩니다.",
            )
            return

        # ── 미리보기 상태 전용 저장 (diff 없이 edited만 있는 경우) ──────────
        if not self._diff_matrix:
            if not self._edited.get(side):
                return
            if not path:
                QMessageBox.warning(self, "경고",
                    f"{'A' if side == 'a' else 'B'} 파일이 지정되지 않았습니다.")
                return
            if _is_file_locked(path):
                QMessageBox.warning(self, "파일 열림",
                    f"변경하고자 하는 파일이 열려 있으므로 저장할 수 없습니다:\n\n"
                    f"{'A' if side == 'a' else 'B'} 파일: {os.path.basename(path)}"
                    "\n\n파일을 닫은 후 다시 시도하세요.")
                return
            data_len = len(self._preview_data.get(side, []))
            meta = [(r if side == "a" else None, r if side == "b" else None)
                    for r in range(data_len)]
            dummy_matrix = [[("same", "", "")] for _ in range(data_len)]
            self._saving_side = side
            self._set_buttons_enabled(False)
            self.status.showMessage("저장 중...")
            edited_side = {side: dict(self._edited[side]), other: {}}
            fa = self._formula_data["a"] if side == "a" else []
            fb = self._formula_data["b"] if side == "b" else []
            path_a = path if side == "a" else ""
            path_b = path if side == "b" else ""
            self._staged_merge_worker = StagedMergeWorker(
                path_a, path_b, dummy_matrix, meta, {}, edited_side, fa, fb,
            )
            self._staged_merge_worker.done.connect(self._on_staged_saved)
            self._staged_merge_worker.error.connect(self._on_error)
            self._staged_merge_worker.finished.connect(self._staged_merge_worker.deleteLater)
            self._staged_merge_worker.start()
            return

        # ── diff 모드 저장 ──────────────────────────────────────────────────
        has_a2b = any(v == "a_to_b" for v in self._staged.values())
        has_b2a = any(v == "b_to_a" for v in self._staged.values())
        has_edit_a = bool(self._edited.get("a"))
        has_edit_b = bool(self._edited.get("b"))

        # 저장 대상 측이 실제로 쓸 내용이 있는지 확인
        if side == "a":
            needs_write = has_b2a or has_edit_a
        else:
            needs_write = has_a2b or has_edit_b

        if not needs_write:
            return

        if not path:
            QMessageBox.warning(self, "경고",
                f"{'A' if side == 'a' else 'B'} 파일이 지정되지 않았습니다.")
            return

        if _is_file_locked(path):
            QMessageBox.warning(
                self, "파일 열림",
                f"변경하고자 하는 파일이 열려 있으므로 저장할 수 없습니다:\n\n"
                f"{'A' if side == 'a' else 'B'} 파일: {os.path.basename(path)}"
                "\n\n파일을 닫은 후 다시 시도하세요."
            )
            return

        # 저장하지 않는 쪽 경로를 빈 문자열로 전달 → Worker가 해당 파일은 건드리지 않음
        path_a = path if side == "a" else ""
        path_b = path if side == "b" else ""
        # staged 병합 시 반대쪽 edited 값도 병합 소스로 필요하므로 양쪽 모두 전달
        edited_side = {
            "a": dict(self._edited["a"]),
            "b": dict(self._edited["b"]),
        }
        # staged 셀 중 이 side에 쓰는 것만 전달
        # a 저장: b_to_a (B→A 방향) staged 셀만
        # b 저장: a_to_b (A→B 방향) staged 셀만
        relevant_direction = "b_to_a" if side == "a" else "a_to_b"
        staged_for_side = {k: v for k, v in self._staged.items() if v == relevant_direction}

        self._saving_side = side
        self._set_buttons_enabled(False)
        self.status.showMessage("저장 중...")

        self._staged_merge_worker = StagedMergeWorker(
            path_a, path_b, list(self._diff_matrix),
            list(self._diff_row_meta),
            staged_for_side,
            edited_side,
            self._formula_data["a"], self._formula_data["b"],
        )
        self._staged_merge_worker.done.connect(self._on_staged_saved)
        self._staged_merge_worker.error.connect(self._on_error)
        self._staged_merge_worker.finished.connect(self._staged_merge_worker.deleteLater)
        self._staged_merge_worker.start()

    def _on_staged_saved(self, count: int):
        side = getattr(self, "_saving_side", None)

        # 미리보기 상태(비교 전) 저장 완료 → 해당 side 재로드
        if not self._diff_matrix:
            self._edited[side] = {}
            self._set_buttons_enabled(True)
            self.status.showMessage(f"저장 완료 — {count}개 셀 저장됨")
            QMessageBox.information(self, "저장 완료", f"{count}개 셀이 파일에 저장됐습니다.")
            path = self.panel_a.get_path() if side == "a" else self.panel_b.get_path()
            if path:
                self._run_preview(side, path)
            return

        # ── diff 모드: 저장한 side에 해당하는 staged/edited만 확정 반영 ──────
        relevant_direction = "b_to_a" if side == "a" else "a_to_b"
        saved_staged = {k: v for k, v in self._staged.items() if v == relevant_direction}
        staged_cells = set(saved_staged.keys())

        # staged 방향대로 diff_matrix 확정 반영
        for (r, c), direction in saved_staged.items():
            if r < len(self._diff_matrix) and c < len(self._diff_matrix[r]):
                _, a_val, b_val = self._diff_matrix[r][c]
                if direction == "a_to_b":
                    b_val = a_val
                else:
                    a_val = b_val
                self._diff_matrix[r][c] = ("same", a_val, b_val)

        # 저장한 side의 staged/edited 제거 (나머지 side는 유지)
        # ※ edited 셀의 diff_matrix는 _on_cell_edited에서 이미 계산값으로 갱신되어 있으므로
        #   여기서 _edited[side] 값(수식 원문 포함)으로 재덮어쓰지 않는다.
        for k in list(self._staged.keys()):
            if self._staged[k] == relevant_direction:
                del self._staged[k]
        self._edited[side] = {}
        self._merged_cells |= staged_cells

        self._refresh_tables()
        self._set_buttons_enabled(True)
        self.status.showMessage(f"저장 완료 — {count}개 셀 저장됨")
        QMessageBox.information(self, "저장 완료", f"{count}개 셀이 파일에 저장됐습니다.")

    def _on_error(self, msg: str):
        self._set_buttons_enabled(True)
        self.status.showMessage(f"오류: {msg}")
        QMessageBox.critical(self, "오류", f"작업 실패:\n{msg}")

    # ── 셀 직접 편집 ──────────────────────────────────────────────────────────

    def _on_cell_edited(self, side: str, r: int, c: int, new_val: str):
        panel = self.panel_a if side == "a" else self.panel_b

        # ── 미리보기 상태 (비교 전) ──────────────────────────────────────────
        if not self._diff_matrix:
            data = self._preview_data.get(side, [])
            if r >= len(data):
                data.extend([[] for _ in range(r - len(data) + 1)])
                self._preview_data[side] = data
            row = data[r]
            if c >= len(row):
                row.extend([""] * (c - len(row) + 1))
            old_val = data[r][c]
            self._undo_stack.append((side, r, c, old_val, "preview"))
            # 수식이면 계산값으로 표시, 수식 자체는 formula_data와 _edited_values에 저장
            if new_val.startswith("="):
                display_val = _eval_formula_with_row(new_val, data[r])
                try:
                    if r < len(self._formula_data.get(side, [])):
                        while len(self._formula_data[side][r]) <= c:
                            self._formula_data[side][r].append("")
                        self._formula_data[side][r][c] = new_val
                except (IndexError, TypeError):
                    pass
                data[r][c] = display_val
            else:
                data[r][c] = new_val
            panel.preview(data)
            panel.cell_edit.setText(new_val)
            self._edited[side][(r, c)] = new_val
            panel._edited_values[(r, c)] = new_val
            self._set_save_btn_state()
            self.status.showMessage(
                f"셀 ({r+1}행, {get_column_letter(c+1)}열) 편집됨 — 저장 버튼을 눌러 파일에 반영하세요."
            )
            return

        # ── 비교 후 상태 ─────────────────────────────────────────────────────
        if r >= len(self._diff_matrix) or c >= len(self._diff_matrix[r]):
            return

        # undo 스택에 이전 값 저장
        _, a_val_cur, b_val_cur = self._diff_matrix[r][c]
        old_val = a_val_cur if side == "a" else b_val_cur
        self._undo_stack.append((side, r, c, old_val, "diff"))

        # 저장용 누적
        self._edited[side][(r, c)] = new_val
        panel._edited_values[(r, c)] = new_val

        # 병합 준비(staged) 상태 셀을 수정하면 준비 해제
        if (r, c) in self._staged:
            del self._staged[(r, c)]

        # diff_matrix 즉시 갱신 (수식이면 계산값으로 표시, 수식 자체는 formula_data에 저장)
        _, a_val, b_val = self._diff_matrix[r][c]
        if side == "a":
            if new_val.startswith("="):
                row_data = [self._diff_matrix[r][cc][1] for cc in range(len(self._diff_matrix[r]))]
                display_val = _eval_formula_with_row(new_val, row_data)
                try:
                    orig = panel._row_meta[r][0]
                    if orig is not None and orig < len(self._formula_data["a"]):
                        while len(self._formula_data["a"][orig]) <= c:
                            self._formula_data["a"][orig].append("")
                        self._formula_data["a"][orig][c] = new_val
                except (IndexError, TypeError):
                    pass
                a_val = display_val
            else:
                a_val = new_val
        else:
            if new_val.startswith("="):
                row_data = [self._diff_matrix[r][cc][2] for cc in range(len(self._diff_matrix[r]))]
                display_val = _eval_formula_with_row(new_val, row_data)
                try:
                    orig = panel._row_meta[r][1]
                    if orig is not None and orig < len(self._formula_data["b"]):
                        while len(self._formula_data["b"][orig]) <= c:
                            self._formula_data["b"][orig].append("")
                        self._formula_data["b"][orig][c] = new_val
                except (IndexError, TypeError):
                    pass
                b_val = display_val
            else:
                b_val = new_val
        status = _cell_status(a_val, b_val)
        self._diff_matrix[r][c] = (status, a_val, b_val)

        # 병합됨 셀을 수정했을 때 값이 달라지면 merged 상태 해제
        if (r, c) in self._merged_cells:
            if status != "same":
                self._merged_cells.discard((r, c))

        self._refresh_tables()

        # _refresh_tables 후 populate가 선택을 초기화하므로 cell_edit 값 복원
        panel.cell_edit.setText(new_val)

        self._set_save_btn_state()
        self.status.showMessage(
            f"셀 ({r+1}행, {get_column_letter(c+1)}열) 편집됨 — 저장 버튼을 눌러 파일에 반영하세요."
        )

    def _undo(self):
        if not self._undo_stack:
            return
        entry = self._undo_stack.pop()

        # 병합 준비(stage) 되돌리기
        if entry[0] == "stage":
            _, cells, _ = entry
            for cell in cells:
                self._staged.pop(cell, None)
                self.panel_a._staged_display.pop(cell, None)
                self.panel_b._staged_display.pop(cell, None)
            self.panel_a.table.clearSelection()
            self.panel_b.table.clearSelection()
            self.panel_a._selected_cell = None
            self.panel_b._selected_cell = None
            self._refresh_tables()
            self._set_save_btn_state()
            return

        side, r, c, old_val, mode = entry
        panel = self.panel_a if side == "a" else self.panel_b

        if mode == "preview":
            data = self._preview_data.get(side, [])
            if r < len(data) and c < len(data[r]):
                data[r][c] = old_val
            self._edited[side].pop((r, c), None)
            panel._edited_values.pop((r, c), None)
            panel.preview(data)
            panel.cell_edit.setText(old_val)
            self._set_save_btn_state()
        else:
            if r >= len(self._diff_matrix) or c >= len(self._diff_matrix[r]):
                return
            self._edited[side].pop((r, c), None)
            panel._edited_values.pop((r, c), None)
            _, a_val, b_val = self._diff_matrix[r][c]
            if side == "a":
                a_val = old_val
            else:
                b_val = old_val
            status = _cell_status(a_val, b_val)
            self._diff_matrix[r][c] = (status, a_val, b_val)
            if (r, c) in self._merged_cells and status != "same":
                self._merged_cells.discard((r, c))
            # _refresh_tables 전에 선택 상태를 초기화해야
            # populate 후 itemSelectionChanged 발화 시 자동 적용 로직이 cell_edit 값을
            # 엉뚱한 셀에 쓰는 것을 막을 수 있다.
            self.panel_a.table.clearSelection()
            self.panel_b.table.clearSelection()
            self.panel_a._selected_cell = None
            self.panel_b._selected_cell = None
            self._refresh_tables()
            panel.cell_edit.clear()
            panel.cell_edit.setEnabled(False)
            self._set_save_btn_state()

    # ── 유틸 ──────────────────────────────────────────────────────────────────

    def _refresh_tables(self):
        self.panel_a.populate(self._diff_matrix, self._merged_cells, self._staged,
                              self._diff_row_meta, self._excluded_cols)
        self.panel_b.populate(self._diff_matrix, self._merged_cells, self._staged,
                              self._diff_row_meta, self._excluded_cols)
        self._apply_diff_filter()

    def _effective_status(self, r: int, c: int) -> str:
        """제외 열은 강제로 'same' 으로 노출 — _diff_matrix 원본은 보존."""
        if c in self._excluded_cols:
            return "same"
        return self._diff_matrix[r][c][0]

    def _on_columns_exclude_set(self, cols: list, exclude: bool):
        """헤더 우클릭 → cols 일괄 제외/해제."""
        if not cols:
            return
        if exclude:
            new_cols = [c for c in cols if c not in self._excluded_cols]
            for c in new_cols:
                self._excluded_cols.add(c)
            # 새로 제외된 열들의 기존 staged 항목 자동 해제.
            new_set = set(new_cols)
            staged_keys = [k for k in self._staged if k[1] in new_set]
            for key in staged_keys:
                del self._staged[key]
                self.panel_a._staged_display.pop(key, None)
                self.panel_b._staged_display.pop(key, None)
        else:
            for c in cols:
                self._excluded_cols.discard(c)
        self.panel_a.table.set_excluded_cols(self._excluded_cols)
        self.panel_b.table.set_excluded_cols(self._excluded_cols)
        self._refresh_tables()
        self._set_save_btn_state()
        changed = self._count_changed()
        excl_letters = ", ".join(get_column_letter(c + 1) for c in sorted(self._excluded_cols))
        excl_msg = excl_letters if excl_letters else "없음"
        self.status.showMessage(
            f"검사 제외 열: {excl_msg}  |  변경된 셀: {changed}개"
        )

    def _sync_selection(self, src: ExcelTableWidget, dst: ExcelTableWidget):
        if self._syncing_selection or src._populating or dst._populating:
            return
        self._syncing_selection = True
        try:
            dst.mirror_selection(src.get_selected_cells())
            # mirror_selection 중 _populating=True라 _on_table_selection_changed가 막히므로
            # 반대쪽 패널의 cell_edit을 수동으로 갱신
            dst_panel = self.panel_a if dst is self.panel_a.table else self.panel_b
            dst_panel._sync_cell_edit()
        finally:
            self._syncing_selection = False

    def _count_changed(self) -> int:
        excl = self._excluded_cols
        return sum(
            1
            for r, row in enumerate(self._diff_matrix)
            for c, (st, *_) in enumerate(row)
            if st != "same" and c not in excl
        )

    def _set_save_btn_state(self, enabled: bool = True):
        # b_to_a staged → A 파일에 쓸 내용 / a_to_b staged → B 파일에 쓸 내용
        has_a = (any(v == "b_to_a" for v in self._staged.values())
                 or bool(self._edited.get("a")))
        has_b = (any(v == "a_to_b" for v in self._staged.values())
                 or bool(self._edited.get("b")))

        # JSON/uasset 등 비-xlsx 파일은 저장 미지원 → 버튼 강제 비활성화 + 툴팁 안내
        def _xlsx_ok(path: str) -> bool:
            return (not path) or os.path.splitext(path)[1].lower() in _EXCEL_EXTS
        path_a = self.panel_a.get_path()
        path_b = self.panel_b.get_path()
        a_writable = _xlsx_ok(path_a)
        b_writable = _xlsx_ok(path_b)

        self.panel_a.save_btn.setEnabled(enabled and has_a and a_writable)
        self.panel_b.save_btn.setEnabled(enabled and has_b and b_writable)
        self.panel_a.save_btn.setToolTip(
            "파일 저장" if a_writable
            else "JSON/uasset은 비교 전용 — 저장은 Excel(.xlsx/.xls/.xlsm)만 지원"
        )
        self.panel_b.save_btn.setToolTip(
            "파일 저장" if b_writable
            else "JSON/uasset은 비교 전용 — 저장은 Excel(.xlsx/.xls/.xlsm)만 지원"
        )

    def _on_file_loaded(self, side: str, path: str):
        either = bool(self.panel_a.get_path()) or bool(self.panel_b.get_path())
        self.refresh_btn.setEnabled(either)
        both = bool(self.panel_a.get_path()) and bool(self.panel_b.get_path())
        if both:
            self._run_compare()
        else:
            self._run_preview(side, path)

    def _set_buttons_enabled(self, enabled: bool):
        has_diff = enabled and bool(self._diff_matrix)
        self.diff_only_btn.setEnabled(has_diff)
        self.prev_diff_btn.setEnabled(has_diff)
        self.next_diff_btn.setEnabled(has_diff)
        self._set_find_enabled(has_diff)
        self._set_save_btn_state(enabled)

    def _set_find_enabled(self, enabled: bool):
        for w in (self.find_edit, self.find_case_btn, self.find_word_btn,
                  self.find_prev_btn, self.find_next_btn):
            w.setEnabled(enabled)


def _parse_args() -> tuple[str, str]:
    """
    P4V diff 호출 형식: -s <원본> -d <수정본>
    위치 인자 2개도 지원: <파일A> <파일B>
    반환: (path_a, path_b)  — 없으면 빈 문자열
    """
    args = sys.argv[1:]
    path_a = path_b = ""
    i = 0
    positional = []
    while i < len(args):
        if args[i] == "-s" and i + 1 < len(args):
            path_a = args[i + 1]; i += 2
        elif args[i] == "-d" and i + 1 < len(args):
            path_b = args[i + 1]; i += 2
        elif not args[i].startswith("-"):
            positional.append(args[i]); i += 1
        else:
            i += 1
    # -s/-d 없이 위치 인자로 넘어온 경우
    if not path_a and len(positional) >= 1:
        path_a = positional[0]
    if not path_b and len(positional) >= 2:
        path_b = positional[1]
    return path_a, path_b


def _set_windows_app_user_model_id():
    """Windows 작업 표시줄이 앱을 python.exe 그룹과 분리하고
    우리가 지정한 아이콘을 사용하도록 AppUserModelID를 등록한다."""
    if sys.platform != "win32":
        return
    try:
        import ctypes
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(
            "Netmarble.ExcelMerge.App")
    except Exception:
        pass


def main():
    _set_windows_app_user_model_id()
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    app.setWindowIcon(_load_app_icon())
    win = MainWindow()

    # P4V 등 외부 도구가 커맨드라인으로 파일을 넘긴 경우 자동 로드
    path_a, path_b = _parse_args()
    if path_a and os.path.isfile(path_a):
        win.panel_a.set_path(path_a)
    if path_b and os.path.isfile(path_b):
        win.panel_b.set_path(path_b)
    # 두 파일 모두 있으면 preview 없이 바로 비교 (load_sheet 2회로 단축)
    if path_a and path_b and os.path.isfile(path_a) and os.path.isfile(path_b):
        win._run_compare()
    elif path_a and os.path.isfile(path_a):
        win._run_preview("a", path_a)
    elif path_b and os.path.isfile(path_b):
        win._run_preview("b", path_b)

    win.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
