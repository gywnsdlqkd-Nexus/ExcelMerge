"""테이블/스크롤/입력 위젯 (excel_diff_merge.py에서 분리)."""
import os

from PyQt5.QtWidgets import (
    QApplication, QTableView, QAbstractItemView, QLineEdit, QPlainTextEdit,
    QHeaderView, QMenu, QStyle, QScrollBar, QStyleOptionSlider,
    QStyledItemDelegate, QStyleOptionViewItem,
    QTabBar, QStylePainter, QStyleOptionTab,
)
from PyQt5.QtCore import (
    Qt, pyqtSignal, QItemSelection, QItemSelectionRange, QItemSelectionModel,
    QRect, QPoint, QSize, QObject,
)
from PyQt5.QtGui import (
    QPainter, QPalette, QTextCursor, QTextCharFormat, QTextDocument, QTextOption,
    QIcon, QPixmap, QFont, QPen, QColor,
)
from PyQt5 import sip
from openpyxl.utils import get_column_letter

from .diff_model import DiffTableModel
from .diff_model import EXTRA_ROWS as EXTRA_ROWS  # smoke_test 재노출
from .loaders import _SUPPORTED_EXTS
from .constants import DIR_A2B, DIR_B2A
from .theme import (
    CELL_DIFF_HL, CELL_DIFF_FG, HATCH_COLOR, MENU_QSS, SHEET_TAB_CHANGED_BG,
    MINIMAP_MARKER_COLOR, ui_font, key_header_icon, exclude_header_icon,
    reset_header_icon, force_active_highlight,
)


def draw_diagonal_hatch(painter, rect, color=HATCH_COLOR, step=6, width=1):
    """rect 안에 등간격 대각선(↗)을 그린다. clip은 이 함수가 rect로 건다.
    신규 셀의 반대쪽 빈 칸에 매칭 표시(Beyond Compare식)로 사용."""
    painter.save()
    painter.setClipRect(rect)
    painter.setPen(QPen(color, width))
    h = rect.height()
    x = rect.left() - h
    while x < rect.right():
        painter.drawLine(x, rect.bottom(), x + h, rect.top())
        x += step
    painter.restore()


def make_find_icon(kind: str) -> QIcon:
    """찾기 버튼용 아이콘을 QPainter로 렌더링 (HiDPI 2배 해상도).
    체크 시 배경이 파란색으로 바뀌므로 Off=진회색 / On=흰색 두 벌을 등록.
    (DiffView·FolderCompareView 공용)"""
    def render(color: QColor) -> QPixmap:
        s = 32
        pm = QPixmap(s * 2, s * 2)
        pm.setDevicePixelRatio(2)
        pm.fill(Qt.transparent)
        p = QPainter(pm)
        p.setRenderHint(QPainter.Antialiasing)
        p.setRenderHint(QPainter.TextAntialiasing)
        if kind == "case":
            p.setPen(color)
            p.setFont(QFont("Segoe UI", 12, QFont.Bold))
            p.drawText(QRect(0, 0, s, s), Qt.AlignCenter, "Aa")
        elif kind == "word":
            p.setPen(color)
            p.setFont(QFont("Segoe UI", 10, QFont.Bold))
            p.drawText(QRect(0, 0, s, s - 8), Qt.AlignCenter, "ab")
            p.setPen(QPen(color, 1.8, Qt.SolidLine, Qt.RoundCap))
            y = s - 7
            p.drawLine(QPoint(7, y), QPoint(s - 7, y))
            p.drawLine(QPoint(7, y), QPoint(7, y - 4))
            p.drawLine(QPoint(s - 7, y), QPoint(s - 7, y - 4))
        else:  # "prev" / "next" — 셰브론 화살표
            p.setPen(QPen(color, 2.6, Qt.SolidLine, Qt.RoundCap, Qt.RoundJoin))
            m = s // 2
            d = -1 if kind == "prev" else 1
            p.drawLine(QPoint(m - 3 * d, m - 7), QPoint(m + 4 * d, m))
            p.drawLine(QPoint(m + 4 * d, m), QPoint(m - 3 * d, m + 7))
        p.end()
        return pm

    ic = QIcon()
    ic.addPixmap(render(QColor("#3b3b3b")), QIcon.Normal, QIcon.Off)
    ic.addPixmap(render(QColor("#ffffff")), QIcon.Normal, QIcon.On)
    ic.addPixmap(render(QColor("#b8b8b8")), QIcon.Disabled, QIcon.Off)
    ic.addPixmap(render(QColor("#b8b8b8")), QIcon.Disabled, QIcon.On)
    return ic


# 자동 컬럼 너비 상한 — 150px
# 데이터가 긴 셀 때문에 열이 화면을 가리지 않도록 제한.
# 사용자가 헤더 드래그로 직접 넓힌 열은 _user_col_widths 에 기록되어 이 상한 무시.
# 새로고침 시에는 _run_refresh()가 _user_col_widths를 비우므로 모든 열이 디폴트로 복귀.
MAX_AUTO_COL_WIDTH_PX = 150


class MinimapScrollBar(QScrollBar):
    """수직 또는 수평 스크롤바 위에 변경된 셀(행/열)의 위치를 색상 마커로 오버레이.
    paintEvent에서 super 호출 후, orientation에 맞춰 트랙(groove) 영역에
    비율 위치(0.0~1.0)별로 가는 막대를 그린다."""
    _MARKER_COLOR = MINIMAP_MARKER_COLOR

    def __init__(self, orientation, parent=None):
        super().__init__(orientation, parent)
        self._ratios: list = []   # 0.0~1.0 사이 변경 위치 목록

    def set_change_ratios(self, ratios):
        # 변경된 경우에만 repaint (불필요한 페인트 방지)
        if list(ratios) != self._ratios:
            self._ratios = list(ratios)
            self.update()

    def paintEvent(self, e):
        super().paintEvent(e)
        if not self._ratios:
            return
        # QStyle을 통해 정확한 trough(groove) 영역을 얻는다
        opt = QStyleOptionSlider()
        self.initStyleOption(opt)
        groove = self.style().subControlRect(
            QStyle.CC_ScrollBar, opt, QStyle.SC_ScrollBarGroove, self)
        if groove.width() <= 0 or groove.height() <= 0:
            return
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing, False)
        painter.setPen(Qt.NoPen)
        painter.setBrush(self._MARKER_COLOR)
        if self.orientation() == Qt.Vertical:
            track_top = groove.top()
            track_h = groove.height()
            x = groove.left() + 2
            w = max(1, groove.width() - 4)
            denom = max(0, track_h - 2)
            for ratio in self._ratios:
                y = track_top + int(ratio * denom)
                painter.fillRect(x, y, w, 2, self._MARKER_COLOR)
        else:
            track_left = groove.left()
            track_w = groove.width()
            y = groove.top() + 2
            h = max(1, groove.height() - 4)
            denom = max(0, track_w - 2)
            for ratio in self._ratios:
                x = track_left + int(ratio * denom)
                painter.fillRect(x, y, 2, h, self._MARKER_COLOR)
        painter.end()


class DiffHighlightDelegate(QStyledItemDelegate):
    """modified 셀에서 A/B가 다른 문자 구간을 CELL_DIFF_HL(핑크)로 강조 렌더한다.
    강조 구간이 없는 셀(대다수)은 기본 델리게이트 렌더 그대로 — 빠른 경로.

    텍스트를 QTextDocument로 그려(셀값란과 동일 방식) 여러 줄('\\n')·문자 인덱스·
    수직 정렬을 그대로 반영한다. 단일 라인 가정으로 위치를 계산하면 멀티라인 셀에서
    강조가 엉뚱한 곳에 찍히므로, 문자 단위 배경은 QTextCharFormat으로 지정한다."""

    def paint(self, painter, option, index):
        model = index.model()
        r, c = index.row(), index.column()
        ranges = (model.diff_char_ranges(r, c)
                  if hasattr(model, "diff_char_ranges") else [])
        # 강조 구간(빨강)이 없으면 기본 델리게이트 렌더 — 빠른 경로.
        # (수식 셀의 파랑 폰트는 모델 ForegroundRole로 처리되어 QTextDocument가 불필요하다.)
        if not ranges:
            # 신규(added) 셀의 빈 쪽 → 대각선 해치로 매칭 표시.
            if (hasattr(model, "is_added_placeholder")
                    and model.is_added_placeholder(r, c)):
                opt = QStyleOptionViewItem(option)
                self.initStyleOption(opt, index)
                widget = opt.widget
                style = widget.style() if widget is not None else QApplication.style()
                opt.text = ""
                style.drawControl(QStyle.CE_ItemViewItem, opt, painter, widget)
                draw_diagonal_hatch(painter, opt.rect)
                return
            super().paint(painter, option, index)
            return

        opt = QStyleOptionViewItem(option)
        self.initStyleOption(opt, index)
        widget = opt.widget
        style = widget.style() if widget is not None else QApplication.style()

        text = opt.text
        # 1) 텍스트 없이 배경/선택/프레임만 그린다 (기본 델리게이트와 동일한 배경/선택색).
        opt.text = ""
        style.drawControl(QStyle.CE_ItemViewItem, opt, painter, widget)

        text_rect = style.subElementRect(QStyle.SE_ItemViewItemText, opt, widget)

        # 2) QTextDocument로 텍스트 구성 — '\n'은 줄바꿈, 소프트랩 없음(뷰와 동일).
        doc = QTextDocument()
        doc.setDocumentMargin(0)
        doc.setDefaultFont(opt.font)
        to = QTextOption(opt.displayAlignment)
        to.setWrapMode(QTextOption.NoWrap)
        doc.setDefaultTextOption(to)
        doc.setPlainText(text)

        # 기본 전경색 (선택 시 흰색, 그 외 ForegroundRole/기본).
        # ForegroundRole은 수식 결과 셀에 파랑을 돌려준다. 선택 셀은 흰색 우선(파란 선택 배경
        # 위 가독), 변경 구간(빨강)은 아래에서 덮어쓴다.
        if opt.state & QStyle.State_Selected:
            fg = opt.palette.color(QPalette.Active, QPalette.HighlightedText)
        else:
            fg = index.data(Qt.ForegroundRole) or opt.palette.color(
                QPalette.Active, QPalette.Text)
        base_fmt = QTextCharFormat()
        base_fmt.setForeground(fg)
        cur = QTextCursor(doc)
        cur.select(QTextCursor.Document)
        cur.mergeCharFormat(base_fmt)

        # 3) 다른 문자 구간에 핑크 배경 + 빨강 폰트. QTextDocument 위치는 '\n'을 1칸으로
        #    세어 flat 인덱스와 1:1 정렬되므로 diff_char_ranges 값을 그대로 쓴다.
        #    빨강 전경은 base_fmt(선택 시 흰색)를 덮어써 선택 셀에서도 강조가 보이게 한다.
        hl_fmt = QTextCharFormat()
        hl_fmt.setBackground(CELL_DIFF_HL)
        hl_fmt.setForeground(CELL_DIFF_FG)
        n = len(text)
        for start, end in ranges:
            start = max(0, start)
            end = min(n, end)
            if start >= end:
                continue
            c = QTextCursor(doc)
            c.setPosition(start)
            c.setPosition(end, QTextCursor.KeepAnchor)
            c.mergeCharFormat(hl_fmt)

        # 4) 텍스트 영역으로 clip + 수직 중앙 정렬해 그린다.
        painter.save()
        painter.setClipRect(text_rect)
        doc_h = doc.size().height()
        y = text_rect.top() + max(0, (text_rect.height() - doc_h) / 2)
        painter.translate(text_rect.left(), y)
        doc.drawContents(painter)
        painter.restore()


class _FrozenView(QTableView):
    """틀 고정 헬퍼 뷰 — 본체 모델/선택모델을 공유한다. 자체 스크롤바 없음.
    휠 이벤트는 본체로 전달해 본체가 스크롤되고 컨트롤러가 헬퍼를 되동기하게 한다.
    선택 모델을 본체와 공유하므로(FreezeController._make_view) 고정 셀(키 열/행)을
    이 뷰에서 클릭하면 본체 선택이 갱신돼 셀값란·A/B 미러 등 기존 로직이 그대로 동작한다."""
    def __init__(self, host):
        super().__init__(host)
        self._host = host

    def wheelEvent(self, event):
        self._host.wheelEvent(event)

    def scrollTo(self, index, hint=QAbstractItemView.EnsureVisible):
        """스크롤 위치는 전적으로 FreezeController가 제어한다. 선택/현재 셀 변경 시
        Qt가 자동 호출하는 scrollTo가 고정 뷰를 제 위치에서 밀어내지 않도록 무력화."""
        return


class FreezeController(QObject):
    """엑셀 '틀 고정'(정식) — 본체(ExcelTableView)를 우하단 사분면으로 쓰고, 예약된 상단/좌측
    여백에 헬퍼 뷰 3개를 배치한다. 본체가 고정 행/열을 setRowHidden/setColumnHidden으로 숨겨
    스크롤 영역만 렌더하므로 '가림'이 없고, corner 헤더가 고정 눈금('2'/'B')을 정확히 표시한다.

    좌표(본체 기준, fr=frameWidth, hdrW=세로헤더폭, hdrH=가로헤더높이, fw=고정열폭합, fh=고정행높이합):
      corner:(fr, fr, hdrW+fw, hdrH+fh)          헤더 표시 → 고정 눈금 + 고정 코너 셀
      top   :(fr+hdrW+fw, fr+hdrH, 본체폭, fh)    헤더 숨김, 가로 동기 → 고정 행 데이터
      left  :(fr+hdrW, fr+hdrH+fh, fw, 본체높이)  헤더 숨김, 세로 동기 → 고정 열 데이터
    본체 여백은 ExcelTableView.updateGeometries() 오버라이드가 reposition()과 함께 적용한다.
    """
    _MIN_BODY_W = 80    # 고정 영역이 본체를 다 먹으면 freeze 중단(잔여 본체 최소치)
    _MIN_BODY_H = 44

    def __init__(self, host):
        super().__init__(host)
        self.host = host
        host._freeze = self
        self._n_rows = 1          # 고정 행 수 = key_row + 1
        self._n_cols = 1          # 고정 열 수 = key_col + 1 (key_col<0 → 0)
        self._fw = 0              # 고정 열 폭합(refresh에서 캡처 — 본체 숨김과 무관하게 안정)
        self._fh = 0              # 고정 행 높이합
        self._active = False
        self.corner = self._make_view(headers=True)
        self.top = self._make_view(headers=False)
        self.left = self._make_view(headers=False)
        self._views = (self.corner, self.top, self.left)
        # 본체를 열/행 단위 스크롤로 고정 → 스크롤바 값이 곧 열/행 인덱스라 헬퍼와 값 동기로 정렬된다.
        host.setHorizontalScrollMode(QAbstractItemView.ScrollPerItem)
        host.setVerticalScrollMode(QAbstractItemView.ScrollPerItem)
        host.horizontalScrollBar().valueChanged.connect(self._on_h_scroll)
        host.verticalScrollBar().valueChanged.connect(self._on_v_scroll)
        host.horizontalHeader().sectionResized.connect(self._on_col_resized)
        host.verticalHeader().sectionResized.connect(self._on_row_resized)
        # corner(고정 눈금) 헤더 우클릭 → 키 열/행 설정·해제 메뉴.
        # ★ 반드시 popup()(비모달)로 띄운다. 자식 오버레이 헤더 이벤트 처리 중 모달 메뉴(exec_ 중첩
        #   이벤트루프)를 쓰면 Qt 상태가 깨져 access violation(Windows fatal exception)이 난다.
        #   본체 헤더는 exec_ 그대로(문제 없음). corner는 최소 키 메뉴만 popup으로 제공.
        ch = self.corner.horizontalHeader()
        cv = self.corner.verticalHeader()
        ch.setContextMenuPolicy(Qt.CustomContextMenu)
        cv.setContextMenuPolicy(Qt.CustomContextMenu)
        ch.customContextMenuRequested.connect(self._corner_col_menu)
        cv.customContextMenuRequested.connect(self._corner_row_menu)

    def _corner_col_menu(self, pos):
        """고정 열 헤더(corner) 우클릭 — 병합 준비/취소 + 키 열 설정/해제(비모달 popup).
        ★ 반드시 popup(비모달). exec_(모달)은 자식 오버레이 이벤트 중 access violation 유발."""
        if not self._alive():
            return
        host = self.host
        ch = self.corner.horizontalHeader()
        col = ch.logicalIndexAt(pos)
        if col < 0:
            return
        m = QMenu(host)
        m.setStyleSheet(MENU_QSS)
        m.setAttribute(Qt.WA_DeleteOnClose)

        # 병합 준비/취소 — 본체 헤더 메뉴와 동일(대상 열 = 선택 반영). 키 열도 변경/스테이징
        # 셀이 있으면 노출(예: 신규 행의 키 값). 없으면 항목 자체가 안 뜬다.
        target_cols = host._selected_header_cols(col)
        cols_label = ", ".join(get_column_letter(c + 1) for c in target_cols)
        has_changed = any(host._col_has_changed(c) for c in target_cols)
        has_staged = host._cols_have_staged(target_cols)
        if has_changed:
            m.addAction(f"A → B  병합 준비  [{cols_label}열]").triggered.connect(
                lambda _=False, cs=target_cols:
                (host._select_cols(cs), host.stage_requested.emit(DIR_A2B)))
            m.addAction(f"B → A  병합 준비  [{cols_label}열]").triggered.connect(
                lambda _=False, cs=target_cols:
                (host._select_cols(cs), host.stage_requested.emit(DIR_B2A)))
        if has_staged:
            if has_changed:
                m.addSeparator()
            m.addAction(f"병합 준비 취소  [{cols_label}열]").triggered.connect(
                lambda _=False, cs=target_cols:
                (host._select_cols(cs), host.unstage_requested.emit()))
        if has_changed or has_staged:
            m.addSeparator()

        if col == host._key_col:
            m.addAction("🔓  키 열 해제 (ROW 순서 기반 비교)").triggered.connect(
                lambda: host.key_col_changed.emit(-1))
        else:
            letter = get_column_letter(col + 1)
            m.addAction(key_header_icon(), f"키 열로 설정  [{letter}열]").triggered.connect(
                lambda: host.key_col_changed.emit(col))
        m.popup(ch.mapToGlobal(pos))

    def _corner_row_menu(self, pos):
        """고정 행 헤더(corner) 우클릭 — 병합 준비/취소 + 키 행 설정(비모달 popup).
        ★ 반드시 popup(비모달). exec_(모달)은 자식 오버레이 이벤트 중 access violation 유발."""
        if not self._alive():
            return
        host = self.host
        cv = self.corner.verticalHeader()
        row = cv.logicalIndexAt(pos)
        if row < 0:
            return
        orig = host.model().orig_row(row)   # display→원본 파일 행
        if orig is None:
            return
        m = QMenu(host)
        m.setStyleSheet(MENU_QSS)
        m.setAttribute(Qt.WA_DeleteOnClose)

        # 병합 준비/취소 — 본체 행 헤더 메뉴와 동일(대상 행 = 선택 반영). 키 행에서도 노출.
        target_rows = host._selected_header_rows(row)
        suffix = f"  [{len(target_rows)}개 행]" if len(target_rows) > 1 else ""
        has_changed = any(host._row_has_changed(r) for r in target_rows)
        has_staged = host._rows_have_staged(target_rows)
        if has_changed:
            m.addAction(f"A → B  병합 준비{suffix}").triggered.connect(
                lambda _=False, rs=target_rows:
                (host._select_rows(rs), host.stage_requested.emit(DIR_A2B)))
            m.addAction(f"B → A  병합 준비{suffix}").triggered.connect(
                lambda _=False, rs=target_rows:
                (host._select_rows(rs), host.stage_requested.emit(DIR_B2A)))
        if has_staged:
            if has_changed:
                m.addSeparator()
            m.addAction(f"병합 준비 취소{suffix}").triggered.connect(
                lambda _=False, rs=target_rows:
                (host._select_rows(rs), host.unstage_requested.emit()))

        # 키 행 지정 — 현재 키 행이 아닌 단일 행에서만(초기화 기능은 제거됨).
        if orig != host._key_row and len(target_rows) == 1:
            if has_changed or has_staged:
                m.addSeparator()
            m.addAction(key_header_icon(), f"키 행으로 설정  [{orig + 1}행]").triggered.connect(
                lambda _=False, o=orig: host.key_row_changed.emit(o))

        if m.isEmpty():
            return   # 키 행인데 병합할 것도 없음 → 메뉴 미표시
        m.popup(cv.mapToGlobal(pos))

    def _make_view(self, headers: bool):
        host = self.host
        v = _FrozenView(host)
        v.setModel(host.model())
        # 선택 모델을 본체와 공유 — 고정 셀(키 열/행)을 이 뷰에서 클릭하면 본체 선택이 갱신되고
        # selectionChanged가 발화해 셀값란·A/B 미러·수식 플래그 등 기존 배선이 그대로 동작한다.
        v.setSelectionModel(host.selectionModel())
        v.setItemDelegate(DiffHighlightDelegate(v))
        v.setFocusPolicy(Qt.NoFocus)
        v.setSelectionBehavior(QAbstractItemView.SelectItems)
        v.setSelectionMode(QAbstractItemView.ExtendedSelection)
        # 선택색 통일(비포커스에도 파랑) — 본체와 동일 팔레트.
        force_active_highlight(v)
        v.setEditTriggers(QAbstractItemView.NoEditTriggers)
        v.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        v.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        v.setHorizontalScrollMode(QAbstractItemView.ScrollPerItem)
        v.setVerticalScrollMode(QAbstractItemView.ScrollPerItem)
        v.setAlternatingRowColors(False)
        v.setFrameShape(QTableView.NoFrame)
        v.setStyleSheet("QTableView { border: none; }")
        v.setFont(ui_font(9))
        v.verticalHeader().setDefaultSectionSize(
            host.verticalHeader().defaultSectionSize())
        v.horizontalHeader().setVisible(headers)
        v.verticalHeader().setVisible(headers)
        if headers:   # corner — 고정 눈금(키 아이콘 포함)
            v.horizontalHeader().setIconSize(QSize(14, 14))
            v.verticalHeader().setIconSize(QSize(14, 14))
        v.hide()
        return v

    def _alive(self) -> bool:
        """teardown 중(헬퍼 C++ 객체 삭제됨) 시그널이 도착해도 크래시하지 않도록 가드."""
        if sip.isdeleted(self) or sip.isdeleted(self.host):
            return False
        return not any(sip.isdeleted(v) for v in self._views)

    @property
    def active(self) -> bool:
        return self._active

    @staticmethod
    def frozen_span(key_row: int, key_col: int, rows: int, cols: int) -> tuple:
        """앵커(key_row/key_col)와 데이터 크기로 고정 (행수, 열수) 계산.
        key_col<0(ROW 순서)면 열 고정 없음(0)."""
        n_rows = min((key_row if key_row and key_row > 0 else 0) + 1, max(rows, 0))
        n_cols = 0 if key_col is None or key_col < 0 else min(key_col + 1, max(cols, 0))
        return n_rows, n_cols

    def frozen_px(self) -> tuple:
        # 본체가 고정 행/열을 숨기면 폭·높이가 0이 되므로, refresh에서 캡처한 값을 쓴다.
        return self._fw, self._fh

    # ── 외부 API ─────────────────────────────────────────────────────────────
    def clear(self):
        """freeze 비활성화 — 헬퍼 숨김 + 본체 열숨김/여백 원복(행 숨김은 필터가 소유)."""
        if not self._alive():
            return
        self._active = False
        for v in self._views:
            v.hide()
        host = self.host
        for c in range(host.model().columnCount()):
            if host.isColumnHidden(c):
                host.setColumnHidden(c, False)
        host.updateGeometries()   # super()가 기본 여백 복원

    def refresh(self):
        if not self._alive():
            return
        host = self.host
        model = host.model()
        rows = getattr(model, "data_rows", 0)
        if not rows or getattr(model, "_mode", None) != "diff":
            self.clear()
            return
        self._n_rows, self._n_cols = self.frozen_span(
            getattr(host, "_key_row", 0), getattr(host, "_key_col", 0),
            rows, getattr(model, "data_cols", 0))
        self._active = True
        ncols = model.columnCount()
        # 벌크 뮤테이션 구간: _applying_sizes로 freeze 핸들러/오버라이드의 per-op 무거운 작업을 억제하고
        # (40k행 O(R) 미러가 폭풍이 되지 않도록) 화면 갱신도 잠시 끈다. 끝난 뒤 1회만 full reposition.
        prev_host = host.updatesEnabled()
        prev_v = [v.updatesEnabled() for v in self._views]
        prev_flag = getattr(host, "_applying_sizes", False)
        host._applying_sizes = True
        host.setUpdatesEnabled(False)
        for v in self._views:
            v.setUpdatesEnabled(False)
        try:
            # 1) 본체 고정 열을 잠시 해제해 실제 폭을 캡처(숨김 열은 폭 0이라 캡처 불가).
            for c in range(ncols):
                if host.isColumnHidden(c):
                    host.setColumnHidden(c, False)
            colw = [host.columnWidth(c) for c in range(ncols)]
            self._fw = sum(colw[c] for c in range(self._n_cols))
            dh = host.verticalHeader().defaultSectionSize()
            urh = getattr(host, "_user_row_heights", {})
            self._fh = sum(urh.get(r, dh) for r in range(self._n_rows))
            # 2) 본체/헬퍼 열 숨김.
            self._apply_col_hidden()
            # 3) 헬퍼 열 폭/행 높이를 '숨김 이후' 캡처값으로 명시 적용 — setColumnHidden 부작용이
            #    폭을 0으로 만드는 것을 되돌린다(마지막에 적용해야 덮어쓰이지 않음).
            for c in range(ncols):
                for v in self._views:
                    if v.columnWidth(c) != colw[c]:
                        v.setColumnWidth(c, colw[c])
            for r, h in urh.items():
                for v in self._views:
                    if v.rowHeight(r) != h:
                        v.setRowHeight(r, h)
            self._mirror_hidden_rows()
        finally:
            host._applying_sizes = prev_flag
            host.setUpdatesEnabled(prev_host)
            for v, pv in zip(self._views, prev_v):
                v.setUpdatesEnabled(pv)
        host.updateGeometries()   # 벌크 종료(_applying_sizes 복원) 후 1회 여백 재적용 + reposition()
        self._sync_scroll()

    # ── 숨김/크기/스크롤 미러 ────────────────────────────────────────────────
    def _apply_col_hidden(self):
        """본체·헬퍼의 열 숨김: 본체/top은 고정 열 숨김(스크롤 열만), left/corner는 스크롤 열 숨김."""
        host = self.host
        for c in range(host.model().columnCount()):
            frozen = c < self._n_cols
            if host.isColumnHidden(c) != frozen:
                host.setColumnHidden(c, frozen)
            if self.top.isColumnHidden(c) != frozen:
                self.top.setColumnHidden(c, frozen)
            if self.left.isColumnHidden(c) != (not frozen):
                self.left.setColumnHidden(c, not frozen)
            if self.corner.isColumnHidden(c) != (not frozen):
                self.corner.setColumnHidden(c, not frozen)

    def _sync_sizes(self):
        host = self.host
        for c in range(host.model().columnCount()):
            if host.isColumnHidden(c):
                continue   # 숨겨진(고정) 열은 폭 0 → 복사하면 헬퍼의 고정 열 폭이 사라짐
            w = host.columnWidth(c)
            for v in self._views:
                if v.columnWidth(c) != w:
                    v.setColumnWidth(c, w)
        for r, h in getattr(host, "_user_row_heights", {}).items():
            for v in self._views:
                if v.rowHeight(r) != h:
                    v.setRowHeight(r, h)

    def _mirror_hidden_rows(self):
        # left만 스크롤 행 정렬 필요(본체와 동일 숨김). top/corner는 고정 행만 노출(높이 클립).
        host = self.host
        left = self.left
        for r in range(host.model().rowCount()):
            hidden = host.isRowHidden(r)
            if left.isRowHidden(r) != hidden:
                left.setRowHidden(r, hidden)

    def _sync_scroll(self):
        if not self._alive():
            return
        self._sync_top_h()
        self.top.verticalScrollBar().setValue(0)
        self.left.verticalScrollBar().setValue(self.host.verticalScrollBar().value())
        self.left.horizontalScrollBar().setValue(0)
        self.corner.horizontalScrollBar().setValue(0)
        self.corner.verticalScrollBar().setValue(0)

    def _sync_top_h(self):
        # 본체·헬퍼 모두 ScrollPerItem이라 스크롤바 값 = 보이는 열 인덱스 → 값 동기로 정렬된다.
        self.top.horizontalScrollBar().setValue(self.host.horizontalScrollBar().value())

    def reposition(self):
        """헬퍼 3개의 위치/크기를 본체 헤더·고정 크기 기준으로 재계산. updateGeometries에서 호출."""
        if not self._alive() or not self._active:
            return
        host = self.host
        fr = host.frameWidth()
        hdrW = host.verticalHeader().width()
        hdrH = host.horizontalHeader().height()
        fw, fh = self.frozen_px()
        W, H = host.width(), host.height()
        body_w = W - 2 * fr - hdrW - fw
        body_h = H - 2 * fr - hdrH - fh
        # 고정 영역이 본체를 다 먹으면 freeze 표시 중단(스크롤 영역 확보 불가)
        if body_w < self._MIN_BODY_W or body_h < self._MIN_BODY_H:
            for v in self._views:
                v.hide()
            return
        # 본체는 스크롤바가 뜨면 그만큼 셀 렌더 영역이 줄어드는데, 고정 헬퍼 뷰는 스크롤바가
        # 항상 꺼져 있어(ScrollBarAlwaysOff) 그만큼 더 많은 행/열을 그린다. 그러면 ScrollPerItem
        # 최댓값이 서로 달라져 스크롤 끝(하단/우단)에서 고정 열/행이 본체와 한 칸 어긋난다.
        # → 헬퍼 뷰 크기에서 본체 스크롤바 두께를 빼 '보이는 행/열 수'를 본체와 맞춘다.
        hbar = host.horizontalScrollBar()
        vbar = host.verticalScrollBar()
        hbar_h = hbar.height() if hbar.isVisible() else 0
        vbar_w = vbar.width() if vbar.isVisible() else 0
        self.corner.verticalHeader().setFixedWidth(hdrW)
        self.corner.horizontalHeader().setFixedHeight(hdrH)
        self.corner.setGeometry(fr, fr, hdrW + fw, hdrH + fh)
        self.corner.setVisible(True)
        self.top.setGeometry(fr + hdrW + fw, fr + hdrH, max(0, body_w - vbar_w), fh)
        self.top.setVisible(True)
        if self._n_cols > 0:
            self.left.setGeometry(fr + hdrW, fr + hdrH + fh, fw, max(0, body_h - hbar_h))
            self.left.setVisible(True)
        else:
            self.left.hide()
        self.corner.raise_()
        self.top.raise_()
        self.left.raise_()
        self._sync_scroll()

    # ── 시그널 핸들러 (teardown 가드 필수) ───────────────────────────────────
    def _on_h_scroll(self, value):
        if self._alive() and self._active:
            self._sync_top_h()

    def _on_v_scroll(self, value):
        if self._alive() and self._active:
            self.left.verticalScrollBar().setValue(value)

    def _on_col_resized(self, idx, old, new):
        # 벌크 작업(_applying_sizes: 필터의 setRowHidden 폭풍, 크기 일괄 적용) 중엔 발화 무시 —
        # 그러지 않으면 숨김 해제되는 행마다 sectionResized→여기서 O(C)+reposition이 돌아 O(R×C) 폭주.
        if not self._alive() or not self._active or getattr(self.host, "_applying_sizes", False):
            return
        self._sync_sizes()
        self.host.updateGeometries()

    def _on_row_resized(self, idx, old, new):
        if new <= 0 or not self._alive() or not self._active \
                or getattr(self.host, "_applying_sizes", False):
            return
        self._sync_sizes()
        self.host.updateGeometries()


class ExcelTableView(QTableView):
    stage_requested   = pyqtSignal(str)   # direction: 'a_to_b' | 'b_to_a'
    unstage_requested = pyqtSignal()
    key_col_changed   = pyqtSignal(int)   # 키 열 변경 요청
    key_row_changed   = pyqtSignal(int)   # 키 행(헤더 행) 변경 요청
    columns_exclude_set = pyqtSignal(list, bool)   # (cols, exclude) — True: 제외 추가, False: 제외 해제
    column_resized    = pyqtSignal(int, int)   # (col, new_width) — 사용자 조작에 의한 변경만
    row_resized       = pyqtSignal(int, int)   # (row, new_height) — 사용자 조작에 의한 변경만

    def __init__(self, side: str, parent=None):
        super().__init__(parent)
        # 틀 고정 — updateGeometries() 오버라이드가 setModel 등에서 조기 호출될 수 있어 먼저 초기화.
        self._freeze = None            # FreezeController(생성 시 자기 자신을 여기 설정)
        self._in_update_geoms = False  # setViewportMargins 재진입 가드
        self.side = side
        self._model = DiffTableModel(side, self)
        self.setModel(self._model)   # selectionModel은 여기서 1회 생성 — 이후 교체 없음
        # A/B 값이 다른 문자 구간을 셀 안에서 핑크로 강조 (modified 셀 한정)
        self.setItemDelegate(DiffHighlightDelegate(self))
        self._populating = False
        self._key_col: int = 0
        self._key_row: int = 0
        self._excluded_cols: set[int] = set()   # 변경 검사 제외 열 (display 인덱스)
        # 사용자가 직접 조정한 열/행 크기 — 세션 동안만 유지 (재로드/저장/새로고침 후 복원)
        self._user_col_widths: dict[int, int] = {}
        self._user_row_heights: dict[int, int] = {}
        # 외부(다른 패널)에서 크기를 강제 적용 중일 때 sectionResized 재방출 방지
        self._applying_sizes: bool = False
        # 헤더 다중 선택의 anchor (Shift+방향 확장의 고정점)
        self._header_anchor_col: int | None = None
        self._header_anchor_row: int | None = None
        self.setFont(ui_font(9))
        self.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        # 키/제외 열 헤더의 PNG 아이콘(DecorationRole) 가시성 확보용 크기.
        self.horizontalHeader().setIconSize(QSize(14, 14))
        self.verticalHeader().setDefaultSectionSize(22)
        # 키 행 헤더의 PNG 아이콘(DecorationRole) 가시성 확보용 크기(가로와 동일).
        self.verticalHeader().setIconSize(QSize(14, 14))
        self.setAlternatingRowColors(False)
        self.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.setSelectionMode(QAbstractItemView.ExtendedSelection)
        # 선택색 통일: 포커스가 없어도(버튼/단축키로 이동 시) 비활성 하이라이트가
        # 회색으로 흐려지지 않고 활성(클릭 선택)과 같은 파랑으로 보이게 한다.
        force_active_highlight(self)
        self.setContextMenuPolicy(Qt.CustomContextMenu)
        self.customContextMenuRequested.connect(self._show_context_menu)
        self.horizontalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        self.horizontalHeader().customContextMenuRequested.connect(
            self._show_header_context_menu)
        self.verticalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        self.verticalHeader().customContextMenuRequested.connect(
            self._show_row_header_context_menu)
        # 헤더 크기 변경 추적 — 사용자 조작 시에만 저장/시그널 발행
        self.horizontalHeader().sectionResized.connect(self._on_section_h_resized)
        self.verticalHeader().sectionResized.connect(self._on_section_v_resized)
        # 헤더 클릭 시 anchor 갱신 (Shift 없는 클릭 → 새 anchor / Shift 클릭 → 기존 유지)
        self.horizontalHeader().sectionPressed.connect(self._on_h_section_pressed)
        self.verticalHeader().sectionPressed.connect(self._on_v_section_pressed)

    # ── QTableWidget 호환 헬퍼 ───────────────────────────────────────────────
    def rowCount(self) -> int:
        return self._model.rowCount()

    def columnCount(self) -> int:
        return self._model.columnCount()

    def _current_cell(self) -> tuple:
        """QTableWidget.currentRow()/currentColumn() 대응 — 무효 시 (-1, -1)."""
        idx = self.currentIndex()
        return (idx.row(), idx.column()) if idx.isValid() else (-1, -1)

    def _set_current_cell(self, r: int, c: int):
        """QTableWidget.setCurrentCell() 대응.
        setCurrentIndex()는 호출 시점의 키보드 수정자에 따라 선택을 확장/클리어하는
        기존 setCurrentCell과 완전히 같은 경로(selectionCommand)를 탄다."""
        if 0 <= r < self.rowCount() and 0 <= c < self.columnCount():
            self.setCurrentIndex(self._model.index(r, c))

    def _set_current_cell_no_update(self, r: int, c: int):
        """선택을 건드리지 않고 currentIndex만 이동.
        헤더/범위 선택 직후에 사용 — setCurrentIndex()는 눌려 있는 Shift를 보고
        SelectCurrent(앵커 사각형으로 선택 대체)를 적용해 방금 만든 범위
        선택을 붕괴시키므로, NoUpdate로 현재 셀만 옮긴다."""
        sm = self.selectionModel()
        if sm is not None and 0 <= r < self.rowCount() and 0 <= c < self.columnCount():
            sm.setCurrentIndex(self._model.index(r, c), QItemSelectionModel.NoUpdate)

    def _move_current_cell(self, r: int, c: int):
        """현재 셀을 단일 선택으로 이동 — Excel의 Ctrl+점프처럼 기존 선택을 비운다.
        setCurrentIndex()는 Ctrl이 눌린 상태에서 Toggle로 동작해 원래 셀 선택이
        남으므로 ClearAndSelect를 명시한다."""
        sm = self.selectionModel()
        if sm is not None and 0 <= r < self.rowCount() and 0 <= c < self.columnCount():
            sm.setCurrentIndex(self._model.index(r, c),
                               QItemSelectionModel.ClearAndSelect)

    # ── 사용자 헤더 크기 추적 ────────────────────────────────────────────────
    def _on_section_h_resized(self, logical_index: int, _old: int, new_size: int):
        # populate 중이거나 다른 패널에서 강제 적용 중인 변경은 무시
        if self._populating or self._applying_sizes:
            return
        self._user_col_widths[logical_index] = new_size
        self.column_resized.emit(logical_index, new_size)

    def _on_section_v_resized(self, logical_index: int, _old: int, new_size: int):
        if self._populating or self._applying_sizes:
            return
        self._user_row_heights[logical_index] = new_size
        self.row_resized.emit(logical_index, new_size)

    def _on_h_section_pressed(self, logical_index: int):
        """열 헤더 클릭: Shift/Ctrl 없으면 anchor 갱신, 동반이면 유지."""
        mods = QApplication.keyboardModifiers()
        if not (mods & (Qt.ShiftModifier | Qt.ControlModifier)):
            self._header_anchor_col = logical_index
        elif self._header_anchor_col is None:
            # Shift+클릭인데 anchor가 없으면 현재 클릭 지점을 anchor로
            self._header_anchor_col = logical_index
        # 행 anchor는 무관 — 열 헤더 클릭은 행 헤더 모드를 종료시킴
        self._header_anchor_row = None

    def _on_v_section_pressed(self, logical_index: int):
        mods = QApplication.keyboardModifiers()
        if not (mods & (Qt.ShiftModifier | Qt.ControlModifier)):
            self._header_anchor_row = logical_index
        elif self._header_anchor_row is None:
            self._header_anchor_row = logical_index
        self._header_anchor_col = None

    def apply_column_width(self, col: int, width: int):
        """반대 패널에서의 열 너비 변경을 동기 적용 (시그널 재방출 안 함)."""
        self._user_col_widths[col] = width
        if 0 <= col < self.columnCount() and self.columnWidth(col) != width:
            self._applying_sizes = True
            try:
                self.setColumnWidth(col, width)
            finally:
                self._applying_sizes = False

    def apply_row_height(self, row: int, height: int):
        """반대 패널에서의 행 높이 변경을 동기 적용."""
        self._user_row_heights[row] = height
        if 0 <= row < self.rowCount() and self.rowHeight(row) != height:
            self._applying_sizes = True
            try:
                self.setRowHeight(row, height)
            finally:
                self._applying_sizes = False

    def _apply_user_sizes(self):
        """저장된 사용자 크기를 현재 테이블에 다시 적용 (populate 후 호출)."""
        # 0-크기 값은 hidden 행/열에 대한 sectionResized 시그널이 남긴 오염값일 수
        # 있으므로 무시한다. UI상 0으로 만드는 사용자 조작은 없다.
        self._applying_sizes = True
        try:
            for col, w in self._user_col_widths.items():
                if 0 <= col < self.columnCount() and w > 0:
                    self.setColumnWidth(col, w)
            for row, h in self._user_row_heights.items():
                if 0 <= row < self.rowCount() and h > 0:
                    self.setRowHeight(row, h)
        finally:
            self._applying_sizes = False

    def _auto_size_columns(self, max_samples: int = 200):
        """resizeColumnsToContents() 대체 — 셀 텍스트를 샘플링해 QFontMetrics로 측정.
        모든 자동 폭이 MAX_AUTO_COL_WIDTH_PX로 클립되므로 상한 도달 시 조기 종료해
        전체 셀 스캔을 피한다. sectionResized가 사용자 변경으로 기록되지 않도록
        _applying_sizes 플래그로 차단."""
        m = self._model
        rows = m.data_rows
        cols = self.columnCount()
        if cols <= 0:
            return
        fm = self.fontMetrics()
        adv = fm.horizontalAdvance
        pad = 14   # 셀 좌우 여백+그리드 근사
        if rows <= max_samples:
            sample_rows = range(rows)
        else:
            stride = max(1, rows // (max_samples - 100))
            sample_rows = list(range(100)) + list(range(100, rows, stride))
        hdr_fm = self.horizontalHeader().fontMetrics()
        self._applying_sizes = True
        try:
            for c in range(cols):
                w = hdr_fm.horizontalAdvance(get_column_letter(c + 1)) + 24
                for r in sample_rows:
                    text = m.display_text(r, c)
                    if not text:
                        continue
                    if "\n" in text:
                        tw = max(adv(line) for line in text.split("\n"))
                    else:
                        tw = adv(text)
                    if tw + pad > w:
                        w = tw + pad
                        if w >= MAX_AUTO_COL_WIDTH_PX:
                            break
                self.setColumnWidth(c, min(w, MAX_AUTO_COL_WIDTH_PX))
        finally:
            self._applying_sizes = False

    def set_key_col(self, col: int):
        self._key_col = col
        self._model.set_key_col(col)

    def set_key_row(self, row: int):
        self._key_row = row
        self._model.set_key_row(row)

    def updateGeometries(self):
        """틀 고정 활성 시 상단/좌측 여백을 예약(setViewportMargins)해 고정 헬퍼 자리를 확보하고
        본체 데이터 뷰포트를 그만큼 밀어낸다(가림 방지). Qt가 지오메트리를 재계산할 때마다
        (리사이즈·헤더폭 변화·스크롤·스플리터) 여백을 다시 적용하고 헬퍼 위치를 갱신한다."""
        if self._in_update_geoms:
            # 재진입(내 setViewportMargins가 유발) — super()를 다시 부르면 방금 설정한 여백이
            # 헤더 크기로 리셋되므로 아무것도 하지 않고 내 여백을 보존한다.
            return
        self._in_update_geoms = True
        try:
            super().updateGeometries()
            fc = self._freeze
            # 벌크(_applying_sizes: 필터의 setRowHidden 폭풍/크기 일괄 적용) 중엔 무거운 freeze
            # 작업(여백·헤더 재배치·reposition)을 건너뛴다 — 벌크가 끝난 뒤 refresh가 1회만 재적용.
            if fc is not None and fc.active and not self._applying_sizes:
                fr = self.frameWidth()
                hdr_w = self.verticalHeader().width()
                hdr_h = self.horizontalHeader().height()
                fw, fh = fc.frozen_px()
                self.setViewportMargins(hdr_w + fw, hdr_h + fh, 0, 0)
                # 본체 헤더를 밀어낸 뷰포트에 맞춰 재배치 — 그래야 고정 눈금(corner) 옆에
                # 스크롤 눈금(본체 헤더)이 정확히 이어진다. QTableView는 추가 여백만큼
                # 헤더를 옮기지 않으므로 수동 정렬한다.
                vp = self.viewport().geometry()
                self.horizontalHeader().setGeometry(
                    fr + hdr_w + fw, fr, vp.width(), hdr_h)
                self.verticalHeader().setGeometry(
                    fr, fr + hdr_h + fh, hdr_w, vp.height())
                fc.reposition()
        finally:
            self._in_update_geoms = False

    def set_excluded_cols(self, cols: set):
        """외부(MainWindow)에서 제외 열 집합을 갱신하고 헤더/셀을 다시 칠한다."""
        self._excluded_cols = set(cols)
        self._model.set_excluded_cols(self._excluded_cols)

    def _col_has_changed(self, col: int) -> bool:
        """지정 열에 changed 셀이 있는가 — 첫 changed에서 조기 종료(O(발견까지의 행))."""
        m = self._model
        return any(m.cell_kind(r, col) == "changed" for r in range(m.data_rows))

    def _row_has_changed(self, row: int) -> bool:
        m = self._model
        return any(m.cell_kind(row, c) == "changed" for c in range(m.data_cols))

    def _cols_have_staged(self, cols) -> bool:
        """대상 열들 중 하나라도 staged 셀을 포함하는가 — staged 집합 기반 O(#staged)."""
        col_set = set(cols)
        return any(c in col_set for (_r, c) in self._model.staged_coords())

    def _rows_have_staged(self, rows) -> bool:
        row_set = set(rows)
        return any(r in row_set for (r, _c) in self._model.staged_coords())

    def _select_col(self, col: int):
        """해당 열 전체 셀을 선택 상태로 설정."""
        sm = self.selectionModel()
        rows = self.rowCount()
        cols = self.columnCount()
        if sm is None or rows == 0 or cols == 0 or not (0 <= col < cols):
            return
        model = self.model()
        sel = QItemSelection(model.index(0, col), model.index(rows - 1, col))
        sm.select(sel, QItemSelectionModel.ClearAndSelect)

    def _select_cols(self, cols) -> None:
        """여러 열 전체 셀을 한 번에 선택 (비연속 지원) — 다중 열 병합 준비용."""
        sm = self.selectionModel()
        rows = self.rowCount()
        col_max = self.columnCount() - 1
        if sm is None or rows == 0 or col_max < 0:
            return
        model = self.model()
        sel = QItemSelection()
        for c in cols:
            if 0 <= c <= col_max:
                sel.append(QItemSelectionRange(
                    model.index(0, c), model.index(rows - 1, c)))
        sm.select(sel, QItemSelectionModel.ClearAndSelect)

    def _select_row(self, row: int):
        """해당 행 전체 셀을 선택 상태로 설정."""
        sm = self.selectionModel()
        rows = self.rowCount()
        cols = self.columnCount()
        if sm is None or rows == 0 or cols == 0 or not (0 <= row < rows):
            return
        model = self.model()
        sel = QItemSelection(model.index(row, 0), model.index(row, cols - 1))
        sm.select(sel, QItemSelectionModel.ClearAndSelect)

    def _select_rows(self, rows) -> None:
        """여러 행 전체 셀을 한 번에 선택 (비연속 지원)."""
        sm = self.selectionModel()
        cols = self.columnCount()
        row_max = self.rowCount() - 1
        if sm is None or cols == 0 or row_max < 0:
            return
        model = self.model()
        sel = QItemSelection()
        for r in rows:
            if 0 <= r <= row_max:
                sel.append(QItemSelectionRange(
                    model.index(r, 0), model.index(r, cols - 1)))
        sm.select(sel, QItemSelectionModel.ClearAndSelect)

    def _touched_rows(self) -> set[int]:
        """선택 range가 닿은 모든 행(부분 선택 포함). O(#range × 평균행폭)."""
        sm = self.selectionModel()
        rows: set[int] = set()
        if sm is not None:
            for rng in sm.selection():
                rows.update(range(rng.top(), rng.bottom() + 1))
        return rows

    def _touched_cols(self) -> set[int]:
        sm = self.selectionModel()
        cols: set[int] = set()
        if sm is not None:
            for rng in sm.selection():
                cols.update(range(rng.left(), rng.right() + 1))
        return cols

    def _selected_header_rows(self, anchor_row: int) -> list[int]:
        """우클릭 시 대상 행 집합 결정 (열 헤더 _selected_header_cols와 대칭).
        - 우클릭한 행이 현재 다중 선택에 포함되어 있으면 그 선택 전체.
        - 아니면 우클릭한 단일 행만.
        range 기반 — 전체 행 선택 우선, 없으면 닿은 행 집합(기존 selectedIndexes 폴백과 동일 의미)."""
        rows = set(self._full_rows_selected()) or self._touched_rows()
        if anchor_row in rows and len(rows) > 1:
            return sorted(rows)
        return [anchor_row]

    def _selected_header_cols(self, anchor_col: int) -> list[int]:
        """우클릭 시 대상 열 집합 결정.
        - 우클릭한 열이 현재 헤더 다중 선택에 포함되어 있으면 그 선택 전체.
        - 아니면 우클릭한 단일 열만.
        range 기반 — 전체 열 선택 우선, 없으면 닿은 열 집합. O(#range)."""
        cols = set(self._full_columns_selected()) or self._touched_cols()
        if anchor_col in cols and len(cols) > 1:
            return sorted(cols)
        return [anchor_col]

    def _show_header_context_menu(self, pos, header=None):
        # header: 클릭된 실제 가로 헤더(기본=본체). 틀 고정 corner 헤더에서도 호출될 수 있다.
        header = header or self.horizontalHeader()
        col = header.logicalIndexAt(pos)
        if col < 0:
            return

        target_cols = self._selected_header_cols(col)
        multi = len(target_cols) > 1
        col_letter = get_column_letter(col + 1)
        cols_label = ", ".join(get_column_letter(c + 1) for c in target_cols)

        # 병합 준비/취소는 단일·다중 모두 지원 — 대상 열들 중 하나라도 변경/스테이징 셀이 있으면 노출.
        # (제외 열은 cell_kind가 'same'이라 has_changed/has_staged에 기여하지 않는다.)
        # has_changed는 열별 조기 종료(_col_has_changed) + any(), has_staged는 staged 집합 기반.
        has_changed = any(self._col_has_changed(c) for c in target_cols)
        has_staged  = self._cols_have_staged(target_cols)

        menu = QMenu(self)
        menu.setStyleSheet(MENU_QSS)

        # ── 병합 준비 항목 (단일·다중 공통) ──
        act_a2b = act_b2a = act_unstage = None
        if has_changed:
            act_a2b = menu.addAction(f"A → B  병합 준비  [{cols_label}열]")
            act_b2a = menu.addAction(f"B → A  병합 준비  [{cols_label}열]")
        if has_staged:
            if has_changed:
                menu.addSeparator()
            act_unstage = menu.addAction(f"병합 준비 취소  [{cols_label}열]")

        if has_changed or has_staged:
            menu.addSeparator()

        # ── 키 열 항목 (단일 선택일 때만) ──
        act_key_clear = act_key_set = None
        if not multi:
            if col == self._key_col:
                act_key_clear = menu.addAction("🔓  키 열 해제 (ROW 순서 기반 비교)")
            else:
                # 키 아이콘을 함께 노출해 키 지정 동작임을 시각적으로 표시.
                act_key_set = menu.addAction(
                    key_header_icon(), f"키 열로 설정  [{col_letter}열]")

        # ── 변경 검사 제외/해제 토글 ──
        # 키 열은 변경 검사에서 제외할 수 없으므로 대상에서 뺀다. 선택에 '제외됨'과 '비제외'가
        # 섞여 있으면 두 항목(제외 / 해제)을 모두 노출한다.
        exclude_cols = [c for c in target_cols if c != self._key_col]
        to_exclude = [c for c in exclude_cols if c not in self._excluded_cols]
        to_unexclude = [c for c in exclude_cols if c in self._excluded_cols]
        act_excl = act_unexcl = None
        if to_exclude or to_unexclude:
            menu.addSeparator()
        if to_exclude:
            lbl = ", ".join(get_column_letter(c + 1) for c in to_exclude)
            act_excl = menu.addAction(
                exclude_header_icon(), f"변경 검사에서 제외  [{lbl}열]")
        if to_unexclude:
            lbl = ", ".join(get_column_letter(c + 1) for c in to_unexclude)
            act_unexcl = menu.addAction(
                reset_header_icon(), f"검사 제외 해제  [{lbl}열]")

        act = menu.exec_(header.mapToGlobal(pos))
        if act is None:
            return

        if act_a2b is not None and act == act_a2b:
            self._select_cols(target_cols)
            self.stage_requested.emit(DIR_A2B)
        elif act_b2a is not None and act == act_b2a:
            self._select_cols(target_cols)
            self.stage_requested.emit(DIR_B2A)
        elif act_unstage is not None and act == act_unstage:
            self._select_cols(target_cols)
            self.unstage_requested.emit()
        elif act_key_clear is not None and act == act_key_clear:
            self.key_col_changed.emit(-1)
        elif act_key_set is not None and act == act_key_set:
            self.key_col_changed.emit(col)
        elif act_excl is not None and act == act_excl:
            self.columns_exclude_set.emit(to_exclude, True)
        elif act_unexcl is not None and act == act_unexcl:
            self.columns_exclude_set.emit(to_unexclude, False)

    def _show_row_header_context_menu(self, pos, header=None):
        # header: 클릭된 실제 세로 헤더(기본=본체). 틀 고정 corner 헤더에서도 호출될 수 있다.
        header = header or self.verticalHeader()
        row = header.logicalIndexAt(pos)
        if row < 0:
            return

        # 다중 선택된 행 헤더 전체를 대상으로 (열 헤더와 대칭)
        target_rows = self._selected_header_rows(row)
        multi = len(target_rows) > 1
        # has_changed는 행별 조기 종료(_row_has_changed) + any(), has_staged는 staged 집합 기반.
        has_changed = any(self._row_has_changed(r) for r in target_rows)
        has_staged  = self._rows_have_staged(target_rows)

        # 키 행(헤더 행) 지정: 단일 행이면서 원본 파일 행으로 매핑 가능한 데이터 행일 때만.
        # display 행은 키 매칭으로 재정렬될 수 있으므로 원본 파일 행 번호로 변환해 emit한다.
        orig_row = self._model.orig_row(row) if not multi else None
        can_key = orig_row is not None
        # 현재 키 행 우클릭 시엔 키 항목을 노출하지 않는다(초기화 기능 제거).
        # 키 행이 아닌 데이터 행에서만 '키 행으로 설정'을 노출.
        show_key_set = can_key and orig_row != self._key_row

        if not has_changed and not has_staged and not show_key_set:
            return

        suffix = f"  [{len(target_rows)}개 행]" if multi else ""
        menu = QMenu(self)
        menu.setStyleSheet(MENU_QSS)

        act_a2b = act_b2a = act_unstage = None
        if has_changed:
            act_a2b = menu.addAction(f"A → B  병합 준비{suffix}")
            act_b2a = menu.addAction(f"B → A  병합 준비{suffix}")
        if has_staged:
            if has_changed:
                menu.addSeparator()
            act_unstage = menu.addAction(f"병합 준비 취소{suffix}")

        # ── 키 행 항목: 키 행이 아닌 단일 데이터 행에만 '키 행으로 설정' 노출 ──
        act_key_row_set = None
        if show_key_set:
            if has_changed or has_staged:
                menu.addSeparator()
            act_key_row_set = menu.addAction(
                key_header_icon(), f"키 행으로 설정  [{orig_row + 1}행]")

        act = menu.exec_(header.mapToGlobal(pos))
        if act is None:
            return

        if act_a2b is not None and act == act_a2b:
            self._select_rows(target_rows)
            self.stage_requested.emit(DIR_A2B)
        elif act_b2a is not None and act == act_b2a:
            self._select_rows(target_rows)
            self.stage_requested.emit(DIR_B2A)
        elif act_unstage is not None and act == act_unstage:
            self._select_rows(target_rows)
            self.unstage_requested.emit()
        elif act_key_row_set is not None and act == act_key_row_set:
            self.key_row_changed.emit(orig_row)

    def _show_context_menu(self, pos):
        # has_changed: any()가 첫 changed에서 조기 종료. has_staged: staged 집합 기반(소수).
        # 과거엔 전 선택 셀을 순회하며 둘 다 찾을 때까지 멈추지 않아, 전체 선택+staged 없음이면
        # 수백만 셀을 훑어 우클릭이 지연됐다.
        has_staged = self._has_staged_selection()
        has_changed = self._has_changed_selection()
        if not has_changed and not has_staged:
            return

        menu = QMenu(self)
        menu.setStyleSheet(MENU_QSS)

        act_a2b     = menu.addAction("A → B  병합 준비") if has_changed else None
        act_b2a     = menu.addAction("B → A  병합 준비") if has_changed else None
        act_unstage = None
        if has_staged:
            if has_changed:
                menu.addSeparator()
            act_unstage = menu.addAction("병합 준비 취소")

        act = menu.exec_(self.viewport().mapToGlobal(pos))
        if act is None:
            return
        if act_a2b is not None and act == act_a2b:
            self.stage_requested.emit(DIR_A2B)
        elif act_b2a is not None and act == act_b2a:
            self.stage_requested.emit(DIR_B2A)
        elif act_unstage is not None and act == act_unstage:
            self.unstage_requested.emit()

    # ── 엑셀식 키보드 네비/선택/병합 단축키 ──────────────────────────────────
    def _is_empty_cell(self, r: int, c: int) -> bool:
        if r < 0 or r >= self.rowCount() or c < 0 or c >= self.columnCount():
            return True
        return self._model.display_text(r, c) == ""

    def _next_visible_row(self, r: int, dr: int) -> int:
        """dr 방향의 다음 '보이는' 행 — 숨겨진 행(변경 행만 보기)은 건너뛴다.
        범위 밖이면 범위 밖 인덱스를 그대로 반환한다."""
        max_r = self.rowCount() - 1
        n = r + dr
        while 0 <= n <= max_r and self.isRowHidden(n):
            n += dr
        return n

    def _jump_target(self, r: int, c: int, dr: int, dc: int) -> tuple:
        """엑셀의 Ctrl+방향키 시맨틱으로 점프 대상 (row, col) 반환.
        세로 이동은 보이는 행만 밟는다 — 숨겨진 행에는 착지하지 않는다."""
        max_r = self.rowCount() - 1
        max_c = self.columnCount() - 1
        if max_r < 0 or max_c < 0:
            return (max(0, r), max(0, c))

        def step(rr, cc):
            if dr:
                return self._next_visible_row(rr, dr), cc
            return rr, cc + dc

        def in_range(rr, cc):
            return 0 <= rr <= max_r and 0 <= cc <= max_c

        nr, nc = step(r, c)
        if not in_range(nr, nc):
            return (max(0, min(r, max_r)), max(0, min(c, max_c)))
        if self._is_empty_cell(r, c) or self._is_empty_cell(nr, nc):
            # 빈 구간 건너 다음 비어있지 않은 셀까지 — 못 찾으면 마지막 보이는 셀
            prev = (r, c)
            while in_range(nr, nc) and self._is_empty_cell(nr, nc):
                prev = (nr, nc)
                nr, nc = step(nr, nc)
            if not in_range(nr, nc):
                return prev
            return (nr, nc)
        # 연속 데이터의 마지막 비어있지 않은 셀까지
        while True:
            r2, c2 = step(nr, nc)
            if not in_range(r2, c2) or self._is_empty_cell(r2, c2):
                return (nr, nc)
            nr, nc = r2, c2

    @staticmethod
    def _header_jump_target(cur: int, d: int, last: int, is_empty) -> int:
        """엑셀 Ctrl+Shift 시맨틱의 열/행 단위 점프 대상 — _jump_target의 1차원 버전.
        매 호출마다 현재 위치(cur) 기준으로 재판정하므로 경계에서 한 번 더 누르면
        다음 값 블록 또는 그리드 끝으로 계속 이동한다."""
        if last < 0:
            return max(0, cur)
        n = cur + d
        if n < 0 or n > last:
            return min(max(cur, 0), last)
        if is_empty(cur) or is_empty(n):
            # 빈 구간 건너 다음 값 블록의 첫 열/행까지 — 없으면 그리드 끝
            while 0 <= n <= last and is_empty(n):
                n += d
            if n < 0 or n > last:
                return 0 if d < 0 else last
            return n
        # 연속 값 블록의 마지막 열/행까지
        while 0 <= n + d <= last and not is_empty(n + d):
            n += d
        return n

    def _select_range(self, r1: int, c1: int, r2: int, c2: int):
        """(r1,c1)~(r2,c2) 직사각형의 셀들을 모두 선택 상태로 설정 (기존 선택은 클리어).
        기존 구현은 아이템이 있는 셀만 선택됐으므로 데이터 영역으로 클램프한다."""
        sm = self.selectionModel()
        m = self._model
        if sm is None:
            return
        rs, re_ = sorted((r1, r2))
        cs, ce_ = sorted((c1, c2))
        rs, cs = max(0, rs), max(0, cs)
        re_ = min(re_, m.data_rows - 1)
        ce_ = min(ce_, m.data_cols - 1)
        if rs > re_ or cs > ce_:
            self.clearSelection()
            return
        sel = QItemSelection(m.index(rs, cs), m.index(re_, ce_))
        sm.select(sel, QItemSelectionModel.ClearAndSelect)

    def _has_changed_selection(self) -> bool:
        # any()가 첫 changed 셀에서 조기 종료 — 변경이 있는 선택(일반)에선 빠르다.
        m = self._model
        return any(m.cell_kind(r, c) == "changed"
                   for (r, c) in self._iter_selected_cells())

    def _has_staged_selection(self) -> bool:
        """선택 안에 staged 셀이 있는가 — staged 집합(소수)을 선택 range에 대해 조회.
        전 선택 셀을 순회하던 과거 O(선택셀수) 대신 O(#staged × #range)."""
        sm = self.selectionModel()
        if sm is None:
            return False
        staged = self._model.staged_coords()
        if not staged:
            return False
        sel = sm.selection()
        model = self._model
        return any(sel.contains(model.index(r, c)) for (r, c) in staged)

    # ── 헤더 다중 선택 지원 ──────────────────────────────────────────────────
    def _full_columns_selected(self) -> list[int]:
        """모든 행에 걸쳐 선택된 열 목록 = '열 헤더 선택' 상태.
        selectedColumns()는 대량 선택에서 셀을 개별 열거해 O(선택셀수)로 느리므로,
        selection() range를 직접 본다: 행 전체(0..rowCount-1)를 덮는 range의 열들 합집합.
        앱은 전체 열 선택을 그 경계 range로 만들므로 selectedColumns()와 동일 결과 — O(#range)."""
        sm = self.selectionModel()
        if sm is None:
            return []
        row_max = self.rowCount() - 1
        if row_max < 0:
            return []
        cols: set[int] = set()
        for rng in sm.selection():
            if rng.top() == 0 and rng.bottom() == row_max:
                cols.update(range(rng.left(), rng.right() + 1))
        return sorted(cols)

    def _full_rows_selected(self) -> list[int]:
        """모든 열에 걸쳐 선택된 행 목록 (열 대칭). O(#range)."""
        sm = self.selectionModel()
        if sm is None:
            return []
        col_max = self.columnCount() - 1
        if col_max < 0:
            return []
        rows: set[int] = set()
        for rng in sm.selection():
            if rng.left() == 0 and rng.right() == col_max:
                rows.update(range(rng.top(), rng.bottom() + 1))
        return sorted(rows)

    def _select_column_range(self, c1: int, c2: int):
        """[c1..c2] 모든 열의 모든 행 셀을 선택."""
        sm = self.selectionModel()
        rows = self.rowCount()
        cols = self.columnCount()
        if sm is None or rows == 0 or cols == 0:
            return
        cs, ce_ = sorted((c1, c2))
        cs = max(0, cs); ce_ = min(cols - 1, ce_)
        if cs > ce_:
            return
        model = self.model()
        sel = QItemSelection(model.index(0, cs), model.index(rows - 1, ce_))
        sm.select(sel, QItemSelectionModel.ClearAndSelect)

    def _select_row_range(self, r1: int, r2: int):
        sm = self.selectionModel()
        rows = self.rowCount()
        cols = self.columnCount()
        if sm is None or rows == 0 or cols == 0:
            return
        rs, re_ = sorted((r1, r2))
        rs = max(0, rs); re_ = min(rows - 1, re_)
        if rs > re_:
            return
        model = self.model()
        sel = QItemSelection(model.index(rs, 0), model.index(re_, cols - 1))
        sm.select(sel, QItemSelectionModel.ClearAndSelect)

    def keyPressEvent(self, event):
        if self._populating:
            return super().keyPressEvent(event)
        key = event.key()
        mods = event.modifiers()
        ctrl = bool(mods & Qt.ControlModifier)
        shift = bool(mods & Qt.ShiftModifier)
        alt = bool(mods & Qt.AltModifier)

        cur_r, cur_c = self._current_cell()

        # ── 헤더 다중 선택 확장 (Shift / Ctrl+Shift + 방향키) ──
        # 열 전체가 선택된 상태에서 Shift+←/→ 는 열 단위 확장,
        # 행 전체가 선택된 상태에서 Shift+↑/↓ 는 행 단위 확장.
        if shift and not alt and key in (Qt.Key_Left, Qt.Key_Right, Qt.Key_Up, Qt.Key_Down):
            full_cols = self._full_columns_selected()
            full_rows = self._full_rows_selected()
            is_col_mode = bool(full_cols) and key in (Qt.Key_Left, Qt.Key_Right)
            is_row_mode = bool(full_rows) and key in (Qt.Key_Up, Qt.Key_Down)

            if is_col_mode and self.columnCount() > 0 and self.rowCount() > 0:
                # anchor 초기화: 단일 열만 선택돼있고 anchor 없음 → 그 열을 anchor로
                if self._header_anchor_col is None:
                    if len(full_cols) == 1:
                        self._header_anchor_col = full_cols[0]
                    else:
                        # 다중 열 이미 선택 — currentIndex와 가장 먼 끝을 anchor로
                        cur_col_idx = cur_c if cur_c >= 0 else full_cols[-1]
                        self._header_anchor_col = (
                            full_cols[0] if cur_col_idx == full_cols[-1] else full_cols[-1]
                        )
                # 현재 확장 끝점 = currentIndex 또는 anchor 반대편 끝
                if cur_c >= 0 and cur_c in full_cols:
                    cur_end = cur_c
                else:
                    cur_end = full_cols[-1] if self._header_anchor_col == full_cols[0] else full_cols[0]
                delta = -1 if key == Qt.Key_Left else 1
                if ctrl:
                    # 엑셀처럼 현재 위치 기준 재판정 — 값 블록 끝 → 다음 블록 → 그리드 끝
                    m = self._model
                    target = self._header_jump_target(
                        cur_end, delta, self.columnCount() - 1,
                        lambda c: not m.col_has_values(c))
                else:
                    target = max(0, min(self.columnCount() - 1, cur_end + delta))
                self._select_column_range(self._header_anchor_col, target)
                self._set_current_cell_no_update(max(0, cur_r if cur_r >= 0 else 0), target)
                event.accept(); return

            if is_row_mode and self.columnCount() > 0 and self.rowCount() > 0:
                if self._header_anchor_row is None:
                    if len(full_rows) == 1:
                        self._header_anchor_row = full_rows[0]
                    else:
                        cur_row_idx = cur_r if cur_r >= 0 else full_rows[-1]
                        self._header_anchor_row = (
                            full_rows[0] if cur_row_idx == full_rows[-1] else full_rows[-1]
                        )
                if cur_r >= 0 and cur_r in full_rows:
                    cur_end = cur_r
                else:
                    cur_end = full_rows[-1] if self._header_anchor_row == full_rows[0] else full_rows[0]
                delta = -1 if key == Qt.Key_Up else 1
                if ctrl:
                    # 엑셀처럼 현재 위치 기준 재판정 — 값 블록 끝 → 다음 블록 → 그리드 끝
                    m = self._model
                    target = self._header_jump_target(
                        cur_end, delta, self.rowCount() - 1,
                        lambda r: not m.row_has_values(r))
                else:
                    target = max(0, min(self.rowCount() - 1, cur_end + delta))
                self._select_row_range(self._header_anchor_row, target)
                self._set_current_cell_no_update(target, max(0, cur_c if cur_c >= 0 else 0))
                event.accept(); return

        # 헤더 anchor 라이프사이클: 헤더 모드 분기에 들어가지 않은 일반 키는 anchor 무효화
        # (단순 Shift 아닌 키, 혹은 헤더가 아닌 일반 셀 선택 상태일 때)
        if key in (Qt.Key_Left, Qt.Key_Right, Qt.Key_Up, Qt.Key_Down):
            if not shift:
                self._header_anchor_col = None
                self._header_anchor_row = None

        # ── 병합 단축키 (Alt+Left / Alt+Right / Alt+Backspace) ──
        if alt and not ctrl and not shift:
            if key == Qt.Key_Right:
                if self._has_changed_selection():
                    self.stage_requested.emit(DIR_A2B)
                event.accept(); return
            if key == Qt.Key_Left:
                if self._has_changed_selection():
                    self.stage_requested.emit(DIR_B2A)
                event.accept(); return
            if key in (Qt.Key_Backspace, Qt.Key_Delete):
                if self._has_staged_selection():
                    self.unstage_requested.emit()
                event.accept(); return

        # ── Enter/Return: 엑셀처럼 아래 칸으로 이동 ──
        if key in (Qt.Key_Return, Qt.Key_Enter) and not ctrl and not alt:
            if cur_r >= 0 and cur_c >= 0 and cur_r + 1 < self.rowCount():
                self._set_current_cell(cur_r + 1, cur_c)
            event.accept(); return

        # ── Shift+Space: 행 전체, Ctrl+Space: 열 전체 ──
        if key == Qt.Key_Space and shift and not ctrl and not alt and cur_r >= 0:
            self._select_row(cur_r)
            event.accept(); return
        if key == Qt.Key_Space and ctrl and not shift and not alt and cur_c >= 0:
            self._select_col(cur_c)
            event.accept(); return

        # ── Ctrl(+Shift)+방향키: 데이터 경계 점프 (Excel 시맨틱) ──
        if ctrl and not alt and key in (Qt.Key_Left, Qt.Key_Right, Qt.Key_Up, Qt.Key_Down):
            if cur_r < 0 or cur_c < 0:
                return super().keyPressEvent(event)
            dr = -1 if key == Qt.Key_Up else (1 if key == Qt.Key_Down else 0)
            dc = -1 if key == Qt.Key_Left else (1 if key == Qt.Key_Right else 0)
            tr, tc = self._jump_target(cur_r, cur_c, dr, dc)
            if shift:
                # anchor = 현재 selection의 처음 시작점 추정 (currentIndex 기준)
                anchor = self.currentIndex()
                ar = anchor.row() if anchor.isValid() else cur_r
                ac = anchor.column() if anchor.isValid() else cur_c
                self._select_range(ar, ac, tr, tc)
                self._set_current_cell_no_update(tr, tc)
            else:
                self._move_current_cell(tr, tc)
            event.accept(); return

        # ── Ctrl+Home / Ctrl+End ──
        if ctrl and not alt and key == Qt.Key_Home:
            tr, tc = 0, 0
            if shift and cur_r >= 0 and cur_c >= 0:
                self._select_range(cur_r, cur_c, tr, tc)
                self._set_current_cell_no_update(tr, tc)
            else:
                self._move_current_cell(tr, tc)
            event.accept(); return
        if ctrl and not alt and key == Qt.Key_End:
            tr, tc = max(0, self.rowCount() - 1), max(0, self.columnCount() - 1)
            if shift and cur_r >= 0 and cur_c >= 0:
                self._select_range(cur_r, cur_c, tr, tc)
                self._set_current_cell_no_update(tr, tc)
            else:
                self._move_current_cell(tr, tc)
            event.accept(); return

        super().keyPressEvent(event)

    def populate(self, diff_matrix: list[list], which: str,
                 merged_set: set = None, staged: dict = None,
                 row_meta: list = None, excluded_cols: set = None):
        if not diff_matrix:
            self._safe_clear()
            return

        self._excluded_cols = set(excluded_cols) if excluded_cols else set()
        self._populating = True
        self._header_anchor_col = None
        self._header_anchor_row = None
        prev_updates = self.updatesEnabled()
        self.setUpdatesEnabled(False)
        try:
            # 모델 리셋 한 번 — 셀 아이템 생성 없음(O(1)). 리셋이 selectionModel
            # 시그널(선택 해제)을 발화시키므로 _populating 플래그 유지가 필수.
            # ※ `staged or {}`처럼 falsy 체크를 쓰면 '빈 dict'일 때 새 객체가
            #   만들어져 모델이 MainWindow 상태와 분리된다(이후 스테이징 색 미반영).
            #   반드시 is None 체크로 원본 참조를 공유해야 한다.
            self._model.set_diff_data(
                diff_matrix, row_meta,
                staged if staged is not None else {},
                merged_set if merged_set is not None else set(),
                self._excluded_cols)
            # 1) 샘플 기반 자동 너비(상한 클립 포함)
            # → 2) 사용자가 직접 조정한 열/행만 그 위에 덮어쓰기 (상한 무시).
            # 새로고침(_run_refresh)은 _user_col_widths/_user_row_heights를 미리
            # 비우므로 그 경로에서는 2)가 건너뛰어져 디폴트로 복귀한다.
            self._auto_size_columns()
            if self._user_col_widths or self._user_row_heights:
                self._apply_user_sizes()
        finally:
            self.setUpdatesEnabled(prev_updates)
            self._populating = False

    def populate_preview(self, data: list[list]):
        if not data:
            self._safe_clear()
            return
        self._populating = True
        self._header_anchor_col = None
        self._header_anchor_row = None
        prev_updates = self.updatesEnabled()
        self.setUpdatesEnabled(False)
        try:
            self._model.set_preview_data(data)
            self._auto_size_columns()
            if self._user_col_widths or self._user_row_heights:
                self._apply_user_sizes()
        finally:
            self.setUpdatesEnabled(prev_updates)
            self._populating = False

    def _safe_clear(self):
        self._populating = True
        try:
            self._model.clear()
        finally:
            self._populating = False
        self._header_anchor_col = None
        self._header_anchor_row = None

    def get_selected_cells(self) -> set:
        return set(self._iter_selected_cells())

    def _iter_selected_cells(self):
        """선택 셀 (r, c)를 range 단위로 순회.
        selectedIndexes()처럼 전체 QModelIndex 리스트를 먼저 만들지 않아
        대량 선택 + 조기 종료 조합에서 훨씬 가볍다."""
        sm = self.selectionModel()
        if sm is None:
            return
        for rng in sm.selection():
            for r in range(rng.top(), rng.bottom() + 1):
                for c in range(rng.left(), rng.right() + 1):
                    yield (r, c)

    def _single_selected_cell(self):
        """정확히 1개 셀이 선택돼 있으면 (r, c), 아니면 None — O(range 수)."""
        sm = self.selectionModel()
        if sm is None:
            return None
        total = 0
        first = None
        for rng in sm.selection():
            if first is None:
                first = (rng.top(), rng.left())
            total += rng.width() * rng.height()
            if total > 1:
                return None
        return first if total == 1 else None

    def mirror_selection_from(self, src: "ExcelTableView"):
        """반대 패널의 선택을 range 단위로 그대로 복제 — O(range 수).
        헤더 클릭(열/행 전체 선택)은 range 1개라 셀 수와 무관하게 즉시 끝난다.
        (셀 집합으로 풀었다 재조립하면 행마다 range가 생겨 이후 모든 선택
        질의/페인팅이 느려진다 — 헤더 클릭 딜레이의 주범이었다.)"""
        src_sm = src.selectionModel()
        sm = self.selectionModel()
        if src_sm is None or sm is None:
            return
        self._populating = True
        prev_updates = self.updatesEnabled()
        self.setUpdatesEnabled(False)
        try:
            row_max = self.rowCount() - 1
            col_max = self.columnCount() - 1
            sel = QItemSelection()
            model = self.model()
            for rng in src_sm.selection():
                if rng.top() > row_max or rng.left() > col_max:
                    continue
                sel.append(QItemSelectionRange(
                    model.index(rng.top(), rng.left()),
                    model.index(min(rng.bottom(), row_max),
                                min(rng.right(), col_max))))
            sm.select(sel, QItemSelectionModel.ClearAndSelect)
        finally:
            self.setUpdatesEnabled(prev_updates)
            self._populating = False

    def mirror_selection(self, cells: set):
        # 셀 집합을 row별 연속 column 구간(span)으로 묶고, 연속 행의 span이
        # 동일하면 직사각형으로 수직 병합해 range 수를 최소화한다.
        # (열 전체 선택 = range 1개. range 수천 개짜리 selection model은
        # 이후 모든 질의·페인팅·선택 병합을 느리게 만든다.)
        self._populating = True
        prev_updates = self.updatesEnabled()
        self.setUpdatesEnabled(False)
        try:
            sm = self.selectionModel()
            if sm is None:
                return
            row_max = self.rowCount() - 1
            col_max = self.columnCount() - 1
            if row_max < 0 or col_max < 0:
                sm.clearSelection()
                return
            by_row: dict[int, list[int]] = {}
            for (r, c) in cells:
                if 0 <= r <= row_max and 0 <= c <= col_max:
                    by_row.setdefault(r, []).append(c)

            def _spans(cs: list) -> tuple:
                cs.sort()
                out = []
                start = prev = cs[0]
                for c in cs[1:]:
                    if c == prev + 1:
                        prev = c
                        continue
                    out.append((start, prev))
                    start = prev = c
                out.append((start, prev))
                return tuple(out)

            sel = QItemSelection()
            model = self.model()
            run_start = prev_r = None
            run_spans = None

            def _flush(end_r):
                for (c1, c2) in run_spans:
                    sel.append(QItemSelectionRange(
                        model.index(run_start, c1), model.index(end_r, c2)))

            for r in sorted(by_row):
                sp = _spans(by_row[r])
                if run_spans is not None and sp == run_spans and r == prev_r + 1:
                    prev_r = r
                    continue
                if run_spans is not None:
                    _flush(prev_r)
                run_start = prev_r = r
                run_spans = sp
            if run_spans is not None:
                _flush(prev_r)
            sm.select(sel, QItemSelectionModel.ClearAndSelect)
        finally:
            self.setUpdatesEnabled(prev_updates)
            self._populating = False




def _extract_supported_path(mime_data) -> str:
    if mime_data.hasUrls():
        for url in mime_data.urls():
            path = url.toLocalFile()
            if os.path.splitext(path)[1].lower() in _SUPPORTED_EXTS:
                return path
    return ""


def _extract_folder_path(mime_data) -> str:
    """드롭된 MIME에서 첫 번째 '폴더' 로컬 경로를 반환(없으면 "")."""
    if mime_data.hasUrls():
        for url in mime_data.urls():
            p = url.toLocalFile()
            if p and os.path.isdir(p):
                return p
    return ""


class DropLineEdit(QLineEdit):
    file_dropped = pyqtSignal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)

    def dragEnterEvent(self, event):
        if _extract_supported_path(event.mimeData()):
            event.acceptProposedAction()
        else:
            event.ignore()

    def dragMoveEvent(self, event):
        if _extract_supported_path(event.mimeData()):
            event.acceptProposedAction()
        else:
            event.ignore()

    def dropEvent(self, event):
        path = _extract_supported_path(event.mimeData())
        if path:
            self.file_dropped.emit(path)
            event.acceptProposedAction()


class CellEditWidget(QPlainTextEdit):
    """선택 셀의 값 표시란 — 읽기전용 뷰어 (직접 수정 기능 제거됨).
    기본 4줄 높이지만, 상하 스플리터 핸들을 드래그해 높이를 조절할 수 있다.
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        line_h = self.fontMetrics().lineSpacing()
        # 기본 4줄이 한눈에 보이도록 — 고정이 아니라 최소/기본 높이만 지정(스플리터로 조절 가능).
        base_h = line_h * 4 + 12
        self.setMinimumHeight(line_h + 10)
        self.resize(self.width(), base_h)
        self._base_height = base_h
        self.setLineWrapMode(QPlainTextEdit.WidgetWidth)
        self.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.setReadOnly(True)   # 값 확인·복사 전용

    def sizeHint(self):
        return QSize(super().sizeHint().width(), getattr(self, "_base_height", 48))

    def text(self):
        return self.toPlainText()

    def _set_plain_default(self, val: str):
        """텍스트를 기본 서식(검정·배경 없음)으로 설정.
        setPlainText는 직전 강조(빨강)의 char format을 새 텍스트에 물려받는 경우가 있어,
        전체를 빈 QTextCharFormat으로 다시 칠해 강조 잔상을 제거한다."""
        self.setPlainText(val if val is not None else "")
        cur = QTextCursor(self.document())
        cur.select(QTextCursor.Document)
        cur.setCharFormat(QTextCharFormat())
        self.setTextCursor(QTextCursor(self.document()))

    def setText(self, val: str):
        self._set_plain_default(val)
        # 텍스트 설정 후 항상 맨 위부터 표시
        self.verticalScrollBar().setValue(0)

    def set_highlighted(self, val: str, ranges, color):
        """val을 표시하되 ranges의 [start, end) 문자 구간 배경을 color로 강조.
        ranges가 비면 일반 텍스트와 동일."""
        self._set_plain_default(val)
        if val and ranges:
            fmt = QTextCharFormat()
            fmt.setBackground(color)
            fmt.setForeground(CELL_DIFF_FG)   # 강조 구간 폰트도 빨강 (배경 위 가독)
            n = len(val)
            for start, end in ranges:
                start = max(0, start)
                end = min(n, end)
                if start >= end:
                    continue
                cur = self.textCursor()
                cur.setPosition(start)
                cur.setPosition(end, QTextCursor.KeepAnchor)
                cur.mergeCharFormat(fmt)
            # 커서/선택을 맨 앞으로 되돌려 강조가 파란 선택색에 가려지지 않게
            self.setTextCursor(QTextCursor(self.document()))
        self.verticalScrollBar().setValue(0)


class SheetTabBar(QTabBar):
    """하단 시트 탭 바 — 변경점이 있는 시트 탭을 노란 배경으로 강조.
    기본 탭을 그린 뒤 변경 탭 위에 반투명 노랑을 덧칠해 텍스트 가독을 유지한다."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._changed_idx: set[int] = set()   # 변경점 있는 탭 인덱스

    def set_changed_indices(self, indices):
        new = set(indices)
        if new != self._changed_idx:
            self._changed_idx = new
            self.update()

    def clear_changed(self):
        self.set_changed_indices(set())

    def paintEvent(self, _event):
        p = QStylePainter(self)
        opt = QStyleOptionTab()
        # 선택 탭을 마지막에 그려 이웃 탭 경계에 가려지지 않게 한다(기본 QTabBar와 동일).
        sel = self.currentIndex()
        order = [i for i in range(self.count()) if i != sel]
        if 0 <= sel < self.count():
            order.append(sel)
        for i in order:
            self.initStyleOption(opt, i)
            p.drawControl(QStyle.CE_TabBarTab, opt)
        # 변경 탭: 노랑을 덧칠하면 기존 글자가 묻히므로, 그 위에 진한 볼드 텍스트를 다시 그린다.
        for i in self._changed_idx:
            if not (0 <= i < self.count()):
                continue
            r = self.tabRect(i)
            p.fillRect(r, SHEET_TAB_CHANGED_BG)
            p.save()
            f = self.font()
            f.setBold(True)
            p.setFont(f)
            p.setPen(QColor(0x20, 0x20, 0x20))   # 진한 먹색 — 노랑 위 가독
            p.drawText(r, Qt.AlignCenter, self.tabText(i))
            p.restore()

