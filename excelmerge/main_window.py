"""메인 윈도우 (excel_diff_merge.py에서 분리)."""
import os
import re

from PyQt5.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QSplitter, QLineEdit, QStatusBar, QMessageBox,
    QShortcut, QAbstractItemView,
)
from PyQt5.QtCore import Qt, QSize, QRect, QPoint
from PyQt5.QtGui import QColor, QFont, QIcon, QPixmap, QKeySequence, QPainter, QPen
from openpyxl.utils import get_column_letter

from .diff_engine import compute_diff
from .loaders import _EXCEL_EXTS
from .panels import FilePanel
from .theme import APP_QSS, DIFF_COLORS, load_app_icon, ui_font
from .widgets import EXTRA_ROWS, ExcelTableView, MinimapScrollBar
from .workers import LoadWorker, PreviewWorker, StagedMergeWorker
from .xlsx_writer import _is_file_locked


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("ExcelMerge")
        # 작업 표시줄·Alt+Tab에서도 타이틀바와 동일한 앱 아이콘이 보이도록 명시 지정.
        self.setWindowIcon(load_app_icon())
        self.resize(1400, 800)
        self._load_worker:          LoadWorker | None         = None
        self._preview_workers: dict[str, PreviewWorker | None] = {"a": None, "b": None}
        self._staged_merge_worker: StagedMergeWorker | None = None
        self._saving_side: str = "a"
        self._diff_matrix: list[list] = []
        self._diff_row_meta: list = []   # [(orig_a_row, orig_b_row), ...]
        self._merged_cells: set = set()
        self._staged: dict = {}          # {(r, c): 'a_to_b' | 'b_to_a'}
        self._preview_data: dict = {"a": [], "b": []}   # 미리보기 raw data
        self._formula_data: dict = {"a": [], "b": []}   # 수식 원문 데이터
        self._diff_only: bool = False
        self._undo_stack: list = []   # [("stage", cells, direction)] — 병합 준비 되돌리기용
        self._raw_data: dict = {"a": [], "b": []}   # 키 열 변경 시 재계산용 캐시
        self._key_col: int = 0
        self._excluded_cols: set[int] = set()   # 변경 검사에서 제외할 (display) 열 인덱스

        self._build_ui()
        self._apply_style()
        undo_sc = QShortcut(QKeySequence("Ctrl+Z"), self)
        undo_sc.activated.connect(self._undo)
        diff_only_sc = QShortcut(QKeySequence("Ctrl+D"), self)
        diff_only_sc.setContext(Qt.ApplicationShortcut)
        diff_only_sc.activated.connect(self._toggle_diff_only_shortcut)
        # F5 — 새로고침 (refresh_btn이 enabled 일 때만 실행)
        refresh_sc = QShortcut(QKeySequence("F5"), self)
        refresh_sc.setContext(Qt.ApplicationShortcut)
        refresh_sc.activated.connect(self._on_refresh_shortcut)
        # Alt+↑ / Alt+↓ — 이전/다음 변경 셀로 이동
        prev_diff_sc = QShortcut(QKeySequence("Alt+Up"), self)
        prev_diff_sc.setContext(Qt.ApplicationShortcut)
        prev_diff_sc.activated.connect(self._on_prev_diff_shortcut)
        next_diff_sc = QShortcut(QKeySequence("Alt+Down"), self)
        next_diff_sc.setContext(Qt.ApplicationShortcut)
        next_diff_sc.activated.connect(self._on_next_diff_shortcut)
        # Ctrl+F — 찾기 입력란으로 포커스 이동
        find_sc = QShortcut(QKeySequence("Ctrl+F"), self)
        find_sc.setContext(Qt.ApplicationShortcut)
        find_sc.activated.connect(self._focus_find)

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(8, 8, 8, 8)
        root.setSpacing(6)

        # ── 툴바 ──
        toolbar = QHBoxLayout()

        self.diff_only_btn = QPushButton("변경 행만 보기")
        self.diff_only_btn.setFixedHeight(36)
        self.diff_only_btn.setFont(ui_font(10))
        self.diff_only_btn.setCheckable(True)
        self.diff_only_btn.setObjectName("toggle_btn")
        self.diff_only_btn.setEnabled(False)
        self.diff_only_btn.setToolTip("변경된 행만 표시 / 전체 표시 전환 (Ctrl+D)")
        self.diff_only_btn.toggled.connect(self._on_diff_only_toggled)
        toolbar.addWidget(self.diff_only_btn)

        self.refresh_btn = QPushButton("새로고침")
        self.refresh_btn.setFixedHeight(36)
        self.refresh_btn.setFont(ui_font(10))
        self.refresh_btn.setToolTip("지정된 경로의 파일을 다시 불러와 비교합니다 (F5)")
        self.refresh_btn.setEnabled(False)
        self.refresh_btn.clicked.connect(self._run_refresh)
        toolbar.addWidget(self.refresh_btn)

        # 이전/다음 변경 셀 이동 버튼
        self.prev_diff_btn = QPushButton("◀ 이전 변경")
        self.prev_diff_btn.setFixedHeight(36)
        self.prev_diff_btn.setFont(ui_font(10))
        self.prev_diff_btn.setEnabled(False)
        self.prev_diff_btn.setToolTip("이전 변경 셀로 이동 (Alt+↑)")
        self.prev_diff_btn.clicked.connect(lambda: self._goto_changed(-1))
        toolbar.addWidget(self.prev_diff_btn)

        self.next_diff_btn = QPushButton("다음 변경 ▶")
        self.next_diff_btn.setFixedHeight(36)
        self.next_diff_btn.setFont(ui_font(10))
        self.next_diff_btn.setEnabled(False)
        self.next_diff_btn.setToolTip("다음 변경 셀로 이동 (Alt+↓)")
        self.next_diff_btn.clicked.connect(lambda: self._goto_changed(+1))
        toolbar.addWidget(self.next_diff_btn)

        # 찾기 — 검색란 + 옵션 토글 + 이전/다음 찾기 버튼
        toolbar.addSpacing(16)
        find_box = QHBoxLayout()
        find_box.setSpacing(4)

        self.find_edit = QLineEdit()
        self.find_edit.setObjectName("find_edit")
        self.find_edit.setPlaceholderText("찾을 내용 (Ctrl+F)")
        self.find_edit.setFixedHeight(36)
        self.find_edit.setFixedWidth(200)
        self.find_edit.setFont(ui_font(10))
        self.find_edit.setClearButtonEnabled(True)
        self.find_edit.setEnabled(False)
        self.find_edit.setToolTip("셀 값 검색 — Enter: 다음 찾기, Shift+Enter: 이전 찾기")
        self.find_edit.returnPressed.connect(lambda: self._goto_find(+1))
        find_prev_sc = QShortcut(QKeySequence("Shift+Return"), self.find_edit)
        find_prev_sc.setContext(Qt.WidgetShortcut)
        find_prev_sc.activated.connect(lambda: self._goto_find(-1))
        find_box.addWidget(self.find_edit)

        def _find_btn(kind: str, tooltip: str, checkable: bool) -> QPushButton:
            btn = QPushButton()
            btn.setObjectName("find_btn")
            btn.setFixedSize(36, 36)
            btn.setIcon(self._make_find_icon(kind))
            btn.setIconSize(QSize(22, 22))
            btn.setCheckable(checkable)
            btn.setEnabled(False)
            btn.setToolTip(tooltip)
            find_box.addWidget(btn)
            return btn

        self.find_case_btn = _find_btn(
            "case",
            "대소문자 무시 (Ignore case)\n"
            "켜짐: 대소문자를 구분하지 않고 검색\n"
            "꺼짐: 대소문자가 정확히 일치할 때만 검색",
            checkable=True,
        )
        self.find_case_btn.setChecked(True)

        self.find_word_btn = _find_btn(
            "word",
            "전체 단어 일치 (Match whole word only)\n"
            "켜짐: 검색어가 독립된 단어로 존재할 때만 찾음\n"
            "꺼짐: 부분 문자열도 찾음",
            checkable=True,
        )

        self.find_prev_btn = _find_btn("prev", "이전 찾기 (Shift+Enter)", checkable=False)
        self.find_prev_btn.clicked.connect(lambda: self._goto_find(-1))

        self.find_next_btn = _find_btn("next", "다음 찾기 (Enter)", checkable=False)
        self.find_next_btn.clicked.connect(lambda: self._goto_find(+1))

        toolbar.addLayout(find_box)

        toolbar.addStretch()

        # 범례
        for lbl, key in [
            ("신규", "added"),
            ("변경", "modified"), ("준비", "staged"), ("병합", "merged"),
        ]:
            dot = QLabel("  ")
            dot.setFixedSize(20, 20)
            dot.setStyleSheet(
                f"background:{DIFF_COLORS[key].name()};"
                "border:1px solid #aaa; border-radius:3px;"
            )
            txt = QLabel(lbl)
            txt.setFont(ui_font(9))
            toolbar.addWidget(dot)
            toolbar.addWidget(txt)
            toolbar.addSpacing(8)

        root.addLayout(toolbar)

        # ── 좌우 패널 ──
        splitter = QSplitter(Qt.Horizontal)
        self.panel_a = FilePanel("A 파일 (원본)", "a")
        self.panel_b = FilePanel("B 파일 (비교)", "b")
        self.panels = {"a": self.panel_a, "b": self.panel_b}   # side 셀렉터 공용 맵
        splitter.addWidget(self.panel_a)
        splitter.addWidget(self.panel_b)
        splitter.setSizes([700, 700])
        root.addWidget(splitter, 1)

        # ── 변경 셀 위치 미니맵: 양쪽 테이블의 세로/가로 스크롤바를 커스텀으로 교체 ──
        # 스크롤 동기화 시그널 연결 전에 교체해야 verticalScrollBar()/horizontalScrollBar()
        # 핸들이 새 객체를 가리킨다.
        for tbl in (self.panel_a.table, self.panel_b.table):
            tbl.setVerticalScrollBar(MinimapScrollBar(Qt.Vertical, tbl))
            tbl.setHorizontalScrollBar(MinimapScrollBar(Qt.Horizontal, tbl))

        # ── 상태바 ──
        self.status = QStatusBar()
        self.setStatusBar(self.status)
        self.status.showMessage(
            "파일을 선택하면 자동으로 비교합니다.  "
            "| 셀 선택 후 우클릭 → 병합 준비 → 저장"
        )

        self._syncing_selection = False
        self._hidden_rows = set()   # '변경 행만 보기' 필터의 델타 캐시

        # A↔B 대칭 배선 — 스크롤/크기/선택 동기화 (src → dst 양방향 루프)
        # ※ 람다의 루프 변수는 기본값 인자(s=..., d=...)로 바인딩해야 한다.
        for src, dst in ((self.panel_a, self.panel_b), (self.panel_b, self.panel_a)):
            src.table.horizontalScrollBar().valueChanged.connect(
                dst.table.horizontalScrollBar().setValue)
            src.table.verticalScrollBar().valueChanged.connect(
                dst.table.verticalScrollBar().setValue)
            # 열/행 크기 동기화 — 사용자 조작에 의한 변경만 (apply_*는 시그널 미발행)
            src.table.column_resized.connect(dst.table.apply_column_width)
            src.table.row_resized.connect(dst.table.apply_row_height)
            # 선택 셀 동기화 — QTableView는 itemSelectionChanged가 없으므로
            # selectionModel().selectionChanged 사용 (selectionModel은 ctor에서 1회 생성)
            src.table.selectionModel().selectionChanged.connect(
                lambda *_, s=src.table, d=dst.table: self._sync_selection(s, d))

        # side별 패널 시그널 배선
        for side, panel in self.panels.items():
            panel.table.stage_requested.connect(self._stage_selected)
            panel.table.unstage_requested.connect(self._unstage_selected)
            panel.table.key_col_changed.connect(self._on_key_col_changed)
            panel.table.columns_exclude_set.connect(self._on_columns_exclude_set)
            panel.save_btn.clicked.connect(
                lambda _=False, s=side: self._save_staged(s))
            panel.file_loaded.connect(
                lambda p, s=side: self._on_file_loaded(s, p))

    def _apply_style(self):
        self.setStyleSheet(APP_QSS)

    # ── 파일 미리보기 (비교 전 단독 표시) ────────────────────────────────────────

    def _reset_compare_state(self):
        """비교 결과를 초기화하고 버튼 상태를 되돌린다."""
        self._diff_matrix = []
        self._diff_row_meta = []
        self._merged_cells = set()
        self._staged = {}
        self._preview_data = {"a": [], "b": []}
        self._formula_data = {"a": [], "b": []}
        self._clear_row_filter()   # diff 모드의 숨김 행이 미리보기에 남지 않도록 해제
        self._excluded_cols.clear()
        self.panel_a.table.set_excluded_cols(self._excluded_cols)
        self.panel_b.table.set_excluded_cols(self._excluded_cols)
        self.panel_a._row_meta = []
        self.panel_b._row_meta = []
        self.panel_a._staged_display = {}
        self.panel_b._staged_display = {}
        self.diff_only_btn.setChecked(False)
        self.diff_only_btn.setEnabled(False)
        self.prev_diff_btn.setEnabled(False)
        self.next_diff_btn.setEnabled(False)
        self._set_find_enabled(False)
        self._update_minimap()
        self._set_save_btn_state()

    def _run_preview(self, side: str, path: str):
        """파일 선택 즉시 해당 패널에 원본 데이터를 색상 없이 표시한다."""
        # 비교 결과가 있으면 비교 상태를 리셋하고 반대쪽 패널도 미리보기로 전환
        if self._diff_matrix:
            self._reset_compare_state()
            other_side = "b" if side == "a" else "a"
            other_path = self.panels[other_side].get_path()
            if other_path:
                self._run_preview(other_side, other_path)

        worker = PreviewWorker(side, path)
        worker.done.connect(self._on_preview_done)
        worker.progress.connect(self.status.showMessage)
        worker.error.connect(lambda msg: self.status.showMessage(f"파일 로드 오류: {msg}"))
        worker.finished.connect(worker.deleteLater)
        self._preview_workers[side] = worker
        worker.start()
        self.status.showMessage(f"{'A' if side == 'a' else 'B'} 파일 로딩 중...")

    def _on_preview_done(self, side: str, data: list[list], formula_data: list[list]):
        self._preview_data[side] = data
        self._formula_data[side] = formula_data
        panel = self.panels[side]
        panel._formula_data = formula_data
        panel._row_meta = []   # 미리보기 모드: row_meta 없음 (행 인덱스 = 원본 인덱스)
        panel.preview(data)
        if data:
            self._set_find_enabled(True)   # 미리보기 상태에서도 Ctrl+F 검색 허용
        rows = len(data)
        cols = max((len(r) for r in data), default=0)
        self.status.showMessage(
            f"{'A' if side == 'a' else 'B'} 파일 로드 완료 — {rows}행 × {cols}열  "
            "| '비교 실행'을 클릭해 두 파일을 비교하세요."
        )
        self._set_save_btn_state()

    # ── 비교 ──────────────────────────────────────────────────────────────────

    def _on_refresh_shortcut(self):
        """F5 단축키 — 새로고침 버튼이 활성 상태일 때만 동작."""
        if self.refresh_btn.isEnabled():
            self._run_refresh()

    def _run_refresh(self):
        # 새로고침은 디폴트(자동너비+MAX_AUTO_COL_WIDTH_PX 상한) 상태로 복귀시킨다.
        # 세션 중 사용자가 헤더를 드래그해 늘려놓은 열/행 크기도 함께 리셋되어야
        # populate() 종료부의 _apply_user_sizes() 분기를 건너뛴다.
        for tbl in (self.panel_a.table, self.panel_b.table):
            tbl._user_col_widths.clear()
            tbl._user_row_heights.clear()

        path_a = self.panel_a.get_path()
        path_b = self.panel_b.get_path()
        if path_a and path_b:
            self._run_compare()
        elif path_a:
            self._run_preview("a", path_a)
        elif path_b:
            self._run_preview("b", path_b)

    def _run_compare(self):
        path_a = self.panel_a.get_path()
        path_b = self.panel_b.get_path()
        if not path_a and not path_b:
            QMessageBox.warning(self, "경고", "A 파일과 B 파일 중 하나 이상 선택하세요.")
            return

        self._set_buttons_enabled(False)
        self._merged_cells = set()
        self._staged = {}
        self._diff_matrix = []
        self._diff_row_meta = []   # 미리보기 잠금 해제
        self._excluded_cols.clear()
        self.panel_a.table.set_excluded_cols(self._excluded_cols)
        self.panel_b.table.set_excluded_cols(self._excluded_cols)
        self.status.showMessage("파일 로딩 중...")

        self._load_worker = LoadWorker(path_a, path_b)
        self._load_worker.done.connect(self._on_loaded)
        self._load_worker.error.connect(self._on_error)
        self._load_worker.progress.connect(self.status.showMessage)
        self._load_worker.finished.connect(self._load_worker.deleteLater)
        self._load_worker.start()

    def _on_loaded(self, a_data, b_data, a_formulas, b_formulas):
        self._raw_data["a"] = a_data
        self._raw_data["b"] = b_data
        self._formula_data["a"] = a_formulas
        self._formula_data["b"] = b_formulas
        self.panel_a._formula_data = a_formulas
        self.panel_b._formula_data = b_formulas
        self._diff_matrix, self._diff_row_meta = compute_diff(
            a_data, b_data, self._key_col)
        self.panel_a._row_meta = self._diff_row_meta
        self.panel_b._row_meta = self._diff_row_meta
        self._refresh_tables()

        rows = len(self._diff_matrix)
        cols = len(self._diff_matrix[0]) if self._diff_matrix else 0
        changed = self._count_changed()

        self._set_buttons_enabled(True)
        # 비교 완료 시 디폴트로 "변경 행만 보기" ON.
        # 이미 ON이었다면 setChecked는 시그널이 발생하지 않으므로 _apply_diff_filter를 직접 호출.
        if self.diff_only_btn.isChecked():
            self._apply_diff_filter()
        else:
            self.diff_only_btn.setChecked(True)
        self.status.showMessage(
            f"비교 완료 — {rows}행 × {cols}열 | 변경된 셀: {changed}개  "
            "| 셀 선택 후 우클릭 → 병합 준비 → 선택 병합 저장"
        )

    # ── 키 열 변경 ────────────────────────────────────────────────────────────

    def _on_key_col_changed(self, col: int):
        if col == self._key_col:
            return
        self._key_col = col
        self.panel_a.table.set_key_col(col)
        self.panel_b.table.set_key_col(col)
        if self._raw_data["a"] or self._raw_data["b"]:
            self._recompute_diff()

    def _recompute_diff(self):
        self._merged_cells = set()
        self._staged = {}
        # 키 열이 바뀌면 동일 인덱스가 다른 의미가 될 수 있으므로 제외 상태도 리셋.
        self._excluded_cols.clear()
        self.panel_a.table.set_excluded_cols(self._excluded_cols)
        self.panel_b.table.set_excluded_cols(self._excluded_cols)
        # 키 열 변경 시 사용자의 현재 토글 상태(diff_only_btn)는 그대로 유지한다.
        self._diff_matrix, self._diff_row_meta = compute_diff(
            self._raw_data["a"], self._raw_data["b"], self._key_col)
        self.panel_a._row_meta = self._diff_row_meta
        self.panel_b._row_meta = self._diff_row_meta
        self._refresh_tables()
        self._set_buttons_enabled(True)
        self._apply_diff_filter()
        rows = len(self._diff_matrix)
        cols = len(self._diff_matrix[0]) if self._diff_matrix else 0
        changed = self._count_changed()
        key_letter = get_column_letter(self._key_col + 1) if self._key_col >= 0 else "없음(ROW 순서)"
        self.status.showMessage(
            f"키 열: {key_letter}  |  {rows}행 × {cols}열  |  변경된 셀: {changed}개  "
            "| 셀 선택 후 우클릭 → 병합 준비 → 선택 병합 저장"
        )

    # ── 변경 행 필터 ──────────────────────────────────────────────────────────

    def _toggle_diff_only_shortcut(self):
        if self.diff_only_btn.isEnabled():
            self.diff_only_btn.setChecked(not self.diff_only_btn.isChecked())

    def _on_diff_only_toggled(self, checked: bool):
        self._diff_only = checked
        self._apply_diff_filter()

    def _apply_diff_filter(self):
        if not self._diff_matrix:
            self._update_minimap()
            return
        # 목표 숨김 집합 계산 후, 캐시(_hidden_rows)와 달라진 행만 토글한다.
        # QTableView는 모델 리셋 후에도 숨김 상태를 유지하므로 캐시는 리셋과 무관하게
        # 유효하다. 현재 뷰 범위 밖 캐시 항목은 나중에 행이 다시 늘어날 때를 위해 보존.
        excl = self._excluded_cols
        desired = set()
        if self._diff_only:
            # 병합됨(연파랑) 셀이 있는 행은 값이 같아졌어도 계속 표시 —
            # 저장 직후 행이 사라져 병합 결과를 확인할 수 없던 문제 방지.
            # (새 비교/새로고침으로 _merged_cells가 리셋되면 일반 규칙으로 복귀)
            merged_rows = {r for (r, c) in self._merged_cells if c not in excl}
            for r, row in enumerate(self._diff_matrix):
                if r == 0:   # 최상단 행(헤더)은 항상 표시
                    continue
                if r in merged_rows:
                    continue
                is_changed = any(
                    status != "same"
                    for c, (status, *_) in enumerate(row)
                    if c not in excl
                )
                if not is_changed:
                    desired.add(r)
        view_rows = len(self._diff_matrix) + EXTRA_ROWS
        flips = {r for r in (self._hidden_rows ^ desired) if r < view_rows}
        if flips:
            # setRowHidden은 sectionResized(_, _, 0)을 emit해 _user_row_heights를
            # 오염시킨다. _applying_sizes 플래그로 _on_section_v_resized 기록을 차단.
            tables = (self.panel_a.table, self.panel_b.table)
            for tbl in tables:
                tbl._applying_sizes = True
            try:
                for r in flips:
                    hidden = r in desired
                    self.panel_a.table.setRowHidden(r, hidden)
                    self.panel_b.table.setRowHidden(r, hidden)
            finally:
                for tbl in tables:
                    tbl._applying_sizes = False
        self._hidden_rows = {r for r in self._hidden_rows if r >= view_rows} | desired
        self._update_minimap()

    def _clear_row_filter(self):
        """숨김 행을 전부 해제 — 미리보기 전환 등 diff 모드를 떠날 때 호출.
        QTableView는 리셋 후에도 숨김을 기억하므로 명시적으로 풀어야 한다."""
        if not self._hidden_rows:
            return
        tables = (self.panel_a.table, self.panel_b.table)
        for tbl in tables:
            tbl._applying_sizes = True
        try:
            for r in self._hidden_rows:
                self.panel_a.table.setRowHidden(r, False)
                self.panel_b.table.setRowHidden(r, False)
        finally:
            for tbl in tables:
                tbl._applying_sizes = False
        self._hidden_rows = set()

    # ── 변경 셀 탐색(이전/다음) ───────────────────────────────────────────────

    def _on_prev_diff_shortcut(self):
        if self.prev_diff_btn.isEnabled():
            self._goto_changed(-1)

    def _on_next_diff_shortcut(self):
        if self.next_diff_btn.isEnabled():
            self._goto_changed(+1)

    def _iter_changed_cells(self):
        """변경된 (r, c) 셀을 행 우선 순서로 yield. 숨겨진 행/제외 열은 제외."""
        if not self._diff_matrix:
            return
        excl = self._excluded_cols
        for r, row in enumerate(self._diff_matrix):
            if self.panel_a.table.isRowHidden(r):
                continue
            for c, cell in enumerate(row):
                if c in excl:
                    continue
                if cell[0] != "same":
                    yield (r, c)

    def _current_anchor(self):
        """현재 선택 셀(우선순위: panel_a → panel_b).
        없으면 (0, -1) 반환 — A1부터 검사하기 위한 sentinel."""
        for tbl in (self.panel_a.table, self.panel_b.table):
            r, c = tbl._current_cell()
            if r >= 0 and c >= 0:
                return (r, c)
        return (0, -1)

    def _goto_changed(self, direction: int):
        """direction=+1: 다음 변경 셀, -1: 이전 변경 셀."""
        if not self._diff_matrix:
            return
        cells = list(self._iter_changed_cells())
        if not cells:
            self.status.showMessage("변경된 셀이 없습니다.")
            return
        anchor = self._current_anchor()
        if direction > 0:
            target = next((p for p in cells if p > anchor), None)
            if target is None:
                self.status.showMessage("마지막 변경 셀입니다.")
                return
        else:
            target = next((p for p in reversed(cells) if p < anchor), None)
            if target is None:
                self.status.showMessage("첫 변경 셀입니다.")
                return
        r, c = target
        # 양쪽 패널 동기 선택 + 화면 중앙으로 스크롤
        for tbl in (self.panel_a.table, self.panel_b.table):
            tbl._set_current_cell(r, c)
            if tbl.model().is_data_cell(r, c):
                tbl.scrollTo(tbl.model().index(r, c), QAbstractItemView.PositionAtCenter)

    # ── 찾기 ──
    def _make_find_icon(self, kind: str) -> QIcon:
        """찾기 버튼용 아이콘을 QPainter로 렌더링 (HiDPI 2배 해상도).
        체크 시 배경이 파란색으로 바뀌므로 Off=진회색 / On=흰색 두 벌을 등록."""
        def render(color: QColor) -> QPixmap:
            s = 32
            pm = QPixmap(s * 2, s * 2)
            pm.setDevicePixelRatio(2)
            pm.fill(Qt.transparent)
            p = QPainter(pm)
            p.setRenderHint(QPainter.Antialiasing)
            p.setRenderHint(QPainter.TextAntialiasing)
            if kind == "case":
                p.setPen(color)
                p.setFont(QFont("Segoe UI", 12, QFont.Bold))
                p.drawText(QRect(0, 0, s, s), Qt.AlignCenter, "Aa")
            elif kind == "word":
                p.setPen(color)
                p.setFont(QFont("Segoe UI", 10, QFont.Bold))
                p.drawText(QRect(0, 0, s, s - 8), Qt.AlignCenter, "ab")
                p.setPen(QPen(color, 1.8, Qt.SolidLine, Qt.RoundCap))
                y = s - 7
                p.drawLine(QPoint(7, y), QPoint(s - 7, y))
                p.drawLine(QPoint(7, y), QPoint(7, y - 4))
                p.drawLine(QPoint(s - 7, y), QPoint(s - 7, y - 4))
            else:  # "prev" / "next" — 셰브론 화살표
                p.setPen(QPen(color, 2.6, Qt.SolidLine, Qt.RoundCap, Qt.RoundJoin))
                m = s // 2
                d = -1 if kind == "prev" else 1
                p.drawLine(QPoint(m - 3 * d, m - 7), QPoint(m + 4 * d, m))
                p.drawLine(QPoint(m + 4 * d, m), QPoint(m - 3 * d, m + 7))
            p.end()
            return pm

        ic = QIcon()
        ic.addPixmap(render(QColor("#3b3b3b")), QIcon.Normal, QIcon.Off)
        ic.addPixmap(render(QColor("#ffffff")), QIcon.Normal, QIcon.On)
        ic.addPixmap(render(QColor("#b8b8b8")), QIcon.Disabled, QIcon.Off)
        ic.addPixmap(render(QColor("#b8b8b8")), QIcon.Disabled, QIcon.On)
        return ic

    def _focus_find(self):
        """Ctrl+F — 찾기 입력란 포커스 + 전체 선택."""
        if self.find_edit.isEnabled():
            self.find_edit.setFocus()
            self.find_edit.selectAll()

    def _make_find_matcher(self, term: str):
        """검색 옵션(대소문자 무시/전체 단어)에 맞는 판별 함수를 반환."""
        ignore_case = self.find_case_btn.isChecked()
        whole_word = self.find_word_btn.isChecked()
        if whole_word:
            # \b는 검색어가 특수문자로 시작/끝나면 동작하지 않으므로 lookaround 사용
            flags = re.IGNORECASE if ignore_case else 0
            pat = re.compile(r"(?<!\w)" + re.escape(term) + r"(?!\w)", flags)
            return lambda text: pat.search(text) is not None
        if ignore_case:
            needle = term.casefold()
            return lambda text: needle in text.casefold()
        return lambda text: term in text

    def _iter_find_matches(self, match):
        """검색어와 일치하는 (r, c) 셀을 행 우선 순서로 yield.
        diff 모드: 숨겨진 행/제외 열 제외, A/B 표시 텍스트 중 한쪽이라도 일치하면 매치.
        미리보기 모드(파일 1개만 로드)에서도 로드된 패널의 데이터를 검색한다."""
        tbl_a = self.panel_a.table
        models = (tbl_a.model(), self.panel_b.table.model())
        rows = max(m.data_rows for m in models)
        cols = max(m.data_cols for m in models)
        if rows == 0:
            return
        diff_mode = bool(self._diff_matrix)
        excl = self._excluded_cols if diff_mode else set()
        for r in range(rows):
            if diff_mode and tbl_a.isRowHidden(r):
                continue
            for c in range(cols):
                if c in excl:
                    continue
                for model in models:
                    if match(model.display_text(r, c)):
                        yield (r, c)
                        break

    def _goto_find(self, direction: int):
        """direction=+1: 다음 찾기, -1: 이전 찾기. 끝에 도달하면 반대편에서 순환.
        diff 모드뿐 아니라 미리보기(파일 1개) 상태에서도 동작한다."""
        term = self.find_edit.text()
        has_data = (self.panel_a.table.model().data_rows
                    or self.panel_b.table.model().data_rows)
        if not term or not has_data:
            return
        cells = list(self._iter_find_matches(self._make_find_matcher(term)))
        if not cells:
            self.status.showMessage(f'"{term}" — 일치 항목이 없습니다.')
            return
        anchor = self._current_anchor()
        wrapped = ""
        if direction > 0:
            target = next((p for p in cells if p > anchor), None)
            if target is None:
                target = cells[0]
                wrapped = " — 처음부터 다시 검색합니다."
        else:
            target = next((p for p in reversed(cells) if p < anchor), None)
            if target is None:
                target = cells[-1]
                wrapped = " — 끝에서부터 다시 검색합니다."
        r, c = target
        # 양쪽 패널 동기 선택 + 화면 중앙으로 스크롤
        for tbl in (self.panel_a.table, self.panel_b.table):
            tbl._set_current_cell(r, c)
            if tbl.model().is_data_cell(r, c):
                tbl.scrollTo(tbl.model().index(r, c), QAbstractItemView.PositionAtCenter)
        idx = cells.index(target) + 1
        self.status.showMessage(f'찾기: "{term}" {idx}/{len(cells)}개 일치{wrapped}')

    def _update_minimap(self):
        """현재 표시 중인 행을 기준으로 변경 행/열의 비율 위치를 양쪽
        세로/가로 스크롤바에 전달한다.
        - 세로: 가시 행 중 변경 셀이 하나라도 있는 행의 위치.
        - 가로: 가시 행 안에서 변경 셀이 있는 열의 위치.
        """
        row_ratios = []
        col_ratios = []
        if self._diff_matrix:
            excl = self._excluded_cols
            visible_rows = [
                r for r in range(len(self._diff_matrix))
                if not self.panel_a.table.isRowHidden(r)
            ]

            def _row_has_changed(r):
                return any(
                    st != "same"
                    for c, (st, *_) in enumerate(self._diff_matrix[r])
                    if c not in excl
                )

            n = len(visible_rows)
            if n == 1:
                r = visible_rows[0]
                if _row_has_changed(r):
                    row_ratios.append(0.0)
            elif n > 1:
                denom = n - 1
                for vi, r in enumerate(visible_rows):
                    if _row_has_changed(r):
                        row_ratios.append(vi / denom)

            cols_total = len(self._diff_matrix[0]) if self._diff_matrix else 0
            if cols_total > 0 and visible_rows:
                if cols_total == 1:
                    if 0 not in excl and any(self._diff_matrix[r][0][0] != "same" for r in visible_rows):
                        col_ratios.append(0.0)
                else:
                    denom_c = cols_total - 1
                    for c in range(cols_total):
                        if c in excl:
                            continue
                        if any(self._diff_matrix[r][c][0] != "same" for r in visible_rows):
                            col_ratios.append(c / denom_c)
        for tbl in (self.panel_a.table, self.panel_b.table):
            v = tbl.verticalScrollBar()
            if isinstance(v, MinimapScrollBar):
                v.set_change_ratios(row_ratios)
            h = tbl.horizontalScrollBar()
            if isinstance(h, MinimapScrollBar):
                h.set_change_ratios(col_ratios)

    # ── 선택 셀 스테이징 (우클릭) ─────────────────────────────────────────────

    def _stage_selected(self, direction: str):
        if not self._diff_matrix:
            return

        cells = (
            self.panel_a.table.get_selected_cells()
            | self.panel_b.table.get_selected_cells()
        )
        cells = {
            (r, c) for (r, c) in cells
            if r < len(self._diff_matrix)
            and c < len(self._diff_matrix[r])
            and c not in self._excluded_cols
            and self._diff_matrix[r][c][0] != "same"
        }
        if not cells:
            QMessageBox.information(self, "알림", "선택한 셀 중 변경된 셀이 없습니다.")
            return

        # undo 스택에 stage 동작 기록 (셀 목록과 방향을 한 번에 저장)
        self._undo_stack.append(("stage", list(cells), direction))

        for cell in cells:
            self._staged[cell] = direction

        # staged 셀에 대해 양쪽 패널의 셀값란 표시값(수식 우선) 미리 계산
        def _resolve_display(fd, meta, side_idx, r, c, fallback):
            try:
                orig = meta[r][side_idx] if r < len(meta) else r
            except (IndexError, TypeError):
                orig = r
            if orig is not None:
                try:
                    v = fd[orig][c]
                    if v:
                        return v
                except (IndexError, TypeError):
                    pass
            return fallback

        for (r, c) in cells:
            dir_ = self._staged[r, c]
            try:
                _, a_val, b_val = self._diff_matrix[r][c]
            except (IndexError, TypeError):
                a_val, b_val = "", ""

            if dir_ == "a_to_b":
                a_display = _resolve_display(self._formula_data["a"], self.panel_a._row_meta, 0, r, c, a_val)
                b_display = a_display
            else:
                b_display = _resolve_display(self._formula_data["b"], self.panel_b._row_meta, 1, r, c, b_val)
                a_display = b_display
            self.panel_a._staged_display[r, c] = a_display
            self.panel_b._staged_display[r, c] = b_display

        self._notify_cells(cells)
        self._silent_clear_selection()
        self.panel_a._selected_cell = None
        self.panel_b._selected_cell = None
        self.panel_a.cell_edit.clear()
        self.panel_a.cell_edit.setEnabled(False)
        self.panel_b.cell_edit.clear()
        self.panel_b.cell_edit.setEnabled(False)
        self._set_save_btn_state()
        self.status.showMessage(
            f"병합 준비 완료 — {len(self._staged)}개 셀 대기 중  | '저장'을 클릭하면 파일에 저장됩니다."
        )

    # ── 선택 셀 언스테이징 (병합 준비 취소) ───────────────────────────────────────

    def _unstage_selected(self):
        if not self._staged:
            return
        cells = (
            self.panel_a.table.get_selected_cells()
            | self.panel_b.table.get_selected_cells()
        )
        removed = {c for c in cells if c in self._staged}
        if not removed:
            QMessageBox.information(self, "알림", "선택한 셀 중 병합 준비된 셀이 없습니다.")
            return
        for cell in removed:
            del self._staged[cell]
            self.panel_a._staged_display.pop(cell, None)
            self.panel_b._staged_display.pop(cell, None)
        sel_a = self.panel_a.table.get_selected_cells()
        sel_b = self.panel_b.table.get_selected_cells()
        self._notify_cells(removed)
        # 부분 갱신이라 선택이 유지되지만, 기존 복원 경로를 그대로 태워
        # cell_edit 동기화까지 동일하게 맞춘다.
        if sel_a:
            self.panel_a.table.mirror_selection(sel_a)
        if sel_b:
            self.panel_b.table.mirror_selection(sel_b)
        self.panel_a._sync_cell_edit()
        self.panel_b._sync_cell_edit()
        self._set_save_btn_state()
        self.status.showMessage(
            f"병합 준비 취소 — {len(removed)}개 셀 제거됨 | 남은 대기 셀: {len(self._staged)}개"
        )

    # ── 저장 ─────────────────────────────────────────────────────────────────

    def _save_staged(self, side: str):
        """side: 'a' 또는 'b' — 해당 파일만 저장. 병합 준비(staged)된 셀만 기록한다."""
        path = self.panels[side].get_path()

        # JSON/uasset 등 비-xlsx 는 저장 미지원 — 사용자에게 안내 후 중단
        if path and os.path.splitext(path)[1].lower() not in _EXCEL_EXTS:
            QMessageBox.information(
                self, "저장 미지원",
                f"{'A' if side == 'a' else 'B'} 파일은 비교 전용 형식입니다.\n"
                f"저장(병합)은 Excel(.xlsx/.xls/.xlsm) 파일에서만 지원됩니다.",
            )
            return

        # 직접 편집 기능이 제거되어 저장할 내용은 병합 준비 셀뿐 — diff 필수
        if not self._diff_matrix:
            return

        # 저장 대상 측이 실제로 쓸 내용이 있는지 확인
        # a 저장: b_to_a (B→A 방향) staged 셀만 / b 저장: a_to_b staged 셀만
        relevant_direction = "b_to_a" if side == "a" else "a_to_b"
        staged_for_side = {k: v for k, v in self._staged.items() if v == relevant_direction}
        if not staged_for_side:
            return

        if not path:
            QMessageBox.warning(self, "경고",
                f"{'A' if side == 'a' else 'B'} 파일이 지정되지 않았습니다.")
            return

        if _is_file_locked(path):
            QMessageBox.warning(
                self, "파일 열림",
                f"변경하고자 하는 파일이 열려 있으므로 저장할 수 없습니다:\n\n"
                f"{'A' if side == 'a' else 'B'} 파일: {os.path.basename(path)}"
                "\n\n파일을 닫은 후 다시 시도하세요."
            )
            return

        # 저장하지 않는 쪽 경로를 빈 문자열로 전달 → Worker가 해당 파일은 건드리지 않음
        path_a = path if side == "a" else ""
        path_b = path if side == "b" else ""

        self._saving_side = side
        self._set_buttons_enabled(False)
        self.status.showMessage("저장 중...")

        self._staged_merge_worker = StagedMergeWorker(
            path_a, path_b, list(self._diff_matrix),
            list(self._diff_row_meta),
            staged_for_side,
            self._formula_data["a"], self._formula_data["b"],
        )
        self._staged_merge_worker.done.connect(self._on_staged_saved)
        self._staged_merge_worker.error.connect(self._on_error)
        self._staged_merge_worker.finished.connect(self._staged_merge_worker.deleteLater)
        self._staged_merge_worker.start()

    def _on_staged_saved(self, count: int):
        side = getattr(self, "_saving_side", None)

        # ── 저장한 side에 해당하는 staged만 확정 반영 ────────────────────────
        relevant_direction = "b_to_a" if side == "a" else "a_to_b"
        saved_staged = {k: v for k, v in self._staged.items() if v == relevant_direction}
        staged_cells = set(saved_staged.keys())

        # staged 방향대로 diff_matrix 확정 반영
        for (r, c), direction in saved_staged.items():
            if r < len(self._diff_matrix) and c < len(self._diff_matrix[r]):
                _, a_val, b_val = self._diff_matrix[r][c]
                if direction == "a_to_b":
                    b_val = a_val
                else:
                    a_val = b_val
                self._diff_matrix[r][c] = ("same", a_val, b_val)

        # 저장한 side의 staged 제거 (나머지 side는 유지)
        for k in list(self._staged.keys()):
            if self._staged[k] == relevant_direction:
                del self._staged[k]
        self._merged_cells |= staged_cells

        self._notify_cells(staged_cells)
        self._silent_clear_selection()
        self._set_buttons_enabled(True)
        self.status.showMessage(f"저장 완료 — {count}개 셀 저장됨")
        QMessageBox.information(self, "저장 완료", f"{count}개 셀이 파일에 저장됐습니다.")

    def _on_error(self, msg: str):
        self._set_buttons_enabled(True)
        self.status.showMessage(f"오류: {msg}")
        QMessageBox.critical(self, "오류", f"작업 실패:\n{msg}")

    def _undo(self):
        """Ctrl+Z — 병합 준비(stage) 동작을 단계별로 되돌린다.
        (셀 직접 편집 기능이 제거되어 undo 대상은 stage 항목뿐)"""
        if not self._undo_stack:
            return
        entry = self._undo_stack.pop()
        if entry[0] != "stage":
            return
        _, cells, _ = entry
        for cell in cells:
            self._staged.pop(cell, None)
            self.panel_a._staged_display.pop(cell, None)
            self.panel_b._staged_display.pop(cell, None)
        self._silent_clear_selection()
        self.panel_a._selected_cell = None
        self.panel_b._selected_cell = None
        self._notify_cells(cells)
        self._set_save_btn_state()

    # ── 유틸 ──────────────────────────────────────────────────────────────────

    def _refresh_tables(self):
        """전체 리셋 경로 — 매트릭스 자체가 재계산됐을 때만 사용.
        stage/unstage/편집/저장확정/undo는 _notify_cells()로 부분 갱신한다."""
        self.panel_a.populate(self._diff_matrix, self._merged_cells, self._staged,
                              self._diff_row_meta, self._excluded_cols)
        self.panel_b.populate(self._diff_matrix, self._merged_cells, self._staged,
                              self._diff_row_meta, self._excluded_cols)
        self._apply_diff_filter()

    def _notify_cells(self, cells):
        """부분 갱신 경로 — 상태(dict/set)는 이미 변형된 뒤 호출된다.
        양쪽 모델에 최소 범위 dataChanged만 방출해 전체 repopulate를 피한다."""
        self.panel_a.table.model().notify_cells(cells)
        self.panel_b.table.model().notify_cells(cells)
        self._apply_diff_filter()   # 행 가시성은 status에 의존 — 델타 방식이라 저렴

    def _silent_clear_selection(self):
        """기존 populate가 선택을 조용히 초기화하던 동작 재현.
        _populating 플래그로 selectionChanged 핸들러(자동 편집 적용)를 차단한다."""
        for tbl in (self.panel_a.table, self.panel_b.table):
            tbl._populating = True
            try:
                tbl.clearSelection()
            finally:
                tbl._populating = False

    def _effective_status(self, r: int, c: int) -> str:
        """제외 열은 강제로 'same' 으로 노출 — _diff_matrix 원본은 보존."""
        if c in self._excluded_cols:
            return "same"
        return self._diff_matrix[r][c][0]

    def _on_columns_exclude_set(self, cols: list, exclude: bool):
        """헤더 우클릭 → cols 일괄 제외/해제."""
        if not cols:
            return
        if exclude:
            new_cols = [c for c in cols if c not in self._excluded_cols]
            for c in new_cols:
                self._excluded_cols.add(c)
            # 새로 제외된 열들의 기존 staged 항목 자동 해제.
            new_set = set(new_cols)
            staged_keys = [k for k in self._staged if k[1] in new_set]
            for key in staged_keys:
                del self._staged[key]
                self.panel_a._staged_display.pop(key, None)
                self.panel_b._staged_display.pop(key, None)
        else:
            for c in cols:
                self._excluded_cols.discard(c)
        # set_excluded_cols가 모델에 열 단위 dataChanged + 헤더 갱신을 방출한다
        self.panel_a.table.set_excluded_cols(self._excluded_cols)
        self.panel_b.table.set_excluded_cols(self._excluded_cols)
        self._silent_clear_selection()
        self._apply_diff_filter()
        self._set_save_btn_state()
        changed = self._count_changed()
        excl_letters = ", ".join(get_column_letter(c + 1) for c in sorted(self._excluded_cols))
        excl_msg = excl_letters if excl_letters else "없음"
        self.status.showMessage(
            f"검사 제외 열: {excl_msg}  |  변경된 셀: {changed}개"
        )

    def _sync_selection(self, src: ExcelTableView, dst: ExcelTableView):
        if self._syncing_selection or src._populating or dst._populating:
            return
        self._syncing_selection = True
        try:
            dst.mirror_selection_from(src)
            # mirror_selection 중 _populating=True라 _on_table_selection_changed가 막히므로
            # 반대쪽 패널의 cell_edit을 수동으로 갱신
            dst_panel = self.panel_a if dst is self.panel_a.table else self.panel_b
            dst_panel._sync_cell_edit()
        finally:
            self._syncing_selection = False

    def _count_changed(self) -> int:
        excl = self._excluded_cols
        return sum(
            1
            for r, row in enumerate(self._diff_matrix)
            for c, (st, *_) in enumerate(row)
            if st != "same" and c not in excl
        )

    def _set_save_btn_state(self, enabled: bool = True):
        # b_to_a staged → A 파일에 쓸 내용 / a_to_b staged → B 파일에 쓸 내용
        has_a = any(v == "b_to_a" for v in self._staged.values())
        has_b = any(v == "a_to_b" for v in self._staged.values())

        # JSON/uasset 등 비-xlsx 파일은 저장 미지원 → 버튼 강제 비활성화 + 툴팁 안내
        def _xlsx_ok(path: str) -> bool:
            return (not path) or os.path.splitext(path)[1].lower() in _EXCEL_EXTS
        path_a = self.panel_a.get_path()
        path_b = self.panel_b.get_path()
        a_writable = _xlsx_ok(path_a)
        b_writable = _xlsx_ok(path_b)

        self.panel_a.save_btn.setEnabled(enabled and has_a and a_writable)
        self.panel_b.save_btn.setEnabled(enabled and has_b and b_writable)
        self.panel_a.save_btn.setToolTip(
            "파일 저장" if a_writable
            else "JSON/uasset은 비교 전용 — 저장은 Excel(.xlsx/.xls/.xlsm)만 지원"
        )
        self.panel_b.save_btn.setToolTip(
            "파일 저장" if b_writable
            else "JSON/uasset은 비교 전용 — 저장은 Excel(.xlsx/.xls/.xlsm)만 지원"
        )

    def _on_file_loaded(self, side: str, path: str):
        either = bool(self.panel_a.get_path()) or bool(self.panel_b.get_path())
        self.refresh_btn.setEnabled(either)
        both = bool(self.panel_a.get_path()) and bool(self.panel_b.get_path())
        if both:
            self._run_compare()
        else:
            self._run_preview(side, path)

    def _set_buttons_enabled(self, enabled: bool):
        has_diff = enabled and bool(self._diff_matrix)
        self.diff_only_btn.setEnabled(has_diff)
        self.prev_diff_btn.setEnabled(has_diff)
        self.next_diff_btn.setEnabled(has_diff)
        self._set_find_enabled(has_diff)
        self._set_save_btn_state(enabled)

    def _set_find_enabled(self, enabled: bool):
        for w in (self.find_edit, self.find_case_btn, self.find_word_btn,
                  self.find_prev_btn, self.find_next_btn):
            w.setEnabled(enabled)

