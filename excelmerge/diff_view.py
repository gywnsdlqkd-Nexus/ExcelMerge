"""파일 비교 뷰 — 파일 쌍 하나의 diff/미리보기/병합/저장을 담당하는 탭 위젯.

구 MainWindow 본문에서 추출됐다. MainWindow는 이제 여러 DiffView/FolderCompareView를
담는 탭 컨테이너 셸이고, 실제 비교 UI·로직은 전부 이 클래스에 있다.

동작은 추출 전과 동일하다. 두 가지만 의도적으로 바뀌었다:
 ① 상태바를 직접 소유하지 않고 MainWindow가 만든 공유 QStatusBar를 주입받는다(self.status).
 ② 단축키 컨텍스트가 ApplicationShortcut → WidgetWithChildrenShortcut 로 바뀌어,
    포커스된(활성) 탭에서만 발화한다(탭 여러 개일 때 중복 발화 방지).
"""
import os
import re

from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QSplitter, QLineEdit, QMessageBox,
    QShortcut, QAbstractItemView,
)
from PyQt5.QtCore import Qt, QSize, pyqtSignal, QThread
from PyQt5.QtGui import QIcon, QKeySequence
from openpyxl.utils import get_column_letter

from .diff_engine import count_changed
from .loaders import _EXCEL_EXTS, list_sheet_names, clear_values_cache
from .panels import FilePanel
from .prefs import load_key_prefs, save_key_prefs, load_last_sheet, save_last_sheet
from .theme import APP_QSS, DIFF_COLORS, ui_font, ext_tab_icon
from .widgets import (
    ExcelTableView, MinimapScrollBar, make_find_icon, SheetTabBar, FreezeController,
)
from .workers import (
    LoadWorker, PreviewWorker, StagedMergeWorker, FormulaFlagWorker, SheetDiffWorker,
    DiffWorker,
)
from .xlsx_writer import _is_file_locked


class DiffView(QWidget):
    # 패널에 폴더가 드롭/선택됨 → MainWindow가 이 탭을 폴더 비교로 전환 (side, path)
    folder_requested = pyqtSignal(str, str)

    def __init__(self, status_bar, parent=None):
        super().__init__(parent)
        # MainWindow가 소유한 공유 상태바. _build_ui 이전에 세팅해야
        # 빌드 중 showMessage 호출이 안전하다.
        self.status = status_bar
        self._load_worker:          LoadWorker | None         = None
        self._preview_workers: dict[str, PreviewWorker | None] = {"a": None, "b": None}
        self._staged_merge_worker: StagedMergeWorker | None = None
        self._formula_flag_worker: FormulaFlagWorker | None = None   # 수식 플래그 지연 로딩
        self._sheet_diff_worker: SheetDiffWorker | None = None        # 시트별 변경 여부(탭 색)
        self._sheet_diff_token: int = 0
        self._pending_sheet_diff = None   # 첫 비교/미리보기 표시 후로 미룬 시트 diff (path_a,path_b,names)
        self._diff_worker: DiffWorker | None = None   # compute_diff 백그라운드
        self._diff_token: int = 0                     # 낡은 diff 결과 폐기용
        self._formula_flags_requested: bool = False   # 이번 비교/미리보기에서 로드 시작 여부
        self._formula_flags: dict = {"a": set(), "b": set()}   # 수식 셀 좌표 {(row0,col0)}
        self._saving_side: str = "a"
        self._diff_matrix: list[list] = []
        self._diff_row_meta: list = []   # [(orig_a_row, orig_b_row), ...]
        self._merged_cells: set = set()
        self._staged: dict = {}          # {(r, c): 'a_to_b' | 'b_to_a'}
        self._preview_data: dict = {"a": [], "b": []}   # 미리보기 raw data
        self._diff_only: bool = False
        self._undo_stack: list = []   # [("stage", cells, direction)] — 병합 준비 되돌리기용
        self._raw_data: dict = {"a": [], "b": []}   # 키 열/행 변경 시 재계산용 캐시
        # 키 헤더 앵커 = (키 행, 키 열). 전역 저장값을 기본으로 로드(없으면 A1 = 0,0).
        self._key_row, self._key_col = load_key_prefs()
        self._excluded_cols: set[int] = set()   # 변경 검사에서 제외할 (display) 열 인덱스
        self._sheet_names: list[str] = []       # 현재 시트 탭에 표시 중인 이름 목록(A∪B)
        self._current_sheet: str | None = None  # 현재 비교/미리보기 중인 시트 이름

        self._build_ui()
        # 로드한 앵커를 두 패널에 반영(헤더 아이콘/하이라이트). 실제 매칭은 로드 후 compute_diff.
        for _panel in (self.panel_a, self.panel_b):
            _panel.table.set_key_col(self._key_col)
            _panel.table.set_key_row(self._key_row)
        self._apply_style()
        # 단축키 — 활성 탭(포커스가 이 뷰나 그 자식에 있을 때)에서만 발화하도록
        # WidgetWithChildrenShortcut 컨텍스트로 스코프. 탭이 여러 개여도 중복되지 않는다.
        undo_sc = QShortcut(QKeySequence("Ctrl+Z"), self)
        undo_sc.setContext(Qt.WidgetWithChildrenShortcut)
        undo_sc.activated.connect(self._undo)
        diff_only_sc = QShortcut(QKeySequence("Ctrl+D"), self)
        diff_only_sc.setContext(Qt.WidgetWithChildrenShortcut)
        diff_only_sc.activated.connect(self._toggle_diff_only_shortcut)
        # F5 — 새로고침 (refresh_btn이 enabled 일 때만 실행)
        refresh_sc = QShortcut(QKeySequence("F5"), self)
        refresh_sc.setContext(Qt.WidgetWithChildrenShortcut)
        refresh_sc.activated.connect(self._on_refresh_shortcut)
        # Alt+↑ / Alt+↓ — 이전/다음 변경 셀로 이동
        prev_diff_sc = QShortcut(QKeySequence("Alt+Up"), self)
        prev_diff_sc.setContext(Qt.WidgetWithChildrenShortcut)
        prev_diff_sc.activated.connect(self._on_prev_diff_shortcut)
        next_diff_sc = QShortcut(QKeySequence("Alt+Down"), self)
        next_diff_sc.setContext(Qt.WidgetWithChildrenShortcut)
        next_diff_sc.activated.connect(self._on_next_diff_shortcut)
        # Ctrl+F — 찾기 입력란으로 포커스 이동
        find_sc = QShortcut(QKeySequence("Ctrl+F"), self)
        find_sc.setContext(Qt.WidgetWithChildrenShortcut)
        find_sc.activated.connect(self._focus_find)

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(8, 8, 8, 8)
        root.setSpacing(6)

        # ── 툴바 ──
        toolbar = QHBoxLayout()

        self.diff_only_btn = QPushButton("변경점만 보기")
        self.diff_only_btn.setFixedHeight(36)
        self.diff_only_btn.setFont(ui_font(10))
        self.diff_only_btn.setCheckable(True)
        self.diff_only_btn.setObjectName("toggle_btn")
        self.diff_only_btn.setEnabled(False)
        self.diff_only_btn.setToolTip("변경된 행만 표시 / 전체 표시 전환 (Ctrl+D)")
        self.diff_only_btn.toggled.connect(self._on_diff_only_toggled)
        toolbar.addWidget(self.diff_only_btn)

        self.refresh_btn = QPushButton("새로고침")
        self.refresh_btn.setFixedHeight(36)
        self.refresh_btn.setFont(ui_font(10))
        self.refresh_btn.setToolTip("지정된 경로의 파일을 다시 불러와 비교합니다 (F5)")
        self.refresh_btn.setEnabled(False)
        self.refresh_btn.clicked.connect(self._run_refresh)
        toolbar.addWidget(self.refresh_btn)

        # 이전/다음 변경 셀 이동 버튼
        self.prev_diff_btn = QPushButton("◀ 이전 변경")
        self.prev_diff_btn.setFixedHeight(36)
        self.prev_diff_btn.setFont(ui_font(10))
        self.prev_diff_btn.setEnabled(False)
        self.prev_diff_btn.setToolTip("이전 변경 셀로 이동 (Alt+↑)")
        self.prev_diff_btn.clicked.connect(lambda: self._goto_changed(-1))
        toolbar.addWidget(self.prev_diff_btn)

        self.next_diff_btn = QPushButton("다음 변경 ▶")
        self.next_diff_btn.setFixedHeight(36)
        self.next_diff_btn.setFont(ui_font(10))
        self.next_diff_btn.setEnabled(False)
        self.next_diff_btn.setToolTip("다음 변경 셀로 이동 (Alt+↓)")
        self.next_diff_btn.clicked.connect(lambda: self._goto_changed(+1))
        toolbar.addWidget(self.next_diff_btn)

        # 찾기 — 검색란 + 옵션 토글 + 이전/다음 찾기 버튼
        toolbar.addSpacing(16)
        find_box = QHBoxLayout()
        find_box.setSpacing(4)

        self.find_edit = QLineEdit()
        self.find_edit.setObjectName("find_edit")
        self.find_edit.setPlaceholderText("찾을 내용 (Ctrl+F)")
        self.find_edit.setFixedHeight(36)
        self.find_edit.setFixedWidth(200)
        self.find_edit.setFont(ui_font(10))
        self.find_edit.setClearButtonEnabled(True)
        self.find_edit.setEnabled(False)
        self.find_edit.setToolTip("셀 값 검색 — Enter: 다음 찾기, Shift+Enter: 이전 찾기")
        self.find_edit.returnPressed.connect(lambda: self._goto_find(+1))
        find_prev_sc = QShortcut(QKeySequence("Shift+Return"), self.find_edit)
        find_prev_sc.setContext(Qt.WidgetShortcut)
        find_prev_sc.activated.connect(lambda: self._goto_find(-1))
        find_box.addWidget(self.find_edit)

        def _find_btn(kind: str, tooltip: str, checkable: bool) -> QPushButton:
            btn = QPushButton()
            btn.setObjectName("find_btn")
            btn.setFixedSize(36, 36)
            btn.setIcon(self._make_find_icon(kind))
            btn.setIconSize(QSize(22, 22))
            btn.setCheckable(checkable)
            btn.setEnabled(False)
            btn.setToolTip(tooltip)
            find_box.addWidget(btn)
            return btn

        self.find_case_btn = _find_btn(
            "case",
            "대소문자 무시 (Ignore case)\n"
            "켜짐: 대소문자를 구분하지 않고 검색\n"
            "꺼짐: 대소문자가 정확히 일치할 때만 검색",
            checkable=True,
        )
        self.find_case_btn.setChecked(True)

        self.find_word_btn = _find_btn(
            "word",
            "전체 단어 일치 (Match whole word only)\n"
            "켜짐: 검색어가 독립된 단어로 존재할 때만 찾음\n"
            "꺼짐: 부분 문자열도 찾음",
            checkable=True,
        )

        self.find_prev_btn = _find_btn("prev", "이전 찾기 (Shift+Enter)", checkable=False)
        self.find_prev_btn.clicked.connect(lambda: self._goto_find(-1))

        self.find_next_btn = _find_btn("next", "다음 찾기 (Enter)", checkable=False)
        self.find_next_btn.clicked.connect(lambda: self._goto_find(+1))

        toolbar.addLayout(find_box)

        toolbar.addStretch()

        # 범례
        for lbl, key in [
            ("신규", "added"),
            ("변경", "modified"), ("준비", "staged"), ("병합", "merged"),
        ]:
            dot = QLabel("  ")
            dot.setFixedSize(20, 20)
            dot.setStyleSheet(
                f"background:{DIFF_COLORS[key].name()};"
                "border:1px solid #aaa; border-radius:3px;"
            )
            txt = QLabel(lbl)
            txt.setFont(ui_font(9))
            toolbar.addWidget(dot)
            toolbar.addWidget(txt)
            toolbar.addSpacing(8)

        root.addLayout(toolbar)

        # ── 좌우 패널 ──
        splitter = QSplitter(Qt.Horizontal)
        self.panel_a = FilePanel("A 파일 (원본)", "a")
        self.panel_b = FilePanel("B 파일 (비교)", "b")
        self.panels = {"a": self.panel_a, "b": self.panel_b}   # side 셀렉터 공용 맵
        splitter.addWidget(self.panel_a)
        splitter.addWidget(self.panel_b)
        splitter.setSizes([700, 700])
        root.addWidget(splitter, 1)

        # ── 하단 시트 탭 (엑셀식) — A/B 공통, 시트 이름 기준 비교 ──
        self.sheet_tabs = SheetTabBar()
        self.sheet_tabs.setObjectName("sheet_tabs")
        self.sheet_tabs.setExpanding(False)
        self.sheet_tabs.setDrawBase(True)
        self.sheet_tabs.setElideMode(Qt.ElideNone)   # 이름 축약 금지 — 폭은 텍스트에 맞춰 자동
        self.sheet_tabs.setUsesScrollButtons(True)
        self.sheet_tabs.setFont(ui_font(9))
        self.sheet_tabs.currentChanged.connect(self._on_sheet_tab_changed)
        self.sheet_tabs.hide()   # 시트가 2개 이상일 때만 표시
        root.addWidget(self.sheet_tabs)

        # ── 변경 셀 위치 미니맵: 양쪽 테이블의 세로/가로 스크롤바를 커스텀으로 교체 ──
        # 스크롤 동기화 시그널 연결 전에 교체해야 verticalScrollBar()/horizontalScrollBar()
        # 핸들이 새 객체를 가리킨다.
        for tbl in (self.panel_a.table, self.panel_b.table):
            tbl.setVerticalScrollBar(MinimapScrollBar(Qt.Vertical, tbl))
            tbl.setHorizontalScrollBar(MinimapScrollBar(Qt.Horizontal, tbl))

        # ── 틀 고정: 각 패널 데이터 뷰포트 위에 키 행/열 오버레이 (스크롤바 교체 직후 생성) ──
        self._freeze = {
            "a": FreezeController(self.panel_a.table),
            "b": FreezeController(self.panel_b.table),
        }

        # ── 상태바 — MainWindow 소유의 공유 상태바에 초기 안내를 표시 ──
        self.status.showMessage(
            "파일을 선택하면 자동으로 비교합니다.  "
            "| 셀 선택 후 우클릭 → 병합 준비 → 저장"
        )

        self._syncing_selection = False

        # A↔B 대칭 배선 — 스크롤/크기/선택 동기화 (src → dst 양방향 루프)
        # ※ 람다의 루프 변수는 기본값 인자(s=..., d=...)로 바인딩해야 한다.
        for src, dst in ((self.panel_a, self.panel_b), (self.panel_b, self.panel_a)):
            src.table.horizontalScrollBar().valueChanged.connect(
                dst.table.horizontalScrollBar().setValue)
            src.table.verticalScrollBar().valueChanged.connect(
                dst.table.verticalScrollBar().setValue)
            # 열/행 크기 동기화 — 사용자 조작에 의한 변경만 (apply_*는 시그널 미발행)
            src.table.column_resized.connect(dst.table.apply_column_width)
            src.table.row_resized.connect(dst.table.apply_row_height)
            # 선택 셀 동기화 — QTableView는 itemSelectionChanged가 없으므로
            # selectionModel().selectionChanged 사용 (selectionModel은 ctor에서 1회 생성)
            src.table.selectionModel().selectionChanged.connect(
                lambda *_, s=src.table, d=dst.table: self._sync_selection(s, d))
            # 셀값란 높이 스플리터 동기화 (A↔B 대칭)
            src.v_split.splitterMoved.connect(
                lambda _pos, _i, s=src, d=dst: self._sync_splitter(s, d))

        # side별 패널 시그널 배선
        for side, panel in self.panels.items():
            panel.table.stage_requested.connect(self._stage_selected)
            panel.table.unstage_requested.connect(self._unstage_selected)
            panel.table.key_col_changed.connect(self._on_key_col_changed)
            panel.table.key_row_changed.connect(self._on_key_row_changed)
            panel.table.columns_exclude_set.connect(self._on_columns_exclude_set)
            panel.save_btn.clicked.connect(
                lambda _=False, s=side: self._save_staged(s))
            panel.file_loaded.connect(
                lambda p, s=side: self._on_file_loaded(s, p))
            panel.folder_loaded.connect(
                lambda p, s=side: self.folder_requested.emit(s, p))
            # 첫 셀 선택 시 수식 플래그 지연 로드(파랑 폰트 표시용) 트리거.
            panel.table.selectionModel().selectionChanged.connect(
                lambda *_: self._maybe_load_formula_flags())

    def _apply_style(self):
        self.setStyleSheet(APP_QSS)

    def shutdown(self):
        """탭/앱 종료 시 실행 중 워커를 안전하게 정리한다.
        QThread가 실행 중인 채로 파괴되면 'QThread destroyed while running' 크래시가 나므로,
        시그널을 막아 종료 중 도착 결과가 슬롯을 건드리지 못하게 한 뒤 완료를 기다린다
        (저장 워커는 중간에 끊기면 데이터 위험이라 특히 끝까지 대기)."""
        workers = [self._load_worker, self._formula_flag_worker,
                   self._staged_merge_worker, self._sheet_diff_worker,
                   self._preview_workers.get("a"), self._preview_workers.get("b")]
        for w in workers:
            if w is None:
                continue
            try:
                w.blockSignals(True)
                if w.isRunning():
                    w.wait()
            except RuntimeError:
                pass   # 이미 삭제된 C++ 객체 — 무시

    # ── MainWindow 진입점 ──────────────────────────────────────────────────────

    def load_pair(self, path_a: str = "", path_b: str = ""):
        """MainWindow가 이 뷰에 파일 쌍을 로드한다(경로 세팅 + 비교/미리보기).
        _on_file_loaded 와 동일한 판단 로직을 한곳에 모은 공개 진입점."""
        if path_a:
            self.panel_a.set_path(path_a)
        if path_b:
            self.panel_b.set_path(path_b)
        pa = self.panel_a.get_path()
        pb = self.panel_b.get_path()
        self.refresh_btn.setEnabled(bool(pa) or bool(pb))
        # 시트 탭을 먼저 재구성해 self._current_sheet를 확정한 뒤 비교/미리보기 실행.
        self._rebuild_sheet_tabs()
        if pa and pb:
            self._run_compare()
        elif pa:
            self._run_preview("a", pa)
        elif pb:
            self._run_preview("b", pb)

    def tab_title(self) -> str:
        """탭 라벨용 이름 — 파일명(형식 아이콘은 tab_icon으로 별도 표시). 둘 다 없으면 '새 비교'."""
        for side in ("a", "b"):
            p = self.panels[side].get_path()
            if p:
                return os.path.basename(p)
        return "새 비교"

    def tab_icon(self):
        """탭 라벨용 파일 형식 PNG 아이콘. 로드된 파일이 없으면 null QIcon."""
        for side in ("a", "b"):
            p = self.panels[side].get_path()
            if p:
                return ext_tab_icon(os.path.splitext(p)[1].lower())
        return QIcon()

    # ── 파일 미리보기 (비교 전 단독 표시) ────────────────────────────────────────

    def _reset_compare_state(self):
        """비교 결과를 초기화하고 버튼 상태를 되돌린다."""
        self._diff_matrix = []
        self._diff_row_meta = []
        self._merged_cells = set()
        self._staged = {}
        self._preview_data = {"a": [], "b": []}
        self._clear_row_filter()   # diff 모드의 숨김 행이 미리보기에 남지 않도록 해제
        self._clear_freeze()       # 틀 고정 오버레이도 숨김(미리보기/빈 상태)
        self._excluded_cols.clear()
        self.panel_a.table.set_excluded_cols(self._excluded_cols)
        self.panel_b.table.set_excluded_cols(self._excluded_cols)
        self.panel_a._row_meta = []
        self.panel_b._row_meta = []
        self.panel_a._staged_display = {}
        self.panel_b._staged_display = {}
        self.diff_only_btn.setChecked(False)
        self.diff_only_btn.setEnabled(False)
        self.prev_diff_btn.setEnabled(False)
        self.next_diff_btn.setEnabled(False)
        self._set_find_enabled(False)
        self._update_minimap()
        self._set_save_btn_state()

    # ── 시트 탭 ────────────────────────────────────────────────────────────────

    def _rebuild_sheet_tabs(self):
        """A/B 파일의 시트 목록(합집합, A 순서 우선)으로 하단 탭을 재구성한다.
        기존 선택 시트가 새 목록에 있으면 유지, 없으면 첫 시트로. 시트가 1개 이하면 숨김.
        currentChanged 재귀 트리거를 막기 위해 시그널을 차단하고 조작한다."""
        path_a = self.panel_a.get_path()
        path_b = self.panel_b.get_path()
        a_names = list_sheet_names(path_a) if path_a else []
        b_names = list_sheet_names(path_b) if path_b else []

        names: list[str] = list(a_names)
        for n in b_names:
            if n not in names:
                names.append(n)

        # 유지할 인덱스 결정 — 기존 선택 시트가 남아 있으면 그 위치, 아니면 스마트 기본값
        if self._current_sheet in names:
            idx = names.index(self._current_sheet)
        else:
            idx = self._default_sheet_index(names, path_a, path_b)

        self._sheet_names = names
        self._current_sheet = names[idx] if names else None

        blocked = self.sheet_tabs.blockSignals(True)
        try:
            while self.sheet_tabs.count() > 0:
                self.sheet_tabs.removeTab(0)
            for n in names:
                self.sheet_tabs.addTab(n)
            if names:
                self.sheet_tabs.setCurrentIndex(idx)
        finally:
            self.sheet_tabs.blockSignals(blocked)

        self.sheet_tabs.setVisible(len(names) > 1)
        # 시트 탭 색칠(변경=노랑)은 모든 시트를 다시 로드하고 큰 값 비교를 하므로 GIL을 오래
        # 잡는다. 첫 비교/미리보기 로드와 동시에 돌리면 그 로드를 2배 가까이 늦춘다(실측).
        # → 색칠은 첫 결과가 화면에 뜬 뒤로 미룬다(_flush_sheet_diff). 색은 부가 정보라 무해.
        self.sheet_tabs.clear_changed()
        self._pending_sheet_diff = (path_a, path_b, names)

    def _flush_sheet_diff(self):
        """미뤄둔 시트 diff(탭 색칠)를 시작한다 — 첫 비교/미리보기가 표시된 뒤 호출."""
        pend = self._pending_sheet_diff
        self._pending_sheet_diff = None
        if pend:
            self._start_sheet_diff(*pend)

    def _default_sheet_index(self, names, path_a, path_b) -> int:
        """새 파일을 열 때 기본 시트 인덱스를 고른다.
        1) 파일별 마지막 선택 시트(A 우선, 없으면 B) — 사용자가 고른 적 있으면 그대로.
        2) '!'로 시작하지 않는 첫 시트 — !Define 등 메타/설정 시트를 건너뛰고 데이터 시트로.
        3) 폴백 0.
        열자마자 데이터 시트가 뜨므로 '작은 시트 로드→데이터 시트 재로드' 왕복을 없앤다."""
        if not names:
            return 0
        for p in (path_a, path_b):
            remembered = load_last_sheet(p) if p else None
            if remembered in names:
                return names.index(remembered)
        for i, n in enumerate(names):
            if not n.startswith("!"):
                return i
        return 0

    def _start_sheet_diff(self, path_a, path_b, names):
        """시트 탭이 여럿이고 A/B 두 파일이 모두 있으면, 각 시트의 변경 여부를
        백그라운드로 판정해 변경 탭을 노랑으로 칠한다(값 매트릭스 비교 — 키 무관)."""
        self.sheet_tabs.clear_changed()
        if not (path_a and path_b) or len(names) <= 1:
            return
        self._sheet_diff_token += 1
        token = self._sheet_diff_token
        ctx = (path_a, path_b)
        w = SheetDiffWorker(path_a, path_b, names)
        w.done.connect(
            lambda changed, t=token, c=ctx, nm=list(names):
            self._on_sheet_diff_done(changed, t, c, nm))
        w.finished.connect(w.deleteLater)
        self._sheet_diff_worker = w
        w.start(QThread.LowestPriority)

    def _on_sheet_diff_done(self, changed_names, token, ctx, names):
        # 낡은 결과(파일 바뀜/새 스캔) 무시.
        if token != self._sheet_diff_token:
            return
        if ctx != (self.panel_a.get_path(), self.panel_b.get_path()):
            return
        idx = [i for i, n in enumerate(names) if n in changed_names]
        self.sheet_tabs.set_changed_indices(idx)

    def _on_sheet_tab_changed(self, index: int):
        """시트 탭 클릭 — 선택 시트로 비교/미리보기를 재실행한다."""
        if index < 0 or index >= len(self._sheet_names):
            return
        self._current_sheet = self._sheet_names[index]
        path_a = self.panel_a.get_path()
        path_b = self.panel_b.get_path()
        # 사용자가 명시적으로 고른 시트를 파일별로 기억 → 다음에 그 파일을 열면 기본 선택.
        for p in (path_a, path_b):
            if p:
                save_last_sheet(p, self._current_sheet)
        if path_a and path_b:
            self._run_compare()
        elif path_a:
            self._run_preview("a", path_a)
        elif path_b:
            self._run_preview("b", path_b)

    def _run_preview(self, side: str, path: str):
        """파일 선택 즉시 해당 패널에 원본 데이터를 색상 없이 표시한다."""
        self._formula_flags_requested = False   # 새 미리보기 — 수식 플래그 지연 로드 재무장
        self._formula_flags = {"a": set(), "b": set()}
        # 비교 결과가 있으면 비교 상태를 리셋하고 반대쪽 패널도 미리보기로 전환
        if self._diff_matrix:
            self._reset_compare_state()
            other_side = "b" if side == "a" else "a"
            other_path = self.panels[other_side].get_path()
            if other_path:
                self._run_preview(other_side, other_path)

        worker = PreviewWorker(side, path, self._current_sheet)
        ctx = (path, self._current_sheet)   # 결과 도착 시 컨텍스트가 여전히 유효한지 판별용
        worker.done.connect(lambda s, d, c=ctx: self._on_preview_done(s, d, c))
        worker.progress.connect(self.status.showMessage)
        worker.progress_n.connect(self.status.set_progress)
        worker.error.connect(self._on_preview_error)
        worker.finished.connect(worker.deleteLater)
        self._preview_workers[side] = worker
        worker.start()
        self.status.showMessage(f"{'A' if side == 'a' else 'B'} 파일 로딩 중...")

    def _on_preview_error(self, msg: str):
        self.status.end_progress()
        self.status.showMessage(f"파일 로드 오류: {msg}")

    def _on_preview_done(self, side: str, data: list[list], ctx=None):
        # 로딩 중 파일/시트가 바뀌었으면(낡은 워커) 무시 — 현재 화면 오염 방지.
        if ctx is not None and ctx != (self.panels[side].get_path(), self._current_sheet):
            return
        self.status.end_progress()
        self._preview_data[side] = data
        panel = self.panels[side]
        panel._row_meta = []   # 미리보기 모드: row_meta 없음 (행 인덱스 = 원본 인덱스)
        panel.preview(data)
        if data:
            self._set_find_enabled(True)   # 미리보기 상태에서도 Ctrl+F 검색 허용
        rows = len(data)
        cols = max((len(r) for r in data), default=0)
        self.status.showMessage(
            f"{'A' if side == 'a' else 'B'} 파일 로드 완료 — {rows}행 × {cols}열  "
            "| '비교 실행'을 클릭해 두 파일을 비교하세요."
        )
        self._set_save_btn_state()
        # 미리보기가 떴으니 미뤄둔 시트 탭 색칠 시작(임계 경로 밖).
        self._flush_sheet_diff()

    # ── 수식 플래그 지연 로딩 (그리드 수식 결과 셀 파랑 표시) ──────────────────────

    def _maybe_load_formula_flags(self):
        """셀을 처음 선택하는 순간 수식 여부 플래그를 백그라운드로 로드(지연 로딩).
        훑어보기/폴더 드릴다운으로 안 보는 파일은 파싱하지 않는다(값 표시는 이미 완료됨)."""
        if self._formula_flags_requested:
            return
        if not self._diff_matrix and not (
                self._preview_data["a"] or self._preview_data["b"]):
            return
        if not (self.panel_a.table.selectionModel().hasSelection()
                or self.panel_b.table.selectionModel().hasSelection()):
            return
        self._formula_flags_requested = True
        path_a, path_b = self.panel_a.get_path(), self.panel_b.get_path()
        if not path_a and not path_b:
            return
        w = FormulaFlagWorker(path_a, path_b, self._current_sheet)
        ctx = (path_a, path_b, self._current_sheet)
        w.done.connect(lambda a, b, c=ctx: self._on_formula_flags_loaded(a, b, c))
        w.finished.connect(w.deleteLater)
        self._formula_flag_worker = w
        w.start(QThread.LowestPriority)

    def _on_formula_flags_loaded(self, a_flags: list, b_flags: list, ctx=None):
        # 로딩 중 파일/시트가 바뀌었으면(낡은 워커) 무시 — 다른 시트 좌표가 잘못 칠해지는 것 방지.
        if ctx is not None and ctx != (
                self.panel_a.get_path(), self.panel_b.get_path(), self._current_sheet):
            return
        # 원본 행 인덱스 기준이라 키 열 변경으로 재계산돼도 유효 — 뷰에 보관 후
        # 모델에 반영한다(키 변경 시 _refresh_tables가 다시 적용).
        self._formula_flags["a"] = a_flags or set()
        self._formula_flags["b"] = b_flags or set()
        self._apply_formula_flags()

    def _apply_formula_flags(self):
        """보관 중인 수식 플래그를 양쪽 모델에 반영(그리드 파랑 폰트 갱신)."""
        if self._formula_flags["a"]:
            self.panel_a.table.model().set_formula_flags(self._formula_flags["a"])
        if self._formula_flags["b"]:
            self.panel_b.table.model().set_formula_flags(self._formula_flags["b"])

    # ── 비교 ──────────────────────────────────────────────────────────────────

    def _on_refresh_shortcut(self):
        """F5 단축키 — 새로고침 버튼이 활성 상태일 때만 동작."""
        if self.refresh_btn.isEnabled():
            self._run_refresh()

    def _run_refresh(self):
        # 새로고침은 디폴트(자동너비+MAX_AUTO_COL_WIDTH_PX 상한) 상태로 복귀시킨다.
        # 세션 중 사용자가 헤더를 드래그해 늘려놓은 열/행 크기도 함께 리셋되어야
        # populate() 종료부의 _apply_user_sizes() 분기를 건너뛴다.
        for tbl in (self.panel_a.table, self.panel_b.table):
            tbl._user_col_widths.clear()
            tbl._user_row_heights.clear()

        # 디스크의 시트 구성이 바뀌었을 수 있으므로 탭 재구성(현재 시트 유지 시도).
        self._rebuild_sheet_tabs()

        path_a = self.panel_a.get_path()
        path_b = self.panel_b.get_path()
        if path_a and path_b:
            self._run_compare()
        elif path_a:
            self._run_preview("a", path_a)
        elif path_b:
            self._run_preview("b", path_b)

    def _run_compare(self):
        path_a = self.panel_a.get_path()
        path_b = self.panel_b.get_path()
        if not path_a and not path_b:
            QMessageBox.warning(self, "경고", "A 파일과 B 파일 중 하나 이상 선택하세요.")
            return

        self._set_buttons_enabled(False)
        self._formula_flags_requested = False   # 새 비교 — 수식 플래그 지연 로드 재무장
        self._formula_flags = {"a": set(), "b": set()}
        self._merged_cells = set()
        self._staged = {}
        self._diff_matrix = []
        self._diff_row_meta = []   # 미리보기 잠금 해제
        self._excluded_cols.clear()
        self.panel_a.table.set_excluded_cols(self._excluded_cols)
        self.panel_b.table.set_excluded_cols(self._excluded_cols)
        self.status.showMessage("파일 로딩 중...")

        self._load_worker = LoadWorker(path_a, path_b, self._current_sheet)
        ctx = (path_a, path_b, self._current_sheet)
        self._load_worker.done.connect(lambda a, b, c=ctx: self._on_loaded(a, b, c))
        self._load_worker.error.connect(self._on_error)
        self._load_worker.progress.connect(self.status.showMessage)
        self._load_worker.progress_n.connect(self.status.set_progress)
        self._load_worker.finished.connect(self._load_worker.deleteLater)
        self._load_worker.start()

    def _on_loaded(self, a_data, b_data, ctx=None):
        # 로딩 중 파일/시트가 바뀌었으면(낡은 워커) 무시.
        if ctx is not None and ctx != (
                self.panel_a.get_path(), self.panel_b.get_path(), self._current_sheet):
            return
        self._raw_data["a"] = a_data
        self._raw_data["b"] = b_data
        # 키 열이 현재 데이터 폭을 벗어나면(예: 넓은 시트→좁은 시트) 기본(A열)으로 리셋 —
        # 그대로 두면 전 행이 빈 키로 드롭돼 '차이 없음'처럼 보이는 착시가 생긴다.
        # (자동 리셋은 전역 저장하지 않는다 — 사용자 조작만 저장.)
        ncols = max((len(r) for r in a_data), default=0)
        ncols = max(ncols, max((len(r) for r in b_data), default=0))
        if self._key_col >= 0 and ncols and self._key_col >= ncols:
            self._key_col = 0
            self.panel_a.table.set_key_col(0)
            self.panel_b.table.set_key_col(0)
            self.status.showMessage("키 열이 현재 시트 범위를 벗어나 A열로 초기화했습니다.")
        # 키 행(헤더 행)이 현재 시트 행 수를 벗어나면 기본(1행)으로 리셋.
        nrows = max(len(a_data), len(b_data))
        if self._key_row > 0 and nrows and self._key_row >= nrows:
            self._key_row = 0
            self.panel_a.table.set_key_row(0)
            self.panel_b.table.set_key_row(0)
            self.status.showMessage("키 행이 현재 시트 범위를 벗어나 1행으로 초기화했습니다.")
        # 비교(compute_diff)와 O(R×C) 카운트는 DiffWorker(백그라운드)에서 — 대형 데이터에서
        # 로드 직후 UI 프리즈를 없앤다. 결과는 _on_diff_ready로 돌아온다.
        self._start_diff("load")

    def _start_diff(self, mode: str):
        """A/B 비교와 무거운 카운트 스캔을 DiffWorker에 위임한다.
        mode='load'=최초 비교(디폴트 필터 ON·드롭 경고), 'recompute'=키 변경 재계산."""
        self.status.showMessage("비교 계산 중...")
        self.status.begin_busy()   # DiffWorker는 진행 숫자가 없어 불확정 표시
        self._diff_token += 1
        token = self._diff_token
        w = DiffWorker(
            self._raw_data["a"], self._raw_data["b"],
            self._key_col, self._key_row, set(self._excluded_cols),
            token, mode, want_dropped=(mode == "load"),
        )
        w.done.connect(self._on_diff_ready)
        w.error.connect(self._on_error)
        w.finished.connect(w.deleteLater)
        self._diff_worker = w
        w.start()

    def _on_diff_ready(self, token, matrix, row_meta, changed, dropped, mode):
        # 낡은 결과(빠른 연속 키 변경/새 비교로 토큰이 밀림) 폐기.
        if token != self._diff_token:
            return
        self.status.end_progress()
        self._diff_matrix = matrix
        self._diff_row_meta = row_meta
        self.panel_a._row_meta = row_meta
        self.panel_b._row_meta = row_meta
        self._refresh_tables()   # populate + 수식플래그 + _apply_diff_filter
        self._set_buttons_enabled(True)
        rows = len(matrix)
        cols = len(matrix[0]) if matrix else 0
        if mode == "load":
            # 비교 완료 시 디폴트로 "변경 행만 보기" ON.
            # 이미 ON이면 toggled가 안 나오므로 _apply_diff_filter를 직접 호출.
            if self.diff_only_btn.isChecked():
                self._apply_diff_filter()
            else:
                self.diff_only_btn.setChecked(True)
            warn = f"  | ⚠ {dropped}개 행이 키 중복/공백으로 비교에서 제외됨" if dropped else ""
            self.status.showMessage(
                f"비교 완료 — {rows}행 × {cols}열 | 변경된 셀: {changed}개{warn}  "
                "| 셀 선택 후 우클릭 → 병합 준비 → 선택 병합 저장"
            )
        else:
            if self._key_col >= 0:
                anchor = f"{self._key_row + 1}행 {get_column_letter(self._key_col + 1)}열"
            else:
                anchor = "ROW 순서(키 없음)"
            self.status.showMessage(
                f"키 헤더: {anchor}  |  {rows}행 × {cols}열  |  변경된 셀: {changed}개  "
                "| 셀 선택 후 우클릭 → 병합 준비 → 선택 병합 저장"
            )
        # 첫 결과가 화면에 떴으니 미뤄둔 시트 탭 색칠을 이제 시작(임계 경로 밖).
        self._flush_sheet_diff()

    # ── 키 열 변경 ────────────────────────────────────────────────────────────

    def _on_key_col_changed(self, col: int):
        if col == self._key_col:
            return
        self._key_col = col
        self.panel_a.table.set_key_col(col)
        self.panel_b.table.set_key_col(col)
        save_key_prefs(self._key_row, self._key_col)   # 사용자 조작 → 전역 저장
        if self._raw_data["a"] or self._raw_data["b"]:
            self._recompute_diff()

    def _on_key_row_changed(self, row: int):
        if row == self._key_row:
            return
        self._key_row = row
        self.panel_a.table.set_key_row(row)
        self.panel_b.table.set_key_row(row)
        save_key_prefs(self._key_row, self._key_col)   # 사용자 조작 → 전역 저장
        if self._raw_data["a"] or self._raw_data["b"]:
            self._recompute_diff()

    def _recompute_diff(self):
        self._merged_cells = set()
        self._staged = {}
        # 키 열이 바뀌면 동일 인덱스가 다른 의미가 될 수 있으므로 제외 상태도 리셋.
        self._excluded_cols.clear()
        self.panel_a.table.set_excluded_cols(self._excluded_cols)
        self.panel_b.table.set_excluded_cols(self._excluded_cols)
        # 키 열 변경 시 사용자의 현재 토글 상태(diff_only_btn)는 그대로 유지한다.
        # 재계산도 DiffWorker(백그라운드)로 — 키 변경 시 프리즈 제거. 결과는 _on_diff_ready.
        self._set_buttons_enabled(False)
        self._start_diff("recompute")

    # ── 변경 행 필터 ──────────────────────────────────────────────────────────

    def _toggle_diff_only_shortcut(self):
        if self.diff_only_btn.isEnabled():
            self.diff_only_btn.setChecked(not self.diff_only_btn.isChecked())

    def _on_diff_only_toggled(self, checked: bool):
        self._diff_only = checked
        self._apply_diff_filter()

    def _apply_diff_filter(self):
        if not self._diff_matrix:
            self._update_minimap()
            self._refresh_freeze()
            return
        # 목표 숨김 집합(desired)을 계산하고, 각 행의 '실제' 뷰 숨김 상태와 다를 때만
        # setRowHidden을 호출한다. 캐시(과거 _hidden_rows) 대신 isRowHidden으로 실제
        # 상태를 조회하므로, 모델 리셋이 숨김 상태를 초기화하는 플랫폼(실 Windows)에서도
        # 항상 정확하다. 키 열 변경으로 매트릭스가 재계산돼도 어긋나지 않음.
        excl = self._excluded_cols
        # 틀 고정: 고정 행(0..key_row)은 본체에서 숨긴다 — 상단 고정 띠(top/corner)에 항상 표시되므로
        # 본체에 중복 렌더하지 않고, 스크롤 영역이 고정 띠 밑으로 가려지는 것도 방지한다.
        desired = set(range(min(self._key_row + 1, len(self._diff_matrix))))
        if self._diff_only:
            # 병합됨(연파랑) 셀이 있는 행은 값이 같아졌어도 계속 표시 —
            # 저장 직후 행이 사라져 병합 결과를 확인할 수 없던 문제 방지.
            # (새 비교/새로고침으로 _merged_cells가 리셋되면 일반 규칙으로 복귀)
            merged_rows = {r for (r, c) in self._merged_cells if c not in excl}
            for r, row in enumerate(self._diff_matrix):
                if r <= self._key_row:   # 고정 행 — 이미 desired에 포함(본체 숨김)
                    continue
                if r in merged_rows:
                    continue
                is_changed = any(
                    status != "same"
                    for c, (status, *_) in enumerate(row)
                    if c not in excl
                )
                if not is_changed:
                    desired.add(r)
        # setRowHidden은 sectionResized(_, _, 0)을 emit해 _user_row_heights를
        # 오염시킨다. _applying_sizes 플래그로 _on_section_v_resized 기록을 차단.
        for tbl in (self.panel_a.table, self.panel_b.table):
            tbl._applying_sizes = True
            prev_upd = tbl.updatesEnabled()
            tbl.setUpdatesEnabled(False)   # 대량 setRowHidden 중 재도색 방지(40k행 성능)
            try:
                nrows = tbl.rowCount()
                for r in range(nrows):
                    want = r in desired   # 여분(EXTRA) 행은 desired에 없으므로 항상 표시
                    if tbl.isRowHidden(r) != want:
                        tbl.setRowHidden(r, want)
            finally:
                tbl._applying_sizes = False
                tbl.setUpdatesEnabled(prev_upd)
        self._update_minimap()
        self._refresh_freeze()

    def _refresh_freeze(self):
        """양 패널 틀 고정 오버레이 갱신(앵커·크기·숨김행·스크롤·지오메트리 재적용)."""
        for fc in getattr(self, "_freeze", {}).values():
            fc.refresh()

    def _clear_freeze(self):
        for fc in getattr(self, "_freeze", {}).values():
            fc.clear()

    def _clear_row_filter(self):
        """숨김 행을 전부 해제 — 미리보기 전환 등 diff 모드를 떠날 때 호출.
        QTableView는 리셋 후에도 숨김을 기억할 수 있어 명시적으로 풀어야 한다."""
        for tbl in (self.panel_a.table, self.panel_b.table):
            tbl._applying_sizes = True
            try:
                for r in range(tbl.rowCount()):
                    if tbl.isRowHidden(r):
                        tbl.setRowHidden(r, False)
            finally:
                tbl._applying_sizes = False

    # ── 변경 셀 탐색(이전/다음) ───────────────────────────────────────────────

    def _on_prev_diff_shortcut(self):
        if self.prev_diff_btn.isEnabled():
            self._goto_changed(-1)

    def _on_next_diff_shortcut(self):
        if self.next_diff_btn.isEnabled():
            self._goto_changed(+1)

    def _iter_changed_cells(self):
        """변경된 (r, c) 셀을 행 우선 순서로 yield. 숨겨진 행/제외 열은 제외."""
        if not self._diff_matrix:
            return
        excl = self._excluded_cols
        for r, row in enumerate(self._diff_matrix):
            if self.panel_a.table.isRowHidden(r):
                continue
            for c, cell in enumerate(row):
                if c in excl:
                    continue
                if cell[0] != "same":
                    yield (r, c)

    def _current_anchor(self):
        """현재 선택 셀(우선순위: panel_a → panel_b).
        없으면 (0, -1) 반환 — A1부터 검사하기 위한 sentinel."""
        for tbl in (self.panel_a.table, self.panel_b.table):
            r, c = tbl._current_cell()
            if r >= 0 and c >= 0:
                return (r, c)
        return (0, -1)

    def _goto_changed(self, direction: int):
        """direction=+1: 다음 변경 셀, -1: 이전 변경 셀."""
        if not self._diff_matrix:
            return
        cells = list(self._iter_changed_cells())
        if not cells:
            self.status.showMessage("변경된 셀이 없습니다.")
            return
        anchor = self._current_anchor()
        if direction > 0:
            target = next((p for p in cells if p > anchor), None)
            if target is None:
                self.status.showMessage("마지막 변경 셀입니다.")
                return
        else:
            target = next((p for p in reversed(cells) if p < anchor), None)
            if target is None:
                self.status.showMessage("첫 변경 셀입니다.")
                return
        r, c = target
        # 양쪽 패널 동기 선택 + 화면 중앙으로 스크롤.
        # 프로그램적 이동(변경점/찾기)은 주변 선택·수정자 상태와 무관하게 항상 깨끗한
        # 단일 선택으로 착지해야 한다 → ClearAndSelect. setCurrentIndex는 호출 시점의
        # 키보드 수정자에 따라 선택을 확장/토글해 이동 결과가 비결정적이 될 수 있다.
        for tbl in (self.panel_a.table, self.panel_b.table):
            tbl._move_current_cell(r, c)
            if tbl.model().is_data_cell(r, c):
                tbl.scrollTo(tbl.model().index(r, c), QAbstractItemView.PositionAtCenter)
        # 이동 후 테이블에 키보드 포커스를 줘 방향키로 이어서 탐색 가능하게 한다
        self.panel_a.table.setFocus()

    # ── 찾기 ──
    def _make_find_icon(self, kind: str) -> QIcon:
        return make_find_icon(kind)   # 공용 구현(widgets)

    def _focus_find(self):
        """Ctrl+F — 찾기 입력란 포커스 + 전체 선택."""
        if self.find_edit.isEnabled():
            self.find_edit.setFocus()
            self.find_edit.selectAll()

    def _make_find_matcher(self, term: str):
        """검색 옵션(대소문자 무시/전체 단어)에 맞는 판별 함수를 반환."""
        ignore_case = self.find_case_btn.isChecked()
        whole_word = self.find_word_btn.isChecked()
        if whole_word:
            # \b는 검색어가 특수문자로 시작/끝나면 동작하지 않으므로 lookaround 사용
            flags = re.IGNORECASE if ignore_case else 0
            pat = re.compile(r"(?<!\w)" + re.escape(term) + r"(?!\w)", flags)
            return lambda text: pat.search(text) is not None
        if ignore_case:
            needle = term.casefold()
            return lambda text: needle in text.casefold()
        return lambda text: term in text

    def _iter_find_matches(self, match):
        """검색어와 일치하는 (r, c) 셀을 행 우선 순서로 yield.
        diff 모드: 숨겨진 행/제외 열 제외, A/B 표시 텍스트 중 한쪽이라도 일치하면 매치.
        미리보기 모드(파일 1개만 로드)에서도 로드된 패널의 데이터를 검색한다."""
        tbl_a = self.panel_a.table
        models = (tbl_a.model(), self.panel_b.table.model())
        rows = max(m.data_rows for m in models)
        cols = max(m.data_cols for m in models)
        if rows == 0:
            return
        diff_mode = bool(self._diff_matrix)
        excl = self._excluded_cols if diff_mode else set()
        for r in range(rows):
            if diff_mode and tbl_a.isRowHidden(r):
                continue
            for c in range(cols):
                if c in excl:
                    continue
                for model in models:
                    if match(model.display_text(r, c)):
                        yield (r, c)
                        break

    def _goto_find(self, direction: int):
        """direction=+1: 다음 찾기, -1: 이전 찾기. 끝에 도달하면 반대편에서 순환.
        diff 모드뿐 아니라 미리보기(파일 1개) 상태에서도 동작한다."""
        term = self.find_edit.text()
        has_data = (self.panel_a.table.model().data_rows
                    or self.panel_b.table.model().data_rows)
        if not term or not has_data:
            return
        cells = list(self._iter_find_matches(self._make_find_matcher(term)))
        if not cells:
            self.status.showMessage(f'"{term}" — 일치 항목이 없습니다.')
            return
        anchor = self._current_anchor()
        wrapped = ""
        if direction > 0:
            target = next((p for p in cells if p > anchor), None)
            if target is None:
                target = cells[0]
                wrapped = " — 처음부터 다시 검색합니다."
        else:
            target = next((p for p in reversed(cells) if p < anchor), None)
            if target is None:
                target = cells[-1]
                wrapped = " — 끝에서부터 다시 검색합니다."
        r, c = target
        # 양쪽 패널 동기 선택 + 화면 중앙으로 스크롤.
        # 프로그램적 이동(변경점/찾기)은 주변 선택·수정자 상태와 무관하게 항상 깨끗한
        # 단일 선택으로 착지해야 한다 → ClearAndSelect. setCurrentIndex는 호출 시점의
        # 키보드 수정자에 따라 선택을 확장/토글해 이동 결과가 비결정적이 될 수 있다.
        for tbl in (self.panel_a.table, self.panel_b.table):
            tbl._move_current_cell(r, c)
            if tbl.model().is_data_cell(r, c):
                tbl.scrollTo(tbl.model().index(r, c), QAbstractItemView.PositionAtCenter)
        idx = cells.index(target) + 1
        self.status.showMessage(f'찾기: "{term}" {idx}/{len(cells)}개 일치{wrapped}')

    def _update_minimap(self):
        """현재 표시 중인 행을 기준으로 변경 행/열의 비율 위치를 양쪽
        세로/가로 스크롤바에 전달한다.
        - 세로: 가시 행 중 변경 셀이 하나라도 있는 행의 위치.
        - 가로: 가시 행 안에서 변경 셀이 있는 열의 위치.
        """
        row_ratios = []
        col_ratios = []
        if self._diff_matrix:
            excl = self._excluded_cols
            visible_rows = [
                r for r in range(len(self._diff_matrix))
                if not self.panel_a.table.isRowHidden(r)
            ]

            def _row_has_changed(r):
                return any(
                    st != "same"
                    for c, (st, *_) in enumerate(self._diff_matrix[r])
                    if c not in excl
                )

            n = len(visible_rows)
            if n == 1:
                r = visible_rows[0]
                if _row_has_changed(r):
                    row_ratios.append(0.0)
            elif n > 1:
                denom = n - 1
                for vi, r in enumerate(visible_rows):
                    if _row_has_changed(r):
                        row_ratios.append(vi / denom)

            cols_total = len(self._diff_matrix[0]) if self._diff_matrix else 0
            if cols_total > 0 and visible_rows:
                if cols_total == 1:
                    if 0 not in excl and any(self._diff_matrix[r][0][0] != "same" for r in visible_rows):
                        col_ratios.append(0.0)
                else:
                    denom_c = cols_total - 1
                    for c in range(cols_total):
                        if c in excl:
                            continue
                        if any(self._diff_matrix[r][c][0] != "same" for r in visible_rows):
                            col_ratios.append(c / denom_c)
        for tbl in (self.panel_a.table, self.panel_b.table):
            v = tbl.verticalScrollBar()
            if isinstance(v, MinimapScrollBar):
                v.set_change_ratios(row_ratios)
            h = tbl.horizontalScrollBar()
            if isinstance(h, MinimapScrollBar):
                h.set_change_ratios(col_ratios)

    # ── 선택 셀 스테이징 (우클릭) ─────────────────────────────────────────────

    def _stage_selected(self, direction: str):
        if not self._diff_matrix:
            return

        cells = (
            self.panel_a.table.get_selected_cells()
            | self.panel_b.table.get_selected_cells()
        )
        cells = {
            (r, c) for (r, c) in cells
            if r < len(self._diff_matrix)
            and c < len(self._diff_matrix[r])
            and c not in self._excluded_cols
            and self._diff_matrix[r][c][0] != "same"
        }
        if not cells:
            QMessageBox.information(self, "알림", "선택한 셀 중 변경된 셀이 없습니다.")
            return

        # undo 스택에 stage 동작 기록 (셀 목록과 방향을 한 번에 저장)
        self._undo_stack.append(("stage", list(cells), direction))

        for cell in cells:
            self._staged[cell] = direction

        # staged 셀에 대해 양쪽 패널의 셀값란 표시값(병합될 값)을 미리 계산.
        # 값으로 병합하므로 소스 side의 계산값을 양쪽 셀값란에 동일하게 표시한다.
        for (r, c) in cells:
            dir_ = self._staged[r, c]
            try:
                _, a_val, b_val = self._diff_matrix[r][c]
            except (IndexError, TypeError):
                a_val, b_val = "", ""

            display = a_val if dir_ == "a_to_b" else b_val
            self.panel_a._staged_display[r, c] = display
            self.panel_b._staged_display[r, c] = display

        self._notify_cells(cells)
        self._silent_clear_selection()
        self.panel_a._selected_cell = None
        self.panel_b._selected_cell = None
        self.panel_a.cell_edit.clear()
        self.panel_a.cell_edit.setEnabled(False)
        self.panel_b.cell_edit.clear()
        self.panel_b.cell_edit.setEnabled(False)
        self._set_save_btn_state()
        self.status.showMessage(
            f"병합 준비 완료 — {len(self._staged)}개 셀 대기 중  | '저장'을 클릭하면 파일에 저장됩니다."
        )

    # ── 선택 셀 언스테이징 (병합 준비 취소) ───────────────────────────────────────

    def _unstage_selected(self):
        if not self._staged:
            return
        cells = (
            self.panel_a.table.get_selected_cells()
            | self.panel_b.table.get_selected_cells()
        )
        removed = {c for c in cells if c in self._staged}
        if not removed:
            QMessageBox.information(self, "알림", "선택한 셀 중 병합 준비된 셀이 없습니다.")
            return
        for cell in removed:
            del self._staged[cell]
            self.panel_a._staged_display.pop(cell, None)
            self.panel_b._staged_display.pop(cell, None)
        self._notify_cells(removed)
        # 병합 준비와 동일하게, 취소 후에도 셀 선택 상태를 완전 초기화(요청).
        self._silent_clear_selection()
        self.panel_a._selected_cell = None
        self.panel_b._selected_cell = None
        self.panel_a.cell_edit.clear()
        self.panel_a.cell_edit.setEnabled(False)
        self.panel_b.cell_edit.clear()
        self.panel_b.cell_edit.setEnabled(False)
        self._set_save_btn_state()
        self.status.showMessage(
            f"병합 준비 취소 — {len(removed)}개 셀 제거됨 | 남은 대기 셀: {len(self._staged)}개"
        )

    # ── 저장 ─────────────────────────────────────────────────────────────────

    def _save_staged(self, side: str):
        """side: 'a' 또는 'b' — 해당 파일만 저장. 병합 준비(staged)된 셀만 기록한다."""
        path = self.panels[side].get_path()

        # JSON/uasset 등 비-xlsx 는 저장 미지원 — 사용자에게 안내 후 중단
        if path and os.path.splitext(path)[1].lower() not in _EXCEL_EXTS:
            QMessageBox.information(
                self, "저장 미지원",
                f"{'A' if side == 'a' else 'B'} 파일은 비교 전용 형식입니다.\n"
                f"저장(병합)은 Excel(.xlsx/.xls/.xlsm) 파일에서만 지원됩니다.",
            )
            return

        # 직접 편집 기능이 제거되어 저장할 내용은 병합 준비 셀뿐 — diff 필수
        if not self._diff_matrix:
            return

        # 저장 대상 측이 실제로 쓸 내용이 있는지 확인
        # a 저장: b_to_a (B→A 방향) staged 셀만 / b 저장: a_to_b staged 셀만
        relevant_direction = "b_to_a" if side == "a" else "a_to_b"
        staged_for_side = {k: v for k, v in self._staged.items() if v == relevant_direction}
        if not staged_for_side:
            return

        if not path:
            QMessageBox.warning(self, "경고",
                f"{'A' if side == 'a' else 'B'} 파일이 지정되지 않았습니다.")
            return

        if _is_file_locked(path):
            QMessageBox.warning(
                self, "파일 열림",
                f"변경하고자 하는 파일이 열려 있으므로 저장할 수 없습니다:\n\n"
                f"{'A' if side == 'a' else 'B'} 파일: {os.path.basename(path)}"
                "\n\n파일을 닫은 후 다시 시도하세요."
            )
            return

        # 저장하지 않는 쪽 경로를 빈 문자열로 전달 → Worker가 해당 파일은 건드리지 않음
        path_a = path if side == "a" else ""
        path_b = path if side == "b" else ""

        self._saving_side = side
        self._set_buttons_enabled(False)
        self.status.showMessage("저장 중...")
        self.status.begin_busy()   # 저장은 진행 숫자가 없어 불확정 표시

        # 서식 병합용 소스 = 저장 대상의 반대편 파일 (a2b는 A, b2a는 B)
        other = "b" if side == "a" else "a"
        src_path = self.panels[other].get_path()

        self._staged_merge_worker = StagedMergeWorker(
            path_a, path_b, list(self._diff_matrix),
            list(self._diff_row_meta),
            staged_for_side,
            self._current_sheet,
            src_path, self._current_sheet,
        )
        self._staged_merge_worker.done.connect(self._on_staged_saved)
        self._staged_merge_worker.error.connect(self._on_error)
        self._staged_merge_worker.finished.connect(self._staged_merge_worker.deleteLater)
        self._staged_merge_worker.start()

    def _on_staged_saved(self, count: int):
        self.status.end_progress()
        # 저장으로 대상 파일이 바뀌었으니 값 캐시를 무효화(다음 로드가 새 내용을 읽도록).
        # mtime 키만으로도 자동 무효화되지만, 정확성을 위한 명시적 방어.
        clear_values_cache()
        side = getattr(self, "_saving_side", None)

        # ── 저장한 side에 해당하는 staged만 확정 반영 ────────────────────────
        relevant_direction = "b_to_a" if side == "a" else "a_to_b"
        saved_staged = {k: v for k, v in self._staged.items() if v == relevant_direction}
        staged_cells = set(saved_staged.keys())

        # staged 방향대로 diff_matrix 확정 반영
        for (r, c), direction in saved_staged.items():
            if r < len(self._diff_matrix) and c < len(self._diff_matrix[r]):
                _, a_val, b_val = self._diff_matrix[r][c]
                if direction == "a_to_b":
                    b_val = a_val
                else:
                    a_val = b_val
                self._diff_matrix[r][c] = ("same", a_val, b_val)

        # 저장한 side의 staged 제거 (나머지 side는 유지)
        for k in list(self._staged.keys()):
            if self._staged[k] == relevant_direction:
                del self._staged[k]
        self._merged_cells |= staged_cells

        # 저장으로 status가 modified→same이 되어 행 가시성·미니맵이 바뀔 수 있어 refilter 필요.
        self._notify_cells(staged_cells, refilter=True)
        self._silent_clear_selection()
        self._set_buttons_enabled(True)
        self.status.showMessage(f"저장 완료 — {count}개 셀 저장됨")
        QMessageBox.information(self, "저장 완료", f"{count}개 셀이 파일에 저장됐습니다.")

    def _on_error(self, msg: str):
        self.status.end_progress()
        self._set_buttons_enabled(True)
        self.status.showMessage(f"오류: {msg}")
        QMessageBox.critical(self, "오류", f"작업 실패:\n{msg}")

    def _undo(self):
        """Ctrl+Z — 병합 준비(stage) 동작을 단계별로 되돌린다.
        (셀 직접 편집 기능이 제거되어 undo 대상은 stage 항목뿐)"""
        if not self._undo_stack:
            return
        entry = self._undo_stack.pop()
        if entry[0] != "stage":
            return
        _, cells, _ = entry
        for cell in cells:
            self._staged.pop(cell, None)
            self.panel_a._staged_display.pop(cell, None)
            self.panel_b._staged_display.pop(cell, None)
        self._silent_clear_selection()
        self.panel_a._selected_cell = None
        self.panel_b._selected_cell = None
        self._notify_cells(cells)
        self._set_save_btn_state()

    # ── 유틸 ──────────────────────────────────────────────────────────────────

    def _refresh_tables(self):
        """전체 리셋 경로 — 매트릭스 자체가 재계산됐을 때만 사용.
        stage/unstage/편집/저장확정/undo는 _notify_cells()로 부분 갱신한다."""
        self.panel_a.populate(self._diff_matrix, self._merged_cells, self._staged,
                              self._diff_row_meta, self._excluded_cols)
        self.panel_b.populate(self._diff_matrix, self._merged_cells, self._staged,
                              self._diff_row_meta, self._excluded_cols)
        # populate가 모델의 수식 플래그를 리셋하므로, 보관본이 있으면 다시 반영.
        self._apply_formula_flags()
        self._apply_diff_filter()

    def _notify_cells(self, cells, refilter=False):
        """부분 갱신 경로 — 상태(dict/set)는 이미 변형된 뒤 호출된다.
        양쪽 모델에 최소 범위 dataChanged만 방출해 전체 repopulate를 피한다.

        refilter: 행 가시성/미니맵 재계산 여부. stage/unstage/undo는 diff_matrix status를
        바꾸지 않아(가시성·미니맵 모두 status에만 의존) 재계산이 불필요 → 기본 False로 O(R×C)
        스캔을 건너뛴다. 저장 확정(status→same)만 refilter=True로 재적용한다."""
        self.panel_a.table.model().notify_cells(cells)
        self.panel_b.table.model().notify_cells(cells)
        if refilter:
            self._apply_diff_filter()

    def _scroll_tables_top_left(self):
        """양쪽 패널 스크롤을 좌상단(0,0)으로 초기화."""
        for tbl in (self.panel_a.table, self.panel_b.table):
            tbl.verticalScrollBar().setValue(0)
            tbl.horizontalScrollBar().setValue(0)

    def _silent_clear_selection(self):
        """기존 populate가 선택을 조용히 초기화하던 동작 재현.
        _populating 플래그로 selectionChanged 핸들러(자동 편집 적용)를 차단한다.
        선택 범위뿐 아니라 현재 인덱스(포커스 셀 테두리)까지 지워 완전 초기화한다 —
        병합 준비/저장/취소 후 해당 셀이 여전히 선택된 것처럼 보이던 문제 방지."""
        for tbl in (self.panel_a.table, self.panel_b.table):
            tbl._populating = True
            try:
                tbl.clearSelection()
                sm = tbl.selectionModel()
                if sm is not None:
                    sm.clearCurrentIndex()
            finally:
                tbl._populating = False

    def _on_columns_exclude_set(self, cols: list, exclude: bool):
        """헤더 우클릭 → cols 일괄 제외/해제.
        키 열은 변경 검사에서 제외할 수 없으므로 제외 추가 경로에서 방어적으로 걸러낸다."""
        if not cols:
            return
        if exclude:
            cols = [c for c in cols if c != self._key_col]
            if not cols:
                return
            new_cols = [c for c in cols if c not in self._excluded_cols]
            for c in new_cols:
                self._excluded_cols.add(c)
            # 새로 제외된 열들의 기존 staged 항목 자동 해제.
            new_set = set(new_cols)
            staged_keys = [k for k in self._staged if k[1] in new_set]
            for key in staged_keys:
                del self._staged[key]
                self.panel_a._staged_display.pop(key, None)
                self.panel_b._staged_display.pop(key, None)
        else:
            for c in cols:
                self._excluded_cols.discard(c)
        # set_excluded_cols가 모델에 열 단위 dataChanged + 헤더 갱신을 방출한다
        self.panel_a.table.set_excluded_cols(self._excluded_cols)
        self.panel_b.table.set_excluded_cols(self._excluded_cols)
        self._silent_clear_selection()
        self._apply_diff_filter()
        # 제외/해제 시 스크롤을 좌상단으로 초기화(변경점 필터로 행 구성이 바뀌므로 위치 재설정).
        self._scroll_tables_top_left()
        self._set_save_btn_state()
        changed = self._count_changed()
        excl_letters = ", ".join(get_column_letter(c + 1) for c in sorted(self._excluded_cols))
        excl_msg = excl_letters if excl_letters else "없음"
        self.status.showMessage(
            f"검사 제외 열: {excl_msg}  |  변경된 셀: {changed}개"
        )

    def _sync_splitter(self, src, dst):
        """한쪽 패널의 셀값란 높이 조절을 반대 패널에 미러."""
        if getattr(self, "_syncing_split", False):
            return
        self._syncing_split = True
        try:
            dst.v_split.setSizes(src.v_split.sizes())
        finally:
            self._syncing_split = False

    def _sync_selection(self, src: ExcelTableView, dst: ExcelTableView):
        if self._syncing_selection or src._populating or dst._populating:
            return
        self._syncing_selection = True
        try:
            dst.mirror_selection_from(src)
            # mirror_selection 중 _populating=True라 _on_table_selection_changed가 막히므로
            # 반대쪽 패널의 cell_edit을 수동으로 갱신
            dst_panel = self.panel_a if dst is self.panel_a.table else self.panel_b
            dst_panel._sync_cell_edit()
        finally:
            self._syncing_selection = False

    def _count_changed(self) -> int:
        return count_changed(self._diff_matrix, self._excluded_cols)

    def _set_save_btn_state(self, enabled: bool = True):
        # b_to_a staged → A 파일에 쓸 내용 / a_to_b staged → B 파일에 쓸 내용
        has_a = any(v == "b_to_a" for v in self._staged.values())
        has_b = any(v == "a_to_b" for v in self._staged.values())

        # JSON/uasset 등 비-xlsx 파일은 저장 미지원 → 버튼 강제 비활성화 + 툴팁 안내
        def _xlsx_ok(path: str) -> bool:
            return (not path) or os.path.splitext(path)[1].lower() in _EXCEL_EXTS
        path_a = self.panel_a.get_path()
        path_b = self.panel_b.get_path()
        a_writable = _xlsx_ok(path_a)
        b_writable = _xlsx_ok(path_b)

        self.panel_a.save_btn.setEnabled(enabled and has_a and a_writable)
        self.panel_b.save_btn.setEnabled(enabled and has_b and b_writable)
        self.panel_a.save_btn.setToolTip(
            "파일 저장" if a_writable
            else "JSON/uasset은 비교 전용 — 저장은 Excel(.xlsx/.xls/.xlsm)만 지원"
        )
        self.panel_b.save_btn.setToolTip(
            "파일 저장" if b_writable
            else "JSON/uasset은 비교 전용 — 저장은 Excel(.xlsx/.xls/.xlsm)만 지원"
        )

    def _on_file_loaded(self, side: str, path: str):
        either = bool(self.panel_a.get_path()) or bool(self.panel_b.get_path())
        self.refresh_btn.setEnabled(either)
        # 시트 탭을 먼저 재구성해 self._current_sheet를 확정한 뒤 비교/미리보기 실행.
        self._rebuild_sheet_tabs()
        both = bool(self.panel_a.get_path()) and bool(self.panel_b.get_path())
        if both:
            self._run_compare()
        else:
            self._run_preview(side, path)

    def _set_buttons_enabled(self, enabled: bool):
        has_diff = enabled and bool(self._diff_matrix)
        self.diff_only_btn.setEnabled(has_diff)
        self.prev_diff_btn.setEnabled(has_diff)
        self.next_diff_btn.setEnabled(has_diff)
        self._set_find_enabled(has_diff)
        self._set_save_btn_state(enabled)

    def _set_find_enabled(self, enabled: bool):
        for w in (self.find_edit, self.find_case_btn, self.find_word_btn,
                  self.find_prev_btn, self.find_next_btn):
            w.setEnabled(enabled)
