"""파일 로더 — xlsx/json 로딩과 확장자 디스패처 (excel_diff_merge.py에서 분리).

값(계산값)만 로드한다. 수식 원문은 UI에서 더 이상 표시하지 않으므로(v163~ 값 전용)
수식 파싱/평가 경로는 제거됐고, 병합도 계산값을 기록한다(값 붙여넣기).
"""
import io
import os
import re
import csv
import json
import html
import zipfile
import threading
import xml.etree.ElementTree as ET
from collections import OrderedDict

import openpyxl

from .uasset_parser import load_uasset_as_matrix
from .logutil import log


_EXCEL_EXTS = {".xlsx", ".xls", ".xlsm", ".xlsb"}
_CSV_EXTS = {".csv", ".tsv"}
_SUPPORTED_EXTS = _EXCEL_EXTS | _CSV_EXTS | {".json", ".uasset"}

# 공개 별칭 — 폴더 비교(folder_compare) 등 외부 모듈이 비교 대상 파일을
# 선별할 때 쓰는 단일 진실 소스. 내부 판별은 계속 _EXCEL_EXTS/_SUPPORTED_EXTS 사용.
EXCEL_EXTS = _EXCEL_EXTS
SUPPORTED_EXTS = _SUPPORTED_EXTS

_calamine_mod = None   # False = import 실패(미설치), None = 미시도
_orjson_mod = None     # 동일 규약 — 빠른 JSON 파서(선택)


def _get_calamine():
    """python-calamine(Rust) 지연 import. 미설치/실패 시 None 반환(openpyxl 폴백)."""
    global _calamine_mod
    if _calamine_mod is None:
        try:
            import python_calamine as _c
            _calamine_mod = _c
        except Exception:
            _calamine_mod = False
    return _calamine_mod or None


def _get_orjson():
    """orjson(Rust) 지연 import. 미설치/실패 시 None 반환(표준 json 폴백)."""
    global _orjson_mod
    if _orjson_mod is None:
        try:
            import orjson as _o
            _orjson_mod = _o
        except Exception:
            _orjson_mod = False
    return _orjson_mod or None


# 값 매트릭스 캐시 — (abspath, mtime, sheet_name) 키. mtime 변경(저장 등) 시 자동 무효화.
# SheetDiffWorker·_run_compare·미리보기→비교·시트 재클릭이 같은 시트를 중복 로드하는 것을 막는다.
# 대형 매트릭스가 상주하므로 maxsize로 메모리를 억제(필요 시 조정). 결과는 읽기 전용으로만
# 소비되므로(compute_diff/populate 모두 미변형) 동일 객체 공유가 안전하다.
# LoadWorker가 A/B를 두 스레드로 동시 로드하므로 캐시 접근은 락으로 보호한다(로드 자체는 락 밖).
_VALUES_CACHE_MAX = 6
_values_cache: "OrderedDict[tuple, list]" = OrderedDict()
_values_cache_lock = threading.Lock()
# 진행 중(in-flight) 로드 — 같은 (path, sheet)를 여러 스레드가 동시에 요청하면(미리보기 +
# 비교 + SheetDiffWorker가 겹침) 한 번만 로드하고 나머지는 그 결과를 기다린다. 이게 없으면
# 동시 요청들이 모두 캐시를 놓쳐 대형 시트를 중복 로드하고, 문자열화(_cell_to_str)가 GIL을
# 잡아 직렬화되며 크게 느려진다(실측: 오픈 시 4중 로드로 ~5s).
_values_loading: "dict[tuple, threading.Event]" = {}


def _values_cache_key(path: str, sheet_name):
    try:
        mtime = os.path.getmtime(path)
    except OSError:
        mtime = None
    return (os.path.abspath(path), mtime, sheet_name)


def clear_values_cache() -> None:
    """값 캐시 비우기(테스트/저장 후 강제 무효화용). 진행 중 로드는 건드리지 않는다."""
    with _values_cache_lock:
        _values_cache.clear()

# 시트 이름 목록 캐시 — (path, mtime) 키. mtime 변경 시 자동 무효화.
_sheet_names_cache: dict = {}

# 빈 <fill/> — openpyxl이 로드 시 거부(TypeError: expected ...Fill)하는 요소
_EMPTY_FILL_RE = re.compile(r"<fill\s*/>|<fill>\s*</fill>")


def _sanitize_xlsx_bytes(path: str) -> io.BytesIO:
    """styles.xml의 빈 <fill/>을 유효한 최소 형태로 치환한 zip 사본을 메모리로 반환.
    비-Excel 툴이 만든 파일에서 openpyxl이 fill 파싱에 실패하는 경우의 폴백.
    numFmt 등 나머지는 그대로라 날짜/숫자 파싱은 정상 유지된다."""
    buf = io.BytesIO()
    with zipfile.ZipFile(path) as zin, \
         zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/styles.xml":
                text = data.decode("utf-8", errors="replace")
                text = _EMPTY_FILL_RE.sub("<fill><patternFill/></fill>", text)
                data = text.encode("utf-8")
            zout.writestr(item, data)
    buf.seek(0)
    return buf


def _safe_close(wb):
    """워크북 close를 조용히 시도(이미 닫혔거나 close 미지원이어도 무시)."""
    try:
        wb.close()
    except Exception:
        pass


def _open_workbook(path: str, data_only: bool):
    """openpyxl 워크북 로드 — 빈 fill 등으로 스타일 파싱이 실패하면 styles.xml을
    정제한 사본으로 1회 재시도한다. 정상 파일은 예외 경로를 타지 않는다."""
    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=data_only)
    except TypeError:
        return openpyxl.load_workbook(
            _sanitize_xlsx_bytes(path), read_only=True, data_only=data_only)


def _cell_to_str(v) -> str:
    """openpyxl 캐시값을 문자열로 변환 (정수형 float은 정수로)."""
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v)


def _pick_worksheet(wb, sheet_name):
    """sheet_name 규칙으로 워크시트를 선택. 선택 불가 시 None 반환.
    - None → 첫 시트(worksheets[0], 하위 호환).
    - 이름이 워크북에 존재 → 해당 시트.
    - 이름이 지정됐으나 없음 → None (호출부가 빈 시트로 처리)."""
    if sheet_name is None:
        return wb.worksheets[0]
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    return None


def _load_values_calamine(path: str, sheet_name=None) -> list[list[str]]:
    """calamine(Rust)로 캐시 계산값 시트를 읽는다 — openpyxl 대비 3~5×.
    sheet_name 규칙은 _pick_worksheet와 동일(None=첫 시트, 없으면 []).
    사용 불가/파싱 실패 시 예외를 던져 openpyxl 폴백을 유도한다.
    skip_empty_area=False 로 openpyxl과 좌표(선행 빈 행/열)를 일치시킨다.
    문자열화는 _cell_to_str로 통일해 openpyxl 경로와 동일한 표기를 보장한다."""
    cal = _get_calamine()
    if cal is None:
        raise RuntimeError("calamine unavailable")
    wb = cal.load_workbook(path)
    try:
        names = list(wb.sheet_names)
        if sheet_name is None:
            name = names[0] if names else None
        elif sheet_name in names:
            name = sheet_name
        else:
            return []
        if name is None:
            return []
        ws = wb.get_sheet_by_name(name)
        rows = ws.to_python(skip_empty_area=False)
        return [[_cell_to_str(v) for v in r] for r in rows]
    finally:
        _safe_close(wb)


def _load_values_pass(path: str, progress=None, sheet_name=None) -> list[list[str]]:
    """캐시 계산값 시트 — calamine 우선, 실패 시 openpyxl 폴백."""
    try:
        vals = _load_values_calamine(path, sheet_name)
        if progress is not None:
            progress(len(vals), len(vals))
        return vals
    except Exception:
        # calamine 실패 → openpyxl 폴백(정상 흐름). 원인 파악용으로만 남긴다.
        log.debug("calamine 로드 실패, openpyxl 폴백: %s", path, exc_info=True)
        return _load_values_pass_openpyxl(path, progress, sheet_name)


def _load_values_pass_openpyxl(path: str, progress=None, sheet_name=None) -> list[list[str]]:
    """data_only=True 패스 — 캐시된 계산값 시트(openpyxl 폴백).
    progress(done_rows, total_rows_or_None)를 500행마다 호출한다."""
    wb = _open_workbook(path, data_only=True)
    try:
        ws = _pick_worksheet(wb, sheet_name)
        if ws is None:
            return []
        # read_only 모드의 max_row는 dimension 헤더에서 읽는다 — 없으면 None(불확정)
        total = ws.max_row
        values: list[list[str]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            values.append([_cell_to_str(v) for v in row])
            if progress is not None and (i + 1) % 500 == 0:
                progress(i + 1, total)
        if progress is not None:
            progress(len(values), total)
        return values
    finally:
        _safe_close(wb)


def _json_value_to_str(v) -> str:
    """JSON 스칼라/구조 값을 셀 표시용 문자열로 변환.
    - dict/list 등 비스칼라는 compact JSON 으로 직렬화 — 셀 한 칸에 들어가도록.
    - 문자열의 제어 공백(개행/탭/CR)은 **이스케이프 형태(\\n·\\t·\\r)로 보존**한다.
      json 파싱이 '\\n' 이스케이프를 실제 개행 문자로 디코드하는데, 게임 텍스트 등에서 '\\n'은
      보존해야 할 제어 코드다. 그대로 두면 그리드가 실제 줄바꿈으로 렌더해 코드가 사라져 보인다.
      → 셀에 코드 그대로 표시되고 비교/스테이징 값도 일관되게 유지된다.
    """
    if v is None:
        return ""
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)):
        return _cell_to_str(v)
    if isinstance(v, str):
        return v.replace("\r\n", "\\n").replace("\n", "\\n").replace(
            "\r", "\\r").replace("\t", "\\t")
    return json.dumps(v, ensure_ascii=False)


def _flatten_json(node, prefix: str = "") -> list[tuple[str, str]]:
    """객체/배열을 점-경로 평탄화 — `[ (경로, 값문자열), ... ]`. 폴백 표기용."""
    rows: list[tuple[str, str]] = []
    if isinstance(node, dict):
        if not node:
            rows.append((prefix or "(empty object)", "{}"))
            return rows
        for k, v in node.items():
            key = f"{prefix}.{k}" if prefix else str(k)
            if isinstance(v, (dict, list)):
                rows.extend(_flatten_json(v, key))
            else:
                rows.append((key, _json_value_to_str(v)))
    elif isinstance(node, list):
        if not node:
            rows.append((prefix or "(empty array)", "[]"))
            return rows
        for i, v in enumerate(node):
            key = f"{prefix}[{i}]" if prefix else f"[{i}]"
            if isinstance(v, (dict, list)):
                rows.extend(_flatten_json(v, key))
            else:
                rows.append((key, _json_value_to_str(v)))
    else:
        rows.append((prefix or "(value)", _json_value_to_str(node)))
    return rows


def _json_records(node) -> list | None:
    """표(행 단위 비교)로 쓸 '객체 배열'을 찾아 반환. 없으면 None.
    - 최상위가 비어있지 않은 dict 리스트 → 그대로.
    - 최상위가 객체(래퍼)면 그 값들 중 첫 번째 '비어있지 않은 dict 리스트'를 언랩.
      (예: {"TablePackage": [ {...}, {...} ]} → 안쪽 배열을 표로.) 한 단계만 본다.
    """
    def _is_records(v) -> bool:
        return isinstance(v, list) and bool(v) and all(isinstance(x, dict) for x in v)

    if _is_records(node):
        return node
    if isinstance(node, dict):
        for v in node.values():
            if _is_records(v):
                return v
    return None


def _json_load(path: str):
    """JSON 파일을 파싱해 파이썬 객체로 반환. orjson(있으면) → 표준 json 폴백.
    orjson은 bytes를 받고 선행 BOM을 거부하므로 BOM을 벗겨 넘긴다."""
    oj = _get_orjson()
    if oj is not None:
        try:
            with open(path, "rb") as f:
                data = f.read()
            if data[:3] == b"\xef\xbb\xbf":   # UTF-8 BOM
                data = data[3:]
            return oj.loads(data)
        except Exception:
            pass   # 손상/비호환 → 표준 json으로 재시도(utf-8-sig가 BOM 처리)
    with open(path, "r", encoding="utf-8-sig") as f:
        return json.load(f)


def load_json_as_matrix(path: str, progress=None) -> list[list[str]]:
    """JSON 파일을 비교용 2D 매트릭스로 변환.
    - 객체 배열([ {...}, ... ]) 또는 그 배열을 감싼 객체({"K": [ {...}, ... ]}) 이면:
      첫 행=키 union(객체 등장 순서 보존), 이후 각 객체를 한 행으로 — 표 형태 행 단위 비교.
    - 그 외(객체/스칼라/스칼라 배열 등)는 점-경로 평탄화로 [경로, 값] 2열 폴백.
    progress(done_rows, total_rows) 콜백을 주면 행 변환 중 주기적으로 호출한다(대형 파일 진행 표시).
    """
    node = _json_load(path)

    records = _json_records(node)
    if records is not None:
        # 키 union — 첫 등장 순서 보존
        keys: list[str] = []
        seen: set[str] = set()
        for obj in records:
            for k in obj.keys():
                if k not in seen:
                    seen.add(k)
                    keys.append(k)
        total = len(records)
        matrix: list[list[str]] = [list(keys)]
        for i, obj in enumerate(records):
            row = [_json_value_to_str(obj.get(k, "")) for k in keys]
            matrix.append(row)
            # 2048행마다 진행 보고(콜백 오버헤드 최소화, 실제 스로틀은 워커가 처리).
            if progress is not None and (i & 0x7FF) == 0:
                progress(i, total)
        if progress is not None:
            progress(total, total)
        return matrix

    # 폴백: 평탄화 2열
    matrix = [["path", "value"]]
    for path_str, val in _flatten_json(node):
        matrix.append([path_str, val])
    return matrix


def load_csv_as_matrix(path: str, progress=None) -> list[list[str]]:
    """CSV/TSV → 문자열 매트릭스(비교/미리보기용).
    - 구분자: 확장자로 결정(.tsv=탭, 그 외=쉼표).
    - 인코딩: utf-8(BOM 허용) 우선, 실패 시 cp949(한국어 파일), 그래도 안 되면 대체 문자.
    - ragged(열 수 불균일) 행은 최대 열 수로 빈칸 패딩 — 그리드/비교가 직사각형을 가정."""
    ext = os.path.splitext(path)[1].lower()
    delim = "\t" if ext == ".tsv" else ","
    rows = None
    for enc in ("utf-8-sig", "cp949"):
        try:
            with open(path, "r", encoding=enc, newline="") as f:
                rows = list(csv.reader(f, delimiter=delim))
            break
        except UnicodeDecodeError:
            rows = None
    if rows is None:   # 인코딩 자동 판별 실패 → 손상 문자 대체로라도 읽는다
        with open(path, "r", encoding="utf-8", errors="replace", newline="") as f:
            rows = list(csv.reader(f, delimiter=delim))
    width = max((len(r) for r in rows), default=0)
    matrix = [r + [""] * (width - len(r)) for r in rows]
    if progress is not None:
        progress(len(matrix), len(matrix))
    return matrix


def load_values_any(path: str, progress=None, sheet_name=None) -> list[list]:
    """확장자별 '값' 매트릭스만 반환 — 비교/미리보기의 단일 경로.
    xlsx는 calamine 캐시값(폴백 openpyxl), json/uasset은 해당 매트릭스.
    (path, mtime, sheet) 캐시로 중복 로드를 피한다. 서로 다른 파일/시트는 여전히 병렬로
    로드되고(각자 다른 키), 같은 (path, sheet)를 동시에 요청한 경우엔 한 번만 로드하고
    나머지는 그 결과를 기다린다(in-flight 디둡). 로드 자체는 락 밖에서 수행한다."""
    key = _values_cache_key(path, sheet_name)
    while True:
        with _values_cache_lock:
            cached = _values_cache.get(key)
            if cached is not None:
                _values_cache.move_to_end(key)
                return cached
            ev = _values_loading.get(key)
            if ev is None:
                ev = threading.Event()
                _values_loading[key] = ev
                break                       # 이 스레드가 로드 담당
        ev.wait(30)                         # 다른 스레드가 로딩 중 → 기다렸다 캐시 재확인

    # 로드 담당 — 실패해도 대기자가 무한 대기하지 않도록 이벤트/플래그를 반드시 정리.
    try:
        ext = os.path.splitext(path)[1].lower()
        if ext == ".json":
            data = load_json_as_matrix(path, progress)
        elif ext == ".uasset":
            data = load_uasset_as_matrix(path)
        elif ext in _CSV_EXTS:
            data = load_csv_as_matrix(path, progress)
        else:
            data = _load_values_pass(path, progress, sheet_name)
    except BaseException:
        with _values_cache_lock:
            _values_loading.pop(key, None)
        ev.set()
        raise

    with _values_cache_lock:
        _values_cache[key] = data
        _values_cache.move_to_end(key)
        while len(_values_cache) > _VALUES_CACHE_MAX:
            _values_cache.popitem(last=False)
        _values_loading.pop(key, None)
    ev.set()
    return data


# 수식 셀 판별용 — 시트 XML에서 <c r="C2" ...><f ...> 형태(셀 여는 태그 직후 <f>)를 찾는다.
# 공유/배열 수식의 후속 셀도 빈 <f .../>를 가지므로 모두 수식으로 잡힌다.
_FORMULA_CELL_RE = re.compile(rb'<c\b[^>]*\br="([A-Z]+)(\d+)"[^>]*>\s*<f[\s>/]')
_OOXML_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_OOXML_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def _col_letters_to_idx(letters: bytes) -> int:
    """열 문자(bytes, 예: b'C')를 0-based 열 인덱스로. (openpyxl.utils 의존 없이)"""
    idx = 0
    for b in letters:
        idx = idx * 26 + (b - 64)   # 'A'=65 → 1
    return idx - 1


def _sheet_xml_path_in_zip(z: zipfile.ZipFile, sheet_name) -> str:
    """sheet_name에 해당하는 워크시트 XML의 zip 내부 경로. 미지정/실패 시 첫 시트."""
    try:
        wb = ET.fromstring(z.read("xl/workbook.xml"))
        sheets = wb.findall(f"{{{_OOXML_MAIN_NS}}}sheets/{{{_OOXML_MAIN_NS}}}sheet")
        rid = None
        if sheet_name:
            for sh in sheets:
                if sh.get("name") == sheet_name:
                    rid = sh.get(f"{{{_OOXML_REL_NS}}}id")
                    break
        if rid is None and sheets:
            rid = sheets[0].get(f"{{{_OOXML_REL_NS}}}id")
        if rid is not None:
            rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
            for rel in rels:
                if rel.get("Id") == rid:
                    tgt = rel.get("Target", "")
                    if tgt.startswith("/"):
                        return tgt[1:]
                    return tgt if tgt.startswith("xl/") else "xl/" + tgt
    except Exception:
        pass
    return "xl/worksheets/sheet1.xml"


def load_formula_flags_any(path: str, sheet_name=None) -> set:
    """'수식으로 계산된 셀' 좌표 집합 {(row0, col0), ...} (0-based, 값 매트릭스와 동일 좌표계).
    - xlsx 계열: 시트 XML을 직접 정규식 스캔해 <f> 요소를 가진 셀을 찾는다.
      openpyxl 객체 모델을 만들지 않아 대형 시트에서도 수십 ms(구 openpyxl 경로는 초 단위).
    - json/uasset: 수식 개념이 없어 빈 집합.
    실패 시 빈 집합 — 수식 표시(파랑 폰트) 없이 동작한다."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".json", ".uasset") or ext in _CSV_EXTS:
        return set()
    try:
        with zipfile.ZipFile(path) as z:
            sheet_path = _sheet_xml_path_in_zip(z, sheet_name)
            try:
                data = z.read(sheet_path)
            except KeyError:
                return set()
    except Exception:
        return set()
    flags = set()
    for m in _FORMULA_CELL_RE.finditer(data):
        row0 = int(m.group(2)) - 1
        col0 = _col_letters_to_idx(m.group(1))
        if row0 >= 0 and col0 >= 0:
            flags.add((row0, col0))
    return flags


def list_sheet_names(path: str) -> list[str]:
    """비교용 시트 탭 표시에 쓸 시트 이름 목록.
    - xlsx 계열/미지 확장자: read_only 워크북의 sheetnames (행 순회 없이 workbook.xml만
      파싱되므로 대용량 파일에도 저렴함).
    - json/uasset: 시트 개념이 없어 빈 목록 반환.
    실패 시 빈 목록 — 시트 탭 없이 기존 단일 시트 동작으로 폴백된다.
    """
    ext = os.path.splitext(path)[1].lower()
    if ext in (".json", ".uasset") or ext in _CSV_EXTS:
        return []
    # (path, mtime) 캐시 — 파일 로드/새로고침/탭 재구성 시 반복 open 제거.
    # mtime이 바뀌면 키가 달라져 자동 무효화된다.
    try:
        mtime = os.path.getmtime(path)
    except OSError:
        mtime = None
    key = (path, mtime)
    cached = _sheet_names_cache.get(key)
    if cached is not None:
        return list(cached)   # 호출부 변형 방지용 방어 복사
    # 빠른 경로: xl/workbook.xml만 읽어 시트명 추출(~5ms). openpyxl은 대형 xlsm에서
    # 공유문자열까지 로드해 수 초 걸리므로(로드 시 A/B 양쪽 호출→블로킹) 회피한다.
    names = _sheet_names_from_zip(path)
    if names is None:
        # 폴백: openpyxl (비표준 zip 구조 등)
        try:
            wb = _open_workbook(path, data_only=True)
            try:
                names = list(wb.sheetnames)
            finally:
                _safe_close(wb)
        except Exception:
            log.debug("시트명 목록 로드 실패(빈 목록 폴백): %s", path, exc_info=True)
            return []
    _sheet_names_cache[key] = names
    return list(names)


_WORKBOOK_SHEET_RE = re.compile(r'<sheet\b[^>]*\bname="([^"]*)"')


def _sheet_names_from_zip(path: str):
    """xl/workbook.xml에서 시트 이름을 문서 순서대로 추출(고속). 실패 시 None."""
    try:
        with zipfile.ZipFile(path) as z:
            data = z.read("xl/workbook.xml").decode("utf-8", errors="replace")
    except Exception:
        return None
    return [html.unescape(n) for n in _WORKBOOK_SHEET_RE.findall(data)]
